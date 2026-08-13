from __future__ import annotations

import csv
import cgi
import hashlib
import hmac
import io
import json
import math
import mimetypes
import os
import re
import secrets
import shutil
import sqlite3
import sys
import tempfile
import functools
import threading
import time
import traceback
import unicodedata
import uuid
from collections import Counter, defaultdict
from contextlib import closing
from dataclasses import dataclass
from datetime import date, datetime, timedelta
from decimal import Decimal, InvalidOperation
from http import HTTPStatus
from http.cookies import SimpleCookie
from http.server import BaseHTTPRequestHandler, ThreadingHTTPServer
from pathlib import Path
from typing import Any
from urllib.parse import parse_qs, urlparse
from zoneinfo import ZoneInfo

import openpyxl
from openpyxl import Workbook
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.platypus import Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle


ROOT_DIR = Path(__file__).resolve().parent
LEGACY_APPDATA_DIR = Path(os.environ.get("LOCALAPPDATA", str(ROOT_DIR / "runtime_data"))) / "PassiniDashboard"
STATIC_DIR = ROOT_DIR / "static"
DEFAULT_COMPANY = "Passini Distribuidora de Autopeças"
DEFAULT_ADMIN_USER = os.environ.get("PASSINI_ADMIN_USER", "admin")
# Semente usada APENAS na criação do banco. Pode ser definida em crm.env
# (PASSINI_ADMIN_PASSWORD). Não é exibida em nenhuma tela.
DEFAULT_ADMIN_PASSWORD = os.environ.get("PASSINI_ADMIN_PASSWORD", "Passini@2026")

# ── Cache para competências fechadas (não mudam após importação) ─────────────
import threading as _threading
_summary_cache: dict[str, tuple[float, dict]] = {}
_summary_cache_lock = _threading.Lock()
_SUMMARY_CACHE_TTL_CLOSED = 3600  # 1h para meses fechados
_SUMMARY_CACHE_TTL_OPEN   = 30    # 30s para mês corrente

def _summary_cache_key(company_id: int, competence: str, filters: dict) -> str:
    import json
    relevant = {k: v for k, v in (filters or {}).items() if k in ("unit_name", "seller_name", "city_name", "allowed_units")}
    return f"{company_id}:{competence}:{json.dumps(relevant, sort_keys=True)}"

def _cached_single_competence_summary(conn, company_id, competence, filters=None):
    """Wraps single_competence_summary com cache por competência fechada."""
    import time
    if not competence:
        return {}
    today = date.today()
    current_comp = today.strftime("%Y-%m")
    is_open = (competence >= current_comp)
    ttl = _SUMMARY_CACHE_TTL_OPEN if is_open else _SUMMARY_CACHE_TTL_CLOSED
    key = _summary_cache_key(company_id, competence, filters or {})
    with _summary_cache_lock:
        if key in _summary_cache:
            ts, cached = _summary_cache[key]
            if time.monotonic() - ts < ttl:
                return cached
    result = single_competence_summary(conn, company_id, competence, filters)
    with _summary_cache_lock:
        _summary_cache[key] = (time.monotonic(), result)
    return result

# ─────────────────────────────────────────────────────────────────────────────

SESSION_COOKIE = "passini_session"
SESSION_TTL_HOURS = 24
# Host/porta configuráveis por variável de ambiente (padrão preserva o comportamento local no Windows).
# No servidor Linux, definir em crm.env: PASSINI_CRM_HOST e PASSINI_CRM_PORT.
DEFAULT_PORT = int(os.environ.get("PASSINI_CRM_PORT", "8876"))
DEFAULT_HOST = os.environ.get("PASSINI_CRM_HOST", "0.0.0.0")
APP_TIMEZONE = ZoneInfo("America/Sao_Paulo")

PASSINI_MAPPING_WORKBOOK = Path(
    r"C:\Users\felip\OneDrive\PASSINI\CARTEIRA DE CLIENTES\UNIDADE X VENDEDOR E CIDADE X UNIDADE.xlsx"
)
SAMPLE_FILES = {
    "faturamento_detalhado": Path(r"C:\Users\felip\Downloads\030-relatorioFaturamento detalhado.csv"),
    "custo_vendedor": Path(r"C:\Users\felip\Downloads\030-relatorioCustoVenda vendedor consolidado.csv"),
    "custo_unidade": Path(r"C:\Users\felip\Downloads\030-relatorioCustoVenda unidade.csv"),
}
SAMPLE_CLIENTS_FILE = Path(r"C:\Users\felip\Downloads\030-relatorioPessoas.csv")


def resolve_data_dir() -> Path:
    # PASSINI_CRM_DATA: diretorio persistente do banco no servidor (ex.: /srv/passini/data/crm).
    # Tem prioridade. No Windows, se nao definida, mantem o comportamento anterior.
    _env_data = os.environ.get("PASSINI_CRM_DATA")
    candidates = [
        Path(_env_data) if _env_data else None,
        Path(os.environ.get("LOCALAPPDATA", "")) / "PassiniDashboardV2" if os.environ.get("LOCALAPPDATA") else None,
        Path(tempfile.gettempdir()) / "PassiniDashboardV2",
        ROOT_DIR / "runtime_data",
    ]
    for candidate in candidates:
        if candidate and (candidate / "passini_dashboard.db").exists():
            return candidate
    for candidate in candidates:
        if not candidate:
            continue
        try:
            candidate.mkdir(parents=True, exist_ok=True)
            probe = candidate / ".write_probe"
            probe.write_text("ok", encoding="utf-8")
            probe.unlink(missing_ok=True)
            return candidate
        except OSError:
            continue
    return ROOT_DIR


DATA_DIR = resolve_data_dir()
DB_PATH = DATA_DIR / "passini_dashboard.db"

AUTO_IMPORT_BASE = ROOT_DIR / "auto-import"
AUTO_IMPORT_INTERVAL = 3600  # 1 hora (era 5 min — reduzia risco de reimport em loop)
# A base de clientes do Alfa vem em vários arquivos complementares. Só importamos
# quando nenhum arquivo foi modificado nos últimos N segundos, evitando processar
# a base pela metade enquanto a segunda exportação ainda está sendo copiada.
CRM_CLIENTS_SETTLE_SECONDS = 120
# Logs de depuração do dashboard (projeção e comparativos). Desligados por padrão —
# eram impressos a cada cálculo, poluindo o journal. Ative com PASSINI_DEBUG_DASHBOARD=1
DASHBOARD_DEBUG_LOG = os.environ.get("PASSINI_DEBUG_DASHBOARD", "").strip() in {"1", "true", "yes"}
AUTO_IMPORT_FOLDERS = [
    {"folder": "faturamento",  "scope": "sales", "label": "Faturamento Detalhado"},
    {"folder": "custo-venda",  "scope": "cost",  "label": "Custo de Venda"},
    # CRM separado em duas pastas — cada base importa de forma independente
    {"folder": "crm/clientes", "scope": "crm_clients",
     "label": "CRM · Cadastro de Clientes",
     "hint": "Base do Alfa, sem competência. A base vem dividida em DOIS arquivos e a "
             "importação SUBSTITUI a base inteira: mandar um arquivo só apaga o outro. "
             "Coloque os dois juntos na pasta antes de importar. Atualizar 1x ao dia."},
    {"folder": "crm/faturamento-consolidado", "scope": "crm_summary",
     "label": "CRM · Faturamento Consolidado",
     "hint": "Competência obrigatória no nome do arquivo, ex: 2026-07_faturamento_cliente.csv"},
    {"folder": "devolucao-garantia", "scope": "warranty",
     "label": "Devolução em Garantia",
     "hint": "Relatório de devoluções em garantia. A competência sai da data de cada devolução — "
             "não precisa do mês no nome do arquivo."},
]

# Ordem canônica das unidades (inclui unidades sem dados ainda, ex: Zona Norte)
CANONICAL_UNITS: list[str] = [
    "MATRIZ",
    "LAJEADO",
    "PELOTAS",
    "ZONA SUL",
    "ZONA NORTE",
    "XANGRILA",
]

UNIT_NORMALIZATION = {
    # Matriz
    "01 MATRIZ": "MATRIZ",
    "MATRIZ": "MATRIZ",
    # Lajeado
    "02 LAJEADO": "LAJEADO",
    "LAJEADO": "LAJEADO",
    # Pelotas
    "03 PELOTAS": "PELOTAS",
    "PELOTAS": "PELOTAS",
    # Porto Alegre → renomeada para Zona Sul (Zona Sul assumiu o lugar de POA)
    "04 POA": "ZONA SUL",
    "04 PORTO ALEGRE": "ZONA SUL",
    "04 PORTOALEGRE": "ZONA SUL",
    "PORTO ALEGRE": "ZONA SUL",
    "POA": "ZONA SUL",
    "PORTOALEGRE": "ZONA SUL",   # compatibilidade com dados históricos no banco
    # Zona Sul (novo nome)
    "04 ZONA SUL": "ZONA SUL",
    "ZONA SUL": "ZONA SUL",
    "ZONASUL": "ZONA SUL",
    "POA ZONA SUL": "ZONA SUL",
    # Zona Norte (nova unidade, ainda sem operação)
    "04 ZONA NORTE": "ZONA NORTE",
    "ZONA NORTE": "ZONA NORTE",
    "ZONANORTE": "ZONA NORTE",
    "POA ZONA NORTE": "ZONA NORTE",
    # Xangrila
    "05 XANGRILA": "XANGRILA",
    "05 XANGRI-LA": "XANGRILA",
    "XANGRI LA": "XANGRILA",
    "XANGRI-LA": "XANGRILA",
    "XANGRILÁ": "XANGRILA",
    "XANGRILA": "XANGRILA",
}

# ─────────────────────────────────────────────────────────────────────────────
# Módulos (telas) disponíveis para montar perfis de acesso.
# A ordem aqui define a ordem de exibição na tela de perfis.
# ─────────────────────────────────────────────────────────────────────────────
ACCESS_MODULES: list[dict[str, str]] = [
    # CRM
    {"id": "crm-agenda",     "label": "Missão do Dia",      "group": "CRM"},
    {"id": "crm-clientes",   "label": "Carteira",           "group": "CRM"},
    {"id": "crm-tarefas",    "label": "Tarefas",            "group": "CRM"},
    {"id": "crm-interacao",  "label": "Registrar interação","group": "CRM"},
    {"id": "meu-placar",     "label": "Meu Placar",         "group": "CRM"},
    {"id": "placar-equipe",  "label": "Placar da Equipe",   "group": "CRM"},
    {"id": "biblioteca",     "label": "Biblioteca de Vendas","group": "CRM"},
    {"id": "sem-vendedor",   "label": "Clientes sem Vendedor","group": "CRM"},
    {"id": "visitas",        "label": "Visitas",            "group": "CRM"},
    {"id": "prospeccao",     "label": "Prospecção",         "group": "CRM"},
    {"id": "contatos",       "label": "Contatos",           "group": "CRM"},
    # Desenvolvimento
    {"id": "reunioes",       "label": "Reuniões e Treinamentos","group": "Desenvolvimento"},
    {"id": "feedback",       "label": "Feedback e PDI",      "group": "Desenvolvimento"},
    # Resultados
    {"id": "executivo",      "label": "Executivo",          "group": "Resultados"},
    {"id": "vendedores",     "label": "Vendedores",         "group": "Resultados"},
    {"id": "unidades",       "label": "Unidades",           "group": "Resultados"},
    {"id": "clientes",       "label": "Clientes",           "group": "Resultados"},
    {"id": "cidades",        "label": "Cidades",            "group": "Resultados"},
    {"id": "descontos",      "label": "Descontos",          "group": "Resultados"},
    {"id": "calendario",     "label": "Calendário",         "group": "Resultados"},
    # Operações
    {"id": "importacoes",    "label": "Importações",        "group": "Operações"},
    {"id": "administracao",  "label": "Administração",      "group": "Operações"},
    {"id": "configuracoes",  "label": "Configurações",      "group": "Operações"},
    {"id": "acessos",        "label": "Usuários e Perfis",  "group": "Operações"},
]
ACCESS_MODULE_IDS = {m["id"] for m in ACCESS_MODULES}

# Escopo de dados que o perfil enxerga
DATA_SCOPES: list[dict[str, str]] = [
    {"id": "todos", "label": "Todas as unidades",
     "hint": "Sem restrição — enxerga a empresa inteira."},
    {"id": "unidade_consolidado", "label": "Unidades vinculadas + consolidado",
     "hint": "Detalhe das unidades do usuário e também os totais da empresa para comparação."},
    {"id": "unidade", "label": "Somente unidades vinculadas",
     "hint": "Restrito às unidades do usuário. Não vê outras unidades nem o consolidado."},
    {"id": "proprio", "label": "Somente a própria carteira",
     "hint": "Vendedor: enxerga apenas os próprios clientes e resultados."},
]
DATA_SCOPE_IDS = {s["id"] for s in DATA_SCOPES}

# Perfis criados automaticamente na primeira execução. is_system impede exclusão,
# mas o conteúdo (telas/escopo) continua editável pela tela.
DEFAULT_ACCESS_PROFILES: list[dict[str, Any]] = [
    {
        "name": "Diretor",
        "description": "Acesso total, incluindo gestão de usuários e perfis.",
        "modules": [m["id"] for m in ACCESS_MODULES],
        "data_scope": "todos",
        "can_manage_users": 1,
    },
    {
        "name": "Administrador",
        "description": "Acesso total, incluindo gestão de usuários e perfis.",
        "modules": [m["id"] for m in ACCESS_MODULES],
        "data_scope": "todos",
        "can_manage_users": 1,
    },
    {
        "name": "Gerente",
        "description": "Gestão da unidade: resultados, carteira e equipe. Sem acesso a configurações.",
        "modules": [
            "crm-agenda", "crm-clientes", "crm-tarefas", "crm-interacao", "placar-equipe", "biblioteca", "sem-vendedor",
            "visitas", "prospeccao", "contatos", "reunioes", "feedback",
            "executivo", "vendedores", "unidades", "clientes", "cidades", "descontos", "calendario",
        ],
        "data_scope": "unidade_consolidado",
        "can_manage_users": 0,
    },
    {
        "name": "Analista",
        "description": "Consulta de resultados e apoio a importações.",
        "modules": [
            "crm-clientes", "executivo", "vendedores", "unidades", "clientes",
            "cidades", "descontos", "calendario", "importacoes",
        ],
        "data_scope": "todos",
        "can_manage_users": 0,
    },
    {
        "name": "Vendedor",
        "description": "Rotina diária de vendas: missão do dia, carteira própria, resultado e placar.",
        # "executivo" mostra o resultado do próprio vendedor — o escopo "proprio"
        # já restringe os dados, então ele vê apenas os números dele.
        "modules": [
            "crm-agenda", "crm-clientes", "crm-tarefas", "crm-interacao",
            "meu-placar", "biblioteca", "visitas", "prospeccao", "contatos", "reunioes", "feedback",
            "executivo", "calendario",
        ],
        "data_scope": "proprio",
        "can_manage_users": 0,
    },
]

# ─────────────────────────────────────────────────────────────────────────────
# Faróis dos indicadores
#
# Decisões de projeto (pesquisadas com o Felipe em 04/08/2026):
#
# 1. RITMO, não valor absoluto. Avaliar "% atingimento" contra 100% no dia 3 de
#    21 pinta tudo de vermelho e treina a equipe a ignorar o alerta. A base é
#    quanto deveria estar HOJE: (dias decorridos / dias totais).
#
# 2. COR NUNCA SOZINHA. Cerca de 8% dos homens não distinguem vermelho de verde.
#    Toda cor vem acompanhada de ícone e rótulo — exigência de acessibilidade
#    (WCAG 1.4.1). Quem não vê a cor lê o símbolo.
#
# 3. POLARIDADE por indicador. Faturamento: maior é melhor. Devolução e
#    desconto: menor é melhor. O motor precisa saber a direção.
# ─────────────────────────────────────────────────────────────────────────────

FAROL_LEVELS = {
    "good":    {"id": "good",    "label": "No ritmo",  "icon": "▲", "color": "#1e8e3e", "bg": "#e6f4ea"},
    "warn":    {"id": "warn",    "label": "Atenção",   "icon": "◆", "color": "#b06000", "bg": "#fef7e0"},
    "bad":     {"id": "bad",     "label": "Crítico",   "icon": "▼", "color": "#c5221f", "bg": "#fce8e6"},
    "neutral": {"id": "neutral", "label": "Sem base",  "icon": "–", "color": "#5f6368", "bg": "#f1f3f4"},
}

# direction: "higher" = quanto maior melhor | "lower" = quanto menor melhor
# basis: "pace" = compara com o ritmo esperado no mês | "absolute" = valor direto
KPI_METRICS: list[dict[str, Any]] = [
    {"id": "goal_attainment", "label": "% Atingimento da meta", "direction": "higher",
     "basis": "pace", "good_at": 95.0, "warn_at": 80.0, "unit": "%",
     "hint": "Percentual do ritmo esperado até hoje. 100% = exatamente em dia."},
    {"id": "projected_attainment", "label": "% Projeção da meta", "direction": "higher",
     "basis": "absolute", "good_at": 100.0, "warn_at": 90.0, "unit": "%",
     "hint": "Projeção de fechamento do mês no ritmo atual."},
    {"id": "return_ratio", "label": "% Devolução comercial", "direction": "lower",
     "basis": "absolute", "good_at": 3.0, "warn_at": 4.5, "unit": "%",
     "hint": "Devolução comercial sobre o líquido. Garantia não entra."},
    {"id": "discount_pct", "label": "% Desconto médio", "direction": "lower",
     "basis": "absolute", "good_at": 20.0, "warn_at": 28.0, "unit": "%",
     "hint": "Desconto médio concedido sobre o bruto."},
    {"id": "margin_value", "label": "Margem", "direction": "higher",
     "basis": "absolute", "good_at": 1.55, "warn_at": 1.50, "unit": "x",
     "hint": "Multiplicador de margem da unidade."},
    {"id": "positivacao", "label": "% Positivação", "direction": "higher",
     "basis": "absolute", "good_at": 85.0, "warn_at": 50.0, "unit": "%",
     "hint": "Clientes da carteira que compraram no mês."},
    {"id": "contacts_day", "label": "Contatos no dia", "direction": "higher",
     "basis": "absolute", "good_at": 5.0, "warn_at": 3.0, "unit": "",
     "hint": "Contatos registrados pelo vendedor hoje."},
]
KPI_METRIC_BY_ID = {m["id"]: m for m in KPI_METRICS}


def seed_kpi_thresholds(conn: sqlite3.Connection, company_id: int) -> None:
    """Cria os limites padrão uma única vez; ajustes feitos na tela são preservados."""
    for metric in KPI_METRICS:
        conn.execute(
            """
            INSERT OR IGNORE INTO kpi_thresholds
                (company_id, metric_id, good_at, warn_at, is_active, created_at)
            VALUES (?, ?, ?, ?, 1, ?)
            """,
            (company_id, metric["id"], metric["good_at"], metric["warn_at"], now_iso()),
        )
    conn.commit()


def load_kpi_thresholds(conn: sqlite3.Connection, company_id: int) -> dict[str, dict[str, Any]]:
    """Limites efetivos: o que está no banco sobrescreve o padrão do código."""
    result: dict[str, dict[str, Any]] = {}
    for metric in KPI_METRICS:
        result[metric["id"]] = {**metric}
    for row in conn.execute(
        "SELECT metric_id, good_at, warn_at, is_active FROM kpi_thresholds WHERE company_id = ?",
        (company_id,),
    ).fetchall():
        mid = row["metric_id"]
        if mid in result:
            result[mid]["good_at"] = float(row["good_at"])
            result[mid]["warn_at"] = float(row["warn_at"])
            result[mid]["is_active"] = bool(row["is_active"])
    for metric in result.values():
        metric.setdefault("is_active", True)
    return result


def evaluate_farol(
    metric_id: str, value: float | None, thresholds: dict[str, dict[str, Any]],
    pace_pct: float | None = None,
) -> dict[str, Any]:
    """Classifica um número em No ritmo / Atenção / Crítico.

    Para métricas com basis="pace", `value` é o % de atingimento acumulado e
    `pace_pct` é o % do mês já decorrido. O que se avalia é a razão entre os dois:
    12,20% de meta no dia 3 de 21 (14,3% decorrido) = 85% do ritmo.
    """
    metric = thresholds.get(metric_id) or KPI_METRIC_BY_ID.get(metric_id)
    if not metric or value is None or not metric.get("is_active", True):
        return {"level": "neutral", **FAROL_LEVELS["neutral"], "metricId": metric_id}

    try:
        numeric = float(value)
    except (TypeError, ValueError):
        return {"level": "neutral", **FAROL_LEVELS["neutral"], "metricId": metric_id}
    if not math.isfinite(numeric):
        return {"level": "neutral", **FAROL_LEVELS["neutral"], "metricId": metric_id}

    compared = numeric
    detail = None
    if metric.get("basis") == "pace":
        if not pace_pct or pace_pct <= 0:
            return {"level": "neutral", **FAROL_LEVELS["neutral"], "metricId": metric_id,
                    "hint": "Mês ainda não começou."}
        compared = numeric / pace_pct * 100
        detail = f"{compared:.0f}% do ritmo esperado ({numeric:.1f}% de {pace_pct:.1f}%)"

    good_at = float(metric["good_at"])
    warn_at = float(metric["warn_at"])
    if metric["direction"] == "higher":
        level = "good" if compared >= good_at else ("warn" if compared >= warn_at else "bad")
    else:
        level = "good" if compared <= good_at else ("warn" if compared <= warn_at else "bad")

    return {
        "level": level,
        **FAROL_LEVELS[level],
        "metricId": metric_id,
        "value": round(numeric, 2),
        "compared": round(compared, 2),
        "detail": detail,
        "hint": metric.get("hint", ""),
    }


CSV_FILE_TYPES = {
    "030-relatorioFaturamento detalhado.csv": "faturamento_detalhado",
    "030-relatorioCustoVenda vendedor consolidado.csv": "custo_vendedor",
    "030-relatorioCustoVenda unidade.csv": "custo_unidade",
    "030-relatorioFaturamento conslidado cliente.csv": "faturamento_cliente_consolidado",
    "030-relatorioPessoas.csv": "cadastro_clientes",
    "030-relatorioDevolucao.csv": "devolucao_garantia",
    "01fat.csv": "faturamento_detalhado",
    "02unidade.csv": "custo_unidade",
    "03vendedor.csv": "custo_vendedor",
}
IMPORT_SCOPE_REQUIREMENTS = {
    "full": {"faturamento_detalhado", "custo_vendedor", "custo_unidade"},
    "sales": {"faturamento_detalhado"},
    "cost": {"custo_vendedor", "custo_unidade"},
    "crm": {"cadastro_clientes", "faturamento_cliente_consolidado"},
    # Escopos individuais do CRM — cada base é importada de forma independente
    "crm_clients": {"cadastro_clientes"},
    "crm_summary": {"faturamento_cliente_consolidado"},
    "warranty": {"devolucao_garantia"},
}
IMPORT_SCOPE_LABELS = {
    "full": "pacote completo",
    "sales": "faturamento detalhado",
    "cost": "custo venda",
    "crm": "crm carteira",
    "crm_clients": "cadastro de clientes",
    "crm_summary": "faturamento consolidado por cliente",
    "warranty": "devolução em garantia",
}
IMPORT_SCOPE_TABLES = {
    "faturamento_detalhado": ("fact_sales_detail",),
    "custo_vendedor": ("fact_vendor_summary",),
    "custo_unidade": ("fact_unit_summary",),
    "cadastro_clientes": ("crm_client_profiles",),
    "faturamento_cliente_consolidado": ("crm_client_summary",),
    "devolucao_garantia": ("fact_warranty_returns",),
}
UPLOAD_FIELD_TYPE_OVERRIDES = {
    "cost_unit_file": "custo_unidade",
    "cost_vendor_file": "custo_vendedor",
    "import-cost-unit-file": "custo_unidade",
    "import-cost-vendor-file": "custo_vendedor",
    "sales_file": "faturamento_detalhado",
    "import-sales-file": "faturamento_detalhado",
    "crm_clients_file": "cadastro_clientes",
    "crm_summary_file": "faturamento_cliente_consolidado",
    "import-crm-clients-file": "cadastro_clientes",
    "import-crm-summary-file": "faturamento_cliente_consolidado",
    "warranty_file": "devolucao_garantia",
    "import-warranty-file": "devolucao_garantia",
    "files": None,
}

NATIONAL_AND_RS_HOLIDAYS = {
    2025: [
        ("2025-01-01", "Confraternização Universal"),
        ("2025-03-04", "Carnaval"),
        ("2025-03-05", "Quarta-feira de Cinzas"),
        ("2025-04-18", "Sexta-feira Santa"),
        ("2025-04-21", "Tiradentes"),
        ("2025-05-01", "Dia do Trabalho"),
        ("2025-06-19", "Corpus Christi"),
        ("2025-09-20", "Revolução Farroupilha"),
        ("2025-10-12", "Nossa Senhora Aparecida"),
        ("2025-11-02", "Finados"),
        ("2025-11-15", "Proclamação da República"),
        ("2025-12-25", "Natal"),
    ],
    2026: [
        ("2026-01-01", "Confraternização Universal"),
        ("2026-02-17", "Carnaval"),
        ("2026-02-18", "Quarta-feira de Cinzas"),
        ("2026-04-03", "Sexta-feira Santa"),
        ("2026-04-21", "Tiradentes"),
        ("2026-05-01", "Dia do Trabalho"),
        ("2026-06-04", "Corpus Christi"),
        ("2026-09-20", "Revolução Farroupilha"),
        ("2026-10-12", "Nossa Senhora Aparecida"),
        ("2026-11-02", "Finados"),
        ("2026-11-15", "Proclamação da República"),
        ("2026-12-25", "Natal"),
    ],
}

CRM_CONTACT_TYPES = [
    ("LIGACAO", "Ligacao"),
    ("WHATSAPP", "WhatsApp"),
    ("VISITA", "Visita"),
    ("ORCAMENTO", "Orcamento/Cotacao"),
    ("OUTRO", "Outro"),
    # Receptivos: o cliente procurou a Passini. Registram histórico e NÃO
    # entram na meta de ligações ativas.
    ("LIGACAO_RECEBIDA", "Ligacao recebida"),
    ("MENSAGEM_RECEBIDA", "Mensagem recebida"),
    ("ANOTACAO", "Anotacao sobre o cliente"),
]

# Tipos que só existem no modo receptivo. A tela usa esta lista para montar o
# formulário curto da ficha do cliente.
CRM_RECEPTIVE_TYPES = ["LIGACAO_RECEBIDA", "MENSAGEM_RECEBIDA", "ANOTACAO"]
INITIATIVE_ACTIVE = "ATIVO"
INITIATIVE_RECEPTIVE = "RECEPTIVO"
# APOIO: vendedor atendendo cliente de OUTRA carteira (férias, almoço, ausência).
# Fica fora da meta de ligações de propósito. Se contasse, bastaria atender a
# carteira alheia para bater meta sem prospectar ninguém — e a meta do MEC mede
# iniciativa, não volume de atendimento.
INITIATIVE_SUPPORT = "APOIO"

CRM_CONTACT_RESULTS = [
    ("FALOU_CLIENTE", "Falou com o cliente", 0, 0),
    ("NAO_ATENDEU", "Nao atendeu", 1, 1),
    ("SEM_SUCESSO", "Sem sucesso / nao consegui falar", 0, 0),
    ("PEDIU_RETORNO", "Pediu retorno", 1, 1),
    ("GEROU_ORCAMENTO", "Gerou orcamento", 1, 1),
    ("GEROU_PEDIDO", "Gerou pedido", 0, 0),
    ("SEM_INTERESSE", "Sem interesse", 0, 0),
    ("CLIENTE_FECHADO", "Cliente inativo / fechado", 0, 0),
    ("OUTRO", "Outro", 0, 0),
]

CRM_PRIORITY_ORDER = [
    "REATIVACAO_INATIVO",
    "PRE_INATIVO",
    "SEM_COMPRA_MES",
    "QUEDA_FATURAMENTO",
    "CLIENTE_CLASSE_ALTA",
    "OPORTUNIDADE_MIX",
    "PROSPECCAO_NOVA",
]

CRM_PRIORITY_LABELS = {
    "REATIVACAO_INATIVO": "Reativacao de inativo",
    "PRE_INATIVO": "Pre-inativo",
    "SEM_COMPRA_MES": "Sem compra no mes",
    "QUEDA_FATURAMENTO": "Queda de faturamento",
    "CLIENTE_CLASSE_ALTA": "Cliente classe alta",
    "OPORTUNIDADE_MIX": "Oportunidade de mix",
    "PROSPECCAO_NOVA": "Prospeccao nova",
}


def ensure_dirs() -> None:
    DATA_DIR.mkdir(parents=True, exist_ok=True)
    STATIC_DIR.mkdir(parents=True, exist_ok=True)


def migrate_legacy_db_if_needed() -> None:
    legacy_db = LEGACY_APPDATA_DIR / "passini_dashboard.db"
    if DB_PATH.exists() or not legacy_db.exists():
        return
    try:
        source = sqlite3.connect(f"file:{legacy_db}?mode=ro", uri=True)
        target = sqlite3.connect(DB_PATH)
        with source, target:
            source.backup(target)
        source.close()
        target.close()
    except sqlite3.Error:
        try:
            shutil.copy2(legacy_db, DB_PATH)
        except OSError:
            pass
    except OSError:
        pass


def now_iso() -> str:
    return datetime.now().isoformat(timespec="seconds")


def today_in_brazil() -> date:
    return datetime.now(APP_TIMEZONE).date()


def dashboard_cutoff_date(today_value: date | None = None) -> date:
    reference_today = today_value or today_in_brazil()
    return reference_today - timedelta(days=1)


def normalize_whitespace(value: str | None) -> str:
    return re.sub(r"\s+", " ", (value or "").strip())


def normalize_upper(value: str | None) -> str:
    return normalize_whitespace(value).upper()


def strip_accents(value: str | None) -> str:
    text = normalize_whitespace(value)
    if not text:
        return ""
    return "".join(char for char in unicodedata.normalize("NFKD", text) if not unicodedata.combining(char))


# Chave de prospect: "P-12". O hífen É significativo e precisa sobreviver à
# normalização — ver o comentário em normalize_client_key.
PROSPECT_KEY_RE = re.compile(r"^P[\s-]*(\d+)$")


def normalize_client_key(value: str | None) -> str:
    """Chave do cliente, tolerante à bagunça do cadastro.

    A limpeza troca qualquer pontuação por espaço — e isso quebrava os
    prospects: "P-1" virava "P 1" na gravação da interação, enquanto a tela
    continuava procurando "P-1". O contato existia, órfão, e a oficina ficava
    "nunca contatada" para sempre, no vendedor e no gerente.

    O caso do prospect é tratado ANTES da limpeza e devolve a forma canônica,
    o que também conserta as linhas gravadas erradas quando forem lidas.
    """
    text = strip_accents(value).upper().strip()
    prospect = PROSPECT_KEY_RE.match(text)
    if prospect:
        return f"P-{int(prospect.group(1))}"
    text = re.sub(r"[^A-Z0-9\s]", " ", text)
    text = re.sub(r"\s+", " ", text).strip()
    return text


def normalize_unit(value: str | None) -> str:
    base = normalize_upper(value)
    return UNIT_NORMALIZATION.get(base, base)


def parse_decimal(value: str | None) -> float:
    if value is None:
        return 0.0
    text = normalize_whitespace(value)
    if not text:
        return 0.0
    if "," in text and "." in text:
        text = text.replace(".", "").replace(",", ".")
    elif "," in text:
        text = text.replace(",", ".")
    try:
        number = float(Decimal(text))
    except (InvalidOperation, ValueError):
        return 0.0
    # Blinda contra "inf"/"nan" vindos do CSV — gravar isso no banco gera JSON inválido
    return number if math.isfinite(number) else 0.0


def parse_int(value: str | None) -> int:
    if value is None:
        return 0
    text = normalize_whitespace(value)
    if not text:
        return 0
    text = text.replace(".", "").replace(",", ".")
    try:
        return int(float(text))
    except ValueError:
        return 0


def parse_datetime_pt(value: str | None) -> datetime | None:
    text = normalize_whitespace(value)
    if not text:
        return None
    for fmt in ("%d/%m/%Y %H:%M:%S", "%d/%m/%Y"):
        try:
            return datetime.strptime(text, fmt)
        except ValueError:
            continue
    return None


def competence_from_date(dt_value: datetime | None) -> str | None:
    if not dt_value:
        return None
    return dt_value.strftime("%Y-%m")


# Nomes aceitos para a coluna de data no faturamento detalhado, em ordem de prioridade.
# A comparação ignora acentos, maiúsculas, espaços e pontuação.
SALES_DATE_COLUMN_CANDIDATES = (
    "DATA EMISSAO", "DT EMISSAO", "EMISSAO", "DATA DA EMISSAO",
    "DATA MOVIMENTO", "DATA MOVIMENTACAO", "DT MOVIMENTO", "MOVIMENTO",
    "DATA", "DATA VENDA", "DT VENDA", "ULT COMPRA", "ULTIMA COMPRA",
)


def _column_lookup_key(value: str | None) -> str:
    """Normaliza um nome de coluna para comparação tolerante."""
    return re.sub(r"[^A-Z0-9]", "", strip_accents(value).upper())


def find_sales_date_value(row: dict[str, str]) -> str | None:
    """Retorna o valor da coluna de data do faturamento, tolerando variações de nome.

    O relatório do Alfa nem sempre exporta o cabeçalho igual (acento em EMISSÃO,
    caixa diferente, ponto em ULT.COMPRA). Buscar pelo nome exato fazia a
    importação falhar com 'não foi possível determinar a competência'.
    """
    if not row:
        return None
    normalized = {_column_lookup_key(k): v for k, v in row.items() if k}
    for candidate in SALES_DATE_COLUMN_CANDIDATES:
        value = normalized.get(_column_lookup_key(candidate))
        if normalize_whitespace(value):
            return value
    # Último recurso: qualquer coluna cujo nome contenha DATA/EMISSAO e tenha valor
    for key, value in normalized.items():
        if ("DATA" in key or "EMISSAO" in key) and normalize_whitespace(value):
            return value
    return None


def parse_excel_serial_date(value: str | None) -> datetime | None:
    """Converte número de série do Excel em data (ex.: '45659,3378' -> 02/01/2025).

    Quando o relatório do Alfa passa pelo Excel antes de virar CSV, a coluna de data
    é exportada como número de série (dias desde 30/12/1899, com a fração indicando
    a hora). Sem esta conversão a competência não é reconhecida.
    """
    text = normalize_whitespace(value)
    if not text or "/" in text or "-" in text or ":" in text:
        return None
    try:
        serial = float(text.replace(",", "."))
    except ValueError:
        return None
    # Faixa segura: 1954 a 2064. Evita interpretar códigos/quantidades como data.
    if not (20000 <= serial <= 60000):
        return None
    try:
        return datetime(1899, 12, 30) + timedelta(days=serial)
    except (OverflowError, ValueError):
        return None


def parse_sales_row_date(row: dict[str, str]) -> datetime | None:
    """Data de uma linha do faturamento detalhado, aceitando vários formatos."""
    raw = find_sales_date_value(row)
    return (
        parse_datetime_pt(raw)
        or parse_excel_serial_date(raw)
        or parse_datetime_flexible(raw)
    )


def parse_datetime_flexible(value: str | None) -> datetime | None:
    text = normalize_whitespace(value)
    if not text:
        return None
    normalized = text.replace("T", " ")
    for fmt in ("%Y-%m-%d %H:%M:%S", "%Y-%m-%d", "%d/%m/%Y %H:%M:%S", "%d/%m/%Y"):
        try:
            return datetime.strptime(normalized, fmt)
        except ValueError:
            continue
    try:
        return datetime.fromisoformat(text)
    except ValueError:
        return None


def crm_status_from_days(days_without_purchase: int | None) -> str:
    days = 9999 if days_without_purchase is None else days_without_purchase
    if days <= 29:
        return "ATIVO"
    if days <= 60:
        return "PRE_INATIVO"
    return "INATIVO"


def crm_class_from_average(avg_revenue: float) -> str:
    if avg_revenue > 10000:
        return "DIAMANTE"
    if avg_revenue >= 6000:
        return "OURO"
    if avg_revenue >= 3000:
        return "PRATA"
    if avg_revenue >= 500:
        return "BRONZE"
    return "NAO_CLASSIFICADO"


def crm_class_rank(class_code: str) -> int:
    order = {
        "DIAMANTE": 0,
        "OURO": 1,
        "PRATA": 2,
        "BRONZE": 3,
        "NAO_CLASSIFICADO": 4,
    }
    return order.get(class_code, 99)


def crm_scope_clause(
    conn: sqlite3.Connection, company_id: int, filters: dict[str, str | None]
) -> tuple[str, list[Any]]:
    clauses: list[str] = ["company_id = ?"]
    params: list[Any] = [company_id]
    seller_name = normalize_whitespace(filters.get("seller_name"))
    if seller_name:
        clauses.append("seller_name = ?")
        params.append(seller_name)
    unit_name = normalize_unit(filters.get("unit_name"))
    allowed_units = normalize_unit_list(filters.get("allowed_units"))
    scoped_units = [unit_name] if unit_name else allowed_units
    if scoped_units:
        cities = active_mapped_cities_for_units(conn, company_id, scoped_units)
        if cities:
            placeholders = ", ".join("?" for _ in cities)
            clauses.append(f"city_name IN ({placeholders})")
            params.extend(cities)
        else:
            clauses.append("1 = 0")
    city_name = normalize_upper(filters.get("city_name"))
    if city_name:
        clauses.append("city_name = ?")
        params.append(city_name)
    return " AND ".join(clauses), params


def init_crm_schema(conn: sqlite3.Connection) -> None:
    conn.executescript(
        """
        CREATE TABLE IF NOT EXISTS crm_contact_types (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            code TEXT NOT NULL UNIQUE,
            label TEXT NOT NULL,
            is_active INTEGER NOT NULL DEFAULT 1
        );

        CREATE TABLE IF NOT EXISTS crm_contact_results (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            code TEXT NOT NULL UNIQUE,
            label TEXT NOT NULL,
            generates_followup INTEGER NOT NULL DEFAULT 0,
            requires_followup_date INTEGER NOT NULL DEFAULT 0,
            is_active INTEGER NOT NULL DEFAULT 1
        );

        CREATE TABLE IF NOT EXISTS crm_interactions (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            company_id INTEGER NOT NULL,
            client_key TEXT NOT NULL,
            client_name TEXT NOT NULL,
            seller_name TEXT NOT NULL,
            unit_name TEXT,
            contact_phone TEXT,
            contact_name TEXT,
            contact_type_code TEXT NOT NULL,
            result_code TEXT NOT NULL,
            occurred_at TEXT NOT NULL,
            notes TEXT NOT NULL,
            question_used TEXT,
            had_progress INTEGER NOT NULL DEFAULT 0,
            offer_title TEXT,
            next_action TEXT,
            followup_due_at TEXT,
            created_at TEXT NOT NULL,
            created_by_user_id INTEGER,
            FOREIGN KEY (company_id) REFERENCES companies(id),
            FOREIGN KEY (created_by_user_id) REFERENCES users(id)
        );

        CREATE TABLE IF NOT EXISTS crm_tasks (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            company_id INTEGER NOT NULL,
            client_key TEXT NOT NULL,
            client_name TEXT NOT NULL,
            seller_name TEXT NOT NULL,
            title TEXT NOT NULL,
            description TEXT,
            due_at TEXT NOT NULL,
            status TEXT NOT NULL DEFAULT 'ABERTA',
            source_interaction_id INTEGER,
            created_at TEXT NOT NULL,
            completed_at TEXT,
            FOREIGN KEY (company_id) REFERENCES companies(id),
            FOREIGN KEY (source_interaction_id) REFERENCES crm_interactions(id)
        );

        CREATE INDEX IF NOT EXISTS idx_crm_interactions_client_date
            ON crm_interactions(company_id, client_key, occurred_at DESC);

        CREATE INDEX IF NOT EXISTS idx_crm_interactions_seller_date
            ON crm_interactions(company_id, seller_name, occurred_at DESC);

        CREATE INDEX IF NOT EXISTS idx_crm_tasks_seller_status_due
            ON crm_tasks(company_id, seller_name, status, due_at);

            CREATE TABLE IF NOT EXISTS crm_agenda_actions (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                company_id INTEGER NOT NULL,
                seller_name TEXT NOT NULL,
            client_key TEXT NOT NULL,
            client_name TEXT NOT NULL,
            action_type TEXT NOT NULL,
            justification TEXT NOT NULL,
            next_visible_at TEXT,
            created_at TEXT NOT NULL,
            created_by_user_id INTEGER,
            FOREIGN KEY (company_id) REFERENCES companies(id),
            FOREIGN KEY (created_by_user_id) REFERENCES users(id)
        );

            CREATE INDEX IF NOT EXISTS idx_crm_agenda_actions_lookup
                ON crm_agenda_actions(company_id, seller_name, client_key, created_at DESC);

            CREATE TABLE IF NOT EXISTS crm_client_profiles (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                company_id INTEGER NOT NULL,
                client_code TEXT NOT NULL,
                client_name TEXT NOT NULL,
                trade_name TEXT,
                document_number TEXT,
                state_registration TEXT,
                address_line TEXT,
                address_number TEXT,
                neighborhood TEXT,
                city_name TEXT,
                state_name TEXT,
                phone TEXT,
                updated_phone TEXT,
                primary_contact_name TEXT,
                contact_notes TEXT,
                contact_updated_at TEXT,
                contact_updated_by_user_id INTEGER,
                postal_code TEXT,
                first_sale_at TEXT,
                last_sale_at TEXT,
                credit_limit REAL NOT NULL DEFAULT 0,
                economic_group TEXT,
                internal_seller_name TEXT,
                external_seller_name TEXT,
                email TEXT,
                source_import_id INTEGER,
                updated_at TEXT NOT NULL,
                created_at TEXT NOT NULL,
                UNIQUE(company_id, client_code),
                FOREIGN KEY (company_id) REFERENCES companies(id),
                FOREIGN KEY (source_import_id) REFERENCES imports(id),
                FOREIGN KEY (contact_updated_by_user_id) REFERENCES users(id)
            );

            CREATE TABLE IF NOT EXISTS crm_client_summary (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                company_id INTEGER NOT NULL,
                competence TEXT NOT NULL,
                import_id INTEGER NOT NULL,
                client_code TEXT NOT NULL,
                client_name TEXT NOT NULL,
                seller_name TEXT,
                city_name TEXT,
                last_purchase_at TEXT,
                gross_value REAL NOT NULL DEFAULT 0,
                discount_value REAL NOT NULL DEFAULT 0,
                freight_value REAL NOT NULL DEFAULT 0,
                return_quantity REAL NOT NULL DEFAULT 0,
                return_value REAL NOT NULL DEFAULT 0,
                net_value REAL NOT NULL DEFAULT 0,
                sale_share REAL NOT NULL DEFAULT 0,
                created_at TEXT NOT NULL,
                UNIQUE(company_id, competence, client_code),
                FOREIGN KEY (company_id) REFERENCES companies(id),
                FOREIGN KEY (import_id) REFERENCES imports(id)
            );

            CREATE INDEX IF NOT EXISTS idx_crm_profiles_company_code
                ON crm_client_profiles(company_id, client_code);

            CREATE INDEX IF NOT EXISTS idx_crm_profiles_company_name
                ON crm_client_profiles(company_id, client_name);

            CREATE INDEX IF NOT EXISTS idx_crm_summary_company_competence_code
                ON crm_client_summary(company_id, competence, client_code);

            CREATE INDEX IF NOT EXISTS idx_crm_summary_company_seller_competence
                ON crm_client_summary(company_id, seller_name, competence);

            CREATE INDEX IF NOT EXISTS idx_city_mappings_company_city
                ON city_mappings(company_id, city_name);

            CREATE INDEX IF NOT EXISTS idx_crm_interactions_company_client
                ON crm_interactions(company_id, client_key);
        """
    )
    interaction_columns = {row["name"] for row in conn.execute("PRAGMA table_info(crm_interactions)").fetchall()}
    if "contact_phone" not in interaction_columns:
        conn.execute("ALTER TABLE crm_interactions ADD COLUMN contact_phone TEXT")
    if "contact_name" not in interaction_columns:
        conn.execute("ALTER TABLE crm_interactions ADD COLUMN contact_name TEXT")
    if "initiative" not in interaction_columns:
        # ATIVO = o vendedor procurou o cliente. É o que conta na meta.
        # RECEPTIVO = o cliente procurou, ou é uma anotação sobre ele. Registra
        # o histórico sem inflar o placar — senão bastaria anotar as ligações
        # recebidas para "bater" a meta sem prospectar ninguém.
        conn.execute("ALTER TABLE crm_interactions ADD COLUMN initiative TEXT NOT NULL DEFAULT 'ATIVO'")
        conn.execute("CREATE INDEX IF NOT EXISTS idx_crm_interactions_initiative "
                     "ON crm_interactions(company_id, initiative, occurred_at)")

    profile_columns = {row["name"] for row in conn.execute("PRAGMA table_info(crm_client_profiles)").fetchall()}
    if "updated_phone" not in profile_columns:
        conn.execute("ALTER TABLE crm_client_profiles ADD COLUMN updated_phone TEXT")
    if "primary_contact_name" not in profile_columns:
        conn.execute("ALTER TABLE crm_client_profiles ADD COLUMN primary_contact_name TEXT")
    if "contact_notes" not in profile_columns:
        conn.execute("ALTER TABLE crm_client_profiles ADD COLUMN contact_notes TEXT")
    if "contact_updated_at" not in profile_columns:
        conn.execute("ALTER TABLE crm_client_profiles ADD COLUMN contact_updated_at TEXT")
    if "contact_updated_by_user_id" not in profile_columns:
        conn.execute("ALTER TABLE crm_client_profiles ADD COLUMN contact_updated_by_user_id INTEGER")

    # Presente ligado à conta de login. O nome sozinho não bastava: o cadastro de
    # pessoas traz "Thielly Henrique da Silva" e o usuário se chama "Thielly
    # Henrique", então a pendência de ciência nunca chegava a ela.
    participant_columns = {row["name"] for row in conn.execute("PRAGMA table_info(meeting_participants)").fetchall()}
    if participant_columns and "user_id" not in participant_columns:
        conn.execute("ALTER TABLE meeting_participants ADD COLUMN user_id INTEGER")

    meeting_columns = {row["name"] for row in conn.execute("PRAGMA table_info(meetings)").fetchall()}
    if meeting_columns and "visibility" not in meeting_columns:
        conn.execute("ALTER TABLE meetings ADD COLUMN visibility TEXT NOT NULL DEFAULT 'UNIDADE'")

    # Tarefa deixou de ser só follow-up de cliente: agora também é direcionamento
    # do gestor, com ou sem cliente vinculado. Precisa saber de onde veio e quem
    # mandou — sem isso o vendedor recebe tarefa sem contexto.
    task_columns = {row["name"] for row in conn.execute("PRAGMA table_info(crm_tasks)").fetchall()}
    if "origin" not in task_columns:
        conn.execute("ALTER TABLE crm_tasks ADD COLUMN origin TEXT NOT NULL DEFAULT 'FOLLOWUP'")
    if "priority" not in task_columns:
        conn.execute("ALTER TABLE crm_tasks ADD COLUMN priority TEXT NOT NULL DEFAULT 'NORMAL'")
    if "created_by_name" not in task_columns:
        conn.execute("ALTER TABLE crm_tasks ADD COLUMN created_by_name TEXT")
    if "created_by_user_id" not in task_columns:
        conn.execute("ALTER TABLE crm_tasks ADD COLUMN created_by_user_id INTEGER")


def seed_crm_catalogs(conn: sqlite3.Connection) -> None:
    for code, label in CRM_CONTACT_TYPES:
        conn.execute(
            """
            INSERT INTO crm_contact_types (code, label)
            VALUES (?, ?)
            ON CONFLICT(code) DO UPDATE SET label = excluded.label, is_active = 1
            """,
            (code, label),
        )
    for code, label, generates_followup, requires_followup_date in CRM_CONTACT_RESULTS:
        conn.execute(
            """
            INSERT INTO crm_contact_results (code, label, generates_followup, requires_followup_date)
            VALUES (?, ?, ?, ?)
            ON CONFLICT(code) DO UPDATE SET
                label = excluded.label,
                generates_followup = excluded.generates_followup,
                requires_followup_date = excluded.requires_followup_date,
                is_active = 1
            """,
            (code, label, generates_followup, requires_followup_date),
        )


def crm_recent_competences(conn: sqlite3.Connection, company_id: int, limit: int = 3) -> list[str]:
    return query_competences(conn, company_id)[:limit]


def crm_latest_competence(conn: sqlite3.Connection, company_id: int) -> str | None:
    competences = crm_recent_competences(conn, company_id, 1)
    return competences[0] if competences else None


def crm_summary_latest_competence(conn: sqlite3.Connection, company_id: int) -> str | None:
    """Competência mais recente nos dados CRM (crm_client_summary).
    Independente da competência do faturamento — permite importar CRM de junho
    mesmo que o faturamento ainda seja de maio."""
    row = conn.execute(
        "SELECT MAX(competence) AS competence FROM crm_client_summary WHERE company_id = ?",
        (company_id,),
    ).fetchone()
    crm_comp = row["competence"] if row else None
    # Fallback para competência do faturamento se não houver dados CRM
    return crm_comp or crm_latest_competence(conn, company_id)


def first_day_of_competence(competence: str) -> date:
    year, month = competence.split("-")
    return date(int(year), int(month), 1)


def last_day_of_competence(competence: str) -> date:
    start = first_day_of_competence(competence)
    if start.month == 12:
        return date(start.year, 12, 31)
    return date(start.year, start.month + 1, 1) - timedelta(days=1)


def shift_competence(competence: str, delta_months: int) -> str:
    start = first_day_of_competence(competence)
    absolute_month = start.year * 12 + (start.month - 1) + delta_months
    year = absolute_month // 12
    month = absolute_month % 12 + 1
    return f"{year:04d}-{month:02d}"


def daterange(start: date, end: date):
    current = start
    while current <= end:
        yield current
        current += timedelta(days=1)


def hash_text(value: str) -> str:
    return hashlib.sha256(value.encode("utf-8")).hexdigest()


def pbkdf2_hash(password: str, salt: str | None = None) -> tuple[str, str]:
    salt = salt or secrets.token_hex(16)
    pwd_hash = hashlib.pbkdf2_hmac("sha256", password.encode("utf-8"), salt.encode("utf-8"), 120000).hex()
    return pwd_hash, salt


def verify_password(password: str, stored_hash: str, salt: str) -> bool:
    candidate, _ = pbkdf2_hash(password, salt)
    return hmac.compare_digest(candidate, stored_hash)


def _json_sanitize(value: Any) -> Any:
    """Substitui Infinity/-Infinity/NaN por None em toda a estrutura.

    O json do Python emite literais `Infinity` e `NaN`, que NÃO são JSON válido:
    o JSON.parse do navegador lança SyntaxError e a tela trava carregando. Isso
    acontecia, por exemplo, em marginValue quando a base de custo tinha divisão
    degenerada. Sanear na serialização resolve a classe inteira do problema.
    """
    if isinstance(value, float):
        return value if math.isfinite(value) else None
    if isinstance(value, dict):
        return {k: _json_sanitize(v) for k, v in value.items()}
    if isinstance(value, (list, tuple)):
        return [_json_sanitize(v) for v in value]
    return value


def json_dumps(data: Any) -> bytes:
    # allow_nan=False faria levantar exceção; preferimos sanear e sempre responder.
    return json.dumps(_json_sanitize(data), ensure_ascii=False, allow_nan=False).encode("utf-8")


def get_connection() -> sqlite3.Connection:
    ensure_dirs()
    migrate_legacy_db_if_needed()
    conn = sqlite3.connect(DB_PATH, timeout=30)
    conn.row_factory = sqlite3.Row
    # PRAGMAs seguros por conexão (sem escrita no DB)
    conn.execute("PRAGMA cache_size=-16000")   # 16 MB page cache por conexão
    conn.execute("PRAGMA temp_store=MEMORY")
    return conn


def audit_log(conn: sqlite3.Connection, company_id: int, user_id: int | None, action: str, entity_type: str, entity_id: str, changes: dict[str, Any]) -> None:
    conn.execute(
        """
        INSERT INTO audit_logs (company_id, user_id, action, entity_type, entity_id, changes_json, created_at)
        VALUES (?, ?, ?, ?, ?, ?, ?)
        """,
        (company_id, user_id, action, entity_type, entity_id, json.dumps(changes, ensure_ascii=False), now_iso()),
    )


# ─────────────────────────────────────────────────────────────────────────────
# Perfis de acesso
# ─────────────────────────────────────────────────────────────────────────────

def normalize_module_list(value: Any) -> list[str]:
    """Aceita lista ou JSON e devolve só ids de módulo válidos, na ordem canônica."""
    if isinstance(value, str):
        try:
            value = json.loads(value or "[]")
        except json.JSONDecodeError:
            value = []
    if not isinstance(value, (list, tuple, set)):
        return []
    wanted = {normalize_whitespace(str(v)) for v in value}
    return [m["id"] for m in ACCESS_MODULES if m["id"] in wanted]


def seed_access_profiles(conn: sqlite3.Connection, company_id: int) -> None:
    """Cria os perfis base uma única vez e vincula usuários existentes pelo nome do papel."""
    for spec in DEFAULT_ACCESS_PROFILES:
        conn.execute(
            """
            INSERT OR IGNORE INTO access_profiles
                (company_id, name, description, modules_json, data_scope, can_manage_users, is_system, created_at)
            VALUES (?, ?, ?, ?, ?, ?, 1, ?)
            """,
            (
                company_id,
                spec["name"],
                spec["description"],
                json.dumps(spec["modules"], ensure_ascii=False),
                spec["data_scope"],
                spec["can_manage_users"],
                now_iso(),
            ),
        )
    conn.commit()

    # Módulos novos criados depois da primeira execução não entram pelo INSERT OR IGNORE
    # acima. Aqui os perfis de sistema recebem as telas que passaram a existir na
    # especificação. Só ADICIONA — nunca remove, para preservar ajustes feitos na tela.
    for spec in DEFAULT_ACCESS_PROFILES:
        row = conn.execute(
            "SELECT id, modules_json FROM access_profiles "
            "WHERE company_id = ? AND name = ? AND is_system = 1",
            (company_id, spec["name"]),
        ).fetchone()
        if not row:
            continue
        current = normalize_module_list(row["modules_json"])
        missing = [m for m in spec["modules"] if m not in current]
        if missing:
            merged = normalize_module_list(current + missing)
            conn.execute(
                "UPDATE access_profiles SET modules_json = ?, updated_at = ? WHERE id = ?",
                (json.dumps(merged, ensure_ascii=False), now_iso(), row["id"]),
            )
            print(f"[perfis] '{spec['name']}' recebeu novas telas: {', '.join(missing)}")
    conn.commit()

    # Usuários criados antes dos perfis: liga cada um ao perfil de mesmo nome do role
    conn.execute(
        """
        UPDATE users
        SET profile_id = (
            SELECT p.id FROM access_profiles p
            WHERE p.company_id = users.company_id AND p.name = users.role
        )
        WHERE profile_id IS NULL
          AND EXISTS (
            SELECT 1 FROM access_profiles p
            WHERE p.company_id = users.company_id AND p.name = users.role
          )
        """
    )
    conn.commit()


def access_profile_row_to_dict(row: sqlite3.Row) -> dict[str, Any]:
    return {
        "id": row["id"],
        "name": row["name"],
        "description": row["description"] or "",
        "modules": normalize_module_list(row["modules_json"]),
        "dataScope": row["data_scope"] or "todos",
        "canManageUsers": bool(row["can_manage_users"]),
        "isSystem": bool(row["is_system"]),
    }


def list_access_profiles(conn: sqlite3.Connection, company_id: int) -> list[dict[str, Any]]:
    rows = conn.execute(
        "SELECT * FROM access_profiles WHERE company_id = ? ORDER BY is_system DESC, name",
        (company_id,),
    ).fetchall()
    return [access_profile_row_to_dict(r) for r in rows]


def get_access_profile_for_user(conn: sqlite3.Connection, user: Any) -> dict[str, Any] | None:
    """Perfil do usuário. Se não houver vínculo, cai no perfil de mesmo nome do role."""
    profile_id = None
    try:
        profile_id = user["profile_id"]
    except (KeyError, IndexError, TypeError):
        profile_id = None
    row = None
    if profile_id:
        row = conn.execute(
            "SELECT * FROM access_profiles WHERE id = ? AND company_id = ?",
            (profile_id, user["company_id"]),
        ).fetchone()
    if row is None:
        row = conn.execute(
            "SELECT * FROM access_profiles WHERE company_id = ? AND name = ?",
            (user["company_id"], user["role"]),
        ).fetchone()
    return access_profile_row_to_dict(row) if row else None


def upsert_access_profile(
    conn: sqlite3.Connection, company_id: int, user_id: int, payload: dict[str, Any]
) -> dict[str, Any]:
    name = normalize_whitespace(payload.get("name"))
    if not name:
        raise ValueError("Nome do perfil é obrigatório.")
    modules = normalize_module_list(payload.get("modules"))
    if not modules:
        raise ValueError("Selecione ao menos uma tela para o perfil.")
    data_scope = normalize_whitespace(payload.get("dataScope")) or "todos"
    if data_scope not in DATA_SCOPE_IDS:
        raise ValueError("Escopo de dados inválido.")
    can_manage = 1 if payload.get("canManageUsers") else 0
    description = normalize_whitespace(payload.get("description"))
    profile_id = payload.get("id")

    if profile_id:
        existing = conn.execute(
            "SELECT * FROM access_profiles WHERE id = ? AND company_id = ?", (profile_id, company_id)
        ).fetchone()
        if not existing:
            raise ValueError("Perfil não encontrado.")
        # Perfis de sistema podem ser editados, mas não renomeados (usuários referenciam pelo nome)
        if existing["is_system"] and name != existing["name"]:
            raise ValueError(f"O perfil '{existing['name']}' é padrão do sistema e não pode ser renomeado.")
        conn.execute(
            """
            UPDATE access_profiles
            SET name = ?, description = ?, modules_json = ?, data_scope = ?,
                can_manage_users = ?, updated_at = ?
            WHERE id = ? AND company_id = ?
            """,
            (name, description, json.dumps(modules, ensure_ascii=False), data_scope,
             can_manage, now_iso(), profile_id, company_id),
        )
        audit_log(conn, company_id, user_id, "editar", "access_profiles", str(profile_id),
                  {"name": name, "modules": len(modules), "dataScope": data_scope})
        created = False
    else:
        duplicate = conn.execute(
            "SELECT id FROM access_profiles WHERE company_id = ? AND name = ?", (company_id, name)
        ).fetchone()
        if duplicate:
            raise ValueError(f"Já existe um perfil chamado '{name}'.")
        cursor = conn.execute(
            """
            INSERT INTO access_profiles
                (company_id, name, description, modules_json, data_scope, can_manage_users, is_system, created_at)
            VALUES (?, ?, ?, ?, ?, ?, 0, ?)
            """,
            (company_id, name, description, json.dumps(modules, ensure_ascii=False),
             data_scope, can_manage, now_iso()),
        )
        profile_id = cursor.lastrowid
        audit_log(conn, company_id, user_id, "criar", "access_profiles", str(profile_id),
                  {"name": name, "modules": len(modules), "dataScope": data_scope})
        created = True
    return {"id": profile_id, "created": created}


def delete_access_profile(conn: sqlite3.Connection, company_id: int, user_id: int, profile_id: Any) -> None:
    row = conn.execute(
        "SELECT * FROM access_profiles WHERE id = ? AND company_id = ?", (profile_id, company_id)
    ).fetchone()
    if not row:
        raise ValueError("Perfil não encontrado.")
    if row["is_system"]:
        raise ValueError(f"O perfil '{row['name']}' é padrão do sistema e não pode ser excluído.")
    in_use = conn.execute(
        "SELECT COUNT(*) AS n FROM users WHERE company_id = ? AND profile_id = ?", (company_id, profile_id)
    ).fetchone()["n"]
    if in_use:
        raise ValueError(f"Há {in_use} usuário(s) usando este perfil. Troque o perfil deles antes de excluir.")
    conn.execute("DELETE FROM access_profiles WHERE id = ? AND company_id = ?", (profile_id, company_id))
    audit_log(conn, company_id, user_id, "excluir", "access_profiles", str(profile_id), {"name": row["name"]})


# ─────────────────────────────────────────────────────────────────────────────
# Biblioteca de conteúdo comercial
# ─────────────────────────────────────────────────────────────────────────────

CONTENT_CATEGORIES = [
    {"id": "ligacao",  "label": "Abordagem por telefone", "icon": "📞"},
    {"id": "whatsapp", "label": "Mensagem de WhatsApp",   "icon": "💬"},
    {"id": "objecao",  "label": "Tratamento de objeção",  "icon": "🛡"},
    {"id": "garantia", "label": "Devolução e garantia",   "icon": "📋"},
    {"id": "prospeccao", "label": "Prospecção",           "icon": "🎯"},
]
CONTENT_CATEGORY_IDS = {c["id"] for c in CONTENT_CATEGORIES}

CONTENT_SITUATIONS = [
    {"id": "GERAL",          "label": "Qualquer situação"},
    {"id": "INATIVO",        "label": "Cliente inativo"},
    {"id": "PRE_INATIVO",    "label": "Cliente pré-inativo"},
    {"id": "SEM_COMPRA_MES", "label": "Sem compra no mês"},
    {"id": "QUEDA",          "label": "Queda de faturamento"},
    {"id": "MIX",            "label": "Ampliar mix"},
    {"id": "NOVO",           "label": "Cliente novo"},
]
CONTENT_SITUATION_IDS = {s["id"] for s in CONTENT_SITUATIONS}


def seed_content_library(conn: sqlite3.Connection, company_id: int) -> None:
    """Adiciona os conteúdos do arquivo que ainda não existem na biblioteca.

    Antes só populava com a tabela vazia — conteúdo novo nunca chegava a quem já
    tinha a biblioteca em uso, que é justamente o caso de todo mundo depois da
    primeira semana. A comparação é pelo título: o que o usuário editou pela
    tela permanece intocado, e material novo entra no restart.
    """
    try:
        from content_seed import CONTENT_SEED
    except ImportError:
        print("[content] content_seed.py não encontrado — biblioteca iniciada vazia")
        return

    existentes = {
        normalize_upper(r["title"])
        for r in conn.execute(
            "SELECT title FROM content_library WHERE company_id = ?", (company_id,)
        ).fetchall()
    }
    novos = 0
    for item in CONTENT_SEED:
        if normalize_upper(item["title"]) in existentes:
            continue
        conn.execute(
            """
            INSERT INTO content_library
                (company_id, category, situation, title, body, hint, sort_order, is_active, is_system, created_at)
            VALUES (?, ?, ?, ?, ?, ?, ?, 1, 1, ?)
            """,
            (company_id, item["category"], item["situation"], item["title"],
             item["body"], item.get("hint"), item.get("sort_order", 0), now_iso()),
        )
        novos += 1
    if novos:
        conn.commit()
        print(f"[content] {novos} conteúdo(s) adicionado(s) à biblioteca")


def content_row_to_dict(row: sqlite3.Row) -> dict[str, Any]:
    return {
        "id": row["id"],
        "category": row["category"],
        "situation": row["situation"],
        "title": row["title"],
        "body": row["body"],
        "hint": row["hint"] or "",
        "sortOrder": row["sort_order"],
        "isActive": bool(row["is_active"]),
        "isSystem": bool(row["is_system"]),
    }


def list_content_library(
    conn: sqlite3.Connection, company_id: int, category: str | None = None,
    situation: str | None = None, only_active: bool = True,
) -> list[dict[str, Any]]:
    where = ["company_id = ?"]
    params: list[Any] = [company_id]
    if category:
        where.append("category = ?")
        params.append(category)
    if situation:
        # GERAL sempre acompanha a situação específica
        where.append("(situation = ? OR situation = 'GERAL')")
        params.append(situation)
    if only_active:
        where.append("is_active = 1")
    rows = conn.execute(
        f"SELECT * FROM content_library WHERE {' AND '.join(where)} "
        "ORDER BY category, CASE situation WHEN 'GERAL' THEN 1 ELSE 0 END, sort_order, id",
        params,
    ).fetchall()
    return [content_row_to_dict(r) for r in rows]


def render_content_text(text: str, context: dict[str, Any]) -> str:
    """Substitui os marcadores {cliente}, {vendedor}, {item}, {dias} etc."""
    result = text or ""
    for key, value in context.items():
        result = result.replace("{" + key + "}", str(value) if value is not None else "")
    # Marcadores sem valor viram um placeholder visível para o vendedor preencher
    result = re.sub(r"\{(\w+)\}", lambda m: f"[{m.group(1)}]", result)
    return result


def content_context_for_client(client: dict[str, Any], seller_name: str | None) -> dict[str, Any]:
    offer = client.get("offerPrimary") or {}
    dias = client.get("daysWithoutPurchase")
    return {
        "cliente": client.get("clientName") or "cliente",
        "vendedor": seller_name or "",
        "item": offer.get("title") or "",
        "dias": dias if dias is not None else "",
        "unidade": client.get("unitName") or "",
    }


def upsert_content_item(
    conn: sqlite3.Connection, company_id: int, user_id: int, payload: dict[str, Any]
) -> dict[str, Any]:
    category = normalize_whitespace(payload.get("category"))
    situation = normalize_upper(payload.get("situation")) or "GERAL"
    title = normalize_whitespace(payload.get("title"))
    body = (payload.get("body") or "").strip()
    hint = normalize_whitespace(payload.get("hint"))
    sort_order = int(payload.get("sortOrder") or 0)
    is_active = 1 if payload.get("isActive", True) else 0
    item_id = payload.get("id")

    if category not in CONTENT_CATEGORY_IDS:
        raise ValueError("Categoria inválida.")
    if situation not in CONTENT_SITUATION_IDS:
        raise ValueError("Situação inválida.")
    if not title or not body:
        raise ValueError("Título e conteúdo são obrigatórios.")

    if item_id:
        exists = conn.execute(
            "SELECT id FROM content_library WHERE id = ? AND company_id = ?", (item_id, company_id)
        ).fetchone()
        if not exists:
            raise ValueError("Conteúdo não encontrado.")
        conn.execute(
            """
            UPDATE content_library
            SET category = ?, situation = ?, title = ?, body = ?, hint = ?,
                sort_order = ?, is_active = ?, updated_at = ?
            WHERE id = ? AND company_id = ?
            """,
            (category, situation, title, body, hint, sort_order, is_active, now_iso(), item_id, company_id),
        )
        audit_log(conn, company_id, user_id, "editar", "content_library", str(item_id), {"title": title})
        return {"id": item_id, "created": False}

    cur = conn.execute(
        """
        INSERT INTO content_library
            (company_id, category, situation, title, body, hint, sort_order, is_active, is_system, created_at)
        VALUES (?, ?, ?, ?, ?, ?, ?, ?, 0, ?)
        """,
        (company_id, category, situation, title, body, hint, sort_order, is_active, now_iso()),
    )
    audit_log(conn, company_id, user_id, "criar", "content_library", str(cur.lastrowid), {"title": title})
    return {"id": cur.lastrowid, "created": True}


def delete_content_item(conn: sqlite3.Connection, company_id: int, user_id: int, item_id: Any) -> None:
    row = conn.execute(
        "SELECT * FROM content_library WHERE id = ? AND company_id = ?", (item_id, company_id)
    ).fetchone()
    if not row:
        raise ValueError("Conteúdo não encontrado.")
    conn.execute("DELETE FROM content_library WHERE id = ? AND company_id = ?", (item_id, company_id))
    audit_log(conn, company_id, user_id, "excluir", "content_library", str(item_id), {"title": row["title"]})


def user_can_manage_users(conn: sqlite3.Connection, user: Any) -> bool:
    profile = get_access_profile_for_user(conn, user)
    if profile:
        return profile["canManageUsers"]
    return user["role"] in {"Administrador", "Diretor"}


def init_db() -> None:
    ensure_dirs()
    with closing(get_connection()) as conn:
        # WAL mode: aplicado apenas uma vez na inicialização (é operação de escrita)
        conn.execute("PRAGMA journal_mode=WAL")
        conn.executescript(
            """
            PRAGMA foreign_keys = ON;

            CREATE TABLE IF NOT EXISTS companies (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                name TEXT NOT NULL UNIQUE,
                code TEXT,
                created_at TEXT NOT NULL
            );

            CREATE TABLE IF NOT EXISTS users (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                company_id INTEGER NOT NULL,
                username TEXT NOT NULL UNIQUE,
                full_name TEXT,
                linked_person_name TEXT,
                linked_units_json TEXT,
                password_hash TEXT NOT NULL,
                password_salt TEXT NOT NULL,
                role TEXT NOT NULL,
                is_active INTEGER NOT NULL DEFAULT 1,
                created_at TEXT NOT NULL,
                FOREIGN KEY (company_id) REFERENCES companies(id)
            );

            CREATE TABLE IF NOT EXISTS sessions (
                id TEXT PRIMARY KEY,
                user_id INTEGER NOT NULL,
                created_at TEXT NOT NULL,
                expires_at TEXT NOT NULL,
                FOREIGN KEY (user_id) REFERENCES users(id)
            );

            -- Devoluções em garantia: chegam misturadas no custo/venda e precisam ser
            -- deduzidas do resultado comercial (defeito de fábrica não é erro de venda)
            CREATE TABLE IF NOT EXISTS fact_warranty_returns (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                company_id INTEGER NOT NULL,
                competence TEXT NOT NULL,
                import_id INTEGER NOT NULL,
                row_hash TEXT NOT NULL,
                unit_name TEXT NOT NULL,
                seller_name TEXT NOT NULL,
                client_name TEXT,
                city_name TEXT,
                return_number TEXT,
                return_date TEXT,
                reason TEXT,
                invoice_number TEXT,
                issue_date TEXT,
                item_code TEXT,
                item_type TEXT,
                item_description TEXT,
                brand_name TEXT,
                supplier_name TEXT,
                quantity REAL NOT NULL DEFAULT 0,
                cost_value REAL NOT NULL DEFAULT 0,
                total_value REAL NOT NULL DEFAULT 0,
                created_at TEXT NOT NULL,
                UNIQUE(company_id, competence, row_hash),
                FOREIGN KEY (company_id) REFERENCES companies(id)
            );

            CREATE INDEX IF NOT EXISTS idx_warranty_company_competence
                ON fact_warranty_returns(company_id, competence);
            CREATE INDEX IF NOT EXISTS idx_warranty_company_comp_seller
                ON fact_warranty_returns(company_id, competence, seller_name);
            CREATE INDEX IF NOT EXISTS idx_warranty_company_comp_unit
                ON fact_warranty_returns(company_id, competence, unit_name);

            -- Faróis: limites de cor por indicador, ajustáveis pela tela
            CREATE TABLE IF NOT EXISTS kpi_thresholds (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                company_id INTEGER NOT NULL,
                metric_id TEXT NOT NULL,
                good_at REAL NOT NULL,
                warn_at REAL NOT NULL,
                is_active INTEGER NOT NULL DEFAULT 1,
                updated_at TEXT,
                created_at TEXT NOT NULL,
                UNIQUE(company_id, metric_id),
                FOREIGN KEY (company_id) REFERENCES companies(id)
            );

            -- Atas de reunião e registros de treinamento
            CREATE TABLE IF NOT EXISTS meetings (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                company_id INTEGER NOT NULL,
                kind TEXT NOT NULL,              -- REUNIAO | TREINAMENTO
                title TEXT NOT NULL,
                topic TEXT,                      -- tema, usado na busca do acervo
                unit_name TEXT,
                occurred_at TEXT NOT NULL,       -- data e hora do encontro
                duration_min INTEGER,
                location TEXT,
                agenda TEXT,                     -- pauta
                summary TEXT,                    -- o que foi tratado
                decisions TEXT,                  -- decisões e encaminhamentos
                organizer_name TEXT NOT NULL,
                visibility TEXT NOT NULL DEFAULT 'UNIDADE',  -- UNIDADE | EMPRESA
                status TEXT NOT NULL DEFAULT 'RASCUNHO',   -- RASCUNHO | PUBLICADA
                published_at TEXT,
                created_by_user_id INTEGER,
                created_at TEXT NOT NULL,
                updated_at TEXT,
                FOREIGN KEY (company_id) REFERENCES companies(id)
            );

            CREATE INDEX IF NOT EXISTS idx_meetings_company_date
                ON meetings(company_id, occurred_at);
            CREATE INDEX IF NOT EXISTS idx_meetings_company_unit
                ON meetings(company_id, unit_name);

            -- Presentes e a ciência de cada um
            CREATE TABLE IF NOT EXISTS meeting_participants (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                meeting_id INTEGER NOT NULL,
                person_name TEXT NOT NULL,
                person_key TEXT NOT NULL,        -- nome normalizado, casa com o login
                user_id INTEGER,                 -- conta resolvida, quando a pessoa tem login
                unit_name TEXT,
                acknowledged_at TEXT,
                feedback TEXT,
                feedback_at TEXT,
                created_at TEXT NOT NULL,
                UNIQUE(meeting_id, person_key),
                FOREIGN KEY (meeting_id) REFERENCES meetings(id) ON DELETE CASCADE
            );

            CREATE INDEX IF NOT EXISTS idx_meeting_participants_person
                ON meeting_participants(person_key, acknowledged_at);

            -- Anexos: o binário fica em disco, a tabela guarda só o ponteiro
            CREATE TABLE IF NOT EXISTS meeting_attachments (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                meeting_id INTEGER NOT NULL,
                file_name TEXT NOT NULL,
                stored_name TEXT NOT NULL,
                content_type TEXT,
                size_bytes INTEGER NOT NULL DEFAULT 0,
                uploaded_by_user_id INTEGER,
                created_at TEXT NOT NULL,
                FOREIGN KEY (meeting_id) REFERENCES meetings(id) ON DELETE CASCADE
            );

            -- Feedback estruturado (gerente x vendedor e diretor x gerente)
            CREATE TABLE IF NOT EXISTS feedbacks (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                company_id INTEGER NOT NULL,
                kind TEXT NOT NULL,                -- VENDEDOR | GERENTE
                competence TEXT NOT NULL,          -- mês de referência (AAAA-MM)
                person_name TEXT NOT NULL,         -- avaliado
                person_key TEXT NOT NULL,
                person_user_id INTEGER,
                unit_name TEXT,
                author_name TEXT NOT NULL,         -- quem conduziu
                author_user_id INTEGER,
                indicators_json TEXT,              -- foto dos números no dia do feedback
                highlights TEXT,                   -- o que foi bem
                improvements TEXT,                 -- o que precisa evoluir
                agreements TEXT,                   -- o que ficou combinado
                tactical_goal TEXT,                -- GROW: objetivo (feedback de gerente)
                tactical_reality TEXT,             -- GROW: realidade
                tactical_options TEXT,             -- GROW: caminhos
                tactical_will TEXT,                -- GROW: compromisso e apoio pedido
                status TEXT NOT NULL DEFAULT 'RASCUNHO',
                published_at TEXT,
                acknowledged_at TEXT,
                person_note TEXT,                  -- observação do avaliado PARA O GESTOR
                confidential_note TEXT,            -- observação do avaliado para RH/Diretoria
                note_at TEXT,
                created_at TEXT NOT NULL,
                updated_at TEXT,
                UNIQUE(company_id, kind, competence, person_key),
                FOREIGN KEY (company_id) REFERENCES companies(id)
            );

            CREATE INDEX IF NOT EXISTS idx_feedbacks_company_person
                ON feedbacks(company_id, person_key, competence);
            CREATE INDEX IF NOT EXISTS idx_feedbacks_company_unit
                ON feedbacks(company_id, unit_name, competence);

            -- Nota de cada item avaliado
            CREATE TABLE IF NOT EXISTS feedback_ratings (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                feedback_id INTEGER NOT NULL,
                item_id TEXT NOT NULL,
                level TEXT NOT NULL,               -- SUPERA | ATENDE | EVOLUIR
                comment TEXT,
                UNIQUE(feedback_id, item_id),
                FOREIGN KEY (feedback_id) REFERENCES feedbacks(id) ON DELETE CASCADE
            );

            -- Fase da unidade. Unidade nova opera meses sem meta e sem carteira:
            -- sem essa marcação, todo indicador de faturamento fica vermelho e a
            -- equipe aprende que o painel não diz nada sobre o trabalho dela.
            CREATE TABLE IF NOT EXISTS unit_phases (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                company_id INTEGER NOT NULL,
                unit_name TEXT NOT NULL,
                phase TEXT NOT NULL DEFAULT 'OPERACAO',   -- IMPLANTACAO | OPERACAO
                opening_date TEXT,                        -- inauguração prevista/realizada
                goal_exempt_until TEXT,                   -- última competência sem meta (AAAA-MM)
                notes TEXT,
                updated_by_user_id INTEGER,
                created_at TEXT NOT NULL,
                updated_at TEXT,
                UNIQUE(company_id, unit_name),
                FOREIGN KEY (company_id) REFERENCES companies(id)
            );

            -- Metas de esforço para quem ainda não tem meta de faturamento.
            -- seller_name vazio = meta da unidade.
            CREATE TABLE IF NOT EXISTS activity_goals (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                company_id INTEGER NOT NULL,
                competence TEXT NOT NULL,
                unit_name TEXT NOT NULL,
                seller_name TEXT,
                metric TEXT NOT NULL,      -- CALLS | PROSPECTS_NEW | PROSPECTS_REGISTERED | FIRST_PURCHASES
                target REAL NOT NULL DEFAULT 0,
                created_at TEXT NOT NULL,
                updated_at TEXT,
                UNIQUE(company_id, competence, unit_name, seller_name, metric),
                FOREIGN KEY (company_id) REFERENCES companies(id)
            );

            -- Prospect: oficina que ainda não existe no cadastro do Alfa.
            -- Vira cliente quando o CNPJ aparece na importação; até lá, o contato
            -- é registrado igual ao de um cliente, na mesma tabela de interações.
            CREATE TABLE IF NOT EXISTS prospects (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                company_id INTEGER NOT NULL,
                unit_name TEXT NOT NULL,
                seller_name TEXT NOT NULL,
                company_name TEXT NOT NULL,
                trade_name TEXT,
                document_number TEXT,      -- CNPJ: é o que permite o casamento automático
                document_digits TEXT,      -- só dígitos, para comparar sem formatação
                phone TEXT,
                contact_name TEXT,
                email TEXT,
                city_name TEXT,
                neighborhood TEXT,
                address_line TEXT,
                origin TEXT,               -- indicação, rua, lista, internet
                status TEXT NOT NULL DEFAULT 'NOVO',  -- NOVO | EM_CONTATO | QUALIFICADO | CADASTRADO | PERDIDO
                -- Qualificação do modelo Passini (4 perguntas)
                q_service_type TEXT,       -- rápida | pesada | ambas
                q_cars_week INTEGER,
                q_main_line TEXT,          -- suspensão | freio | motor | outra
                q_payment TEXT,            -- à vista | cartão | faturado
                closing_trigger TEXT,      -- ORCAMENTO | COTACAO | DIA_COMPRA
                notes TEXT,
                client_code TEXT,          -- preenchido quando vira cliente
                converted_at TEXT,
                first_purchase_at TEXT,
                lost_reason TEXT,
                created_by_user_id INTEGER,
                created_at TEXT NOT NULL,
                updated_at TEXT,
                FOREIGN KEY (company_id) REFERENCES companies(id)
            );

            CREATE INDEX IF NOT EXISTS idx_prospects_unit_status
                ON prospects(company_id, unit_name, status);
            CREATE INDEX IF NOT EXISTS idx_prospects_document
                ON prospects(company_id, document_digits);

            -- Pedido de visita feito pelo vendedor ao registrar uma ligação.
            -- É o que garante a ordem certa: o vendedor tenta por telefone e,
            -- quando o problema passa do alcance dele, pede a presença do gestor.
            CREATE TABLE IF NOT EXISTS visit_requests (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                company_id INTEGER NOT NULL,
                client_key TEXT NOT NULL,
                client_name TEXT NOT NULL,
                unit_name TEXT,
                city_name TEXT,
                seller_name TEXT NOT NULL,
                reason TEXT NOT NULL,
                status TEXT NOT NULL DEFAULT 'PENDENTE',   -- PENDENTE | ACEITA | RECUSADA
                manager_note TEXT,
                interaction_id INTEGER,
                requested_by_user_id INTEGER,
                resolved_by_user_id INTEGER,
                resolved_at TEXT,
                created_at TEXT NOT NULL,
                FOREIGN KEY (company_id) REFERENCES companies(id)
            );

            CREATE INDEX IF NOT EXISTS idx_visit_requests_status
                ON visit_requests(company_id, status, unit_name);

            -- Visita gerencial. O endereço é copiado no momento do registro:
            -- o cadastro do cliente muda, e o roteiro de ontem precisa continuar
            -- mostrando onde a visita realmente aconteceu.
            CREATE TABLE IF NOT EXISTS visits (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                company_id INTEGER NOT NULL,
                client_key TEXT NOT NULL,
                client_name TEXT NOT NULL,
                unit_name TEXT,
                city_name TEXT,
                neighborhood TEXT,
                address_line TEXT,
                visit_type TEXT NOT NULL,        -- REATIVACAO | RELACIONAMENTO | SOLICITADA | NEGOCIACAO
                status TEXT NOT NULL DEFAULT 'PLANEJADA',  -- PLANEJADA | REALIZADA | CANCELADA
                scheduled_for TEXT,
                occurred_at TEXT,
                manager_name TEXT NOT NULL,
                seller_name TEXT,                -- vendedor que foi junto, quando foi
                objective TEXT,
                outcome TEXT,                    -- o que aconteceu
                agreement TEXT,                  -- ação combinada com o cliente
                next_action TEXT,
                next_action_due TEXT,
                request_id INTEGER,
                -- Efeito medido: faturamento nos 60 dias antes e depois da visita
                revenue_before REAL,
                revenue_after REAL,
                effect_measured_at TEXT,
                created_by_user_id INTEGER,
                created_at TEXT NOT NULL,
                updated_at TEXT,
                FOREIGN KEY (company_id) REFERENCES companies(id)
            );

            CREATE INDEX IF NOT EXISTS idx_visits_client
                ON visits(company_id, client_key, occurred_at);
            CREATE INDEX IF NOT EXISTS idx_visits_unit_status
                ON visits(company_id, unit_name, status);

            -- Base de conhecimento do assistente
            CREATE TABLE IF NOT EXISTS help_articles (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                company_id INTEGER NOT NULL,
                category TEXT NOT NULL,
                question TEXT NOT NULL,
                answer TEXT NOT NULL,
                keywords TEXT,
                roles TEXT,                       -- vazio = todos os perfis
                source TEXT NOT NULL DEFAULT 'SEED',   -- SEED | MANUAL
                is_active INTEGER NOT NULL DEFAULT 1,
                sort_order INTEGER NOT NULL DEFAULT 0,
                created_by_user_id INTEGER,
                created_at TEXT NOT NULL,
                updated_at TEXT,
                UNIQUE(company_id, question),
                FOREIGN KEY (company_id) REFERENCES companies(id)
            );

            -- Dicas do assistente
            CREATE TABLE IF NOT EXISTS assistant_tips (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                company_id INTEGER NOT NULL,
                kind TEXT NOT NULL,               -- MENSAGEM | MEC | DESEMPENHO | LEMBRETE
                title TEXT NOT NULL,
                body TEXT NOT NULL,
                roles TEXT,
                trigger_code TEXT,                -- vazio = sempre elegível
                source TEXT NOT NULL DEFAULT 'SEED',
                is_active INTEGER NOT NULL DEFAULT 1,
                created_by_user_id INTEGER,
                created_at TEXT NOT NULL,
                updated_at TEXT,
                UNIQUE(company_id, title),
                FOREIGN KEY (company_id) REFERENCES companies(id)
            );

            -- Perguntas que a busca não respondeu. É daqui que o FAQ cresce:
            -- a dúvida real da equipe vale mais que a que a gente imagina.
            CREATE TABLE IF NOT EXISTS help_questions (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                company_id INTEGER NOT NULL,
                user_id INTEGER,
                user_name TEXT,
                user_role TEXT,
                question TEXT NOT NULL,
                status TEXT NOT NULL DEFAULT 'PENDENTE',   -- PENDENTE | RESPONDIDA | DESCARTADA
                answer TEXT,
                answered_by_user_id INTEGER,
                answered_at TEXT,
                created_at TEXT NOT NULL,
                FOREIGN KEY (company_id) REFERENCES companies(id)
            );

            CREATE INDEX IF NOT EXISTS idx_help_questions_status
                ON help_questions(company_id, status);

            -- Quem já viu o tutorial
            CREATE TABLE IF NOT EXISTS user_onboarding (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                user_id INTEGER NOT NULL,
                tour_key TEXT NOT NULL,
                completed_at TEXT,
                skipped INTEGER NOT NULL DEFAULT 0,
                UNIQUE(user_id, tour_key)
            );

            -- Registro pontual: a conversa que não espera o fechamento do mês.
            -- O MEC (item 7) prevê "orienta e combina a correção; se repetir,
            -- registra o fato e o combinado". É isso — leve, datado, e vira
            -- memória do feedback mensal para o gerente não escrever de cabeça.
            CREATE TABLE IF NOT EXISTS feedback_notes (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                company_id INTEGER NOT NULL,
                person_name TEXT NOT NULL,
                person_key TEXT NOT NULL,
                person_user_id INTEGER,
                unit_name TEXT,
                occurred_at TEXT NOT NULL,        -- data do fato
                competence TEXT NOT NULL,         -- derivada da data, para casar com o mensal
                kind TEXT NOT NULL,               -- RECONHECIMENTO | ORIENTACAO | CORRECAO | ACOMPANHAMENTO
                summary TEXT NOT NULL,            -- o que aconteceu
                agreement TEXT,                   -- o que ficou combinado
                requires_ack INTEGER NOT NULL DEFAULT 0,
                acknowledged_at TEXT,
                person_note TEXT,
                author_name TEXT NOT NULL,
                author_user_id INTEGER,
                created_at TEXT NOT NULL,
                FOREIGN KEY (company_id) REFERENCES companies(id)
            );

            CREATE INDEX IF NOT EXISTS idx_feedback_notes_person
                ON feedback_notes(company_id, person_key, competence);

            -- PDI vivo: o item pertence à PESSOA e atravessa os feedbacks.
            -- Amarrar o plano a um único feedback faria o desenvolvimento
            -- recomeçar do zero todo mês.
            CREATE TABLE IF NOT EXISTS pdi_items (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                company_id INTEGER NOT NULL,
                person_name TEXT NOT NULL,
                person_key TEXT NOT NULL,
                unit_name TEXT,
                title TEXT NOT NULL,               -- o que desenvolver
                why TEXT,                          -- por que importa
                action TEXT,                       -- como, na prática
                support TEXT,                      -- quem apoia
                due_date TEXT,
                status TEXT NOT NULL DEFAULT 'ABERTO',   -- ABERTO | EVOLUINDO | CONCLUIDO | CANCELADO
                progress_note TEXT,
                origin_feedback_id INTEGER,
                created_by_user_id INTEGER,
                created_at TEXT NOT NULL,
                updated_at TEXT,
                closed_at TEXT,
                FOREIGN KEY (company_id) REFERENCES companies(id)
            );

            CREATE INDEX IF NOT EXISTS idx_pdi_company_person
                ON pdi_items(company_id, person_key, status);

            -- Biblioteca de conteúdo comercial: scripts, mensagens e orientações
            CREATE TABLE IF NOT EXISTS content_library (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                company_id INTEGER NOT NULL,
                category TEXT NOT NULL,        -- ligacao | whatsapp | objecao | garantia
                situation TEXT NOT NULL,       -- INATIVO | PRE_INATIVO | SEM_COMPRA_MES | QUEDA | MIX | NOVO | GERAL
                title TEXT NOT NULL,
                body TEXT NOT NULL,
                hint TEXT,
                sort_order INTEGER NOT NULL DEFAULT 0,
                is_active INTEGER NOT NULL DEFAULT 1,
                is_system INTEGER NOT NULL DEFAULT 0,
                created_at TEXT NOT NULL,
                updated_at TEXT,
                FOREIGN KEY (company_id) REFERENCES companies(id)
            );

            CREATE INDEX IF NOT EXISTS idx_content_library_lookup
                ON content_library(company_id, category, situation, is_active);

            -- Perfis de acesso configuráveis pela tela (antes as permissões eram fixas no código)
            CREATE TABLE IF NOT EXISTS access_profiles (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                company_id INTEGER NOT NULL,
                name TEXT NOT NULL,
                description TEXT,
                modules_json TEXT NOT NULL DEFAULT '[]',
                data_scope TEXT NOT NULL DEFAULT 'todos',
                can_manage_users INTEGER NOT NULL DEFAULT 0,
                is_system INTEGER NOT NULL DEFAULT 0,
                created_at TEXT NOT NULL,
                updated_at TEXT,
                UNIQUE(company_id, name),
                FOREIGN KEY (company_id) REFERENCES companies(id)
            );

            CREATE TABLE IF NOT EXISTS people_records (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                company_id INTEGER NOT NULL,
                person_name TEXT NOT NULL,
                role_classification TEXT NOT NULL,
                base_unit TEXT,
                valid_from TEXT NOT NULL,
                valid_to TEXT,
                source TEXT NOT NULL,
                is_active INTEGER NOT NULL DEFAULT 1,
                created_at TEXT NOT NULL,
                UNIQUE(company_id, person_name, valid_from, source),
                FOREIGN KEY (company_id) REFERENCES companies(id)
            );

            CREATE TABLE IF NOT EXISTS city_mappings (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                company_id INTEGER NOT NULL,
                city_name TEXT NOT NULL,
                principal_unit TEXT NOT NULL,
                state_name TEXT,
                country_name TEXT,
                valid_from TEXT NOT NULL,
                valid_to TEXT,
                source TEXT NOT NULL,
                created_at TEXT NOT NULL,
                UNIQUE(company_id, city_name, valid_from, source),
                FOREIGN KEY (company_id) REFERENCES companies(id)
            );

            CREATE TABLE IF NOT EXISTS portfolio_coverage (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                company_id INTEGER NOT NULL,
                covering_seller TEXT NOT NULL,   -- quem cobre
                covered_seller TEXT NOT NULL,    -- carteira coberta
                unit_name TEXT,
                start_date TEXT NOT NULL,
                end_date TEXT,
                reason TEXT,
                created_by_user_id INTEGER,
                created_at TEXT NOT NULL,
                FOREIGN KEY (company_id) REFERENCES companies(id)
            );

            CREATE INDEX IF NOT EXISTS idx_coverage_lookup
                ON portfolio_coverage(company_id, covering_seller, start_date);

            CREATE TABLE IF NOT EXISTS client_name_aliases (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                company_id INTEGER NOT NULL,
                sales_name_key TEXT NOT NULL,
                sales_name TEXT NOT NULL,
                client_code TEXT NOT NULL,
                client_name TEXT NOT NULL,
                created_by_user_id INTEGER,
                created_at TEXT NOT NULL,
                UNIQUE(company_id, sales_name_key),
                FOREIGN KEY (company_id) REFERENCES companies(id)
            );

            CREATE TABLE IF NOT EXISTS territory_mappings (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                company_id INTEGER NOT NULL,
                city_name TEXT NOT NULL,
                neighborhood TEXT NOT NULL,
                unit_name TEXT NOT NULL,
                valid_from TEXT NOT NULL,
                valid_to TEXT,
                source TEXT NOT NULL,
                notes TEXT,
                created_at TEXT NOT NULL,
                UNIQUE(company_id, city_name, neighborhood, valid_from),
                FOREIGN KEY (company_id) REFERENCES companies(id)
            );

            CREATE INDEX IF NOT EXISTS idx_territory_lookup
                ON territory_mappings(company_id, city_name, neighborhood);

            CREATE TABLE IF NOT EXISTS client_registry (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                company_id INTEGER NOT NULL,
                client_name TEXT NOT NULL,
                normalized_client_name TEXT NOT NULL,
                document_number TEXT,
                document_digits TEXT,
                person_type TEXT NOT NULL,
                source TEXT NOT NULL,
                confidence_score REAL NOT NULL DEFAULT 1,
                notes TEXT,
                created_at TEXT NOT NULL,
                updated_at TEXT NOT NULL,
                UNIQUE(company_id, normalized_client_name),
                FOREIGN KEY (company_id) REFERENCES companies(id)
            );

            CREATE TABLE IF NOT EXISTS holidays (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                company_id INTEGER NOT NULL,
                holiday_date TEXT NOT NULL,
                holiday_name TEXT NOT NULL,
                scope TEXT NOT NULL DEFAULT 'NACIONAL_RS',
                created_at TEXT NOT NULL,
                UNIQUE(company_id, holiday_date, holiday_name),
                FOREIGN KEY (company_id) REFERENCES companies(id)
            );

            CREATE TABLE IF NOT EXISTS vacations (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                company_id INTEGER NOT NULL,
                person_name TEXT NOT NULL,
                start_date TEXT NOT NULL,
                end_date TEXT NOT NULL,
                notes TEXT,
                created_at TEXT NOT NULL,
                UNIQUE(company_id, person_name, start_date, end_date),
                FOREIGN KEY (company_id) REFERENCES companies(id)
            );

            CREATE TABLE IF NOT EXISTS score_configs (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                company_id INTEGER NOT NULL,
                valid_from_competence TEXT NOT NULL,
                valid_to_competence TEXT,
                weight_goal REAL NOT NULL,
                weight_ticket REAL NOT NULL,
                weight_clients REAL NOT NULL,
                weight_mix REAL NOT NULL,
                weight_returns REAL NOT NULL,
                created_at TEXT NOT NULL,
                UNIQUE(company_id, valid_from_competence),
                FOREIGN KEY (company_id) REFERENCES companies(id)
            );

            CREATE TABLE IF NOT EXISTS goals_seller (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                company_id INTEGER NOT NULL,
                competence TEXT NOT NULL,
                seller_name TEXT NOT NULL,
                base_unit TEXT,
                revenue_goal REAL NOT NULL DEFAULT 0,
                returns_goal REAL NOT NULL DEFAULT 0,
                created_at TEXT NOT NULL,
                UNIQUE(company_id, competence, seller_name),
                FOREIGN KEY (company_id) REFERENCES companies(id)
            );

            CREATE TABLE IF NOT EXISTS goals_unit (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                company_id INTEGER NOT NULL,
                competence TEXT NOT NULL,
                unit_name TEXT NOT NULL,
                revenue_goal REAL NOT NULL DEFAULT 0,
                returns_goal REAL NOT NULL DEFAULT 0,
                created_at TEXT NOT NULL,
                UNIQUE(company_id, competence, unit_name),
                FOREIGN KEY (company_id) REFERENCES companies(id)
            );


            CREATE TABLE IF NOT EXISTS imports (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                company_id INTEGER NOT NULL,
                competence TEXT NOT NULL,
                import_action TEXT NOT NULL,
                suggested_competence TEXT,
                imported_by INTEGER,
                imported_at TEXT NOT NULL,
                duplicate_rows_skipped INTEGER NOT NULL DEFAULT 0,
                notes TEXT,
                FOREIGN KEY (company_id) REFERENCES companies(id),
                FOREIGN KEY (imported_by) REFERENCES users(id)
            );

            CREATE TABLE IF NOT EXISTS import_files (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                import_id INTEGER NOT NULL,
                file_type TEXT NOT NULL,
                original_name TEXT NOT NULL,
                file_hash TEXT NOT NULL,
                row_count INTEGER NOT NULL DEFAULT 0,
                UNIQUE(import_id, file_type),
                FOREIGN KEY (import_id) REFERENCES imports(id)
            );

            CREATE TABLE IF NOT EXISTS import_issues (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                company_id INTEGER NOT NULL,
                import_id INTEGER,
                competence TEXT NOT NULL,
                issue_type TEXT NOT NULL,
                reference_value TEXT NOT NULL,
                status TEXT NOT NULL DEFAULT 'pendente',
                details_json TEXT,
                created_at TEXT NOT NULL,
                FOREIGN KEY (company_id) REFERENCES companies(id),
                FOREIGN KEY (import_id) REFERENCES imports(id)
            );

            CREATE TABLE IF NOT EXISTS fact_sales_detail (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                company_id INTEGER NOT NULL,
                competence TEXT NOT NULL,
                import_id INTEGER NOT NULL,
                row_hash TEXT NOT NULL,
                seller_name TEXT NOT NULL,
                client_name TEXT NOT NULL,
                city_name TEXT,
                gtin_value TEXT,
                manufacturer_sku TEXT,
                sku_key TEXT,
                issue_date TEXT,
                quantity REAL NOT NULL DEFAULT 0,
                gross_value REAL NOT NULL DEFAULT 0,
                discount_value REAL NOT NULL DEFAULT 0,
                freight_value REAL NOT NULL DEFAULT 0,
                return_quantity REAL NOT NULL DEFAULT 0,
                return_value REAL NOT NULL DEFAULT 0,
                net_value REAL NOT NULL DEFAULT 0,
                sale_share REAL NOT NULL DEFAULT 0,
                created_at TEXT NOT NULL,
                UNIQUE(company_id, competence, row_hash),
                FOREIGN KEY (company_id) REFERENCES companies(id),
                FOREIGN KEY (import_id) REFERENCES imports(id)
            );

            CREATE TABLE IF NOT EXISTS fact_vendor_summary (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                company_id INTEGER NOT NULL,
                competence TEXT NOT NULL,
                import_id INTEGER NOT NULL,
                row_hash TEXT NOT NULL,
                seller_name TEXT NOT NULL,
                qty_sold REAL NOT NULL DEFAULT 0,
                cost_value REAL NOT NULL DEFAULT 0,
                sale_value REAL NOT NULL DEFAULT 0,
                profit_value REAL NOT NULL DEFAULT 0,
                net_profit_value REAL NOT NULL DEFAULT 0,
                profit_pct REAL NOT NULL DEFAULT 0,
                return_cost REAL NOT NULL DEFAULT 0,
                return_value REAL NOT NULL DEFAULT 0,
                net_value REAL NOT NULL DEFAULT 0,
                margin_value REAL NOT NULL DEFAULT 0,
                created_at TEXT NOT NULL,
                UNIQUE(company_id, competence, row_hash),
                FOREIGN KEY (company_id) REFERENCES companies(id),
                FOREIGN KEY (import_id) REFERENCES imports(id)
            );

            CREATE TABLE IF NOT EXISTS fact_unit_summary (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                company_id INTEGER NOT NULL,
                competence TEXT NOT NULL,
                import_id INTEGER NOT NULL,
                row_hash TEXT NOT NULL,
                unit_name TEXT NOT NULL,
                qty_sold REAL NOT NULL DEFAULT 0,
                cost_value REAL NOT NULL DEFAULT 0,
                sale_value REAL NOT NULL DEFAULT 0,
                profit_value REAL NOT NULL DEFAULT 0,
                net_profit_value REAL NOT NULL DEFAULT 0,
                profit_pct REAL NOT NULL DEFAULT 0,
                return_cost REAL NOT NULL DEFAULT 0,
                return_value REAL NOT NULL DEFAULT 0,
                net_value REAL NOT NULL DEFAULT 0,
                margin_value REAL NOT NULL DEFAULT 0,
                created_at TEXT NOT NULL,
                UNIQUE(company_id, competence, row_hash),
                FOREIGN KEY (company_id) REFERENCES companies(id),
                FOREIGN KEY (import_id) REFERENCES imports(id)
            );

            CREATE TABLE IF NOT EXISTS audit_logs (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                company_id INTEGER NOT NULL,
                user_id INTEGER,
                action TEXT NOT NULL,
                entity_type TEXT NOT NULL,
                entity_id TEXT NOT NULL,
                changes_json TEXT NOT NULL,
                created_at TEXT NOT NULL,
                FOREIGN KEY (company_id) REFERENCES companies(id),
                FOREIGN KEY (user_id) REFERENCES users(id)
            );

            CREATE INDEX IF NOT EXISTS idx_sales_company_competence
                ON fact_sales_detail(company_id, competence);

            CREATE INDEX IF NOT EXISTS idx_sales_company_seller_competence_client
                ON fact_sales_detail(company_id, seller_name, competence, client_name);

            CREATE INDEX IF NOT EXISTS idx_sales_company_client_competence
                ON fact_sales_detail(company_id, client_name, competence);

            CREATE INDEX IF NOT EXISTS idx_sales_company_seller_issue
                ON fact_sales_detail(company_id, seller_name, issue_date);

            CREATE INDEX IF NOT EXISTS idx_vendor_summary_company_competence_seller
                ON fact_vendor_summary(company_id, competence, seller_name);

            CREATE INDEX IF NOT EXISTS idx_unit_summary_company_competence_unit
                ON fact_unit_summary(company_id, competence, unit_name);
            """
        )
        init_crm_schema(conn)
        # Migração: Porto Alegre → Zona Sul (renomeação de unidade)
        _OLD_UNIT = "PORTOALEGRE"
        _NEW_UNIT = "ZONA SUL"
        try:
            conn.execute("UPDATE fact_sales_detail   SET unit_name  = ? WHERE unit_name  = ?", (_NEW_UNIT, _OLD_UNIT))
            conn.execute("UPDATE fact_vendor_summary SET unit_name  = ? WHERE unit_name  = ?", (_NEW_UNIT, _OLD_UNIT))
            conn.execute("UPDATE fact_unit_summary   SET unit_name  = ? WHERE unit_name  = ?", (_NEW_UNIT, _OLD_UNIT))
            conn.execute("UPDATE people_records      SET base_unit  = ? WHERE base_unit  = ?", (_NEW_UNIT, _OLD_UNIT))
            conn.execute("UPDATE goals_seller        SET unit_name  = ? WHERE unit_name  = ?", (_NEW_UNIT, _OLD_UNIT))
            conn.execute("UPDATE goals_unit          SET unit_name  = ? WHERE unit_name  = ?", (_NEW_UNIT, _OLD_UNIT))
            conn.execute("UPDATE city_unit_mapping   SET principal_unit = ? WHERE principal_unit = ?", (_NEW_UNIT, _OLD_UNIT))
            conn.commit()
        except sqlite3.OperationalError:
            pass  # tabela ainda não existe em ambientes novos
        conn.executescript("""
            CREATE TABLE IF NOT EXISTS auto_import_log (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                ran_at TEXT NOT NULL,
                folder TEXT NOT NULL,
                scope TEXT NOT NULL,
                competence TEXT,
                status TEXT NOT NULL,
                message TEXT,
                files_json TEXT
            );
        """)
        user_columns = {row["name"] for row in conn.execute("PRAGMA table_info(users)").fetchall()}
        if "linked_person_name" not in user_columns:
            conn.execute("ALTER TABLE users ADD COLUMN linked_person_name TEXT")
            conn.commit()
        if "linked_units_json" not in user_columns:
            conn.execute("ALTER TABLE users ADD COLUMN linked_units_json TEXT")
            conn.commit()
        if "profile_id" not in user_columns:
            conn.execute("ALTER TABLE users ADD COLUMN profile_id INTEGER")
            conn.commit()
        try:
            conn.execute("UPDATE fact_vendor_summary SET margin_value = margin_value / 100 WHERE margin_value > 100")
            conn.execute("UPDATE fact_unit_summary SET margin_value = margin_value / 100 WHERE margin_value > 100")
        except sqlite3.OperationalError:
            # Some local test environments mount the sqlite file read-only.
            pass
        conn.commit()
        company = conn.execute("SELECT id FROM companies WHERE name = ?", (DEFAULT_COMPANY,)).fetchone()
        if not company:
            conn.execute("INSERT INTO companies (name, code, created_at) VALUES (?, ?, ?)", (DEFAULT_COMPANY, "PASSINI", now_iso()))
            conn.commit()
            company = conn.execute("SELECT id FROM companies WHERE name = ?", (DEFAULT_COMPANY,)).fetchone()
        company_id = company["id"]

        seed_access_profiles(conn, company_id)
        seed_content_library(conn, company_id)
        seed_kpi_thresholds(conn, company_id)
        backfill_meeting_participant_users(conn, company_id)
        seed_help_content(conn, company_id)
        seed_territory_mappings(conn, company_id)
        repair_prospect_keys(conn, company_id)

        user = conn.execute("SELECT id FROM users WHERE username = ?", (DEFAULT_ADMIN_USER,)).fetchone()
        if not user:
            pwd_hash, salt = pbkdf2_hash(DEFAULT_ADMIN_PASSWORD)
            conn.execute(
                """
                INSERT INTO users (company_id, username, full_name, linked_person_name, password_hash, password_salt, role, created_at)
                VALUES (?, ?, ?, ?, ?, ?, ?, ?)
                """,
                (company_id, DEFAULT_ADMIN_USER, "Administrador padrão", None, pwd_hash, salt, "Administrador", now_iso()),
            )
            conn.commit()

        score = conn.execute("SELECT id FROM score_configs WHERE company_id = ?", (company_id,)).fetchone()
        if not score:
            conn.execute(
                """
                INSERT INTO score_configs (
                    company_id, valid_from_competence, valid_to_competence,
                    weight_goal, weight_ticket, weight_clients, weight_mix, weight_returns, created_at
                ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)
                """,
                (company_id, "2025-01", None, 30, 15, 15, 15, 25, now_iso()),
            )
            conn.commit()

        seed_holidays(conn, company_id)
        seed_mapping_workbook(conn, company_id)
        seed_crm_catalogs(conn)
        sanitize_unit_goals(conn, company_id)
        conn.commit()


def seed_holidays(conn: sqlite3.Connection, company_id: int) -> None:
    existing = conn.execute("SELECT COUNT(*) AS total FROM holidays WHERE company_id = ?", (company_id,)).fetchone()["total"]
    if existing:
        return
    for year, rows in NATIONAL_AND_RS_HOLIDAYS.items():
        for holiday_date, holiday_name in rows:
            conn.execute(
                """
                INSERT OR IGNORE INTO holidays (company_id, holiday_date, holiday_name, scope, created_at)
                VALUES (?, ?, ?, 'NACIONAL_RS', ?)
                """,
                (company_id, holiday_date, holiday_name, now_iso()),
            )
    conn.commit()


def infer_role_from_name(person_name: str) -> str:
    upper = normalize_upper(person_name)
    if "(VENDAS)" in upper or "(TELEVENDAS)" in upper or "(VAREJO)" in upper:
        return "Vendedor"
    if "GERENTE" in upper:
        return "Gerente"
    return "Outro"


def seed_mapping_workbook(conn: sqlite3.Connection, company_id: int) -> None:
    if not PASSINI_MAPPING_WORKBOOK.exists():
        return
    people_count = conn.execute("SELECT COUNT(*) AS total FROM people_records WHERE company_id = ?", (company_id,)).fetchone()["total"]
    city_count = conn.execute("SELECT COUNT(*) AS total FROM city_mappings WHERE company_id = ?", (company_id,)).fetchone()["total"]
    if people_count and city_count:
        return
    workbook = openpyxl.load_workbook(PASSINI_MAPPING_WORKBOOK, data_only=True)
    if not people_count and "VENDEDOR X UNIDADE" in workbook.sheetnames:
        sheet = workbook["VENDEDOR X UNIDADE"]
        for row in sheet.iter_rows(min_row=2, values_only=True):
            person_name = normalize_whitespace(row[1] if len(row) > 1 else None)
            if not person_name:
                continue
            base_unit = normalize_unit(row[3] if len(row) > 3 and row[3] else row[2] if len(row) > 2 else "")
            conn.execute(
                """
                INSERT OR IGNORE INTO people_records
                    (company_id, person_name, role_classification, base_unit, valid_from, valid_to, source, created_at)
                VALUES (?, ?, ?, ?, ?, ?, ?, ?)
                """,
                (company_id, person_name, infer_role_from_name(person_name), base_unit, "2025-01-01", None, "planilha_apoio", now_iso()),
            )
    if not city_count and "cidade x unidade" in workbook.sheetnames:
        sheet = workbook["cidade x unidade"]
        for row in sheet.iter_rows(min_row=3, values_only=True):
            city_name = normalize_upper(row[0] if len(row) > 0 else None)
            if not city_name:
                continue
            principal_unit = normalize_unit(row[1] if len(row) > 1 else None)
            state_name = normalize_upper(row[2] if len(row) > 2 else None)
            country_name = normalize_upper(row[3] if len(row) > 3 else None)
            conn.execute(
                """
                INSERT OR IGNORE INTO city_mappings
                    (company_id, city_name, principal_unit, state_name, country_name, valid_from, valid_to, source, created_at)
                VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)
                """,
                (company_id, city_name, principal_unit, state_name, country_name, "2025-01-01", None, "planilha_apoio", now_iso()),
            )
    conn.commit()


def get_company_id(conn: sqlite3.Connection) -> int:
    return conn.execute("SELECT id FROM companies WHERE name = ?", (DEFAULT_COMPANY,)).fetchone()["id"]


def current_role_and_unit(conn: sqlite3.Connection, company_id: int, person_name: str, competence: str | None = None) -> tuple[str | None, str | None]:
    competence = competence or date.today().strftime("%Y-%m")
    target = first_day_of_competence(competence).isoformat()
    row = conn.execute(
        """
        SELECT role_classification, base_unit
        FROM people_records
        WHERE company_id = ? AND person_name = ? AND date(valid_from) <= date(?)
          AND (valid_to IS NULL OR date(valid_to) >= date(?))
        ORDER BY date(valid_from) DESC
        LIMIT 1
        """,
        (company_id, person_name, target, target),
    ).fetchone()
    if row:
        return row["role_classification"], row["base_unit"]
    return None, None


# ─── Territórios: bairro decide, cidade é o segundo nível ───────────────────

TERRITORIO_COMPARTILHADO = "COMPARTILHADA"
TERRITORIO_CIDADE_INTEIRA = "*"


def territory_default_valid_from() -> str:
    try:
        import territorio_seed as semente
        return semente.TERRITORIO_VIGENCIA
    except (ModuleNotFoundError, AttributeError):
        return today_in_brazil().isoformat()


def seed_territory_mappings(conn: sqlite3.Connection, company_id: int) -> None:
    """Carrega o mapa de territórios aprovado, sem sobrescrever ajuste manual.

    Roda a cada boot. O INSERT OR IGNORE somado ao UNIQUE(cidade, bairro,
    vigência) garante que uma linha já editada pela tela de Administração não
    volte ao valor original — a semente preenche o que falta e nada mais.
    """
    try:
        import territorio_seed as semente
    except ModuleNotFoundError:
        print("[territorio] territorio_seed.py não encontrado — mapa não carregado")
        return

    vigencia = semente.TERRITORIO_VIGENCIA
    agora = now_iso()
    linhas: list[tuple[Any, ...]] = []

    for bairro, unidade in semente.BAIRROS_POA.items():
        linhas.append((company_id, "PORTO ALEGRE", normalize_upper(strip_accents(bairro)),
                       unidade, vigencia, None, "semente_2026_09", None, agora))
    for cidade, unidade in semente.CIDADES_EXCLUSIVAS.items():
        linhas.append((company_id, normalize_upper(strip_accents(cidade)), TERRITORIO_CIDADE_INTEIRA,
                       unidade, vigencia, None, "semente_2026_09",
                       "Cidade inteira da unidade", agora))
    for cidade in semente.CIDADES_COMPARTILHADAS:
        linhas.append((company_id, normalize_upper(strip_accents(cidade)), TERRITORIO_CIDADE_INTEIRA,
                       TERRITORIO_COMPARTILHADO, vigencia, None, "semente_2026_09",
                       "Sem mapa de propósito: a unidade vem do vendedor que atende", agora))

    # ORDEM IMPORTA. A antecipação de vigência roda ANTES do INSERT: senão o
    # INSERT não veria conflito (a chave única inclui a vigência), criaria uma
    # segunda linha para o mesmo bairro e o mapa passaria a ter duas respostas.
    movidas = conn.execute(
        """
        UPDATE territory_mappings SET valid_from = ?
        WHERE company_id = ? AND source = 'semente_2026_09' AND valid_from <> ?
          AND NOT EXISTS (
              SELECT 1 FROM territory_mappings t2
              WHERE t2.company_id = territory_mappings.company_id
                AND t2.city_name = territory_mappings.city_name
                AND t2.neighborhood = territory_mappings.neighborhood
                AND t2.valid_from = ?)
        """,
        (vigencia, company_id, vigencia, vigencia),
    ).rowcount
    if movidas:
        print(f"[territorio] {movidas} território(s) passaram a valer desde {vigencia}")

    # Insere só o que não existe em NENHUMA vigência: linha já ajustada à mão
    # na tela de Territórios não pode ser sombreada por uma cópia da semente.
    novos = 0
    for linha in linhas:
        cursor = conn.execute(
            """
            INSERT INTO territory_mappings
                (company_id, city_name, neighborhood, unit_name, valid_from, valid_to,
                 source, notes, created_at)
            SELECT ?,?,?,?,?,?,?,?,?
            WHERE NOT EXISTS (
                SELECT 1 FROM territory_mappings
                WHERE company_id = ? AND city_name = ? AND neighborhood = ?)
            """,
            (*linha, linha[0], linha[1], linha[2]),
        )
        novos += cursor.rowcount
    if novos:
        print(f"[territorio] {novos} território(s) carregado(s)")
    conn.commit()


def resolve_territory_unit(
    conn: sqlite3.Connection, company_id: int,
    city_name: str | None, neighborhood: str | None,
    competence: str | None = None,
) -> str | None:
    """Unidade dona do território. None = decide o vendedor que atende.

    Cascata, do mais específico para o mais genérico:
      1. cidade + bairro   (Porto Alegre, dividida entre Zona Norte e Sul)
      2. cidade inteira    (Gravataí, Alvorada, Cachoeirinha → Zona Norte)
      3. mapa antigo de cidades (city_mappings), que segue valendo no interior
      4. None

    Uma cidade marcada COMPARTILHADA devolve None de propósito: Canoas e Viamão
    são atendidas pelas duas unidades e quem manda é o vendedor com histórico,
    não o mapa. Devolver uma unidade ali tiraria carteira de quem já atende.
    """
    cidade = normalize_upper(strip_accents(city_name))
    if not cidade:
        return None
    competence = competence or date.today().strftime("%Y-%m")
    alvo = first_day_of_competence(competence).isoformat()
    bairro = normalize_upper(strip_accents(neighborhood))

    for chave in ([bairro] if bairro else []) + [TERRITORIO_CIDADE_INTEIRA]:
        row = conn.execute(
            """
            SELECT unit_name FROM territory_mappings
            WHERE company_id = ? AND city_name = ? AND neighborhood = ?
              AND date(valid_from) <= date(?)
              AND (valid_to IS NULL OR valid_to = '' OR date(valid_to) >= date(?))
            ORDER BY date(valid_from) DESC LIMIT 1
            """,
            (company_id, cidade, chave, alvo, alvo),
        ).fetchone()
        if row:
            unidade = normalize_unit(row["unit_name"])
            return None if unidade == TERRITORIO_COMPARTILHADO else unidade

    return resolve_city_unit(conn, company_id, city_name, competence)


def territory_for_client(
    conn: sqlite3.Connection, company_id: int, client_code: str,
    competence: str | None = None,
) -> dict[str, Any]:
    """Território do cliente já pronto para a tela, com o motivo da decisão."""
    perfil = conn.execute(
        "SELECT city_name, neighborhood FROM crm_client_profiles "
        "WHERE company_id = ? AND client_code = ?",
        (company_id, client_code),
    ).fetchone()
    if not perfil:
        return {"unit": None, "city": "", "neighborhood": "", "reason": "cliente sem cadastro"}
    cidade = normalize_whitespace(perfil["city_name"])
    bairro = normalize_whitespace(perfil["neighborhood"])
    unidade = resolve_territory_unit(conn, company_id, cidade, bairro, competence)
    if unidade:
        motivo = f"bairro {bairro}" if bairro and territory_has_neighborhood(
            conn, company_id, cidade, bairro) else f"cidade {cidade}"
    elif territory_city_is_shared(conn, company_id, cidade):
        motivo = "cidade compartilhada — vale o vendedor que atende"
    elif bairro:
        # Diferente de compartilhado: aqui ninguém decidiu ainda. Aparece no
        # relatório "bairros sem dono" para o mapa não envelhecer calado.
        motivo = f"bairro {bairro} ainda sem dono no mapa"
    else:
        motivo = "cliente sem bairro no cadastro"
    return {"unit": unidade, "city": cidade, "neighborhood": bairro, "reason": motivo}


def territory_city_is_shared(
    conn: sqlite3.Connection, company_id: int, city_name: str
) -> bool:
    row = conn.execute(
        "SELECT unit_name FROM territory_mappings WHERE company_id = ? AND city_name = ? "
        "AND neighborhood = ?",
        (company_id, normalize_upper(strip_accents(city_name)), TERRITORIO_CIDADE_INTEIRA),
    ).fetchone()
    return bool(row) and normalize_unit(row["unit_name"]) == TERRITORIO_COMPARTILHADO


def territory_has_neighborhood(
    conn: sqlite3.Connection, company_id: int, city_name: str, neighborhood: str
) -> bool:
    return bool(conn.execute(
        "SELECT 1 FROM territory_mappings WHERE company_id = ? AND city_name = ? AND neighborhood = ?",
        (company_id, normalize_upper(strip_accents(city_name)),
         normalize_upper(strip_accents(neighborhood))),
    ).fetchone())


def list_territory_mappings(
    conn: sqlite3.Connection, company_id: int, city: str = ""
) -> list[dict[str, Any]]:
    sql = ("SELECT id, city_name, neighborhood, unit_name, valid_from, valid_to, source, notes "
           "FROM territory_mappings WHERE company_id = ?")
    params: list[Any] = [company_id]
    if city:
        sql += " AND city_name = ?"
        params.append(normalize_upper(strip_accents(city)))
    sql += " ORDER BY city_name, (neighborhood = '*') DESC, neighborhood"
    return [dict(r) for r in conn.execute(sql, params).fetchall()]


def save_territory_mapping(
    conn: sqlite3.Connection, company_id: int, user_id: int, payload: dict[str, Any]
) -> dict[str, Any]:
    cidade = normalize_upper(strip_accents(payload.get("cityName")))
    if not cidade:
        raise ValueError("Informe a cidade.")
    bairro = normalize_upper(strip_accents(payload.get("neighborhood"))) or TERRITORIO_CIDADE_INTEIRA
    unidade = normalize_unit(payload.get("unitName"))
    if not unidade:
        raise ValueError("Informe a unidade.")
    if unidade != TERRITORIO_COMPARTILHADO and unidade not in CANONICAL_UNITS:
        raise ValueError(f"Unidade desconhecida: {unidade}")
    vigencia = normalize_whitespace(payload.get("validFrom")) or today_in_brazil().isoformat()
    registro_id = payload.get("id")

    if registro_id:
        conn.execute(
            "UPDATE territory_mappings SET city_name = ?, neighborhood = ?, unit_name = ?, "
            "valid_from = ?, valid_to = ?, notes = ? WHERE company_id = ? AND id = ?",
            (cidade, bairro, unidade, vigencia,
             normalize_whitespace(payload.get("validTo")) or None,
             normalize_whitespace(payload.get("notes")) or None, company_id, int(registro_id)),
        )
        acao = "atualizado"
    else:
        conn.execute(
            """
            INSERT INTO territory_mappings
                (company_id, city_name, neighborhood, unit_name, valid_from, valid_to,
                 source, notes, created_at)
            VALUES (?,?,?,?,?,?,'manual',?,?)
            ON CONFLICT(company_id, city_name, neighborhood, valid_from)
            DO UPDATE SET unit_name = excluded.unit_name, notes = excluded.notes, valid_to = NULL
            """,
            (company_id, cidade, bairro, unidade, vigencia,
             normalize_whitespace(payload.get("validTo")) or None,
             normalize_whitespace(payload.get("notes")) or None, now_iso()),
        )
        acao = "salvo"

    audit_log(conn, company_id, user_id, "salvar", "territory_mappings",
              f"{cidade}/{bairro}", {"unidade": unidade, "vigencia": vigencia})
    conn.commit()
    invalidate_crm_cache(company_id)
    rotulo = "toda a cidade" if bairro == TERRITORIO_CIDADE_INTEIRA else bairro
    return {"message": f"Território {acao}: {rotulo} ({cidade}) → {unidade}."}


def delete_territory_mapping(
    conn: sqlite3.Connection, company_id: int, user_id: int, registro_id: Any
) -> dict[str, Any]:
    conn.execute("DELETE FROM territory_mappings WHERE company_id = ? AND id = ?",
                 (company_id, int(registro_id)))
    audit_log(conn, company_id, user_id, "excluir", "territory_mappings", str(registro_id), {})
    conn.commit()
    invalidate_crm_cache(company_id)
    return {"message": "Território removido. A cidade volta a valer pela regra antiga."}


def territory_coverage_report(
    conn: sqlite3.Connection, company_id: int, competence: str | None = None
) -> dict[str, Any]:
    """Bairros com cliente na base que ainda não têm dono no mapa.

    É o relatório que impede o mapa de envelhecer em silêncio: bairro novo
    aparece no cadastro do Alfa e ninguém percebe que ele caiu na regra da
    cidade.
    """
    faltando: list[dict[str, Any]] = []
    for r in conn.execute(
        """
        SELECT UPPER(TRIM(city_name)) AS cidade, UPPER(TRIM(COALESCE(neighborhood,''))) AS bairro,
               COUNT(*) AS clientes
        FROM crm_client_profiles
        WHERE company_id = ? AND city_name IS NOT NULL AND TRIM(city_name) <> ''
        GROUP BY cidade, bairro
        ORDER BY clientes DESC
        """,
        (company_id,),
    ).fetchall():
        cidade = normalize_upper(strip_accents(r["cidade"]))
        bairro = normalize_upper(strip_accents(r["bairro"]))
        if not bairro:
            continue
        tem_cidade = bool(conn.execute(
            "SELECT 1 FROM territory_mappings WHERE company_id = ? AND city_name = ?",
            (company_id, cidade)).fetchone())
        if not tem_cidade:
            continue  # cidade fora do mapa novo: segue na regra antiga, tudo bem
        if territory_city_is_shared(conn, company_id, cidade):
            continue  # Canoas e Viamão são compartilhadas de propósito, não é lacuna
        if not territory_has_neighborhood(conn, company_id, cidade, bairro):
            faltando.append({"city": cidade, "neighborhood": bairro, "clients": r["clientes"]})
    return {"missing": faltando[:200], "total": len(faltando)}


# Cidade não muda de unidade com o tempo: a loja que atende Gravataí hoje é a
# que atendia em 2024. Vincular o mapa a uma competência só criava buraco no
# passado — o mês anterior ficava sem unidade e voltava como pendência.
CITY_MAPPING_EPOCH = "2000-01-01"


def map_city_to_unit(
    conn: sqlite3.Connection, company_id: int, city_name: str, unit_name: str,
    state_name: str | None = None, country_name: str | None = None,
) -> None:
    """Aponta a cidade para a unidade em TODOS os períodos, passados inclusive.

    Apaga as faixas antigas daquela cidade antes de gravar. Sem isso sobrava
    uma linha com vigência recente competindo com a nova, e o resultado
    dependia da ordem de leitura.
    """
    cidade = normalize_upper(city_name)
    unidade = normalize_unit(unit_name)
    if not cidade or not unidade:
        raise ValueError("Informe cidade e unidade.")
    conn.execute("DELETE FROM city_mappings WHERE company_id = ? AND city_name = ?",
                 (company_id, cidade))
    conn.execute(
        """
        INSERT INTO city_mappings
            (company_id, city_name, principal_unit, state_name, country_name,
             valid_from, valid_to, source, created_at)
        VALUES (?, ?, ?, ?, ?, ?, NULL, 'resolucao_pendencia', ?)
        """,
        (company_id, cidade, unidade, state_name, country_name, CITY_MAPPING_EPOCH, now_iso()),
    )


def resolve_city_issues_bulk(
    conn: sqlite3.Connection, company_id: int, user_id: int, payload: dict[str, Any]
) -> dict[str, Any]:
    """Direciona várias cidades pendentes para a mesma unidade de uma vez.

    Depois de uma importação grande a lista de cidades novas vem com dezenas de
    nomes, quase todos da mesma região. Resolver uma a uma, com três campos
    cada, é o tipo de trabalho que faz a pendência ser ignorada — e cidade sem
    unidade tira o faturamento do painel.
    """
    unidade = normalize_unit(payload.get("unitName"))
    if not unidade:
        raise ValueError("Escolha a unidade.")
    if unidade not in CANONICAL_UNITS:
        raise ValueError(f"Unidade desconhecida: {unidade}")
    cidades = [normalize_upper(c) for c in (payload.get("cities") or []) if normalize_upper(c)]
    if not cidades:
        raise ValueError("Selecione ao menos uma cidade.")

    for cidade in cidades:
        map_city_to_unit(conn, company_id, cidade, unidade)
        conn.execute(
            """
            UPDATE import_issues SET status = 'resolvida'
            WHERE company_id = ? AND issue_type = 'cidade_sem_correspondencia'
              AND reference_value = ? AND status = 'pendente'
            """,
            (company_id, cidade),
        )
    audit_log(conn, company_id, user_id, "resolver_lote", "import_issue",
              ", ".join(cidades[:10]), {"unidade": unidade, "cidades": len(cidades)})
    conn.commit()
    invalidate_crm_cache(company_id)
    return {"message": f"{len(cidades)} cidade(s) direcionada(s) para {unidade}, "
                       f"valendo para todos os períodos.",
            "resolved": len(cidades), "unitName": unidade}


def resolve_city_unit(conn: sqlite3.Connection, company_id: int, city_name: str | None, competence: str | None = None) -> str | None:
    normalized_city = normalize_upper(city_name)
    if not normalized_city:
        return None
    competence = competence or date.today().strftime("%Y-%m")
    target = first_day_of_competence(competence).isoformat()
    row = conn.execute(
        """
        SELECT principal_unit
        FROM city_mappings
        WHERE company_id = ? AND city_name = ? AND date(valid_from) <= date(?)
          AND (valid_to IS NULL OR date(valid_to) >= date(?))
        ORDER BY date(valid_from) DESC
        LIMIT 1
        """,
        (company_id, normalized_city, target, target),
    ).fetchone()
    if row:
        return normalize_unit(row["principal_unit"])
    return None


def build_seller_unit_map(
    conn: sqlite3.Connection, company_id: int, competence: str | None = None
) -> dict[str, str]:
    """Vendedor → unidade base, na vigência da competência. Uma query só.

    Indexado pela chave normalizada do nome, porque o cadastro de clientes
    escreve "CLEBER ALVES OLIVEIRA (VENDAS)" e o cadastro de pessoas às vezes
    escreve só "CLEBER ALVES OLIVEIRA".
    """
    competence = competence or date.today().strftime("%Y-%m")
    alvo = first_day_of_competence(competence).isoformat()
    mapa: dict[str, str] = {}
    for row in conn.execute(
        """
        SELECT person_name, base_unit FROM people_records
        WHERE company_id = ? AND base_unit IS NOT NULL AND TRIM(base_unit) <> ''
          AND date(valid_from) <= date(?)
          AND (valid_to IS NULL OR valid_to = '' OR date(valid_to) >= date(?))
        ORDER BY date(valid_from) DESC
        """,
        (company_id, alvo, alvo),
    ).fetchall():
        for chave in (person_key(row["person_name"]), short_person_key(row["person_name"])):
            if chave and chave not in mapa:
                mapa[chave] = normalize_unit(row["base_unit"])
    return mapa


def build_territory_map(
    conn: sqlite3.Connection, company_id: int, competence: str | None = None
) -> dict[tuple[str, str], str]:
    """(cidade, bairro) → unidade, em um dict, para não consultar linha a linha."""
    competence = competence or date.today().strftime("%Y-%m")
    alvo = first_day_of_competence(competence).isoformat()
    mapa: dict[tuple[str, str], str] = {}
    for row in conn.execute(
        """
        SELECT city_name, neighborhood, unit_name FROM territory_mappings
        WHERE company_id = ? AND date(valid_from) <= date(?)
          AND (valid_to IS NULL OR valid_to = '' OR date(valid_to) >= date(?))
        ORDER BY date(valid_from) DESC
        """,
        (company_id, alvo, alvo),
    ).fetchall():
        chave = (normalize_upper(strip_accents(row["city_name"])),
                 normalize_upper(strip_accents(row["neighborhood"])))
        if chave not in mapa:
            mapa[chave] = normalize_unit(row["unit_name"])
    return mapa


def unit_for_client_row(
    seller_name: str | None, city_name: str | None, neighborhood: str | None,
    seller_units: dict[str, str], territory: dict[tuple[str, str], str],
    city_units: dict[str, str | None],
) -> str | None:
    """Unidade dona do cliente na carteira, em cascata.

      1. UNIDADE DO VENDEDOR que atende. É a regra combinada — a venda conta
         para a unidade de quem vendeu, então a carteira tem de mostrar o mesmo.
         Antes a unidade vinha só da cidade: um vendedor da Zona Norte com
         cliente em Canoas aparecia como Matriz, e a carteira dele sumia da
         unidade nova.
      2. Sem vendedor, vale o TERRITÓRIO: bairro, depois cidade inteira. É o que
         faz o cliente de um bairro da Zona Norte, ainda sem dono, aparecer na
         carteira da Zona Norte esperando alguém assumir.
      3. Por último, o mapa antigo de cidades, que segue valendo no interior.
    """
    vendedor = normalize_whitespace(seller_name)
    if vendedor:
        unidade = (seller_units.get(person_key(vendedor))
                   or seller_units.get(short_person_key(vendedor)))
        if unidade:
            return unidade

    cidade = normalize_upper(strip_accents(city_name))
    bairro = normalize_upper(strip_accents(neighborhood))
    if cidade:
        for chave in ([(cidade, bairro)] if bairro else []) + [(cidade, TERRITORIO_CIDADE_INTEIRA)]:
            unidade = territory.get(chave)
            if unidade and unidade != TERRITORIO_COMPARTILHADO:
                return unidade
            if unidade == TERRITORIO_COMPARTILHADO:
                break  # cidade compartilhada e sem vendedor: cai no mapa antigo

    return city_units.get(normalize_upper(city_name))


def build_city_unit_map(conn: sqlite3.Connection, company_id: int, competence: str | None = None) -> dict[str, str | None]:
    """Precarrega todos os mapeamentos cidade→unidade em um dict para evitar N+1 queries."""
    competence = competence or date.today().strftime("%Y-%m")
    target = first_day_of_competence(competence).isoformat()
    rows = conn.execute(
        """
        SELECT city_name, principal_unit
        FROM city_mappings
        WHERE company_id = ? AND date(valid_from) <= date(?)
          AND (valid_to IS NULL OR date(valid_to) >= date(?))
        ORDER BY city_name, date(valid_from) DESC
        """,
        (company_id, target, target),
    ).fetchall()
    # Mantém apenas o mapeamento mais recente por cidade (ORDER BY garante isso)
    city_map: dict[str, str | None] = {}
    for row in rows:
        city = normalize_upper(row["city_name"])
        if city and city not in city_map:
            city_map[city] = normalize_unit(row["principal_unit"])
    return city_map


def projection_metrics(realized_value: float, elapsed_days: int, total_days: int) -> tuple[float, float]:
    if elapsed_days <= 0 or total_days <= 0:
        return realized_value, 0.0
    projected = realized_value / elapsed_days * total_days
    pace_pct = projected / realized_value * 100 if realized_value else 0.0
    return projected, pace_pct


def dashboard_metric_projection(realized_value: float, elapsed_days: int, total_days: int) -> tuple[float, float]:
    if elapsed_days <= 0 or total_days <= 0:
        return 0.0, 0.0
    daily_actual = safe_div(realized_value, elapsed_days)
    projected = daily_actual * total_days
    return daily_actual, projected


def load_goal_maps(
    conn: sqlite3.Connection,
    company_id: int,
    competence: str,
    table_name: str,
    key_field: str,
    normalizer,
) -> tuple[dict[str, dict[str, Any]], list[dict[str, Any]]]:
    rows = conn.execute(
        f"""
        SELECT id, competence, {key_field} AS goal_key, revenue_goal, returns_goal
        FROM {table_name}
        WHERE company_id = ? AND competence = ?
        ORDER BY id DESC
        """,
        (company_id, competence),
    ).fetchall()
    latest_by_raw: dict[str, sqlite3.Row] = {}
    duplicate_raw: list[dict[str, Any]] = []
    raw_counts: dict[str, int] = defaultdict(int)
    raw_totals: dict[str, float] = defaultdict(float)
    for row in rows:
        raw_key = normalize_whitespace(row["goal_key"])
        raw_counts[raw_key] += 1
        raw_totals[raw_key] += float(row["revenue_goal"] or 0.0)
        if raw_key not in latest_by_raw:
            latest_by_raw[raw_key] = row
    for raw_key, count in raw_counts.items():
        if count > 1:
            duplicate_raw.append(
                {
                    "competence": competence,
                    "key": raw_key,
                    "count": count,
                    "summedRevenueGoal": round(raw_totals[raw_key], 2),
                    "table": table_name,
                }
            )
    aggregated: dict[str, dict[str, Any]] = {}
    for raw_key, row in latest_by_raw.items():
        normalized_key = normalizer(raw_key)
        if not normalized_key:
            continue
        bucket = aggregated.setdefault(
            normalized_key,
            {
                "revenueGoal": 0.0,
                "returnsGoal": 0.0,
                "sourceKeys": [],
            },
        )
        bucket["revenueGoal"] += float(row["revenue_goal"] or 0.0)
        bucket["returnsGoal"] += float(row["returns_goal"] or 0.0)
        bucket["sourceKeys"].append(raw_key)
    for bucket in aggregated.values():
        bucket["revenueGoal"] = round(bucket["revenueGoal"], 2)
        bucket["returnsGoal"] = round(bucket["returnsGoal"], 2)
        bucket["sourceKeys"].sort()
    return aggregated, duplicate_raw


def normalized_goal_duplicates(
    conn: sqlite3.Connection,
    company_id: int,
    table_name: str,
    key_field: str,
    normalizer,
    competence: str | None = None,
) -> list[dict[str, Any]]:
    sql = f"""
        SELECT id, competence, {key_field} AS goal_key, revenue_goal, created_at
        FROM {table_name}
        WHERE company_id = ?
    """
    params: list[Any] = [company_id]
    if competence:
        sql += " AND competence = ?"
        params.append(competence)
    sql += " ORDER BY competence DESC, id DESC"
    rows = conn.execute(sql, params).fetchall()
    grouped: dict[tuple[str, str], list[dict[str, Any]]] = defaultdict(list)
    for row in rows:
        raw_key = normalize_whitespace(row["goal_key"])
        normalized_key = normalizer(raw_key)
        if not normalized_key:
            continue
        grouped[(row["competence"], normalized_key)].append(
            {
                "id": row["id"],
                "goalKey": raw_key,
                "revenueGoal": audit_round(row["revenue_goal"]),
                "createdAt": row["created_at"],
            }
        )
    duplicates: list[dict[str, Any]] = []
    for (row_competence, normalized_key), duplicate_rows in grouped.items():
        if len(duplicate_rows) <= 1:
            continue
        duplicates.append(
            {
                "competence": row_competence,
                "normalizedKey": normalized_key,
                "rows": duplicate_rows,
            }
        )
    duplicates.sort(key=lambda item: (item["competence"], item["normalizedKey"]), reverse=True)
    return duplicates


def sanitize_unit_goals(conn: sqlite3.Connection, company_id: int, actor_user_id: int | None = None) -> list[dict[str, Any]]:
    rows = conn.execute(
        """
        SELECT id, competence, unit_name, revenue_goal, returns_goal, created_at
        FROM goals_unit
        WHERE company_id = ?
        ORDER BY competence DESC, datetime(created_at) DESC, id DESC
        """,
        (company_id,),
    ).fetchall()
    grouped: dict[tuple[str, str], list[sqlite3.Row]] = defaultdict(list)
    for row in rows:
        grouped[(row["competence"], normalize_unit(row["unit_name"]))].append(row)

    changes_made: list[dict[str, Any]] = []
    for (competence, normalized_unit), group_rows in grouped.items():
        if not normalized_unit:
            continue
        distinct_values = {
            (audit_round(row["revenue_goal"]), audit_round(row["returns_goal"]))
            for row in group_rows
        }
        prefer_latest = len(distinct_values) > 1
        sorted_rows = sorted(
            group_rows,
            key=lambda row: (
                row["created_at"] or "",
                row["id"],
            )
            if prefer_latest
            else (
                1 if normalize_whitespace(row["unit_name"]) == normalized_unit else 0,
                row["created_at"] or "",
                row["id"],
            ),
            reverse=True,
        )
        kept_row = sorted_rows[0]
        deleted_rows = sorted_rows[1:]
        needs_update = normalize_whitespace(kept_row["unit_name"]) != normalized_unit
        if not deleted_rows and not needs_update:
            continue
        if needs_update:
            conn.execute(
                "UPDATE goals_unit SET unit_name = ? WHERE id = ?",
                (normalized_unit, kept_row["id"]),
            )
        deleted_ids: list[int] = []
        if deleted_rows:
            deleted_ids = [int(row["id"]) for row in deleted_rows]
            placeholders = ", ".join("?" for _ in deleted_ids)
            conn.execute(f"DELETE FROM goals_unit WHERE id IN ({placeholders})", deleted_ids)
        values = [
            {
                "id": int(row["id"]),
                "unitName": row["unit_name"],
                "revenueGoal": audit_round(row["revenue_goal"]),
            }
            for row in group_rows
        ]
        audit_payload = {
            "competence": competence,
            "normalizedUnit": normalized_unit,
            "keptId": int(kept_row["id"]),
            "deletedIds": deleted_ids,
            "values": values,
        }
        audit_log(
            conn,
            company_id,
            actor_user_id,
            "sanitize_unit_goals",
            "goals_unit",
            f"{competence}:{normalized_unit}",
            audit_payload,
        )
        changes_made.append(audit_payload)
    return changes_made


def delete_goal_unit_record(
    conn: sqlite3.Connection,
    company_id: int,
    actor_user_id: int | None,
    competence: str,
    unit_name: str,
) -> None:
    normalized_unit = normalize_unit(unit_name)
    row = conn.execute(
        """
        SELECT id, revenue_goal, returns_goal
        FROM goals_unit
        WHERE company_id = ? AND competence = ? AND unit_name = ?
        """,
        (company_id, competence, normalized_unit),
    ).fetchone()
    if not row:
        raise ValueError("Meta da unidade não encontrada.")
    conn.execute(
        "DELETE FROM goals_unit WHERE company_id = ? AND competence = ? AND unit_name = ?",
        (company_id, competence, normalized_unit),
    )
    audit_log(
        conn,
        company_id,
        actor_user_id,
        "delete_goal_unit",
        "goals_unit",
        f"{competence}:{normalized_unit}",
        {
            "competence": competence,
            "unitName": normalized_unit,
            "deletedId": int(row["id"]),
            "revenueGoal": audit_round(row["revenue_goal"]),
            "returnsGoal": audit_round(row["returns_goal"]),
        },
    )


def delete_goal_seller_record(
    conn: sqlite3.Connection,
    company_id: int,
    actor_user_id: int | None,
    competence: str,
    seller_name: str,
) -> None:
    normalized_seller = normalize_whitespace(seller_name)
    row = conn.execute(
        """
        SELECT id, revenue_goal, returns_goal
        FROM goals_seller
        WHERE company_id = ? AND competence = ? AND seller_name = ?
        """,
        (company_id, competence, normalized_seller),
    ).fetchone()
    if not row:
        raise ValueError("Meta do vendedor não encontrada.")
    conn.execute(
        "DELETE FROM goals_seller WHERE company_id = ? AND competence = ? AND seller_name = ?",
        (company_id, competence, normalized_seller),
    )
    audit_log(
        conn,
        company_id,
        actor_user_id,
        "delete_goal_seller",
        "goals_seller",
        f"{competence}:{normalized_seller}",
        {
            "competence": competence,
            "sellerName": normalized_seller,
            "deletedId": int(row["id"]),
            "revenueGoal": audit_round(row["revenue_goal"]),
            "returnsGoal": audit_round(row["returns_goal"]),
        },
    )


def get_score_config(conn: sqlite3.Connection, company_id: int, competence: str) -> sqlite3.Row:
    row = conn.execute(
        """
        SELECT *
        FROM score_configs
        WHERE company_id = ? AND valid_from_competence <= ?
          AND (valid_to_competence IS NULL OR valid_to_competence >= ?)
        ORDER BY valid_from_competence DESC
        LIMIT 1
        """,
        (company_id, competence, competence),
    ).fetchone()
    if row:
        return row
    return conn.execute("SELECT * FROM score_configs WHERE company_id = ? ORDER BY valid_from_competence DESC LIMIT 1", (company_id,)).fetchone()


def decode_text_content(content: bytes) -> str:
    encodings = ["utf-8-sig", "utf-8", "cp1252", "latin-1"]
    last_error: UnicodeDecodeError | None = None
    for encoding in encodings:
        try:
            return content.decode(encoding)
        except UnicodeDecodeError as exc:
            last_error = exc
    if last_error:
        raise last_error
    return content.decode("utf-8", errors="replace")


def parse_csv_bytes(content: bytes) -> list[dict[str, str]]:
    text = decode_text_content(content)
    reader = csv.DictReader(io.StringIO(text, newline=""), delimiter=";")
    return [dict(row) for row in reader]


def file_hash(content: bytes) -> str:
    return hashlib.sha256(content).hexdigest()


def detect_file_type(filename: str) -> str | None:
    clean = Path(filename).name
    clean_lower = clean.lower()
    if clean_lower in {key.lower() for key in CSV_FILE_TYPES}:
        for key, value in CSV_FILE_TYPES.items():
            if key.lower() == clean_lower:
                return value
    stem_lower = Path(filename).stem.lower()
    aliases = {
        "01fat": "faturamento_detalhado",
        "02unidade": "custo_unidade",
        "03vendedor": "custo_vendedor",
        "030-relatoriofaturamento detalhado": "faturamento_detalhado",
        "030-relatoriofaturamento conslidado cliente": "faturamento_cliente_consolidado",
        "030-relatoriofaturamento consolidado cliente": "faturamento_cliente_consolidado",
        "030-relatoriopessoas": "cadastro_clientes",
        "030-relatoriodevolucao": "devolucao_garantia",
        "030-relatoriocustovenda unidade": "custo_unidade",
        "030-relatoriocustovenda vendedor consolidado": "custo_vendedor",
    }
    alias_match = aliases.get(stem_lower)
    if alias_match:
        return alias_match
    if re.fullmatch(r"030-relatoriocustovenda\(\d+\)", stem_lower):
        # Não distingue unidade de vendedor pelo nome — campo do formulário já faz isso via override
        return "custo_unidade"  # fallback conservador; fieldName override tem prioridade
    if re.fullmatch(r"030-relatoriofaturamento\(\d+\)", stem_lower):
        return "faturamento_detalhado"
    if re.fullmatch(r"030-relatoriofaturamento detalhado\(\d+\)", stem_lower):
        return "faturamento_detalhado"
    if re.fullmatch(r"030-relatoriofaturamento detalhado \(\d+\)", stem_lower):
        return "faturamento_detalhado"
    if re.fullmatch(r"030-relatoriofaturamento conslidado cliente\(\d+\)", stem_lower):
        return "faturamento_cliente_consolidado"
    if re.fullmatch(r"030-relatoriofaturamento conslidado cliente \(\d+\)", stem_lower):
        return "faturamento_cliente_consolidado"
    if re.fullmatch(r"030-relatoriofaturamento consolidado cliente\(\d+\)", stem_lower):
        return "faturamento_cliente_consolidado"
    if re.fullmatch(r"030-relatoriofaturamento consolidado cliente \(\d+\)", stem_lower):
        return "faturamento_cliente_consolidado"
    if re.fullmatch(r"030-relatoriopessoas\(\d+\)", stem_lower):
        return "cadastro_clientes"
    if re.fullmatch(r"030-relatoriopessoas \(\d+\)", stem_lower):
        return "cadastro_clientes"
    return None


def normalize_upload_entries(files_payload: dict[str, bytes] | list[dict[str, Any]]) -> list[dict[str, Any]]:
    if isinstance(files_payload, dict):
        return [{"fieldName": "files", "fileName": filename, "content": content} for filename, content in files_payload.items()]
    return files_payload


def detect_upload_file_type(file_name: str, field_name: str | None = None) -> str | None:
    override = UPLOAD_FIELD_TYPE_OVERRIDES.get(field_name or "")
    if override:
        return override
    return detect_file_type(file_name)


def suggest_competence(rows: list[dict[str, str]]) -> str | None:
    counts = Counter()
    for row in rows:
        dt_value = parse_sales_row_date(row)
        competence = competence_from_date(dt_value)
        if competence:
            counts[competence] += 1
    if not counts:
        return None
    return counts.most_common(1)[0][0]


def normalize_sku(gtin_value: str | None, manufacturer_sku: str | None) -> str:
    gtin = normalize_whitespace(gtin_value)
    manufacturer = normalize_whitespace(manufacturer_sku)
    return gtin or manufacturer or "SKU_DESCONHECIDO"


def seller_should_be_counted(conn: sqlite3.Connection, company_id: int, seller_name: str, competence: str) -> bool:
    role, _ = current_role_and_unit(conn, company_id, seller_name, competence)
    if role is None:
        return True
    return role == "Vendedor"


def register_issue(conn: sqlite3.Connection, company_id: int, import_id: int, competence: str, issue_type: str, reference_value: str, details: dict[str, Any]) -> None:
    exists = conn.execute(
        """
        SELECT id FROM import_issues
        WHERE company_id = ? AND competence = ? AND issue_type = ? AND reference_value = ? AND status = 'pendente'
        """,
        (company_id, competence, issue_type, reference_value),
    ).fetchone()
    if exists:
        return
    conn.execute(
        """
        INSERT INTO import_issues (company_id, import_id, competence, issue_type, reference_value, details_json, created_at)
        VALUES (?, ?, ?, ?, ?, ?, ?)
        """,
        (company_id, import_id, competence, issue_type, reference_value, json.dumps(details, ensure_ascii=False), now_iso()),
    )


def normalize_import_scope(value: str | None) -> str:
    scope = normalize_whitespace(value).lower()
    return scope if scope in IMPORT_SCOPE_REQUIREMENTS else "full"


def preview_import_package(files_payload: dict[str, bytes] | list[dict[str, Any]], import_scope: str = "full") -> dict[str, Any]:
    import_scope = normalize_import_scope(import_scope)
    required_file_types = IMPORT_SCOPE_REQUIREMENTS[import_scope]
    file_types = {}
    row_counts = {}
    files_meta = []
    suggested_values = []
    missing = []
    unsupported_files = []

    for entry in normalize_upload_entries(files_payload):
        filename = entry["fileName"]
        field_name = entry.get("fieldName")
        content = entry["content"]
        extension = Path(filename).suffix.lower()
        if extension and extension != ".csv":
            unsupported_files.append({"fileName": filename, "fieldName": field_name, "reason": "Formato inválido. Use CSV."})
            continue
        kind = detect_upload_file_type(filename, field_name)
        if not kind:
            continue
        rows = parse_csv_bytes(content)
        # Detecção pelo conteúdo: custo venda com nome genérico (NNNN) distingue por coluna
        if kind == "custo_unidade" and rows:
            first_col = (list(rows[0].keys()) or [""])[0].strip().upper()
            if first_col in {"VENDEDOR", "VENDEDOR CONSOLIDADO"}:
                kind = "custo_vendedor"
        file_types[kind] = rows
        row_counts[kind] = row_counts.get(kind, 0) + len(rows)
        files_meta.append({"fileName": filename, "fieldName": field_name, "fileType": kind, "rows": len(rows), "hash": file_hash(content)})
        if import_scope == "cost":
            suggested_values.append(date.today().strftime("%Y-%m"))
        elif kind == "faturamento_detalhado" and import_scope in {"full", "sales"}:
            suggestion = suggest_competence(rows)
            if suggestion:
                suggested_values.append(suggestion)
        elif kind == "devolucao_garantia":
            # Competência mais frequente pela data da devolução
            counts = Counter()
            for r in rows:
                dt = parse_datetime_pt(r.get("Data")) or parse_datetime_flexible(r.get("Data"))
                comp = competence_from_date(dt)
                if comp:
                    counts[comp] += 1
            if counts:
                suggested_values.append(counts.most_common(1)[0][0])

    detected_file_types = set(file_types)
    if import_scope == "crm":
        if not detected_file_types:
            missing = sorted(required_file_types)
    else:
        for required in required_file_types:
            if required not in file_types:
                missing.append(required)

    suggestion = Counter(suggested_values).most_common(1)[0][0] if suggested_values else None
    return {
        "isValid": len(missing) == 0 and len(unsupported_files) == 0 and bool(detected_file_types),
        "importScope": import_scope,
        "importScopeLabel": IMPORT_SCOPE_LABELS[import_scope],
        "missingFileTypes": missing,
        "unsupportedFiles": unsupported_files,
        "suggestedCompetence": suggestion,
        "rowCounts": row_counts,
        "files": files_meta,
        "detectedFileTypes": sorted(detected_file_types),
    }


def delete_competence_data(
    conn: sqlite3.Connection,
    company_id: int,
    competence: str,
    file_types: set[str] | None = None,
) -> None:
    selected_file_types = file_types or set(IMPORT_SCOPE_TABLES)
    # Cadastro de clientes: base mestre sem competência — substitui TODA a base anterior.
    if "cadastro_clientes" in selected_file_types:
        conn.execute("DELETE FROM crm_client_profiles WHERE company_id = ?", (company_id,))
    # Faturamento consolidado por cliente: competência vem do nome do arquivo — substitui só o mês.
    if "faturamento_cliente_consolidado" in selected_file_types:
        conn.execute(
            "DELETE FROM crm_client_summary WHERE company_id = ? AND competence = ?",
            (company_id, competence),
        )
    target_tables = set()
    for file_type in selected_file_types:
        # APPEND-ONLY: competência derivada da data de cada linha e dedupe por row_hash.
        # Nunca apagar o histórico dessas tabelas.
        if file_type in {
            "cadastro_clientes", "faturamento_cliente_consolidado",
            "faturamento_detalhado", "devolucao_garantia",
        }:
            continue
        target_tables.update(IMPORT_SCOPE_TABLES.get(file_type, ()))
    for table in target_tables:
        conn.execute(f"DELETE FROM {table} WHERE company_id = ? AND competence = ?", (company_id, competence))


def import_package(
    conn: sqlite3.Connection,
    company_id: int,
    user_id: int,
    competence: str,
    import_action: str,
    import_scope: str,
    preview: dict[str, Any],
    files_payload: dict[str, bytes] | list[dict[str, Any]],
) -> dict[str, Any]:
    import_scope = normalize_import_scope(import_scope)
    selected_file_types = {
        item["fileType"]
        for item in preview.get("files", [])
        if item.get("fileType")
    }
    actual_action = import_action or "substituir"
    # Guarda a contagem anterior da base de clientes para detectar exportação incompleta
    clients_before = 0
    if "cadastro_clientes" in selected_file_types:
        clients_before = int(conn.execute(
            "SELECT COUNT(*) FROM crm_client_profiles WHERE company_id = ?", (company_id,)
        ).fetchone()[0] or 0)
    if actual_action == "substituir":
        delete_competence_data(conn, company_id, competence, selected_file_types)

    import_cursor = conn.execute(
        """
        INSERT INTO imports (company_id, competence, import_action, suggested_competence, imported_by, imported_at)
        VALUES (?, ?, ?, ?, ?, ?)
        """,
        (company_id, competence, actual_action, preview.get("suggestedCompetence"), user_id, now_iso()),
    )
    import_id = import_cursor.lastrowid

    duplicate_rows_skipped = 0
    sales_competences_seen: set[str] = set()
    sales_rows_by_competence: Counter = Counter()
    sales_rows_without_date = 0
    warranty_competences_seen: set[str] = set()
    warranty_total_value = 0.0

    for entry in normalize_upload_entries(files_payload):
        filename = entry["fileName"]
        field_name = entry.get("fieldName")
        content = entry["content"]
        kind = detect_upload_file_type(filename, field_name)
        if not kind:
            continue
        rows = parse_csv_bytes(content)
        # Um mesmo import pode receber VÁRIOS arquivos do mesmo tipo (ex.: a base de
        # clientes do Alfa vem em duas exportações complementares). O UNIQUE
        # (import_id, file_type) é preservado acumulando nomes e contagem de linhas.
        conn.execute(
            """
            INSERT INTO import_files (import_id, file_type, original_name, file_hash, row_count)
            VALUES (?, ?, ?, ?, ?)
            ON CONFLICT(import_id, file_type) DO UPDATE SET
                original_name = original_name || ' + ' || excluded.original_name,
                file_hash = excluded.file_hash,
                row_count = row_count + excluded.row_count
            """,
            (import_id, kind, filename, file_hash(content), len(rows)),
        )

        if kind == "faturamento_detalhado":
            # Duas vendas idênticas no mesmo dia são possíveis (mesma peça, mesmo
            # cliente, dois atendimentos). O contador de ocorrência evita descartar
            # a segunda como duplicata, mantendo o dedupe de reimportação.
            sales_occurrence: Counter = Counter()
            for row in rows:
                seller_name = normalize_whitespace(row.get("Vendedor"))
                client_name = normalize_whitespace(row.get("Cliente") or row.get("CLIENTE") or row.get("Razao Social/Nome")) or "CLIENTE NÃO INFORMADO"
                city_name = normalize_upper(row.get("Cidade"))
                gtin_value = normalize_whitespace(row.get(""))
                manufacturer_sku = normalize_whitespace(row.get("Fabricante"))
                dt_value = parse_sales_row_date(row)
                # Competência POR LINHA, derivada da data de emissão. Um único arquivo
                # pode conter vários meses; cada linha vai para o mês correto.
                row_competence = competence_from_date(dt_value) or competence
                if not competence_from_date(dt_value):
                    sales_rows_without_date += 1
                sales_competences_seen.add(row_competence)
                sales_rows_by_competence[row_competence] += 1
                sku_key = normalize_sku(gtin_value, manufacturer_sku)
                payload = {
                    "seller": seller_name,
                    "client": client_name,
                    "city": city_name,
                    "gtin": gtin_value,
                    "manufacturer": manufacturer_sku,
                    "issue_date": dt_value.isoformat() if dt_value else "",
                    "quantity": parse_decimal(row.get("Quant.")),
                    "gross": parse_decimal(row.get("Bruto")),
                    "discount": parse_decimal(row.get("Desconto")),
                    "freight": parse_decimal(row.get("Frete")),
                    "qty_return": parse_decimal(row.get("QTD. Dev.")),
                    "value_return": parse_decimal(row.get("vlr. dev.")),
                    "net": parse_decimal(row.get("Liquido")),
                    "sale_share": parse_decimal(row.get("%venda")),
                }
                _sig = json.dumps(payload, ensure_ascii=False, sort_keys=True)
                sales_occurrence[_sig] += 1
                row_hash = hash_text(f"{_sig}#{sales_occurrence[_sig]}")
                try:
                    conn.execute(
                        """
                        INSERT INTO fact_sales_detail (
                            company_id, competence, import_id, row_hash, seller_name, client_name, city_name,
                            gtin_value, manufacturer_sku, sku_key, issue_date, quantity, gross_value,
                            discount_value, freight_value, return_quantity, return_value, net_value, sale_share, created_at
                        ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                        """,
                        (
                            company_id,
                            row_competence,
                            import_id,
                            row_hash,
                            seller_name,
                            client_name,
                            city_name,
                            gtin_value,
                            manufacturer_sku,
                            sku_key,
                            dt_value.isoformat() if dt_value else None,
                            payload["quantity"],
                            payload["gross"],
                            payload["discount"],
                            payload["freight"],
                            payload["qty_return"],
                            payload["value_return"],
                            payload["net"],
                            payload["sale_share"],
                            now_iso(),
                        ),
                    )
                except sqlite3.IntegrityError:
                    duplicate_rows_skipped += 1
                role, _ = current_role_and_unit(conn, company_id, seller_name, row_competence)
                if role is None:
                    register_issue(conn, company_id, import_id, row_competence, "vendedor_sem_vinculo", seller_name, {"kind": "seller"})
                if city_name:
                    _comp_day = first_day_of_competence(row_competence).isoformat()
                    city_match = conn.execute(
                        """
                        SELECT id FROM city_mappings
                        WHERE company_id = ? AND city_name = ? AND date(valid_from) <= date(?)
                          AND (valid_to IS NULL OR date(valid_to) >= date(?))
                        LIMIT 1
                        """,
                        (company_id, city_name, _comp_day, _comp_day),
                    ).fetchone()
                    if not city_match:
                        register_issue(conn, company_id, import_id, row_competence, "cidade_sem_correspondencia", city_name, {"kind": "city"})
        elif kind == "cadastro_clientes":
            for row in rows:
                client_code = normalize_whitespace(row.get("Codigo"))
                client_name = normalize_whitespace(row.get("Razao Social/Nome"))
                if not client_code or not client_name:
                    continue
                conn.execute(
                    """
                    INSERT INTO crm_client_profiles (
                        company_id, client_code, client_name, trade_name, document_number, state_registration,
                        address_line, address_number, neighborhood, city_name, state_name, phone, postal_code,
                        first_sale_at, last_sale_at, credit_limit, economic_group, internal_seller_name,
                        external_seller_name, email, source_import_id, updated_at, created_at
                    ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                    ON CONFLICT(company_id, client_code) DO UPDATE SET
                        client_name = excluded.client_name,
                        trade_name = excluded.trade_name,
                        document_number = excluded.document_number,
                        state_registration = excluded.state_registration,
                        address_line = excluded.address_line,
                        address_number = excluded.address_number,
                        neighborhood = excluded.neighborhood,
                        city_name = excluded.city_name,
                        state_name = excluded.state_name,
                        phone = excluded.phone,
                        postal_code = excluded.postal_code,
                        first_sale_at = excluded.first_sale_at,
                        last_sale_at = excluded.last_sale_at,
                        credit_limit = excluded.credit_limit,
                        economic_group = excluded.economic_group,
                        internal_seller_name = excluded.internal_seller_name,
                        external_seller_name = excluded.external_seller_name,
                        email = excluded.email,
                        source_import_id = excluded.source_import_id,
                        updated_at = excluded.updated_at
                    """,
                    (
                        company_id,
                        client_code,
                        client_name,
                        normalize_whitespace(row.get("Fantasia")),
                        normalize_whitespace(row.get("CNPJ/CPF")),
                        normalize_whitespace(row.get("IE")),
                        normalize_whitespace(row.get("Endereco")),
                        normalize_whitespace(row.get("Num")),
                        normalize_whitespace(row.get("Bairro")),
                        normalize_upper(row.get("Cidade")),
                        normalize_upper(row.get("Uf")),
                        normalize_whitespace(row.get("Telefone")),
                        normalize_whitespace(row.get("CEP")),
                        parse_datetime_flexible(row.get("Primeira Venda")).isoformat() if parse_datetime_flexible(row.get("Primeira Venda")) else None,
                        parse_datetime_flexible(row.get("Ultima Venda")).isoformat() if parse_datetime_flexible(row.get("Ultima Venda")) else None,
                        parse_decimal(row.get("Limite Credito")),
                        normalize_whitespace(row.get("Grupo Economico")),
                        normalize_whitespace(row.get("Vend. Interno")),
                        normalize_whitespace(row.get("Vend. Externo")),
                        normalize_whitespace(row.get("Email")),
                        import_id,
                        now_iso(),
                        now_iso(),
                    ),
                )
        elif kind == "faturamento_cliente_consolidado":
            for row in rows:
                client_code = normalize_whitespace(row.get("CODIGO"))
                client_name = normalize_whitespace(row.get("CLIENTE"))
                if not client_code or not client_name:
                    continue
                dt_value = parse_datetime_flexible(row.get("ULT.COMPRA"))
                try:
                    conn.execute(
                        """
                        INSERT INTO crm_client_summary (
                            company_id, competence, import_id, client_code, client_name, seller_name,
                            city_name, last_purchase_at, gross_value, discount_value, freight_value,
                            return_quantity, return_value, net_value, sale_share, created_at
                        ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                        """,
                        (
                            company_id,
                            competence,
                            import_id,
                            client_code,
                            client_name,
                            normalize_whitespace(row.get("Vendedor")),
                            normalize_upper(row.get("Cidade")),
                            dt_value.isoformat() if dt_value else None,
                            parse_decimal(row.get("Bruto")),
                            parse_decimal(row.get("Desconto")),
                            parse_decimal(row.get("Frete")),
                            parse_decimal(row.get("QTD. Dev.")),
                            parse_decimal(row.get("vlr. dev.")),
                            parse_decimal(row.get("Liquido")),
                            parse_decimal(row.get("%venda")),
                            now_iso(),
                        ),
                    )
                except sqlite3.IntegrityError:
                    duplicate_rows_skipped += 1
        elif kind == "custo_vendedor":
            for row in rows:
                seller_name = normalize_whitespace(row.get("VENDEDOR"))
                payload = {
                    "seller": seller_name,
                    "qty_sold": parse_decimal(row.get("QTD VENDIDA")),
                    "cost": parse_decimal(row.get("CUSTO")),
                    "sale": parse_decimal(row.get("VENDA")),
                    "profit": parse_decimal(row.get("R$ LUCRO")),
                    "net_profit": parse_decimal(row.get("R$ LUCRO LIQUIDO")),
                    "profit_pct": parse_decimal(row.get("% LUCRO")),
                    "return_cost": parse_decimal(row.get("CUSTO DEVOLUCAO")),
                    "return_value": parse_decimal(row.get("VALOR DA DEVOLUCAO")),
                    "net": parse_decimal(row.get("VALOR LIQUIDO")),
                    "margin": parse_decimal(row.get("MARGEM")),
                }
                row_hash = hash_text(json.dumps(payload, ensure_ascii=False, sort_keys=True))
                try:
                    conn.execute(
                        """
                        INSERT INTO fact_vendor_summary (
                            company_id, competence, import_id, row_hash, seller_name, qty_sold, cost_value,
                            sale_value, profit_value, net_profit_value, profit_pct, return_cost,
                            return_value, net_value, margin_value, created_at
                        ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                        """,
                        (
                            company_id,
                            competence,
                            import_id,
                            row_hash,
                            seller_name,
                            payload["qty_sold"],
                            payload["cost"],
                            payload["sale"],
                            payload["profit"],
                            payload["net_profit"],
                            payload["profit_pct"],
                            payload["return_cost"],
                            payload["return_value"],
                            payload["net"],
                            payload["margin"],
                            now_iso(),
                        ),
                    )
                except sqlite3.IntegrityError:
                    duplicate_rows_skipped += 1
                role, _ = current_role_and_unit(conn, company_id, seller_name, competence)
                if role is None:
                    register_issue(conn, company_id, import_id, competence, "vendedor_sem_vinculo", seller_name, {"kind": "seller"})
        elif kind == "devolucao_garantia":
            # A competência sai da DATA de cada devolução, linha a linha — o relatório
            # pode cobrir mais de um mês e cada devolução pertence ao mês em que ocorreu.
            # Linhas idênticas são LEGÍTIMAS (2 unidades da mesma peça na mesma
            # devolução vêm em linhas separadas). O contador de ocorrência distingue
            # essas linhas sem quebrar o dedupe de reimportação do mesmo arquivo.
            occurrence_counter: Counter = Counter()
            for row in rows:
                unit_name = normalize_unit(row.get("Empresa"))
                seller_name = normalize_whitespace(row.get("Vendedor"))
                return_dt = parse_datetime_pt(row.get("Data")) or parse_datetime_flexible(row.get("Data"))
                row_competence = competence_from_date(return_dt) or competence
                total_value = parse_decimal(row.get("Total"))
                if not unit_name and not seller_name and not total_value:
                    continue
                payload = {
                    "unit": unit_name,
                    "seller": seller_name,
                    "client": normalize_whitespace(row.get("Cliente")),
                    "return_number": normalize_whitespace(row.get("Devolucao")),
                    "return_date": return_dt.isoformat() if return_dt else "",
                    "invoice": normalize_whitespace(row.get("Nota")),
                    "item": normalize_whitespace(row.get("Codigo")),
                    "qty": parse_decimal(row.get("Qtd.")),
                    "cost": parse_decimal(row.get("P. Custo.")),
                    "total": total_value,
                }
                base_signature = json.dumps(payload, ensure_ascii=False, sort_keys=True)
                occurrence_counter[base_signature] += 1
                row_hash = hash_text(f"{base_signature}#{occurrence_counter[base_signature]}")
                issue_dt = parse_datetime_pt(row.get("Emissao")) or parse_datetime_flexible(row.get("Emissao"))
                try:
                    conn.execute(
                        """
                        INSERT INTO fact_warranty_returns (
                            company_id, competence, import_id, row_hash, unit_name, seller_name,
                            client_name, city_name, return_number, return_date, reason,
                            invoice_number, issue_date, item_code, item_type, item_description,
                            brand_name, supplier_name, quantity, cost_value, total_value, created_at
                        ) VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)
                        """,
                        (
                            company_id, row_competence, import_id, row_hash, unit_name, seller_name,
                            normalize_whitespace(row.get("Cliente")),
                            normalize_upper(row.get("Cidade")),
                            normalize_whitespace(row.get("Devolucao")),
                            return_dt.isoformat() if return_dt else None,
                            normalize_upper(row.get("Motivo")),
                            normalize_whitespace(row.get("Nota")),
                            issue_dt.isoformat() if issue_dt else None,
                            normalize_whitespace(row.get("Codigo")),
                            normalize_whitespace(row.get("Tipo")),
                            normalize_whitespace(row.get("Descricao")),
                            normalize_whitespace(row.get("Marca")),
                            normalize_whitespace(row.get("Forn. ref.")),
                            parse_decimal(row.get("Qtd.")),
                            parse_decimal(row.get("P. Custo.")),
                            total_value,
                            now_iso(),
                        ),
                    )
                    warranty_competences_seen.add(row_competence)
                    warranty_total_value += total_value
                except sqlite3.IntegrityError:
                    duplicate_rows_skipped += 1
        elif kind == "custo_unidade":
            for row in rows:
                unit_name = normalize_unit(row.get("EMPRESA"))
                payload = {
                    "unit": unit_name,
                    "qty_sold": parse_decimal(row.get("QTD VENDIDA")),
                    "cost": parse_decimal(row.get("CUSTO")),
                    "sale": parse_decimal(row.get("VENDA")),
                    "profit": parse_decimal(row.get("R$ LUCRO")),
                    "net_profit": parse_decimal(row.get("R$ LUCRO LIQUIDO")),
                    "profit_pct": parse_decimal(row.get("% LUCRO")),
                    "return_cost": parse_decimal(row.get("CUSTO DEVOLUCAO")),
                    "return_value": parse_decimal(row.get("VALOR DA DEVOLUCAO")),
                    "net": parse_decimal(row.get("VALOR LIQUIDO")),
                    "margin": parse_decimal(row.get("MARGEM")),
                }
                row_hash = hash_text(json.dumps(payload, ensure_ascii=False, sort_keys=True))
                try:
                    conn.execute(
                        """
                        INSERT INTO fact_unit_summary (
                            company_id, competence, import_id, row_hash, unit_name, qty_sold, cost_value,
                            sale_value, profit_value, net_profit_value, profit_pct, return_cost,
                            return_value, net_value, margin_value, created_at
                        ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                        """,
                        (
                            company_id,
                            competence,
                            import_id,
                            row_hash,
                            unit_name,
                            payload["qty_sold"],
                            payload["cost"],
                            payload["sale"],
                            payload["profit"],
                            payload["net_profit"],
                            payload["profit_pct"],
                            payload["return_cost"],
                            payload["return_value"],
                            payload["net"],
                            payload["margin"],
                            now_iso(),
                        ),
                    )
                except sqlite3.IntegrityError:
                    duplicate_rows_skipped += 1

    conn.execute("UPDATE imports SET duplicate_rows_skipped = ? WHERE id = ?", (duplicate_rows_skipped, import_id))
    audit_log(
        conn,
        company_id,
        user_id,
        "importar",
        "import",
        str(import_id),
        {"competence": competence, "action": actual_action, "scope": import_scope},
    )
    ensure_client_registry_for_sales(conn, company_id)
    conn.commit()
    # O cadastro do Alfa acabou de mudar: é o momento em que um prospect pode
    # ter virado cliente. Casa pelo CNPJ e migra o histórico de prospecção.
    try:
        reconcile_prospects(conn, company_id)
        refresh_prospect_first_purchases(conn, company_id)
    except Exception as exc:
        print(f"[prospeccao] reconciliação após importação falhou: {exc}", flush=True)
    # Dados mudaram: derruba caches derivados (dashboard e CRM)
    invalidate_crm_cache(company_id)
    result: dict[str, Any] = {
        "importId": import_id,
        "duplicateRowsSkipped": duplicate_rows_skipped,
        "importAction": actual_action,
        "importScope": import_scope,
        "importedFileTypes": sorted(selected_file_types),
    }
    if sales_competences_seen:
        _ordered = sorted(sales_competences_seen)
        result["salesCompetences"] = _ordered
        result["salesRowsByCompetence"] = dict(sorted(sales_rows_by_competence.items()))
        result["salesRowsWithoutDate"] = sales_rows_without_date
        # Mostra as competências com mais linhas; resume o restante
        _top = sales_rows_by_competence.most_common(8)
        _detail = ", ".join(f"{c}:{n}" for c, n in sorted(_top))
        _rest = len(_ordered) - len(_top)
        _msg = f"Faturamento detalhado — {len(_ordered)} competência(s) ({_ordered[0]}…{_ordered[-1]}): {_detail}"
        if _rest > 0:
            _msg += f" +{_rest} outra(s)"
        _msg += f" | {duplicate_rows_skipped} linha(s) já existentes ignoradas"
        if sales_rows_without_date:
            _msg += f" | {sales_rows_without_date} sem data usaram {competence}"
        result["message"] = _msg

    if warranty_competences_seen:
        _wc = sorted(warranty_competences_seen)
        result["warrantyCompetences"] = _wc
        result["warrantyTotalValue"] = round(warranty_total_value, 2)
        result["message"] = (
            f"Devolução em garantia — {', '.join(_wc)}: {brl(warranty_total_value)} "
            f"| {duplicate_rows_skipped} linha(s) já existentes ignoradas"
        )

    # Cadastro de clientes: reporta o tamanho da base e alerta se encolheu muito
    # (indício de exportação parcial do Alfa — a base vem em vários arquivos).
    if "cadastro_clientes" in selected_file_types:
        clients_after = int(conn.execute(
            "SELECT COUNT(*) FROM crm_client_profiles WHERE company_id = ?", (company_id,)
        ).fetchone()[0] or 0)
        _files_count = len([
            e for e in normalize_upload_entries(files_payload)
            if detect_upload_file_type(e["fileName"], e.get("fieldName")) == "cadastro_clientes"
        ])
        result["clientsBefore"] = clients_before
        result["clientsAfter"] = clients_after
        result["clientsFileCount"] = _files_count
        _msg = f"Cadastro de clientes: {clients_after} clientes de {_files_count} arquivo(s)"
        if clients_before:
            _delta = clients_after - clients_before
            _msg += f" (antes: {clients_before}, variação: {_delta:+d})"
            if clients_after < clients_before * 0.7:
                result["warning"] = (
                    f"ATENÇÃO: a base caiu de {clients_before} para {clients_after} clientes. "
                    "Verifique se todas as exportações do Alfa estavam na pasta."
                )
                _msg += " ⚠ possível exportação incompleta"

        # Cobertura sobre o faturamento. É o alarme que faltava: comparar com a
        # importação anterior não adianta quando a base SEMPRE veio pela metade —
        # foi o que aconteceu aqui, com 57% dos clientes faturados fora do
        # cadastro e nenhum aviso. Comparar com o faturamento é absoluto: cliente
        # que compra tem de existir no cadastro.
        nomes_cadastro = {
            normalize_client_key(r["client_name"])
            for r in conn.execute(
                "SELECT client_name FROM crm_client_profiles WHERE company_id = ?", (company_id,)
            ).fetchall()
        }
        nomes_faturados = {
            normalize_client_key(r["client_name"])
            for r in conn.execute(
                "SELECT DISTINCT client_name FROM fact_sales_detail WHERE company_id = ?",
                (company_id,),
            ).fetchall()
        }
        if nomes_faturados:
            fora = len(nomes_faturados - nomes_cadastro)
            cobertura = 100 * (1 - fora / len(nomes_faturados))
            result["coveragePct"] = round(cobertura, 1)
            result["clientsMissingFromRegistry"] = fora
            _msg += f" · cobertura do faturamento: {cobertura:.1f}%"
            if cobertura < 85:
                result["warning"] = (
                    f"ATENÇÃO: {fora} cliente(s) que compraram não estão no cadastro "
                    f"({cobertura:.1f}% de cobertura). A base do Alfa vem em DOIS arquivos — "
                    "confirme que os dois estavam na pasta antes da importação."
                )
                _msg += " ⚠ cadastro incompleto"
        result["message"] = _msg
    return result


def query_competences(conn: sqlite3.Connection, company_id: int) -> list[str]:
    rows = conn.execute(
        """
        SELECT competence FROM (
            SELECT competence FROM fact_sales_detail WHERE company_id = ?
            UNION
            SELECT competence FROM fact_vendor_summary WHERE company_id = ?
            UNION
            SELECT competence FROM fact_unit_summary WHERE company_id = ?
        )
        ORDER BY competence DESC
        """,
        (company_id, company_id, company_id),
    ).fetchall()
    return [row["competence"] for row in rows]


def build_filters_from_query(query: dict[str, list[str]]) -> dict[str, str | None]:
    return {
        "competence_start": query.get("competenceStart", [None])[0],
        "competence_end": query.get("competenceEnd", [None])[0],
        "unit_name": normalize_unit(query.get("unit", [None])[0]),
        "seller_name": normalize_whitespace(query.get("seller", [None])[0]),
        "city_name": normalize_upper(query.get("city", [None])[0]),
        "status": normalize_upper(query.get("status", [None])[0]),
        "purchaseMonth": normalize_upper(query.get("purchaseMonth", [None])[0]),
        "growth": normalize_upper(query.get("growth", [None])[0]),
        "classCode": normalize_upper(query.get("classCode", [None])[0]),
        "personType": normalize_upper(query.get("personType", [None])[0]),
        # Busca por item: aceita código do fabricante ou interno
        "itemCode": normalize_whitespace(query.get("itemCode", [None])[0]),
        "search": normalize_whitespace(query.get("search", [None])[0]),
        # Carteira coberta que o vendedor pediu para enxergar (cobertura de férias)
        "coverage_of": normalize_whitespace(query.get("coverageOf", [None])[0]),
    }


def competence_range_clause(filters: dict[str, str | None]) -> tuple[str, list[Any]]:
    params: list[Any] = []
    clauses: list[str] = []
    if filters["competence_start"]:
        clauses.append("competence >= ?")
        params.append(filters["competence_start"])
    if filters["competence_end"]:
        clauses.append("competence <= ?")
        params.append(filters["competence_end"])
    return " AND ".join(clauses), params


def selected_primary_competence(filters: dict[str, str | None], competences: list[str]) -> str | None:
    if filters["competence_end"]:
        return filters["competence_end"]
    if filters["competence_start"]:
        return filters["competence_start"]
    return competences[0] if competences else None


def dashboard_competence_state(competence: str, today_value: date | None = None) -> dict[str, Any]:
    today_local = today_value or today_in_brazil()
    current_competence = today_local.strftime("%Y-%m")
    return {
        "today": today_local,
        "cutoffDate": dashboard_cutoff_date(today_local),
        "isCurrentCompetence": competence == current_competence,
        "isPastCompetence": competence < current_competence,
        "isFutureCompetence": competence > current_competence,
    }


# Cache de feriados e férias por (empresa, competência). Estes dados quase nunca mudam
# e get_business_calendar é chamado uma vez POR VENDEDOR dentro do dashboard — sem cache
# eram 2 queries × dezenas de vendedores a cada troca de competência.
_calendar_cache: dict[tuple, tuple[dict[str, str], dict[str, list[dict[str, Any]]]]] = {}
_calendar_cache_lock = threading.Lock()


def invalidate_calendar_cache(company_id: int | None = None) -> None:
    with _calendar_cache_lock:
        if company_id is None:
            _calendar_cache.clear()
        else:
            for k in list(_calendar_cache.keys()):
                if k[0] == company_id:
                    del _calendar_cache[k]


def _calendar_reference_data(
    conn: sqlite3.Connection, company_id: int, competence: str, start: date, end: date
) -> tuple[dict[str, str], dict[str, list[dict[str, Any]]]]:
    """Feriados da competência e férias agrupadas por pessoa, com cache."""
    key = (company_id, competence)
    with _calendar_cache_lock:
        hit = _calendar_cache.get(key)
    if hit is not None:
        return hit
    holidays = {
        row["holiday_date"]: row["holiday_name"]
        for row in conn.execute(
            "SELECT holiday_date, holiday_name FROM holidays WHERE company_id = ? AND holiday_date BETWEEN ? AND ?",
            (company_id, start.isoformat(), end.isoformat()),
        ).fetchall()
    }
    vacations_by_person: dict[str, list[dict[str, Any]]] = defaultdict(list)
    for row in conn.execute(
        """
        SELECT person_name, start_date, end_date, notes
        FROM vacations
        WHERE company_id = ? AND date(end_date) >= date(?) AND date(start_date) <= date(?)
        """,
        (company_id, start.isoformat(), end.isoformat()),
    ).fetchall():
        # Chave em maiúsculas: registros antigos podem ter sido digitados com
        # capitalização diferente da base de vendedores.
        vacations_by_person[normalize_upper(row["person_name"])].append(dict(row))
    result = (holidays, dict(vacations_by_person))
    with _calendar_cache_lock:
        _calendar_cache[key] = result
    return result


def get_business_calendar(
    conn: sqlite3.Connection,
    company_id: int,
    competence: str,
    seller_name: str | None = None,
    reference_today: date | None = None,
    include_current_day: bool = True,
) -> dict[str, Any]:
    start = first_day_of_competence(competence)
    end = last_day_of_competence(competence)
    holidays, vacations_by_person = _calendar_reference_data(conn, company_id, competence, start, end)
    vacation_dates = set()
    vacations = []
    if seller_name:
        for row in vacations_by_person.get(normalize_upper(seller_name), []):
            vacations.append(dict(row))
            try:
                _vs = date.fromisoformat(row["start_date"])
                _ve = date.fromisoformat(row["end_date"])
            except (TypeError, ValueError):
                continue
            for day in daterange(max(start, _vs), min(end, _ve)):
                vacation_dates.add(day.isoformat())

    actual_today = reference_today or today_in_brazil()
    effective_today = actual_today if include_current_day else actual_today - timedelta(days=1)
    is_past_competence = effective_today > end
    is_future_competence = effective_today < start
    total_working = 0
    elapsed_working = 0
    seller_working = 0
    elapsed_seller_working = 0
    for day in daterange(start, end):
        if day.weekday() >= 5 or day.isoformat() in holidays:
            continue
        total_working += 1
        if is_past_competence or (not is_future_competence and day <= effective_today):
            elapsed_working += 1
        if day.isoformat() in vacation_dates:
            continue
        seller_working += 1
        if is_past_competence or (not is_future_competence and day <= effective_today):
            elapsed_seller_working += 1
    return {
        "competence": competence,
        "holidays": [{"date": key, "name": value} for key, value in sorted(holidays.items())],
        "vacations": vacations,
        "totalWorkingDays": total_working,
        "elapsedWorkingDays": min(elapsed_working, total_working),
        "remainingWorkingDays": max(total_working - min(elapsed_working, total_working), 0),
        "sellerWorkingDays": seller_working if seller_name else total_working,
        "sellerElapsedWorkingDays": min(elapsed_seller_working, seller_working) if seller_name else min(elapsed_working, total_working),
        "referenceToday": actual_today.isoformat(),
        "effectiveToday": effective_today.isoformat(),
    }


def safe_div(numerator: float, denominator: float) -> float:
    if not denominator:
        return 0.0
    result = numerator / denominator
    return result if math.isfinite(result) else 0.0


def brl(value: float | int | None) -> str:
    """Formata em real brasileiro para textos exibidos ao vendedor."""
    try:
        number = float(value or 0)
    except (TypeError, ValueError):
        number = 0.0
    inteiro = f"{number:,.2f}"
    return "R$ " + inteiro.replace(",", "@").replace(".", ",").replace("@", ".")


def finite_or_none(value: Any) -> float | None:
    """Converte para float finito; devolve None para inf/NaN/valor inválido."""
    if value is None:
        return None
    try:
        number = float(value)
    except (TypeError, ValueError):
        return None
    return number if math.isfinite(number) else None


def official_cost_net(sale_value: float | int | None, return_value: float | int | None) -> float:
    return float(sale_value or 0) - float(return_value or 0)


def weighted_margin_average(rows: Iterable[dict[str, Any]]) -> float | None:
    weighted_total = 0.0
    base_total = 0.0
    for row in rows:
        revenue_net = float(row.get("net_value") or official_cost_net(row.get("sale_value"), row.get("return_value")) or 0.0)
        margin_value = row.get("margin_value")
        if margin_value is None:
            continue
        weighted_total += float(margin_value or 0.0) * revenue_net
        base_total += revenue_net
    if base_total <= 0:
        return None
    return weighted_total / base_total


def aggregate_official_summary_rows(rows: Iterable[dict[str, Any]]) -> dict[str, float | None]:
    rows_list = list(rows)
    revenue_net = float(sum(float(row.get("net_value") or 0.0) for row in rows_list))
    revenue_gross = float(sum(float(row.get("sale_value") or 0.0) for row in rows_list))
    returns_value = float(sum(float(row.get("return_value") or 0.0) for row in rows_list))
    return_cost = float(sum(float(row.get("return_cost") or 0.0) for row in rows_list))
    qty_sold = float(sum(float(row.get("qty_sold") or 0.0) for row in rows_list))
    cost_value = float(sum(float(row.get("cost_value") or 0.0) for row in rows_list))
    profit_value = float(sum(float(row.get("profit_value") or 0.0) for row in rows_list))
    net_profit_value = float(sum(float(row.get("net_profit_value") or 0.0) for row in rows_list))
    margin_average = weighted_margin_average(rows_list)
    return {
        "revenueNet": revenue_net,
        "revenueGross": revenue_gross,
        "returnsValue": returns_value,
        "returnCost": return_cost,
        "qtySold": qty_sold,
        "costValue": cost_value,
        "profitValue": profit_value,
        "netProfitValue": net_profit_value,
        "marginAverage": margin_average,
        "ticketPerPiece": safe_div(revenue_net, qty_sold),
        "returnRatioPct": safe_div(returns_value, revenue_net) * 100 if revenue_net else 0.0,
    }


def seller_identity_for_user(user: sqlite3.Row) -> str:
    return normalize_whitespace(user["linked_person_name"] or user["full_name"] or user["username"])


def normalize_unit_list(values: Any) -> list[str]:
    if values is None:
        return []
    if isinstance(values, str):
        text = values.strip()
        if not text:
            return []
        try:
            parsed = json.loads(text)
        except json.JSONDecodeError:
            parsed = re.split(r"[;,|]", text)
        values = parsed
    if not isinstance(values, (list, tuple, set)):
        values = [values]
    normalized: list[str] = []
    seen: set[str] = set()
    for value in values:
        unit = normalize_unit(str(value))
        if not unit or unit in seen:
            continue
        normalized.append(unit)
        seen.add(unit)
    return normalized


def linked_units_for_user(user: sqlite3.Row | dict[str, Any]) -> list[str]:
    raw_value = user["linked_units_json"] if isinstance(user, sqlite3.Row) else user.get("linked_units_json")
    return normalize_unit_list(raw_value)


def active_mapped_cities_for_units(conn: sqlite3.Connection, company_id: int, unit_names: list[str]) -> list[str]:
    normalized_units = normalize_unit_list(unit_names)
    if not normalized_units:
        return []
    placeholders = ", ".join("?" for _ in normalized_units)
    rows = conn.execute(
        f"""
        SELECT DISTINCT city_name
        FROM city_mappings
        WHERE company_id = ?
          AND principal_unit IN ({placeholders})
          AND city_name IS NOT NULL
          AND city_name <> ''
          AND (valid_to IS NULL OR valid_to = '')
        ORDER BY city_name
        """,
        [company_id, *normalized_units],
    ).fetchall()
    return [row["city_name"] for row in rows if row["city_name"]]


# ─────────────────────────────────────────────────────────────────────────────
# Atas de reunião e treinamento
#
# Fluxo: o gestor registra o encontro, marca quem participou e publica. A partir
# da publicação cada participante recebe a pendência de ciência dentro do CRM,
# com espaço para feedback que volta para o organizador.
#
# Duas decisões que moldam o resto:
#
# 1. RASCUNHO antes de PUBLICADA. A ata só vira pendência para a equipe quando o
#    gestor confirma. Sem isso, quem digita a ata em duas sessões dispararia
#    cobrança para todo mundo com o texto pela metade.
#
# 2. Participante identificado por NOME NORMALIZADO, não por user_id. Nem todo
#    vendedor tem login (o piloto é 1 por unidade), e a lista de presença precisa
#    valer como registro mesmo para quem não acessa o sistema. Quem tem login
#    encontra a pendência pelo nome; quem não tem fica registrado como presente.
# ─────────────────────────────────────────────────────────────────────────────

MEETING_KINDS = [
    {"id": "REUNIAO", "label": "Reunião", "icon": "🗓️"},
    {"id": "TREINAMENTO", "label": "Treinamento", "icon": "🎓"},
]
MEETING_KIND_IDS = {k["id"] for k in MEETING_KINDS}

# Limite por arquivo. Ata carrega apresentação e planilha, não vídeo — acima
# disso o servidor guarda peso que ninguém abre.
MEETING_ATTACHMENT_MAX_BYTES = 15 * 1024 * 1024
MEETING_ATTACHMENT_ALLOWED_EXT = {
    ".pdf", ".doc", ".docx", ".xls", ".xlsx", ".ppt", ".pptx",
    ".csv", ".txt", ".png", ".jpg", ".jpeg", ".webp", ".zip",
}


def meeting_files_dir() -> Path:
    caminho = DATA_DIR / "meeting_files"
    caminho.mkdir(parents=True, exist_ok=True)
    return caminho


def person_key(value: Any) -> str:
    """Chave de comparação de pessoa: maiúsculas, sem acento e sem pontuação.

    O mesmo vendedor aparece como 'Thielly Henrique', 'THIELLY HENRIQUE' e
    'Thielly Henrique (VENDAS)' dependendo da origem. Sem normalizar, a ciência
    de um nunca casaria com a pendência do outro.
    """
    texto = normalize_upper(strip_accents(str(value or "")))
    texto = re.sub(r"\(.*?\)", " ", texto)
    return re.sub(r"[^A-Z0-9]+", " ", texto).strip()


def short_person_key(value: Any) -> str:
    """Nome + primeiro sobrenome, normalizado.

    Existe porque o cadastro de pessoas e o cadastro de usuários quase nunca
    escrevem o nome igual: "THIELLY HENRIQUE DA SILVA" no faturamento e
    "Thielly Henrique" no login. Comparar o nome inteiro dava sempre falso e a
    vendedora não recebia a pendência de ciência. Dois tokens é o recorte que
    resolve o caso comum sem misturar pessoas diferentes.
    """
    partes = person_key(value).split()
    return " ".join(partes[:2]) if len(partes) >= 2 else (partes[0] if partes else "")


def user_person_keys(user: sqlite3.Row | dict[str, Any]) -> list[str]:
    """Todas as chaves pelas quais este usuário pode aparecer numa lista de presença.

    `linked_person_name` vem primeiro porque é o vínculo que o gestor declarou
    ao criar a conta — é informação explícita, não dedução por semelhança.
    """
    def valor(campo: str) -> Any:
        try:
            return user[campo]
        except (KeyError, IndexError):
            return None

    chaves: list[str] = []
    for campo in ("linked_person_name", "full_name", "username"):
        bruto = valor(campo)
        if not normalize_whitespace(bruto):
            continue
        chaves.append(person_key(bruto))
        chaves.append(short_person_key(bruto))
    return [k for k in dict.fromkeys(chaves) if k]


def resolve_user_for_person(conn: sqlite3.Connection, company_id: int, nome: str) -> int | None:
    """Conta de login correspondente a um nome da lista de presença, se houver.

    Ordem de tentativa, da informação mais confiável para a menos:
      1. `linked_person_name` — o vínculo explícito feito no cadastro do usuário.
      2. nome completo do usuário ou o próprio login.
      3. nome + primeiro sobrenome, para as divergências de nome do meio.

    Em caso de empate no passo 3, devolve None: melhor não notificar ninguém do
    que mandar a pendência para a pessoa errada.
    """
    alvo_completo = person_key(nome)
    alvo_curto = short_person_key(nome)
    if not alvo_completo:
        return None

    usuarios = conn.execute(
        "SELECT id, full_name, username, linked_person_name FROM users "
        "WHERE company_id = ? AND COALESCE(is_active, 1) = 1",
        (company_id,),
    ).fetchall()

    vinculados = [u["id"] for u in usuarios if person_key(u["linked_person_name"]) == alvo_completo]
    if len(vinculados) == 1:
        return vinculados[0]

    exatos = [u["id"] for u in usuarios
              if alvo_completo in {person_key(u["full_name"]), person_key(u["username"])}]
    if len(exatos) == 1:
        return exatos[0]

    if not alvo_curto:
        return None
    curtos = [u["id"] for u in usuarios
              if alvo_curto in {short_person_key(u["linked_person_name"]),
                                short_person_key(u["full_name"]),
                                short_person_key(u["username"])}]
    return curtos[0] if len(curtos) == 1 else None


def user_can_manage_meetings(conn: sqlite3.Connection, user: sqlite3.Row) -> bool:
    """Registra ata quem tem visão de equipe. Vendedor só dá ciência."""
    return data_scope_for_user(conn, user) != "proprio"


def meeting_person_identity(user: sqlite3.Row) -> str:
    """Nome com que o usuário aparece nas listas de presença."""
    return normalize_whitespace(user["full_name"]) or normalize_whitespace(user["username"])


def list_meeting_people(conn: sqlite3.Connection, company_id: int, user: sqlite3.Row) -> list[dict[str, Any]]:
    """Quem o gestor pode marcar como presente: pessoas ativas das unidades dele.

    Inclui quem não é vendedor (balcão, telemarketing, administrativo) — reunião
    e treinamento alcançam a equipe inteira, não só quem emite pedido.
    """
    competence = crm_latest_competence(conn, company_id) or date.today().strftime("%Y-%m")
    comp_day = first_day_of_competence(competence).isoformat()
    allowed = crm_allowed_units_for_user(conn, user)
    rows = conn.execute(
        """
        SELECT person_name, base_unit, role_classification
        FROM people_records
        WHERE company_id = ?
          AND date(valid_from) <= date(?)
          AND (valid_to IS NULL OR valid_to = '' OR date(valid_to) >= date(?))
        ORDER BY person_name
        """,
        (company_id, comp_day, comp_day),
    ).fetchall()
    vistos: set[str] = set()
    pessoas: list[dict[str, Any]] = []
    for r in rows:
        nome = normalize_whitespace(r["person_name"])
        chave = person_key(nome)
        if not nome or not chave or chave in vistos:
            continue
        unidade = normalize_unit(r["base_unit"])
        # `allowed is not None` = usuário restrito. Nesse caso, pessoa sem unidade
        # cadastrada também fica de fora: antes ela escapava do filtro e o gerente
        # via gente de outras equipes na lista de presença.
        if allowed is not None and unidade not in allowed:
            continue
        vistos.add(chave)
        pessoas.append({
            "personName": nome,
            "personKey": chave,
            "unitName": unidade,
            "role": normalize_whitespace(r["role_classification"]) or "Outros",
            # O gestor precisa ver, na hora de montar a lista, quem realmente vai
            # receber a pendência. Sem isso a falha é silenciosa.
            "hasLogin": resolve_user_for_person(conn, company_id, nome) is not None,
        })
    return pessoas


def backfill_meeting_participant_users(conn: sqlite3.Connection, company_id: int) -> None:
    """Liga à conta de login os presentes gravados antes de existir a coluna user_id.

    Roda no boot e é barato: só toca linhas com user_id nulo.
    """
    pendentes = conn.execute(
        """
        SELECT p.id, p.person_name FROM meeting_participants p
        JOIN meetings m ON m.id = p.meeting_id
        WHERE m.company_id = ? AND p.user_id IS NULL
        """,
        (company_id,),
    ).fetchall()
    vinculados = 0
    for linha in pendentes:
        user_id = resolve_user_for_person(conn, company_id, linha["person_name"])
        if user_id:
            conn.execute("UPDATE meeting_participants SET user_id = ? WHERE id = ?", (user_id, linha["id"]))
            vinculados += 1
    if vinculados:
        conn.commit()
        print(f"[reunioes] {vinculados} presente(s) vinculado(s) a contas de login")


def meeting_units_for_user(conn: sqlite3.Connection, user: sqlite3.Row) -> dict[str, Any]:
    """Unidades que o usuário pode atribuir a uma ata.

    Gerente registra reunião só para a equipe dele — não escolhe outra unidade
    nem marca a ata como corporativa. Diretoria e admin, sem restrição, podem os
    dois. Devolve também a unidade padrão, para o formulário já abrir certo.
    """
    permitidas = crm_allowed_units_for_user(conn, user)
    if permitidas is None:
        todas = [
            normalize_unit(r["base_unit"])
            for r in conn.execute(
                "SELECT DISTINCT base_unit FROM people_records "
                "WHERE company_id = ? AND base_unit IS NOT NULL AND base_unit <> '' "
                "ORDER BY base_unit",
                (user["company_id"],),
            ).fetchall()
        ]
        return {"units": [u for u in dict.fromkeys(todas) if u], "canBeCorporate": True, "defaultUnit": ""}
    return {
        "units": list(permitidas),
        "canBeCorporate": False,
        "defaultUnit": permitidas[0] if permitidas else "",
    }


def meeting_row_to_dict(row: sqlite3.Row) -> dict[str, Any]:
    return {
        "id": int(row["id"]),
        "kind": row["kind"],
        "kindLabel": next((k["label"] for k in MEETING_KINDS if k["id"] == row["kind"]), row["kind"]),
        "title": row["title"],
        "topic": row["topic"] or "",
        "unitName": row["unit_name"] or "",
        "occurredAt": row["occurred_at"] or "",
        "durationMin": int(row["duration_min"] or 0),
        "location": row["location"] or "",
        "agenda": row["agenda"] or "",
        "summary": row["summary"] or "",
        "decisions": row["decisions"] or "",
        "organizerName": row["organizer_name"] or "",
        "visibility": row["visibility"] if "visibility" in row.keys() else "UNIDADE",
        "status": row["status"],
        "publishedAt": row["published_at"] or "",
        "createdAt": row["created_at"],
    }


def load_meeting(conn: sqlite3.Connection, company_id: int, meeting_id: int) -> dict[str, Any] | None:
    row = conn.execute(
        "SELECT * FROM meetings WHERE company_id = ? AND id = ?", (company_id, meeting_id)
    ).fetchone()
    if not row:
        return None
    ata = meeting_row_to_dict(row)
    ata["participants"] = [
        {
            "id": int(p["id"]),
            "personName": p["person_name"],
            "personKey": p["person_key"],
            "userId": p["user_id"],
            "hasLogin": p["user_id"] is not None,
            "unitName": p["unit_name"] or "",
            "acknowledgedAt": p["acknowledged_at"] or "",
            "feedback": p["feedback"] or "",
            "feedbackAt": p["feedback_at"] or "",
        }
        for p in conn.execute(
            "SELECT * FROM meeting_participants WHERE meeting_id = ? ORDER BY person_name",
            (meeting_id,),
        ).fetchall()
    ]
    ata["attachments"] = [
        {
            "id": int(a["id"]),
            "fileName": a["file_name"],
            "sizeBytes": int(a["size_bytes"] or 0),
            "contentType": a["content_type"] or "",
            "createdAt": a["created_at"],
        }
        for a in conn.execute(
            "SELECT * FROM meeting_attachments WHERE meeting_id = ? ORDER BY id", (meeting_id,)
        ).fetchall()
    ]
    total = len(ata["participants"])
    cientes = sum(1 for p in ata["participants"] if p["acknowledgedAt"])
    ata["participantCount"] = total
    ata["acknowledgedCount"] = cientes
    ata["pendingCount"] = total - cientes
    ata["feedbackCount"] = sum(1 for p in ata["participants"] if p["feedback"])
    return ata


def list_meetings(
    conn: sqlite3.Connection, company_id: int, user: sqlite3.Row,
    search: str = "", kind: str = "", date_from: str = "", date_to: str = "",
    only_mine: bool = False,
) -> list[dict[str, Any]]:
    """Acervo de atas com o recorte do usuário.

    Gestor vê as atas das unidades dele; diretoria vê tudo. Vendedor só enxerga
    as reuniões em que esteve presente — rascunho nunca aparece para ele.
    """
    scope = data_scope_for_user(conn, user)
    sql = "SELECT m.* FROM meetings m WHERE m.company_id = ?"
    params: list[Any] = [company_id]

    minhas_chaves = user_person_keys(user)
    marcadores_chave = ",".join("?" for _ in minhas_chaves) or "''"

    if scope == "proprio":
        sql += (
            " AND m.status = 'PUBLICADA'"
            " AND EXISTS (SELECT 1 FROM meeting_participants p"
            f"             WHERE p.meeting_id = m.id"
            f"               AND (p.user_id = ? OR p.person_key IN ({marcadores_chave})))"
        )
        params.append(user["id"])
        params.extend(minhas_chaves)
    else:
        allowed = crm_allowed_units_for_user(conn, user)
        # allowed None = diretoria, sem restrição. Lista VAZIA significa gestor
        # sem unidade vinculada: nesse caso ele só enxerga o que é dele, nunca
        # a empresa inteira. O `if allowed:` anterior deixava a lista vazia
        # desligar o filtro e abria tudo — foi o que aconteceu com a Matriz.
        if allowed is not None:
            eu = normalize_upper(meeting_person_identity(user))
            if allowed:
                marcadores = ",".join("?" for _ in allowed)
                sql += (
                    " AND ("
                    f"    m.unit_name IN ({marcadores})"          # a unidade dele
                    "     OR m.visibility = 'EMPRESA'"            # liberada pelo autor
                    "     OR UPPER(m.organizer_name) = ?"         # ata que ele mesmo conduziu
                    ")"
                )
                params.extend(allowed)
                params.append(eu)
            else:
                sql += " AND (m.visibility = 'EMPRESA' OR UPPER(m.organizer_name) = ?)"
                params.append(eu)
            # Rascunho é trabalho em andamento: só o autor vê.
            sql += " AND (m.status <> 'RASCUNHO' OR UPPER(m.organizer_name) = ?)"
            params.append(eu)
        if only_mine:
            sql += " AND UPPER(m.organizer_name) = ?"
            params.append(normalize_upper(meeting_person_identity(user)))

    if kind and kind in MEETING_KIND_IDS:
        sql += " AND m.kind = ?"
        params.append(kind)
    if date_from:
        sql += " AND date(m.occurred_at) >= date(?)"
        params.append(date_from)
    if date_to:
        sql += " AND date(m.occurred_at) <= date(?)"
        params.append(date_to)
    termo = normalize_whitespace(search)
    if termo:
        alvo = f"%{termo.upper()}%"
        sql += (
            " AND (UPPER(m.title) LIKE ? OR UPPER(COALESCE(m.topic,'')) LIKE ?"
            "   OR UPPER(COALESCE(m.summary,'')) LIKE ? OR UPPER(COALESCE(m.agenda,'')) LIKE ?"
            "   OR UPPER(COALESCE(m.decisions,'')) LIKE ?"
            "   OR EXISTS (SELECT 1 FROM meeting_participants p2"
            "              WHERE p2.meeting_id = m.id AND UPPER(p2.person_name) LIKE ?))"
        )
        params.extend([alvo] * 6)

    sql += " ORDER BY datetime(m.occurred_at) DESC, m.id DESC LIMIT 300"
    linhas = conn.execute(sql, params).fetchall()

    resultado: list[dict[str, Any]] = []
    for row in linhas:
        ata = meeting_row_to_dict(row)
        contagem = conn.execute(
            "SELECT COUNT(*) total,"
            "       SUM(CASE WHEN acknowledged_at IS NOT NULL AND acknowledged_at <> '' THEN 1 ELSE 0 END) cientes,"
            "       SUM(CASE WHEN feedback IS NOT NULL AND feedback <> '' THEN 1 ELSE 0 END) feedbacks"
            " FROM meeting_participants WHERE meeting_id = ?",
            (ata["id"],),
        ).fetchone()
        ata["participantCount"] = int(contagem["total"] or 0)
        ata["acknowledgedCount"] = int(contagem["cientes"] or 0)
        ata["pendingCount"] = ata["participantCount"] - ata["acknowledgedCount"]
        ata["feedbackCount"] = int(contagem["feedbacks"] or 0)
        ata["attachmentCount"] = int(conn.execute(
            "SELECT COUNT(*) n FROM meeting_attachments WHERE meeting_id = ?", (ata["id"],)
        ).fetchone()["n"] or 0)
        minha = conn.execute(
            "SELECT acknowledged_at, feedback FROM meeting_participants "
            f"WHERE meeting_id = ? AND (user_id = ? OR person_key IN ({marcadores_chave}))",
            (ata["id"], user["id"], *minhas_chaves),
        ).fetchone()
        ata["iAmParticipant"] = bool(minha)
        ata["myAcknowledgedAt"] = (minha["acknowledged_at"] or "") if minha else ""
        ata["myFeedback"] = (minha["feedback"] or "") if minha else ""
        resultado.append(ata)
    return resultado


def count_pending_acknowledgements(conn: sqlite3.Connection, company_id: int, user: sqlite3.Row) -> int:
    """Quantas atas publicadas ainda esperam a ciência deste usuário."""
    chaves = user_person_keys(user)
    marcadores = ",".join("?" for _ in chaves) or "''"
    row = conn.execute(
        f"""
        SELECT COUNT(*) n
        FROM meeting_participants p
        JOIN meetings m ON m.id = p.meeting_id
        WHERE m.company_id = ? AND m.status = 'PUBLICADA'
          AND (p.user_id = ? OR p.person_key IN ({marcadores}))
          AND (p.acknowledged_at IS NULL OR p.acknowledged_at = '')
        """,
        (company_id, user["id"], *chaves),
    ).fetchone()
    return int(row["n"] or 0)


def save_meeting(
    conn: sqlite3.Connection, company_id: int, user: sqlite3.Row, payload: dict[str, Any]
) -> dict[str, Any]:
    """Cria ou atualiza a ata. A lista de presença é substituída pela enviada.

    Presença mantida por chave: quem já deu ciência e continua na lista NÃO perde
    a confirmação quando o gestor edita a ata para corrigir um texto.
    """
    if not user_can_manage_meetings(conn, user):
        raise PermissionError("Apenas gestão registra atas.")

    meeting_id = payload.get("id")
    kind = normalize_upper(payload.get("kind")) or "REUNIAO"
    if kind not in MEETING_KIND_IDS:
        raise ValueError("Tipo inválido: use Reunião ou Treinamento.")
    title = normalize_whitespace(payload.get("title"))
    if not title:
        raise ValueError("Informe o assunto da reunião.")
    occurred_at = normalize_whitespace(payload.get("occurredAt")).replace("T", " ")
    if not occurred_at:
        raise ValueError("Informe a data do encontro.")

    # Unidade validada no servidor: o gerente não registra ata de outra equipe
    # nem publica ata corporativa, mesmo forjando a requisição.
    escopo_unidades = meeting_units_for_user(conn, user)
    unidade = normalize_unit(payload.get("unitName")) or ""
    if not escopo_unidades["canBeCorporate"]:
        if not escopo_unidades["units"]:
            raise ValueError(
                "Seu usuário não tem unidade vinculada. Peça ao administrador para vincular antes de registrar atas."
            )
        if unidade not in escopo_unidades["units"]:
            unidade = escopo_unidades["defaultUnit"]

    campos = (
        kind, title,
        normalize_whitespace(payload.get("topic")),
        unidade or None,
        occurred_at,
        int(payload.get("durationMin") or 0),
        normalize_whitespace(payload.get("location")),
        normalize_whitespace(payload.get("agenda")),
        normalize_whitespace(payload.get("summary")),
        normalize_whitespace(payload.get("decisions")),
        normalize_whitespace(payload.get("organizerName")) or meeting_person_identity(user),
        "EMPRESA" if normalize_upper(payload.get("visibility")) == "EMPRESA" else "UNIDADE",
    )

    if meeting_id:
        existente = conn.execute(
            "SELECT id FROM meetings WHERE company_id = ? AND id = ?", (company_id, int(meeting_id))
        ).fetchone()
        if not existente:
            raise ValueError("Ata não encontrada.")
        conn.execute(
            """
            UPDATE meetings SET kind=?, title=?, topic=?, unit_name=?, occurred_at=?,
                   duration_min=?, location=?, agenda=?, summary=?, decisions=?,
                   organizer_name=?, visibility=?, updated_at=?
            WHERE id = ?
            """,
            (*campos, now_iso(), int(meeting_id)),
        )
        meeting_id = int(meeting_id)
    else:
        cursor = conn.execute(
            """
            INSERT INTO meetings (company_id, kind, title, topic, unit_name, occurred_at,
                duration_min, location, agenda, summary, decisions, organizer_name, visibility,
                status, created_by_user_id, created_at)
            VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,'RASCUNHO',?,?)
            """,
            (company_id, *campos, user["id"], now_iso()),
        )
        meeting_id = int(cursor.lastrowid)

    enviados = payload.get("participants") or []
    chaves_enviadas: set[str] = set()
    for item in enviados:
        nome = normalize_whitespace(item.get("personName") if isinstance(item, dict) else item)
        chave = person_key(nome)
        if not nome or not chave or chave in chaves_enviadas:
            continue
        chaves_enviadas.add(chave)
        unidade = normalize_unit(item.get("unitName")) if isinstance(item, dict) else ""
        conn.execute(
            """
            INSERT INTO meeting_participants (meeting_id, person_name, person_key, user_id, unit_name, created_at)
            VALUES (?,?,?,?,?,?)
            ON CONFLICT(meeting_id, person_key) DO UPDATE SET
                person_name = excluded.person_name,
                user_id     = excluded.user_id,
                unit_name   = excluded.unit_name
            """,
            (meeting_id, nome, chave, resolve_user_for_person(conn, company_id, nome),
             unidade or None, now_iso()),
        )
    if chaves_enviadas:
        marcadores = ",".join("?" for _ in chaves_enviadas)
        conn.execute(
            f"DELETE FROM meeting_participants WHERE meeting_id = ? AND person_key NOT IN ({marcadores})",
            (meeting_id, *chaves_enviadas),
        )
    else:
        conn.execute("DELETE FROM meeting_participants WHERE meeting_id = ?", (meeting_id,))

    audit_log(conn, company_id, user["id"], "salvar", "meetings", str(meeting_id),
              {"titulo": title, "presentes": len(chaves_enviadas)})
    conn.commit()
    return {"meetingId": meeting_id}


def publish_meeting(
    conn: sqlite3.Connection, company_id: int, user: sqlite3.Row, meeting_id: int
) -> dict[str, Any]:
    if not user_can_manage_meetings(conn, user):
        raise PermissionError("Apenas gestão publica atas.")
    ata = conn.execute(
        "SELECT * FROM meetings WHERE company_id = ? AND id = ?", (company_id, meeting_id)
    ).fetchone()
    if not ata:
        raise ValueError("Ata não encontrada.")
    presentes = conn.execute(
        "SELECT COUNT(*) n FROM meeting_participants WHERE meeting_id = ?", (meeting_id,)
    ).fetchone()["n"]
    if not presentes:
        raise ValueError("Marque pelo menos um presente antes de publicar.")
    if not normalize_whitespace(ata["summary"]):
        raise ValueError("Preencha o que foi tratado antes de publicar.")
    conn.execute(
        "UPDATE meetings SET status='PUBLICADA', published_at=?, updated_at=? WHERE id=?",
        (now_iso(), now_iso(), meeting_id),
    )
    audit_log(conn, company_id, user["id"], "publicar", "meetings", str(meeting_id),
              {"presentes": int(presentes)})
    conn.commit()
    return {"published": True, "participants": int(presentes)}


def acknowledge_meeting(
    conn: sqlite3.Connection, company_id: int, user: sqlite3.Row,
    meeting_id: int, feedback: str = "",
) -> dict[str, Any]:
    """Registra a ciência do participante e, se houver, o feedback dele.

    A ciência é definitiva; o feedback pode ser complementado depois — o vendedor
    costuma confirmar na hora e lembrar da sugestão mais tarde.
    """
    ata = conn.execute(
        "SELECT status FROM meetings WHERE company_id = ? AND id = ?", (company_id, meeting_id)
    ).fetchone()
    if not ata:
        raise ValueError("Ata não encontrada.")
    if ata["status"] != "PUBLICADA":
        raise ValueError("Esta ata ainda não foi publicada.")

    chaves = user_person_keys(user)
    marcadores = ",".join("?" for _ in chaves) or "''"
    participante = conn.execute(
        "SELECT id, acknowledged_at FROM meeting_participants "
        f"WHERE meeting_id = ? AND (user_id = ? OR person_key IN ({marcadores}))",
        (meeting_id, user["id"], *chaves),
    ).fetchone()
    if not participante:
        raise ValueError("Você não consta na lista de presença desta reunião.")

    texto = normalize_whitespace(feedback)
    conn.execute(
        """
        UPDATE meeting_participants
        SET acknowledged_at = COALESCE(NULLIF(acknowledged_at, ''), ?),
            feedback        = CASE WHEN ? <> '' THEN ? ELSE feedback END,
            feedback_at     = CASE WHEN ? <> '' THEN ? ELSE feedback_at END
        WHERE id = ?
        """,
        (now_iso(), texto, texto, texto, now_iso(), participante["id"]),
    )
    audit_log(conn, company_id, user["id"], "ciencia", "meetings", str(meeting_id),
              {"pessoa": meeting_person_identity(user), "comFeedback": bool(texto)})
    conn.commit()
    return {"acknowledged": True, "hasFeedback": bool(texto)}


def save_meeting_attachment(
    conn: sqlite3.Connection, company_id: int, user: sqlite3.Row,
    meeting_id: int, file_name: str, content: bytes,
) -> dict[str, Any]:
    if not user_can_manage_meetings(conn, user):
        raise PermissionError("Apenas gestão anexa documentos.")
    if not conn.execute(
        "SELECT 1 FROM meetings WHERE company_id = ? AND id = ?", (company_id, meeting_id)
    ).fetchone():
        raise ValueError("Ata não encontrada.")

    nome_limpo = Path(normalize_whitespace(file_name) or "arquivo").name
    extensao = Path(nome_limpo).suffix.lower()
    if extensao not in MEETING_ATTACHMENT_ALLOWED_EXT:
        raise ValueError(f"Tipo de arquivo não aceito ({extensao or 'sem extensão'}).")
    if len(content) > MEETING_ATTACHMENT_MAX_BYTES:
        limite = MEETING_ATTACHMENT_MAX_BYTES // (1024 * 1024)
        raise ValueError(f"Arquivo acima de {limite} MB.")
    if not content:
        raise ValueError("Arquivo vazio.")

    # Nome em disco é gerado: o nome original vai para o banco. Evita colisão e
    # impede que um nome com "../" escape da pasta de anexos.
    stored = f"{meeting_id}_{uuid.uuid4().hex}{extensao}"
    (meeting_files_dir() / stored).write_bytes(content)

    cursor = conn.execute(
        """
        INSERT INTO meeting_attachments
            (meeting_id, file_name, stored_name, content_type, size_bytes, uploaded_by_user_id, created_at)
        VALUES (?,?,?,?,?,?,?)
        """,
        (meeting_id, nome_limpo, stored,
         mimetypes.guess_type(nome_limpo)[0] or "application/octet-stream",
         len(content), user["id"], now_iso()),
    )
    conn.commit()
    return {"attachmentId": int(cursor.lastrowid), "fileName": nome_limpo, "sizeBytes": len(content)}


def delete_meeting_attachment(
    conn: sqlite3.Connection, company_id: int, user: sqlite3.Row, attachment_id: int
) -> None:
    if not user_can_manage_meetings(conn, user):
        raise PermissionError("Apenas gestão remove anexos.")
    row = conn.execute(
        """
        SELECT a.id, a.stored_name FROM meeting_attachments a
        JOIN meetings m ON m.id = a.meeting_id
        WHERE a.id = ? AND m.company_id = ?
        """,
        (attachment_id, company_id),
    ).fetchone()
    if not row:
        raise ValueError("Anexo não encontrado.")
    caminho = meeting_files_dir() / row["stored_name"]
    if caminho.exists():
        caminho.unlink()
    conn.execute("DELETE FROM meeting_attachments WHERE id = ?", (attachment_id,))
    conn.commit()


def delete_meeting(
    conn: sqlite3.Connection, company_id: int, user: sqlite3.Row, meeting_id: int
) -> None:
    if not user_can_manage_meetings(conn, user):
        raise PermissionError("Apenas gestão exclui atas.")
    for row in conn.execute(
        "SELECT stored_name FROM meeting_attachments WHERE meeting_id = ?", (meeting_id,)
    ).fetchall():
        caminho = meeting_files_dir() / row["stored_name"]
        if caminho.exists():
            caminho.unlink()
    conn.execute("DELETE FROM meeting_attachments WHERE meeting_id = ?", (meeting_id,))
    conn.execute("DELETE FROM meeting_participants WHERE meeting_id = ?", (meeting_id,))
    conn.execute("DELETE FROM meetings WHERE company_id = ? AND id = ?", (company_id, meeting_id))
    audit_log(conn, company_id, user["id"], "excluir", "meetings", str(meeting_id), {})
    conn.commit()


# ─────────────────────────────────────────────────────────────────────────────
# Feedback e PDI
#
# Conceito: o feedback não inventa critério. Ele lê os números que o CRM já tem
# e os confronta com o que o MEC define como execução esperada. O gerente marca
# três níveis por item, escreve pouco e sai com no máximo 2 pontos de PDI.
#
# Decisões que moldam o resto:
#
# 1. UM feedback por pessoa por competência (UNIQUE). Editar o do mês corrige;
#    não cria um segundo registro. Histórico limpo para comparar evolução.
#
# 2. PDI é da PESSOA, não do feedback. Item aberto em março continua vivo em
#    maio, com o histórico de evolução. Amarrar ao feedback faria o
#    desenvolvimento recomeçar do zero todo mês.
#
# 3. Duas caixas de observação para quem recebe: uma vai para o gestor, outra
#    NÃO. A confidencial só é lida por quem gerencia usuários (Diretor/Admin).
#    Sem esse canal separado, ninguém escreve o que realmente pensa.
# ─────────────────────────────────────────────────────────────────────────────

FEEDBACK_KINDS = [
    {"id": "VENDEDOR", "label": "Vendedor", "icon": "👤",
     "hint": "Gerente × vendedor. Execução do MEC, indicadores do mês e comportamento."},
    {"id": "GERENTE", "label": "Gerente", "icon": "🎯",
     "hint": "Diretor × gerente. Resultado da unidade, gestão da equipe e direcionamento tático."},
]
FEEDBACK_KIND_IDS = {k["id"] for k in FEEDBACK_KINDS}

PDI_STATUSES = [
    {"id": "ABERTO",    "label": "Aberto",     "color": "#5f6368", "bg": "#f1f3f4"},
    {"id": "EVOLUINDO", "label": "Evoluindo",  "color": "#b06000", "bg": "#fef7e0"},
    {"id": "CONCLUIDO", "label": "Concluído",  "color": "#1e8e3e", "bg": "#e6f4ea"},
    {"id": "CANCELADO", "label": "Cancelado",  "color": "#5f6368", "bg": "#f1f3f4"},
]
PDI_STATUS_IDS = {s["id"] for s in PDI_STATUSES}

# Teto de itens ativos por pessoa. Plano com cinco frentes não sai do lugar —
# o limite é a própria regra de gestão, não uma restrição técnica.
PDI_MAX_ACTIVE = 3


def _mec_content() -> Any:
    """Carrega o catálogo do MEC. Falha explícita: sem ele o módulo não faz sentido."""
    import mec_feedback
    return mec_feedback


def feedback_items_for_kind(kind: str) -> list[dict[str, Any]]:
    mec = _mec_content()
    return mec.FEEDBACK_ITEMS_MANAGER if kind == "GERENTE" else mec.FEEDBACK_ITEMS_SELLER


def feedback_groups_for_kind(kind: str) -> list[str]:
    mec = _mec_content()
    return mec.FEEDBACK_ITEM_GROUPS_MANAGER if kind == "GERENTE" else mec.FEEDBACK_ITEM_GROUPS_SELLER


def user_can_give_feedback(conn: sqlite3.Connection, user: sqlite3.Row) -> bool:
    """Dá feedback quem tem visão de equipe. Vendedor apenas recebe."""
    return data_scope_for_user(conn, user) != "proprio"


def user_can_read_confidential(conn: sqlite3.Connection, user: sqlite3.Row) -> bool:
    """Lê a observação confidencial quem gerencia usuários — Diretor e Admin."""
    return user_can_manage_users(conn, user)


# ── Indicadores do avaliado no mês ───────────────────────────────────────────

def safe_feedback_indicators(
    conn: sqlite3.Connection, company_id: int, kind: str,
    person_name: str, unit_name: str, competence: str,
) -> dict[str, Any]:
    """Indicadores com rede de proteção.

    Falha ao ler número NÃO pode impedir o gerente de registrar a conversa —
    o feedback continua valendo sem o painel. Devolve o motivo para a tela
    mostrar em vez de ficar em branco.
    """
    try:
        if kind == "GERENTE":
            return unit_indicators_for_feedback(conn, company_id, unit_name, competence)
        return seller_indicators_for_feedback(conn, company_id, person_name, competence)
    except Exception as exc:
        print(f"[feedback] indicadores indisponíveis ({kind} / {person_name} / {competence}): {exc}", flush=True)
        traceback.print_exc()
        return {
            "found": False,
            "competence": competence,
            "sellerName": person_name,
            "reason": f"Não foi possível calcular os indicadores ({exc}).",
        }


def seller_name_variants(conn: sqlite3.Connection, company_id: int, *nomes: str) -> list[str]:
    """Todas as grafias do mesmo vendedor que existem nas tabelas.

    O faturamento grava "THIELLY HENRIQUES ROCHA (VENDAS)", o CRM grava o nome
    que está no cadastro do usuário, e o cadastro de pessoas pode ter uma
    terceira variação. Contar ligações comparando uma grafia só devolvia zero.
    A comparação usa person_key, que ignora acento, pontuação e o sufixo entre
    parênteses — sem recorte curto, para não somar a atividade de outra pessoa.
    """
    chaves = {person_key(n) for n in nomes if normalize_whitespace(n)}
    chaves.discard("")
    if not chaves:
        return []
    encontrados = {normalize_whitespace(n) for n in nomes if normalize_whitespace(n)}
    for tabela in ("crm_interactions", "crm_client_summary", "fact_sales_detail"):
        try:
            linhas = conn.execute(
                f"SELECT DISTINCT seller_name FROM {tabela} WHERE company_id = ?", (company_id,)
            ).fetchall()
        except sqlite3.OperationalError:
            continue
        for r in linhas:
            nome = normalize_whitespace(r["seller_name"])
            if nome and person_key(nome) in chaves:
                encontrados.add(nome)
    return sorted(encontrados)


def seller_indicators_for_feedback(
    conn: sqlite3.Connection, company_id: int, seller_name: str, competence: str
) -> dict[str, Any]:
    """Números do vendedor na competência, já comparados com a unidade dele.

    A comparação com a unidade importa mais que o valor absoluto: ticket de
    R$ 300 pode ser bom ou ruim dependendo da praça. Sem referência, o gerente
    discute o número em vez de discutir a causa.
    """
    filtros = build_filters_from_query({})
    filtros["competence_start"] = competence
    filtros["competence_end"] = competence
    dados = get_dashboard_data_cached(conn, company_id, filtros)

    # O nome do cadastro de pessoas e o nome que aparece no faturamento raramente
    # são idênticos ("THIELLY HENRIQUES ROCHA (VENDAS)" x "THIELLY HENRIQUES
    # ROCHA"). Comparar string crua devolvia "sem dados" mesmo com o vendedor
    # faturando no mês. A comparação usa a chave normalizada, com o recorte de
    # dois tokens como último recurso.
    ranking = dados.get("sellerRanking", [])
    alvo_completo = person_key(seller_name)
    alvo_curto = short_person_key(seller_name)

    linha = next((r for r in ranking if person_key(r["sellerName"]) == alvo_completo), None)
    if not linha and alvo_curto:
        candidatos = [r for r in ranking if short_person_key(r["sellerName"]) == alvo_curto]
        linha = candidatos[0] if len(candidatos) == 1 else None

    if not linha:
        # Vendedor de unidade nova ainda não fatura, então não existe no ranking.
        # Em vez de devolver "sem dados", devolve a fase e os números de esforço:
        # é o que permite dar feedback antes da primeira venda.
        unidade_pessoa = seller_unit_name(conn, company_id, seller_name, competence)
        em_implantacao = unit_is_in_deployment(conn, company_id, unidade_pessoa, competence)
        if em_implantacao:
            inicio_i = first_day_of_competence(competence).isoformat()
            fim_i = last_day_of_competence(competence).isoformat()
            variantes_i = seller_name_variants(conn, company_id, seller_name)
            marcadores_i = ",".join("?" for _ in variantes_i) or "''"
            ligacoes_i = conn.execute(
                f"""SELECT COUNT(*) n FROM crm_interactions
                    WHERE company_id = ? AND contact_type_code = 'LIGACAO' AND initiative = 'ATIVO'
                      AND UPPER(seller_name) IN ({marcadores_i})
                      AND date(substr(replace(occurred_at,'T',' '),1,10)) BETWEEN date(?) AND date(?)""",
                (company_id, *[normalize_upper(v) for v in variantes_i], inicio_i, fim_i),
            ).fetchone()["n"]
            return {
                "found": False,
                "inDeployment": True,
                "competence": competence,
                "sellerName": seller_name,
                "unitName": unidade_pessoa,
                "calls": int(ligacoes_i or 0),
                "activity": activity_progress(conn, company_id, competence, unidade_pessoa, seller_name),
                "reason": "Unidade em implantação — ainda sem faturamento e sem meta.",
            }
        return {
            "found": False,
            "competence": competence,
            "sellerName": seller_name,
            # Devolve com quem tentou casar: sem isso o gerente vê "sem dados"
            # e não tem como descobrir que o problema é divergência de nome.
            "reason": "Este nome não aparece no faturamento desta competência.",
            "availableNames": sorted(r["sellerName"] for r in ranking
                                     if float(r.get("revenueNet") or 0) > 0)[:40],
        }

    # A partir daqui, usa o nome COMO ESTÁ NO FATURAMENTO para as demais
    # consultas — é ele que está gravado em interações e na carteira.
    nome_vendas = linha["sellerName"]
    unidade = normalize_unit(linha.get("baseUnit"))
    pares = [r for r in dados.get("sellerRanking", [])
             if normalize_unit(r.get("baseUnit")) == unidade and float(r.get("revenueNet") or 0) > 0]

    def media(campo: str) -> float:
        valores = [float(r.get(campo) or 0) for r in pares]
        return round(sum(valores) / len(valores), 2) if valores else 0.0

    # Ligações registradas no mês (mínimo do MEC: 3/dia, 60/mês)
    inicio = first_day_of_competence(competence).isoformat()
    fim = last_day_of_competence(competence).isoformat()
    variantes = seller_name_variants(conn, company_id, seller_name, nome_vendas)
    marcadores = ",".join("?" for _ in variantes) or "''"
    nomes_upper = [normalize_upper(v) for v in variantes]

    ligacoes = conn.execute(
        f"""
        SELECT COUNT(*) n FROM crm_interactions
        WHERE company_id = ? AND UPPER(seller_name) IN ({marcadores})
          AND contact_type_code = 'LIGACAO' AND initiative = 'ATIVO'
          AND date(substr(replace(occurred_at,'T',' '),1,10)) BETWEEN date(?) AND date(?)
        """,
        (company_id, *nomes_upper, inicio, fim),
    ).fetchone()["n"]
    contatos = conn.execute(
        f"""
        SELECT COUNT(*) n FROM crm_interactions
        WHERE company_id = ? AND UPPER(seller_name) IN ({marcadores})
          AND initiative = 'ATIVO'
          AND date(substr(replace(occurred_at,'T',' '),1,10)) BETWEEN date(?) AND date(?)
        """,
        (company_id, *nomes_upper, inicio, fim),
    ).fetchone()["n"]

    # Situação da carteira. O status (ATIVO / PRÉ-INATIVO / INATIVO) é CALCULADO
    # a partir dos dias sem compra — não existe coluna no banco. Reaproveita a
    # mesma função que monta a tela de Carteira, para o feedback mostrar
    # exatamente o número que o vendedor vê lá.
    filtros_carteira = build_filters_from_query({})
    filtros_carteira["competence_start"] = competence
    filtros_carteira["competence_end"] = competence
    filtros_carteira["seller_name"] = nome_vendas
    por_status: dict[str, int] = {}
    try:
        for cliente in list_crm_clients(conn, company_id, filtros_carteira, attach_context=False):
            codigo = normalize_upper(cliente.get("statusCode")) or "SEM_STATUS"
            por_status[codigo] = por_status.get(codigo, 0) + 1
    except Exception as exc:  # carteira indisponível não pode derrubar o feedback
        print(f"[feedback] carteira indisponível para {nome_vendas}: {exc}", flush=True)
    total_carteira = sum(por_status.values())

    # Meta de ligações proporcional ao mês decorrido. O piso do MEC é 60 NO MÊS;
    # cobrar 60 no dia 6 marcaria todo vendedor como irregular.
    dias_decorridos = int(linha.get("sellerElapsedWorkingDays") or 0)
    dias_totais = int(linha.get("sellerWorkingDays") or 0)
    meta_ligacoes_mes = 60
    meta_ligacoes_ate_hoje = (round(meta_ligacoes_mes * dias_decorridos / dias_totais)
                              if dias_decorridos and dias_totais else meta_ligacoes_mes)

    return {
        "found": True,
        "competence": competence,
        "inDeployment": bool(linha.get("inDeployment")),
        "sellerName": nome_vendas,
        "matchedName": nome_vendas if person_key(nome_vendas) != person_key(seller_name) else "",
        "unitName": unidade,
        "revenueNet": linha.get("revenueNet"),
        "revenueGoal": linha.get("revenueGoal"),
        "goalAttainmentPct": linha.get("goalAttainmentPct"),
        # Atingimento projetado para o fim do mês. É o número que diz se a
        # pessoa está NO RITMO — comparar o acumulado do dia 6 com 100% da meta
        # acusaria todo mundo no começo do mês.
        "projectedGoalAttainmentPct": linha.get("projectedGoalAttainmentPct"),
        "elapsedWorkingDays": linha.get("sellerElapsedWorkingDays"),
        "totalWorkingDays": linha.get("sellerWorkingDays"),
        "ticketAverage": linha.get("ticketAverage"),
        "ticketAverageUnit": media("ticketAverage"),
        "distinctClients": linha.get("distinctClients"),
        "distinctClientsUnit": round(media("distinctClients")),
        "mixSku": linha.get("mixSku"),
        "returnsPct": round(safe_div(float(linha.get("returnsValue") or 0),
                                     float(linha.get("revenueNet") or 0)) * 100, 2),
        "discountPct": linha.get("discountPct"),
        "discountPctUnit": media("discountPct"),
        "calls": int(ligacoes),
        "callsTarget": meta_ligacoes_mes,
        "callsTargetToDate": meta_ligacoes_ate_hoje,
        "contacts": int(contatos),
        "portfolioTotal": total_carteira,
        "portfolioActive": por_status.get("ATIVO", 0),
        "portfolioPreInactive": por_status.get("PRE_INATIVO", 0),
        "portfolioInactive": por_status.get("INATIVO", 0),
    }


def feedback_guidance(indicadores: dict[str, Any]) -> list[dict[str, Any]]:
    """Orientações do guia que se aplicam a ESTE vendedor neste mês.

    Só entra o que está fora do esperado. Guia completo em toda conversa vira
    manual que ninguém lê; três pontos relevantes o gerente usa.
    """
    mec = _mec_content()
    guia = mec.INDICATOR_GUIDE
    if not indicadores.get("found"):
        # Sem faturamento o guia normal não se aplica, mas unidade em implantação
        # tem conversa própria — e é justamente ela que precisa acontecer agora.
        if indicadores.get("inDeployment") and "deployment" in guia:
            return [{"id": "deployment", **guia["deployment"]}]
        return []

    alertas: list[str] = []
    # Mede RITMO, não acumulado. No dia 6 de 21 ninguém tem 100% da meta nem 60
    # ligações — comparar com o mês fechado acusaria a equipe inteira e o alerta
    # perderia o sentido, que é o mesmo raciocínio dos faróis.
    # Unidade em implantação não tem meta de faturamento. Cobrar meta de quem
    # ainda está montando carteira é injusto e destrói a credibilidade do guia.
    em_implantacao = bool(indicadores.get("inDeployment"))
    ritmo_meta = indicadores.get("projectedGoalAttainmentPct")
    if ritmo_meta is None:
        ritmo_meta = indicadores.get("goalAttainmentPct")
    if not em_implantacao and float(ritmo_meta or 0) < 90:
        alertas.append("goal_low")
    meta_ligacoes = int(indicadores.get("callsTargetToDate")
                        or indicadores.get("callsTarget") or 60)
    if int(indicadores.get("calls") or 0) < meta_ligacoes:
        alertas.append("calls_low")
    if float(indicadores.get("returnsPct") or 0) > 3:
        alertas.append("returns_high")
    desconto = float(indicadores.get("discountPct") or 0)
    desconto_unidade = float(indicadores.get("discountPctUnit") or 0)
    if desconto_unidade and desconto > desconto_unidade * 1.15:
        alertas.append("discount_high")
    clientes = int(indicadores.get("distinctClients") or 0)
    clientes_unidade = int(indicadores.get("distinctClientsUnit") or 0)
    if clientes_unidade and clientes < clientes_unidade * 0.8:
        alertas.append("clients_low")
    if int(indicadores.get("portfolioTotal") or 0):
        parados = int(indicadores.get("portfolioInactive") or 0) + int(indicadores.get("portfolioPreInactive") or 0)
        if safe_div(parados, indicadores["portfolioTotal"]) > 0.35:
            alertas.append("inactive_high")
    ticket = float(indicadores.get("ticketAverage") or 0)
    ticket_unidade = float(indicadores.get("ticketAverageUnit") or 0)
    if ticket_unidade and ticket < ticket_unidade * 0.85:
        alertas.append("ticket_low")

    if em_implantacao:
        alertas.append("deployment")
    if not alertas:
        alertas.append("good_overall")

    return [{"id": chave, **guia[chave]} for chave in dict.fromkeys(alertas) if chave in guia][:4]


def unit_indicators_for_feedback(
    conn: sqlite3.Connection, company_id: int, unit_name: str, competence: str
) -> dict[str, Any]:
    """Números da unidade para o feedback do gerente, com o retrato dos ritos."""
    filtros = build_filters_from_query({})
    filtros["competence_start"] = competence
    filtros["competence_end"] = competence
    filtros["unit_name"] = normalize_unit(unit_name)
    dados = get_dashboard_data_cached(conn, company_id, filtros)
    resumo = dados.get("summary") or {}

    inicio = first_day_of_competence(competence).isoformat()
    fim = last_day_of_competence(competence).isoformat()
    reunioes = conn.execute(
        "SELECT COUNT(*) n FROM meetings WHERE company_id = ? AND status = 'PUBLICADA' "
        "AND date(occurred_at) BETWEEN date(?) AND date(?) AND (unit_name = ? OR unit_name IS NULL)",
        (company_id, inicio, fim, normalize_unit(unit_name)),
    ).fetchone()["n"]
    feedbacks_feitos = conn.execute(
        "SELECT COUNT(*) n FROM feedbacks WHERE company_id = ? AND kind = 'VENDEDOR' "
        "AND competence = ? AND unit_name = ? AND status = 'PUBLICADO'",
        (company_id, competence, normalize_unit(unit_name)),
    ).fetchone()["n"]
    equipe = conn.execute(
        "SELECT COUNT(DISTINCT person_name) n FROM people_records "
        "WHERE company_id = ? AND base_unit = ? AND role_classification = 'Vendedor' "
        "AND (valid_to IS NULL OR valid_to = '' OR date(valid_to) >= date(?))",
        (company_id, normalize_unit(unit_name), inicio),
    ).fetchone()["n"]
    pdis = conn.execute(
        "SELECT COUNT(*) n FROM pdi_items WHERE company_id = ? AND unit_name = ? "
        "AND status IN ('ABERTO','EVOLUINDO')",
        (company_id, normalize_unit(unit_name)),
    ).fetchone()["n"]
    pdis_vencidos = conn.execute(
        "SELECT COUNT(*) n FROM pdi_items WHERE company_id = ? AND unit_name = ? "
        "AND status IN ('ABERTO','EVOLUINDO') AND due_date IS NOT NULL AND due_date <> '' "
        "AND date(due_date) < date(?)",
        (company_id, normalize_unit(unit_name), today_in_brazil().isoformat()),
    ).fetchone()["n"]

    return {
        "found": bool(resumo),
        "competence": competence,
        "unitName": normalize_unit(unit_name),
        "revenueNet": resumo.get("revenueNet"),
        "revenueGoal": resumo.get("revenueGoal"),
        "goalAttainmentPct": resumo.get("goalAttainmentPct"),
        "returnsPct": resumo.get("returnsPct"),
        "discountPct": resumo.get("discountPct"),
        "ticketAverage": resumo.get("ticketAverage"),
        "teamSize": int(equipe),
        "meetingsPublished": int(reunioes),
        "feedbacksDone": int(feedbacks_feitos),
        "feedbacksExpected": int(equipe),
        "pdiActive": int(pdis),
        "pdiOverdue": int(pdis_vencidos),
    }


# ── Leitura e escrita do feedback ────────────────────────────────────────────

def feedback_row_to_dict(row: sqlite3.Row) -> dict[str, Any]:
    return {
        "id": int(row["id"]),
        "kind": row["kind"],
        "kindLabel": next((k["label"] for k in FEEDBACK_KINDS if k["id"] == row["kind"]), row["kind"]),
        "competence": row["competence"],
        "personName": row["person_name"],
        "personKey": row["person_key"],
        "unitName": row["unit_name"] or "",
        "authorName": row["author_name"],
        "highlights": row["highlights"] or "",
        "improvements": row["improvements"] or "",
        "agreements": row["agreements"] or "",
        "tacticalGoal": row["tactical_goal"] or "",
        "tacticalReality": row["tactical_reality"] or "",
        "tacticalOptions": row["tactical_options"] or "",
        "tacticalWill": row["tactical_will"] or "",
        "status": row["status"],
        "publishedAt": row["published_at"] or "",
        "acknowledgedAt": row["acknowledged_at"] or "",
        "personNote": row["person_note"] or "",
        "noteAt": row["note_at"] or "",
        "createdAt": row["created_at"],
        "indicators": json.loads(row["indicators_json"]) if row["indicators_json"] else {},
    }


def load_feedback(
    conn: sqlite3.Connection, company_id: int, feedback_id: int, user: sqlite3.Row
) -> dict[str, Any] | None:
    row = conn.execute(
        "SELECT * FROM feedbacks WHERE company_id = ? AND id = ?", (company_id, feedback_id)
    ).fetchone()
    if not row:
        return None
    fb = feedback_row_to_dict(row)

    fb["ratings"] = {
        r["item_id"]: {"level": r["level"], "comment": r["comment"] or ""}
        for r in conn.execute(
            "SELECT item_id, level, comment FROM feedback_ratings WHERE feedback_id = ?", (feedback_id,)
        ).fetchall()
    }
    fb["items"] = feedback_items_for_kind(fb["kind"])
    fb["groups"] = feedback_groups_for_kind(fb["kind"])
    fb["guidance"] = feedback_guidance(fb["indicators"]) if fb["kind"] == "VENDEDOR" else []
    fb["script"] = (_mec_content().FEEDBACK_SCRIPT_MANAGER if fb["kind"] == "GERENTE"
                    else _mec_content().FEEDBACK_SCRIPT_SELLER)
    fb["pdi"] = list_pdi_items(conn, company_id, fb["personKey"])
    # Registros pontuais do mesmo mês: viram a memória da conversa. Sem isso o
    # gerente escreve de cabeça e só lembra da última semana.
    fb["notes"] = list_feedback_notes(conn, company_id, user, fb["personKey"], fb["competence"])

    # A confidencial NÃO acompanha o objeto por padrão. Só quem tem direito recebe.
    if user_can_read_confidential(conn, user):
        fb["confidentialNote"] = row["confidential_note"] or ""
        fb["canReadConfidential"] = True
    else:
        fb["canReadConfidential"] = False
        fb["hasConfidentialNote"] = bool(normalize_whitespace(row["confidential_note"]))

    minhas = set(user_person_keys(user))
    fb["isMe"] = fb["personKey"] in minhas or (
        row["person_user_id"] is not None and row["person_user_id"] == user["id"]
    )
    return fb


def list_feedbacks(
    conn: sqlite3.Connection, company_id: int, user: sqlite3.Row,
    kind: str = "", competence: str = "", person: str = "",
) -> list[dict[str, Any]]:
    scope = data_scope_for_user(conn, user)
    minhas = user_person_keys(user)
    marcadores = ",".join("?" for _ in minhas) or "''"

    sql = "SELECT * FROM feedbacks WHERE company_id = ?"
    params: list[Any] = [company_id]

    if scope == "proprio":
        # O vendedor só enxerga o próprio feedback, e só depois de publicado.
        sql += f" AND status = 'PUBLICADO' AND (person_user_id = ? OR person_key IN ({marcadores}))"
        params.append(user["id"])
        params.extend(minhas)
    else:
        permitidas = crm_allowed_units_for_user(conn, user)
        # Lista vazia = gestor sem unidade vinculada. Vê só o próprio feedback,
        # nunca a empresa inteira.
        if permitidas is not None:
            if permitidas:
                unidades = ",".join("?" for _ in permitidas)
                sql += f" AND (unit_name IN ({unidades}) OR person_key IN ({marcadores}))"
                params.extend(permitidas)
            else:
                sql += f" AND person_key IN ({marcadores})"
            params.extend(minhas)

    if kind in FEEDBACK_KIND_IDS:
        sql += " AND kind = ?"
        params.append(kind)
    if competence:
        sql += " AND competence = ?"
        params.append(competence)
    if person:
        sql += " AND UPPER(person_name) LIKE ?"
        params.append(f"%{person.upper()}%")

    sql += " ORDER BY competence DESC, person_name LIMIT 300"

    resultado = []
    for row in conn.execute(sql, params).fetchall():
        fb = feedback_row_to_dict(row)
        fb["isMe"] = fb["personKey"] in set(minhas) or (
            row["person_user_id"] is not None and row["person_user_id"] == user["id"]
        )
        fb["ratingSummary"] = {
            nivel["id"]: conn.execute(
                "SELECT COUNT(*) n FROM feedback_ratings WHERE feedback_id = ? AND level = ?",
                (fb["id"], nivel["id"]),
            ).fetchone()["n"]
            for nivel in _mec_content().FEEDBACK_LEVELS
        }
        fb["hasNote"] = bool(fb["personNote"])
        fb["hasConfidentialNote"] = bool(normalize_whitespace(row["confidential_note"]))
        resultado.append(fb)
    return resultado


def count_pending_feedback_ack(conn: sqlite3.Connection, company_id: int, user: sqlite3.Row) -> int:
    chaves = user_person_keys(user)
    marcadores = ",".join("?" for _ in chaves) or "''"
    row = conn.execute(
        f"""
        SELECT COUNT(*) n FROM feedbacks
        WHERE company_id = ? AND status = 'PUBLICADO'
          AND (person_user_id = ? OR person_key IN ({marcadores}))
          AND (acknowledged_at IS NULL OR acknowledged_at = '')
        """,
        (company_id, user["id"], *chaves),
    ).fetchone()
    return int(row["n"] or 0)


def save_feedback(
    conn: sqlite3.Connection, company_id: int, user: sqlite3.Row, payload: dict[str, Any]
) -> dict[str, Any]:
    if not user_can_give_feedback(conn, user):
        raise PermissionError("Apenas gestão registra feedback.")

    kind = normalize_upper(payload.get("kind")) or "VENDEDOR"
    if kind not in FEEDBACK_KIND_IDS:
        raise ValueError("Tipo de feedback inválido.")
    # Feedback de gerente é conversa de diretoria — gerente não avalia gerente.
    if kind == "GERENTE" and not user_can_manage_users(conn, user):
        raise PermissionError("Feedback de gerente é conduzido pela diretoria.")

    person_name = normalize_whitespace(payload.get("personName"))
    if not person_name:
        raise ValueError("Selecione quem vai receber o feedback.")
    competence = normalize_whitespace(payload.get("competence"))
    if not competence or len(competence) != 7:
        raise ValueError("Informe a competência no formato AAAA-MM.")

    chave = person_key(person_name)
    unidade = normalize_unit(payload.get("unitName")) or None
    permitidas = crm_allowed_units_for_user(conn, user)
    if permitidas is not None:
        if not permitidas:
            raise ValueError("Seu usuário não tem unidade vinculada. Peça ao administrador para vincular.")
        if unidade not in permitidas:
            unidade = permitidas[0]

    # Foto dos indicadores no momento do feedback. Guardada junto porque o
    # dashboard muda com novas importações — e a conversa precisa continuar
    # fazendo sentido daqui a seis meses.
    indicadores = safe_feedback_indicators(conn, company_id, kind, person_name, unidade or "", competence)

    existente = conn.execute(
        "SELECT id, status FROM feedbacks WHERE company_id = ? AND kind = ? AND competence = ? AND person_key = ?",
        (company_id, kind, competence, chave),
    ).fetchone()

    campos = (
        normalize_whitespace(payload.get("highlights")),
        normalize_whitespace(payload.get("improvements")),
        normalize_whitespace(payload.get("agreements")),
        normalize_whitespace(payload.get("tacticalGoal")),
        normalize_whitespace(payload.get("tacticalReality")),
        normalize_whitespace(payload.get("tacticalOptions")),
        normalize_whitespace(payload.get("tacticalWill")),
        json.dumps(indicadores, ensure_ascii=False),
    )

    if existente:
        feedback_id = int(existente["id"])
        conn.execute(
            """
            UPDATE feedbacks SET highlights=?, improvements=?, agreements=?,
                   tactical_goal=?, tactical_reality=?, tactical_options=?, tactical_will=?,
                   indicators_json=?, unit_name=?, updated_at=?
            WHERE id = ?
            """,
            (*campos, unidade, now_iso(), feedback_id),
        )
    else:
        cursor = conn.execute(
            """
            INSERT INTO feedbacks (company_id, kind, competence, person_name, person_key,
                person_user_id, unit_name, author_name, author_user_id,
                highlights, improvements, agreements,
                tactical_goal, tactical_reality, tactical_options, tactical_will,
                indicators_json, status, created_at)
            VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?, 'RASCUNHO', ?)
            """,
            (company_id, kind, competence, person_name, chave,
             resolve_user_for_person(conn, company_id, person_name), unidade,
             meeting_person_identity(user), user["id"], *campos, now_iso()),
        )
        feedback_id = int(cursor.lastrowid)

    validos = {i["id"] for i in feedback_items_for_kind(kind)}
    for item_id, dados in (payload.get("ratings") or {}).items():
        if item_id not in validos:
            continue
        nivel = normalize_upper((dados or {}).get("level"))
        if nivel not in _mec_content().FEEDBACK_LEVEL_IDS:
            continue
        conn.execute(
            """
            INSERT INTO feedback_ratings (feedback_id, item_id, level, comment)
            VALUES (?,?,?,?)
            ON CONFLICT(feedback_id, item_id) DO UPDATE SET
                level = excluded.level, comment = excluded.comment
            """,
            (feedback_id, item_id, nivel, normalize_whitespace((dados or {}).get("comment"))),
        )

    audit_log(conn, company_id, user["id"], "salvar", "feedbacks", str(feedback_id),
              {"pessoa": person_name, "competencia": competence, "tipo": kind})
    conn.commit()
    return {"feedbackId": feedback_id}


def publish_feedback(
    conn: sqlite3.Connection, company_id: int, user: sqlite3.Row, feedback_id: int
) -> dict[str, Any]:
    if not user_can_give_feedback(conn, user):
        raise PermissionError("Apenas gestão publica feedback.")
    fb = conn.execute(
        "SELECT * FROM feedbacks WHERE company_id = ? AND id = ?", (company_id, feedback_id)
    ).fetchone()
    if not fb:
        raise ValueError("Feedback não encontrado.")

    avaliados = conn.execute(
        "SELECT COUNT(*) n FROM feedback_ratings WHERE feedback_id = ?", (feedback_id,)
    ).fetchone()["n"]
    total = len(feedback_items_for_kind(fb["kind"]))
    if avaliados < total:
        raise ValueError(f"Faltam {total - avaliados} item(ns) para avaliar antes de publicar.")
    if not normalize_whitespace(fb["agreements"]):
        raise ValueError("Preencha o que ficou combinado — é o que o vendedor leva da conversa.")

    conn.execute(
        "UPDATE feedbacks SET status='PUBLICADO', published_at=?, updated_at=? WHERE id=?",
        (now_iso(), now_iso(), feedback_id),
    )
    audit_log(conn, company_id, user["id"], "publicar", "feedbacks", str(feedback_id), {})
    conn.commit()
    return {"published": True}


def acknowledge_feedback(
    conn: sqlite3.Connection, company_id: int, user: sqlite3.Row,
    feedback_id: int, note: str = "", confidential: str = "",
) -> dict[str, Any]:
    """Ciência do avaliado, com dois canais de retorno separados.

    `note` vai para o gestor. `confidential` NÃO — fica visível apenas para quem
    gerencia usuários. Quem escreve precisa ter certeza de para onde vai, ou não
    escreve nada de útil.
    """
    fb = conn.execute(
        "SELECT * FROM feedbacks WHERE company_id = ? AND id = ?", (company_id, feedback_id)
    ).fetchone()
    if not fb:
        raise ValueError("Feedback não encontrado.")
    if fb["status"] != "PUBLICADO":
        raise ValueError("Este feedback ainda não foi publicado.")

    minhas = set(user_person_keys(user))
    sou_eu = fb["person_key"] in minhas or (
        fb["person_user_id"] is not None and fb["person_user_id"] == user["id"]
    )
    if not sou_eu:
        raise ValueError("Este feedback não é seu.")

    texto = normalize_whitespace(note)
    sigilo = normalize_whitespace(confidential)
    conn.execute(
        """
        UPDATE feedbacks
        SET acknowledged_at   = COALESCE(NULLIF(acknowledged_at, ''), ?),
            person_note       = CASE WHEN ? <> '' THEN ? ELSE person_note END,
            confidential_note = CASE WHEN ? <> '' THEN ? ELSE confidential_note END,
            note_at           = CASE WHEN ? <> '' OR ? <> '' THEN ? ELSE note_at END
        WHERE id = ?
        """,
        (now_iso(), texto, texto, sigilo, sigilo, texto, sigilo, now_iso(), feedback_id),
    )
    # A auditoria registra que existe uma confidencial, nunca o conteúdo dela.
    audit_log(conn, company_id, user["id"], "ciencia", "feedbacks", str(feedback_id),
              {"comObservacao": bool(texto), "comConfidencial": bool(sigilo)})
    conn.commit()
    return {"acknowledged": True, "hasNote": bool(texto), "hasConfidential": bool(sigilo)}


def delete_feedback(conn: sqlite3.Connection, company_id: int, user: sqlite3.Row, feedback_id: int) -> None:
    if not user_can_give_feedback(conn, user):
        raise PermissionError("Apenas gestão exclui feedback.")
    conn.execute("DELETE FROM feedback_ratings WHERE feedback_id = ?", (feedback_id,))
    conn.execute("DELETE FROM feedbacks WHERE company_id = ? AND id = ?", (company_id, feedback_id))
    audit_log(conn, company_id, user["id"], "excluir", "feedbacks", str(feedback_id), {})
    conn.commit()


# ── Registro pontual ─────────────────────────────────────────────────────────
#
# Complementa o feedback mensal, não substitui. Dura um minuto de digitação e
# resolve o caso do dia 12: algo aconteceu, precisa ser dito e combinado agora.
#
# A ciência é opcional por decisão de projeto: reconhecimento e orientação do
# dia a dia não precisam virar pendência. Correção precisa — é o que transforma
# a conversa em comprovação de que a orientação foi dada.

FEEDBACK_NOTE_KINDS = [
    {"id": "RECONHECIMENTO", "label": "Reconhecimento", "icon": "👏", "color": "#1e8e3e", "bg": "#e6f4ea",
     "hint": "Algo que merece ser dito na hora, não daqui a três semanas.",
     "defaultAck": False},
    {"id": "ORIENTACAO", "label": "Orientação", "icon": "🧭", "color": "#1a5276", "bg": "#e8f0fe",
     "hint": "Ajuste de rota do dia a dia. Ensina sem precisar formalizar.",
     "defaultAck": False},
    {"id": "CORRECAO", "label": "Correção", "icon": "⚠️", "color": "#c5221f", "bg": "#fce8e6",
     "hint": "Falha que precisa parar de acontecer. Já vem com ciência marcada.",
     "defaultAck": True},
    {"id": "ACOMPANHAMENTO", "label": "Acompanhamento", "icon": "🔁", "color": "#b06000", "bg": "#fef7e0",
     "hint": "Checagem do que foi combinado ou do PDI, no meio do caminho.",
     "defaultAck": False},
]
FEEDBACK_NOTE_KIND_IDS = {k["id"] for k in FEEDBACK_NOTE_KINDS}


def feedback_note_row_to_dict(row: sqlite3.Row) -> dict[str, Any]:
    cfg = next((k for k in FEEDBACK_NOTE_KINDS if k["id"] == row["kind"]), None)
    return {
        "id": int(row["id"]),
        "personName": row["person_name"],
        "personKey": row["person_key"],
        "unitName": row["unit_name"] or "",
        "occurredAt": row["occurred_at"],
        "competence": row["competence"],
        "kind": row["kind"],
        "kindLabel": cfg["label"] if cfg else row["kind"],
        "kindIcon": cfg["icon"] if cfg else "•",
        "summary": row["summary"],
        "agreement": row["agreement"] or "",
        "requiresAck": bool(row["requires_ack"]),
        "acknowledgedAt": row["acknowledged_at"] or "",
        "personNote": row["person_note"] or "",
        "authorName": row["author_name"],
        "createdAt": row["created_at"],
    }


def list_feedback_notes(
    conn: sqlite3.Connection, company_id: int, user: sqlite3.Row,
    person_key_value: str = "", competence: str = "", limit: int = 200,
) -> list[dict[str, Any]]:
    scope = data_scope_for_user(conn, user)
    minhas = user_person_keys(user)
    marcadores = ",".join("?" for _ in minhas) or "''"

    sql = "SELECT * FROM feedback_notes WHERE company_id = ?"
    params: list[Any] = [company_id]

    if scope == "proprio":
        sql += f" AND (person_user_id = ? OR person_key IN ({marcadores}))"
        params.append(user["id"])
        params.extend(minhas)
    else:
        permitidas = crm_allowed_units_for_user(conn, user)
        if permitidas is not None:
            if permitidas:
                unidades = ",".join("?" for _ in permitidas)
                sql += f" AND (unit_name IN ({unidades}) OR person_key IN ({marcadores}))"
                params.extend(permitidas)
            else:
                sql += f" AND person_key IN ({marcadores})"
            params.extend(minhas)

    if person_key_value:
        sql += " AND person_key = ?"
        params.append(person_key_value)
    if competence:
        sql += " AND competence = ?"
        params.append(competence)

    sql += " ORDER BY date(occurred_at) DESC, id DESC LIMIT ?"
    params.append(int(limit))

    resultado = []
    for row in conn.execute(sql, params).fetchall():
        nota = feedback_note_row_to_dict(row)
        nota["isMe"] = nota["personKey"] in set(minhas) or (
            row["person_user_id"] is not None and row["person_user_id"] == user["id"]
        )
        resultado.append(nota)
    return resultado


def count_pending_note_ack(conn: sqlite3.Connection, company_id: int, user: sqlite3.Row) -> int:
    chaves = user_person_keys(user)
    marcadores = ",".join("?" for _ in chaves) or "''"
    row = conn.execute(
        f"""
        SELECT COUNT(*) n FROM feedback_notes
        WHERE company_id = ? AND requires_ack = 1
          AND (person_user_id = ? OR person_key IN ({marcadores}))
          AND (acknowledged_at IS NULL OR acknowledged_at = '')
        """,
        (company_id, user["id"], *chaves),
    ).fetchone()
    return int(row["n"] or 0)


def save_feedback_note(
    conn: sqlite3.Connection, company_id: int, user: sqlite3.Row, payload: dict[str, Any]
) -> dict[str, Any]:
    if not user_can_give_feedback(conn, user):
        raise PermissionError("Apenas gestão registra acompanhamento.")

    person_name = normalize_whitespace(payload.get("personName"))
    if not person_name:
        raise ValueError("Selecione a pessoa.")
    resumo = normalize_whitespace(payload.get("summary"))
    if not resumo:
        raise ValueError("Descreva o que aconteceu.")
    kind = normalize_upper(payload.get("kind")) or "ORIENTACAO"
    if kind not in FEEDBACK_NOTE_KIND_IDS:
        raise ValueError("Tipo de registro inválido.")

    ocorrido = normalize_whitespace(payload.get("occurredAt")) or today_in_brazil().isoformat()
    competence = ocorrido[:7]
    unidade = normalize_unit(payload.get("unitName")) or None
    permitidas = crm_allowed_units_for_user(conn, user)
    if permitidas is not None:
        if not permitidas:
            raise ValueError("Seu usuário não tem unidade vinculada. Peça ao administrador para vincular.")
        if unidade not in permitidas:
            unidade = permitidas[0]

    exige = 1 if payload.get("requiresAck") else 0
    note_id = payload.get("id")

    if note_id:
        conn.execute(
            """
            UPDATE feedback_notes SET person_name=?, person_key=?, unit_name=?, occurred_at=?,
                   competence=?, kind=?, summary=?, agreement=?, requires_ack=?
            WHERE company_id = ? AND id = ?
            """,
            (person_name, person_key(person_name), unidade, ocorrido, competence, kind, resumo,
             normalize_whitespace(payload.get("agreement")), exige, company_id, int(note_id)),
        )
        novo_id = int(note_id)
    else:
        cursor = conn.execute(
            """
            INSERT INTO feedback_notes (company_id, person_name, person_key, person_user_id,
                unit_name, occurred_at, competence, kind, summary, agreement, requires_ack,
                author_name, author_user_id, created_at)
            VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?)
            """,
            (company_id, person_name, person_key(person_name),
             resolve_user_for_person(conn, company_id, person_name), unidade, ocorrido, competence,
             kind, resumo, normalize_whitespace(payload.get("agreement")), exige,
             meeting_person_identity(user), user["id"], now_iso()),
        )
        novo_id = int(cursor.lastrowid)

    audit_log(conn, company_id, user["id"], "salvar", "feedback_notes", str(novo_id),
              {"pessoa": person_name, "tipo": kind, "exigeCiencia": bool(exige)})
    conn.commit()
    return {"noteId": novo_id, "requiresAck": bool(exige)}


def acknowledge_feedback_note(
    conn: sqlite3.Connection, company_id: int, user: sqlite3.Row, note_id: int, note: str = ""
) -> dict[str, Any]:
    row = conn.execute(
        "SELECT * FROM feedback_notes WHERE company_id = ? AND id = ?", (company_id, note_id)
    ).fetchone()
    if not row:
        raise ValueError("Registro não encontrado.")
    minhas = set(user_person_keys(user))
    if not (row["person_key"] in minhas
            or (row["person_user_id"] is not None and row["person_user_id"] == user["id"])):
        raise ValueError("Este registro não é seu.")

    texto = normalize_whitespace(note)
    conn.execute(
        """
        UPDATE feedback_notes
        SET acknowledged_at = COALESCE(NULLIF(acknowledged_at, ''), ?),
            person_note     = CASE WHEN ? <> '' THEN ? ELSE person_note END
        WHERE id = ?
        """,
        (now_iso(), texto, texto, note_id),
    )
    conn.commit()
    return {"acknowledged": True, "hasNote": bool(texto)}


def delete_feedback_note(
    conn: sqlite3.Connection, company_id: int, user: sqlite3.Row, note_id: int
) -> None:
    if not user_can_give_feedback(conn, user):
        raise PermissionError("Apenas gestão exclui registro.")
    conn.execute("DELETE FROM feedback_notes WHERE company_id = ? AND id = ?", (company_id, note_id))
    audit_log(conn, company_id, user["id"], "excluir", "feedback_notes", str(note_id), {})
    conn.commit()


# ── PDI ──────────────────────────────────────────────────────────────────────

def list_pdi_items(
    conn: sqlite3.Connection, company_id: int, person_key_value: str, include_closed: bool = True
) -> list[dict[str, Any]]:
    sql = "SELECT * FROM pdi_items WHERE company_id = ? AND person_key = ?"
    params: list[Any] = [company_id, person_key_value]
    if not include_closed:
        sql += " AND status IN ('ABERTO','EVOLUINDO')"
    sql += " ORDER BY CASE status WHEN 'EVOLUINDO' THEN 0 WHEN 'ABERTO' THEN 1 ELSE 2 END, date(due_date)"
    hoje = today_in_brazil().isoformat()
    return [
        {
            "id": int(r["id"]),
            "personName": r["person_name"],
            "title": r["title"],
            "why": r["why"] or "",
            "action": r["action"] or "",
            "support": r["support"] or "",
            "dueDate": r["due_date"] or "",
            "status": r["status"],
            "progressNote": r["progress_note"] or "",
            "createdAt": r["created_at"],
            "closedAt": r["closed_at"] or "",
            "overdue": bool(r["due_date"] and r["status"] in ("ABERTO", "EVOLUINDO")
                            and r["due_date"] < hoje),
        }
        for r in conn.execute(sql, params).fetchall()
    ]


def save_pdi_item(
    conn: sqlite3.Connection, company_id: int, user: sqlite3.Row, payload: dict[str, Any]
) -> dict[str, Any]:
    if not user_can_give_feedback(conn, user):
        raise PermissionError("Apenas gestão mantém o PDI.")
    titulo = normalize_whitespace(payload.get("title"))
    if not titulo:
        raise ValueError("Informe o que precisa ser desenvolvido.")
    person_name = normalize_whitespace(payload.get("personName"))
    if not person_name:
        raise ValueError("Informe a pessoa do PDI.")
    chave = person_key(person_name)
    status = normalize_upper(payload.get("status")) or "ABERTO"
    if status not in PDI_STATUS_IDS:
        status = "ABERTO"

    item_id = payload.get("id")
    if not item_id:
        ativos = conn.execute(
            "SELECT COUNT(*) n FROM pdi_items WHERE company_id = ? AND person_key = ? "
            "AND status IN ('ABERTO','EVOLUINDO')",
            (company_id, chave),
        ).fetchone()["n"]
        if ativos >= PDI_MAX_ACTIVE:
            raise ValueError(
                f"{person_name} já tem {ativos} pontos de desenvolvimento em aberto. "
                f"Conclua um antes de abrir outro — plano com mais de {PDI_MAX_ACTIVE} frentes não sai do lugar."
            )

    campos = (
        person_name, chave,
        normalize_unit(payload.get("unitName")) or None,
        titulo,
        normalize_whitespace(payload.get("why")),
        normalize_whitespace(payload.get("action")),
        normalize_whitespace(payload.get("support")),
        normalize_whitespace(payload.get("dueDate")) or None,
        status,
        normalize_whitespace(payload.get("progressNote")),
    )
    fechado = now_iso() if status in ("CONCLUIDO", "CANCELADO") else None

    if item_id:
        conn.execute(
            """
            UPDATE pdi_items SET person_name=?, person_key=?, unit_name=?, title=?, why=?,
                   action=?, support=?, due_date=?, status=?, progress_note=?,
                   updated_at=?, closed_at=?
            WHERE company_id = ? AND id = ?
            """,
            (*campos, now_iso(), fechado, company_id, int(item_id)),
        )
        novo_id = int(item_id)
    else:
        cursor = conn.execute(
            """
            INSERT INTO pdi_items (person_name, person_key, unit_name, title, why, action,
                support, due_date, status, progress_note, company_id, origin_feedback_id,
                created_by_user_id, created_at, closed_at)
            VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)
            """,
            (*campos, company_id, payload.get("originFeedbackId"), user["id"], now_iso(), fechado),
        )
        novo_id = int(cursor.lastrowid)

    audit_log(conn, company_id, user["id"], "salvar", "pdi_items", str(novo_id),
              {"pessoa": person_name, "status": status})
    conn.commit()
    return {"pdiId": novo_id}


def delete_pdi_item(conn: sqlite3.Connection, company_id: int, user: sqlite3.Row, pdi_id: int) -> None:
    if not user_can_give_feedback(conn, user):
        raise PermissionError("Apenas gestão mantém o PDI.")
    conn.execute("DELETE FROM pdi_items WHERE company_id = ? AND id = ?", (company_id, pdi_id))
    conn.commit()


# ─────────────────────────────────────────────────────────────────────────────
# Visitas gerenciais
#
# Regras que definem o módulo:
#
# 1. LIGAÇÃO ANTES DE VISITA. Só entra na sugestão quem já teve tentativa de
#    contato registrada. Visita é o recurso mais caro que a operação tem —
#    gastar em cliente que nem foi procurado por telefone é desperdício, e
#    inverte a ordem: o vendedor precisa fazer a parte dele primeiro.
#
# 2. ROTEIRO POR PROXIMIDADE. A lista é agrupada por cidade, bairro e rua. Um
#    gerente que atravessa a cidade três vezes faz metade das visitas do dia.
#    Por isso entram também os clientes de relacionamento que estão no caminho,
#    mesmo sem indicador ruim — o custo marginal de parar ali é quase zero.
#
# 3. EFEITO MEDIDO SOZINHO. O faturamento 60 dias antes e 60 dias depois já
#    está no banco. Pedir para alguém avaliar "se funcionou" produz opinião;
#    comparar o faturamento produz resposta.
# ─────────────────────────────────────────────────────────────────────────────

VISIT_TYPES = [
    {"id": "SOLICITADA", "label": "Pedida pelo vendedor", "icon": "🙋", "color": "#c5221f", "bg": "#fce8e6",
     "hint": "O vendedor tentou por telefone e pediu a presença do gestor."},
    {"id": "REATIVACAO", "label": "Reativação", "icon": "🔄", "color": "#b06000", "bg": "#fef7e0",
     "hint": "Cliente parou de comprar ou caiu forte. A visita busca entender o motivo."},
    {"id": "RELACIONAMENTO", "label": "Relacionamento", "icon": "🤝", "color": "#1e8e3e", "bg": "#e6f4ea",
     "hint": "Cliente bom, sem problema aparente. Presença para sustentar a conta."},
    {"id": "NEGOCIACAO", "label": "Negociação", "icon": "💼", "color": "#1a5276", "bg": "#e8f0fe",
     "hint": "Fechar acordo, tabela ou volume que não se resolve por telefone."},
]
VISIT_TYPE_IDS = {t["id"] for t in VISIT_TYPES}

VISIT_STATUSES = [
    {"id": "PLANEJADA", "label": "Planejada", "color": "#5f6368", "bg": "#f1f3f4"},
    {"id": "REALIZADA", "label": "Realizada", "color": "#1e8e3e", "bg": "#e6f4ea"},
    {"id": "CANCELADA", "label": "Cancelada", "color": "#5f6368", "bg": "#f1f3f4"},
]

# Parâmetros do motor de sugestão — juntos e nomeados, são regra comercial.
VISIT_CALL_WINDOW_DAYS = 30      # janela em que a ligação do vendedor vale como pré-requisito
VISIT_COOLDOWN_DAYS = 60         # não sugere de novo quem foi visitado há menos que isso
VISIT_RELATIONSHIP_DAYS = 180    # cliente bom sem visita há mais que isso entra por relacionamento
VISIT_EFFECT_WINDOW_DAYS = 60    # janela de comparação antes/depois
VISIT_SUGGESTION_LIMIT = 60


def user_can_manage_visits(conn: sqlite3.Connection, user: sqlite3.Row) -> bool:
    """Registra visita quem tem visão de equipe. O vendedor solicita e acompanha."""
    return data_scope_for_user(conn, user) != "proprio"


def client_sales_names(conn: sqlite3.Connection, company_id: int, client_key: str) -> list[str]:
    """Nomes com que este cliente aparece no faturamento.

    `client_key` é o CÓDIGO do cliente; o faturamento grava o NOME. Sem essa
    tradução a comparação nunca casa — e o efeito da visita saía sempre zero.
    """
    nomes: set[str] = set()
    perfil = conn.execute(
        "SELECT client_name, trade_name FROM crm_client_profiles WHERE company_id = ? AND client_code = ?",
        (company_id, client_key),
    ).fetchone()
    if perfil:
        for campo in ("client_name", "trade_name"):
            valor = normalize_whitespace(perfil[campo])
            if valor:
                nomes.add(valor)
    for r in conn.execute(
        "SELECT DISTINCT client_name FROM crm_client_summary WHERE company_id = ? AND client_code = ?",
        (company_id, client_key),
    ).fetchall():
        valor = normalize_whitespace(r["client_name"])
        if valor:
            nomes.add(valor)
    return sorted(nomes)


def client_revenue_between(
    conn: sqlite3.Connection, company_id: int, client_key: str, inicio: str, fim: str
) -> float:
    """Faturamento líquido do cliente no intervalo, pela data da nota.

    Linha sem data de emissão usa o primeiro dia da competência — melhor do que
    descartar, porque o mês inteiro cairia fora da janela.
    """
    nomes = client_sales_names(conn, company_id, client_key)
    if not nomes:
        return 0.0
    # Filtra por nome no SQL: a tabela de faturamento tem milhões de linhas e
    # trazer a competência inteira para a memória só para descartar não escala.
    marcadores = ",".join("?" for _ in nomes)
    linhas = conn.execute(
        f"""
        SELECT client_name, competence, issue_date, net_value
        FROM fact_sales_detail
        WHERE company_id = ? AND net_value > 0
          AND competence >= ? AND competence <= ?
          AND UPPER(client_name) IN ({marcadores})
        """,
        (company_id, inicio[:7], fim[:7], *[normalize_upper(n) for n in nomes]),
    ).fetchall()
    total = 0.0
    for r in linhas:
        data = normalize_whitespace(r["issue_date"])[:10] or f"{r['competence']}-01"
        if inicio <= data <= fim:
            total += float(r["net_value"] or 0.0)
    return round(total, 2)


def measure_visit_effect(
    conn: sqlite3.Connection, company_id: int, visit_id: int, force: bool = False
) -> dict[str, Any] | None:
    """Calcula o antes e depois da visita, quando a janela já fechou.

    Só mede depois que os 60 dias posteriores passaram — antes disso o número
    seria parcial e daria a impressão errada de que a visita não funcionou.
    """
    visita = conn.execute(
        "SELECT * FROM visits WHERE company_id = ? AND id = ?", (company_id, visit_id)
    ).fetchone()
    if not visita or visita["status"] != "REALIZADA" or not visita["occurred_at"]:
        return None
    if visita["effect_measured_at"] and not force:
        return {"revenueBefore": visita["revenue_before"], "revenueAfter": visita["revenue_after"]}

    dt = parse_datetime_flexible(visita["occurred_at"][:10])
    if not dt:
        return None
    dia = dt.date()
    fim_depois = dia + timedelta(days=VISIT_EFFECT_WINDOW_DAYS)
    if fim_depois > today_in_brazil() and not force:
        return None   # janela ainda aberta

    antes = client_revenue_between(
        conn, company_id, visita["client_key"],
        (dia - timedelta(days=VISIT_EFFECT_WINDOW_DAYS)).isoformat(),
        (dia - timedelta(days=1)).isoformat(),
    )
    depois = client_revenue_between(
        conn, company_id, visita["client_key"],
        dia.isoformat(), fim_depois.isoformat(),
    )
    conn.execute(
        "UPDATE visits SET revenue_before = ?, revenue_after = ?, effect_measured_at = ? WHERE id = ?",
        (antes, depois, now_iso(), visit_id),
    )
    conn.commit()
    return {"revenueBefore": antes, "revenueAfter": depois}


def client_addresses(conn: sqlite3.Connection, company_id: int, codigos: list[str]) -> dict[str, dict[str, str]]:
    """Endereço por código de cliente, para montar o roteiro."""
    if not codigos:
        return {}
    resultado: dict[str, dict[str, str]] = {}
    for bloco in range(0, len(codigos), 400):
        parte = codigos[bloco:bloco + 400]
        marcadores = ",".join("?" for _ in parte)
        for r in conn.execute(
            f"""
            SELECT client_code, address_line, address_number, neighborhood, city_name, postal_code, phone
            FROM crm_client_profiles
            WHERE company_id = ? AND client_code IN ({marcadores})
            """,
            (company_id, *parte),
        ).fetchall():
            rua = normalize_whitespace(r["address_line"])
            numero = normalize_whitespace(r["address_number"])
            resultado[r["client_code"]] = {
                "street": rua,
                "addressLine": f"{rua}, {numero}" if rua and numero else rua,
                "neighborhood": normalize_upper(r["neighborhood"]) or "SEM BAIRRO",
                "cityName": normalize_upper(r["city_name"]),
                "postalCode": normalize_whitespace(r["postal_code"]),
                "phone": normalize_whitespace(r["phone"]),
            }
    return resultado


def search_clients_for_visit(
    conn: sqlite3.Connection, company_id: int, user: sqlite3.Row, termo: str, limite: int = 25
) -> list[dict[str, Any]]:
    """Busca cliente por código ou nome para vincular a uma visita.

    Traz o endereço junto: o gerente escolhe o cliente e a visita já nasce com
    rua, bairro e cidade preenchidos, sem digitação e sem risco de errar o vínculo.
    """
    texto = normalize_whitespace(termo)
    if len(texto) < 2:
        return []
    alvo = f"%{texto.upper()}%"

    sql = """
        SELECT client_code, client_name, trade_name, city_name, neighborhood,
               address_line, address_number, postal_code, phone, updated_phone,
               NULLIF(TRIM(internal_seller_name), '') AS seller_name
        FROM crm_client_profiles
        WHERE company_id = ?
          AND (UPPER(client_code) LIKE ? OR UPPER(client_name) LIKE ? OR UPPER(COALESCE(trade_name,'')) LIKE ?)
    """
    params: list[Any] = [company_id, alvo, alvo, alvo]

    # Restringe às cidades atendidas pelas unidades do gestor. Sem isso a busca
    # abriria a base inteira da empresa para um gerente de unidade.
    permitidas = crm_allowed_units_for_user(conn, user)
    if permitidas is not None:
        cidades = active_mapped_cities_for_units(conn, company_id, permitidas) if permitidas else []
        if not cidades:
            # Sem unidade vinculada (ou sem cidade mapeada) a busca não devolve
            # nada, em vez de abrir a base inteira da empresa.
            return []
        marcadores = ",".join("?" for _ in cidades)
        sql += f" AND UPPER(COALESCE(city_name,'')) IN ({marcadores})"
        params.extend([normalize_upper(c) for c in cidades])

    sql += " ORDER BY client_name LIMIT ?"
    params.append(int(limite))

    resultado = []
    for r in conn.execute(sql, params).fetchall():
        rua = normalize_whitespace(r["address_line"])
        numero = normalize_whitespace(r["address_number"])
        resultado.append({
            "clientKey": r["client_code"],
            "clientName": normalize_whitespace(r["client_name"]),
            "tradeName": normalize_whitespace(r["trade_name"]),
            "cityName": normalize_upper(r["city_name"]),
            "neighborhood": normalize_upper(r["neighborhood"]),
            "addressLine": f"{rua}, {numero}" if rua and numero else rua,
            "postalCode": normalize_whitespace(r["postal_code"]),
            "phone": normalize_whitespace(r["updated_phone"]) or normalize_whitespace(r["phone"]),
            "assignedSeller": normalize_whitespace(r["seller_name"]),
        })
    return resultado


def suggest_visits(
    conn: sqlite3.Connection, company_id: int, user: sqlite3.Row,
    city: str = "", include_relationship: bool = True,
) -> dict[str, Any]:
    """Roteiro sugerido, agrupado por cidade → bairro → rua."""
    competence = crm_latest_competence(conn, company_id) or date.today().strftime("%Y-%m")
    permitidas = crm_allowed_units_for_user(conn, user)
    hoje = today_in_brazil()

    # Usa exatamente o mesmo recorte da tela de Carteira. Se o gerente vê o
    # cliente lá, ele pode aparecer no roteiro — e vice-versa. Regra própria aqui
    # criaria duas verdades sobre "quais clientes são meus".
    filtros = build_filters_from_query({})
    filtros["competence_start"] = competence
    filtros["competence_end"] = competence
    filtros = crm_scoped_filters_for_user(conn, company_id, user, filtros)
    clientes = list_crm_clients(conn, company_id, filtros, attach_context=False)

    # Pré-requisito: ligação registrada na janela
    limite_ligacao = (hoje - timedelta(days=VISIT_CALL_WINDOW_DAYS)).isoformat()
    com_ligacao = {
        normalize_client_key(r["client_key"])
        for r in conn.execute(
            """
            SELECT DISTINCT client_key FROM crm_interactions
            WHERE company_id = ? AND contact_type_code = 'LIGACAO' AND initiative = 'ATIVO'
              AND date(substr(replace(occurred_at,'T',' '),1,10)) >= date(?)
            """,
            (company_id, limite_ligacao),
        ).fetchall()
    }

    # Visitas recentes: evita repetir o mesmo cliente
    limite_visita = (hoje - timedelta(days=VISIT_COOLDOWN_DAYS)).isoformat()
    visitados = {
        normalize_client_key(r["client_key"]): r["ultima"]
        for r in conn.execute(
            """
            SELECT client_key, MAX(date(occurred_at)) ultima FROM visits
            WHERE company_id = ? AND status = 'REALIZADA'
            GROUP BY client_key
            """,
            (company_id,),
        ).fetchall()
    }

    # Pedidos pendentes do vendedor entram sempre, e na frente
    pedidos = {
        normalize_client_key(r["client_key"]): dict(r)
        for r in conn.execute(
            "SELECT * FROM visit_requests WHERE company_id = ? AND status = 'PENDENTE'",
            (company_id,),
        ).fetchall()
    }

    codigos = [c["clientKey"] for c in clientes]
    enderecos = client_addresses(conn, company_id, codigos)
    cidade_filtro = normalize_upper(city)

    candidatos: list[dict[str, Any]] = []
    for c in clientes:
        chave = normalize_client_key(c["clientKey"])
        endereco = enderecos.get(c["clientKey"], {})
        cidade = endereco.get("cityName") or normalize_upper(c.get("cityName"))
        if cidade_filtro and cidade != cidade_filtro:
            continue

        pedido = pedidos.get(chave)
        ultima_visita = visitados.get(chave)
        dias_visita = None
        if ultima_visita:
            d = parse_datetime_flexible(ultima_visita)
            dias_visita = (hoje - d.date()).days if d else None
            # Cooldown: quem foi visitado há pouco sai, exceto se o vendedor pediu.
            if dias_visita is not None and dias_visita < VISIT_COOLDOWN_DAYS and not pedido:
                continue

        status = normalize_upper(c.get("statusCode"))
        queda = float(c.get("dropPct") or 0.0)
        media = float(c.get("averageRevenue") or 0.0)
        atual = float(c.get("currentRevenue") or 0.0)

        if pedido:
            # Pedido do vendedor vem SEMPRE na frente, qualquer que seja o porte do
            # cliente. Quem está na rua todo dia viu algo que o indicador não mostra;
            # ignorar isso ensina a equipe a parar de pedir.
            tipo = "SOLICITADA"
            motivo = normalize_whitespace(pedido["reason"])
            base = 1_000_000.0 + media
        elif not com_ligacao and not pedido:
            continue
        elif chave not in com_ligacao:
            # Sem ligação registrada não vira sugestão. É a regra que mantém a ordem.
            continue
        elif status in ("INATIVO", "PRE_INATIVO"):
            tipo = "REATIVACAO"
            motivo = f"{'Inativo' if status == 'INATIVO' else 'Pré-inativo'} há {c.get('daysWithoutPurchase') or '?'} dias"
            base = media * 2
        elif queda <= -0.30:
            tipo = "REATIVACAO"
            motivo = f"Queda de {abs(round(queda * 100))}% contra a média do trimestre"
            base = media * 1.5
        elif atual <= 0 and media > 0:
            tipo = "REATIVACAO"
            motivo = "Sem compra no mês"
            base = media
        elif include_relationship and media > 0 and (dias_visita is None or dias_visita >= VISIT_RELATIONSHIP_DAYS):
            tipo = "RELACIONAMENTO"
            motivo = ("Nunca visitado" if dias_visita is None
                      else f"Sem visita há {dias_visita} dias")
            base = media * 0.4
        else:
            continue

        candidatos.append({
            "clientKey": c["clientKey"],
            "clientName": c.get("clientName"),
            "cityName": cidade,
            "neighborhood": endereco.get("neighborhood") or "SEM BAIRRO",
            "street": endereco.get("street") or "",
            "addressLine": endereco.get("addressLine") or "",
            "postalCode": endereco.get("postalCode") or "",
            "phone": c.get("updatedPhone") or c.get("phone") or endereco.get("phone") or "",
            "assignedSeller": c.get("assignedSeller") or "",
            "statusCode": status,
            "classCode": c.get("classCode") or "",
            "daysWithoutPurchase": c.get("daysWithoutPurchase"),
            "currentRevenue": atual,
            "averageRevenue": media,
            "dropPct": queda,
            "lastVisitDays": dias_visita,
            "visitType": tipo,
            "reason": motivo,
            "requestId": pedido["id"] if pedido else None,
            "requestedBy": pedido["seller_name"] if pedido else "",
            "score": round(base, 2),
        })

    candidatos.sort(key=lambda x: x["score"], reverse=True)
    candidatos = candidatos[:VISIT_SUGGESTION_LIMIT]

    # Agrupa por cidade → bairro → rua. A ordem dentro do bairro segue a rua,
    # para o gerente andar por quarteirão em vez de pular de ponta a ponta.
    rotas: dict[str, dict[str, list[dict[str, Any]]]] = {}
    for item in candidatos:
        rotas.setdefault(item["cityName"] or "SEM CIDADE", {}).setdefault(
            item["neighborhood"], []).append(item)

    roteiro = []
    for cidade_nome, bairros in rotas.items():
        blocos = []
        for bairro, itens in bairros.items():
            itens.sort(key=lambda x: (x["street"], -x["score"]))
            blocos.append({
                "neighborhood": bairro,
                "clients": itens,
                "count": len(itens),
                "potential": round(sum(i["averageRevenue"] for i in itens), 2),
            })
        blocos.sort(key=lambda b: b["potential"], reverse=True)
        roteiro.append({
            "cityName": cidade_nome,
            "neighborhoods": blocos,
            "count": sum(b["count"] for b in blocos),
            "potential": round(sum(b["potential"] for b in blocos), 2),
        })
    roteiro.sort(key=lambda r: r["potential"], reverse=True)

    return {
        "route": roteiro,
        "total": len(candidatos),
        "cities": sorted({r["cityName"] for r in roteiro if r["cityName"]}),
        "params": {
            "callWindowDays": VISIT_CALL_WINDOW_DAYS,
            "cooldownDays": VISIT_COOLDOWN_DAYS,
            "relationshipDays": VISIT_RELATIONSHIP_DAYS,
        },
    }


# ── Pedido de visita (vendedor → gerente) ────────────────────────────────────

def visit_request_eligibility(
    conn: sqlite3.Connection, company_id: int, client_key: str
) -> dict[str, Any]:
    """Se o cliente já teve ligação registrada dentro da janela.

    É a condição para pedir visita. Sem ela o pedido vira atalho: o vendedor
    empurra para o gerente o contato que era dele fazer.
    """
    limite = (today_in_brazil() - timedelta(days=VISIT_CALL_WINDOW_DAYS)).isoformat()
    row = conn.execute(
        """
        SELECT COUNT(*) n, MAX(occurred_at) ultima
        FROM crm_interactions
        WHERE company_id = ? AND client_key = ? AND contact_type_code = 'LIGACAO' AND initiative = 'ATIVO'
          AND date(substr(replace(occurred_at,'T',' '),1,10)) >= date(?)
        """,
        (company_id, client_key, limite),
    ).fetchone()
    total = conn.execute(
        "SELECT COUNT(*) n FROM crm_interactions WHERE company_id = ? AND client_key = ? "
        "AND contact_type_code = 'LIGACAO' AND initiative = 'ATIVO'",
        (company_id, client_key),
    ).fetchone()["n"]
    return {
        "eligible": int(row["n"] or 0) > 0,
        "callsInWindow": int(row["n"] or 0),
        "callsTotal": int(total or 0),
        "lastCallAt": row["ultima"] or "",
        "windowDays": VISIT_CALL_WINDOW_DAYS,
    }


def create_visit_request(
    conn: sqlite3.Connection, company_id: int, user: sqlite3.Row, payload: dict[str, Any]
) -> dict[str, Any]:
    client_key = normalize_whitespace(payload.get("clientKey"))
    reason = normalize_whitespace(payload.get("reason"))
    if not client_key:
        raise ValueError("Cliente inválido.")
    if not reason:
        raise ValueError("Diga por que a visita é necessária — é o que o gerente lê para decidir.")

    # Condição do módulo: só pede visita quem já tentou por telefone. A checagem
    # é no servidor porque a tela pode ser contornada, e essa regra é o que
    # mantém a ordem entre o trabalho do vendedor e o do gestor.
    elegivel = visit_request_eligibility(conn, company_id, client_key)
    if not elegivel["eligible"]:
        raise ValueError(
            f"Registre uma ligação para este cliente antes de pedir a visita. "
            f"Nenhuma ligação registrada nos últimos {elegivel['windowDays']} dias."
        )

    ja_existe = conn.execute(
        "SELECT id FROM visit_requests WHERE company_id = ? AND client_key = ? AND status = 'PENDENTE'",
        (company_id, client_key),
    ).fetchone()
    if ja_existe:
        return {"requestId": int(ja_existe["id"]), "duplicated": True,
                "message": "Já existe um pedido de visita aberto para este cliente."}

    perfil = conn.execute(
        "SELECT client_name, city_name FROM crm_client_profiles WHERE company_id = ? AND client_code = ?",
        (company_id, client_key),
    ).fetchone()
    seller = seller_identity_for_user(user)
    _, unidade = current_role_and_unit(
        conn, company_id, seller,
        crm_latest_competence(conn, company_id) or date.today().strftime("%Y-%m"),
    )

    cursor = conn.execute(
        """
        INSERT INTO visit_requests (company_id, client_key, client_name, unit_name, city_name,
            seller_name, reason, interaction_id, requested_by_user_id, created_at)
        VALUES (?,?,?,?,?,?,?,?,?,?)
        """,
        (company_id, client_key,
         normalize_whitespace(payload.get("clientName")) or (perfil["client_name"] if perfil else client_key),
         normalize_unit(unidade), normalize_upper(perfil["city_name"]) if perfil else None,
         seller, reason, payload.get("interactionId"), user["id"], now_iso()),
    )
    audit_log(conn, company_id, user["id"], "criar", "visit_requests", client_key, {"vendedor": seller})
    conn.commit()
    return {"requestId": int(cursor.lastrowid), "duplicated": False,
            "message": "Pedido de visita enviado ao gerente."}


def list_visit_requests(
    conn: sqlite3.Connection, company_id: int, user: sqlite3.Row, status: str = "PENDENTE"
) -> list[dict[str, Any]]:
    sql = "SELECT * FROM visit_requests WHERE company_id = ?"
    params: list[Any] = [company_id]
    if data_scope_for_user(conn, user) == "proprio":
        sql += " AND UPPER(seller_name) = ?"
        params.append(normalize_upper(seller_identity_for_user(user)))
    else:
        permitidas = crm_allowed_units_for_user(conn, user)
        if permitidas is not None:
            if permitidas:
                marcadores = ",".join("?" for _ in permitidas)
                sql += f" AND unit_name IN ({marcadores})"
                params.extend(permitidas)
            else:
                sql += " AND 1 = 0"   # sem unidade vinculada, sem fila
    if status:
        sql += " AND status = ?"
        params.append(normalize_upper(status))
    sql += " ORDER BY datetime(created_at) DESC LIMIT 200"
    return [
        {
            "id": int(r["id"]),
            "clientKey": r["client_key"],
            "clientName": r["client_name"],
            "cityName": r["city_name"] or "",
            "unitName": r["unit_name"] or "",
            "sellerName": r["seller_name"],
            "reason": r["reason"],
            "status": r["status"],
            "managerNote": r["manager_note"] or "",
            "createdAt": r["created_at"],
            "resolvedAt": r["resolved_at"] or "",
        }
        for r in conn.execute(sql, params).fetchall()
    ]


def resolve_visit_request(
    conn: sqlite3.Connection, company_id: int, user: sqlite3.Row,
    request_id: int, accept: bool, note: str = "",
) -> dict[str, Any]:
    if not user_can_manage_visits(conn, user):
        raise PermissionError("Apenas gestão responde pedidos de visita.")
    conn.execute(
        "UPDATE visit_requests SET status = ?, manager_note = ?, resolved_by_user_id = ?, resolved_at = ? "
        "WHERE company_id = ? AND id = ?",
        ("ACEITA" if accept else "RECUSADA", normalize_whitespace(note), user["id"], now_iso(),
         company_id, request_id),
    )
    conn.commit()
    return {"resolved": True}


# ── Visita ───────────────────────────────────────────────────────────────────

def visit_row_to_dict(row: sqlite3.Row) -> dict[str, Any]:
    cfg = next((t for t in VISIT_TYPES if t["id"] == row["visit_type"]), None)
    antes = row["revenue_before"]
    depois = row["revenue_after"]
    variacao = None
    if antes is not None and depois is not None:
        variacao = round(safe_div(float(depois) - float(antes), float(antes)) * 100, 1) if antes else None
    return {
        "id": int(row["id"]),
        "clientKey": row["client_key"],
        "clientName": row["client_name"],
        "unitName": row["unit_name"] or "",
        "cityName": row["city_name"] or "",
        "neighborhood": row["neighborhood"] or "",
        "addressLine": row["address_line"] or "",
        "visitType": row["visit_type"],
        "visitTypeLabel": cfg["label"] if cfg else row["visit_type"],
        "visitTypeIcon": cfg["icon"] if cfg else "📍",
        "status": row["status"],
        "scheduledFor": row["scheduled_for"] or "",
        "occurredAt": row["occurred_at"] or "",
        "managerName": row["manager_name"],
        "sellerName": row["seller_name"] or "",
        "objective": row["objective"] or "",
        "outcome": row["outcome"] or "",
        "agreement": row["agreement"] or "",
        "nextAction": row["next_action"] or "",
        "nextActionDue": row["next_action_due"] or "",
        "revenueBefore": antes,
        "revenueAfter": depois,
        "effectPct": variacao,
        "effectMeasuredAt": row["effect_measured_at"] or "",
        "createdAt": row["created_at"],
    }


def list_visits(
    conn: sqlite3.Connection, company_id: int, user: sqlite3.Row,
    client_key: str = "", status: str = "", limit: int = 200,
) -> list[dict[str, Any]]:
    sql = "SELECT * FROM visits WHERE company_id = ?"
    params: list[Any] = [company_id]
    if data_scope_for_user(conn, user) == "proprio":
        # O vendedor vê as visitas dos clientes dele e aquelas em que participou.
        sql += " AND (UPPER(seller_name) = ? OR client_key IN (SELECT client_code FROM crm_client_profiles "
        sql += "     WHERE company_id = ? AND UPPER(TRIM(COALESCE(internal_seller_name,''))) = ?))"
        eu = normalize_upper(seller_identity_for_user(user))
        params.extend([eu, company_id, eu])
    else:
        permitidas = crm_allowed_units_for_user(conn, user)
        if permitidas is not None:
            if permitidas:
                marcadores = ",".join("?" for _ in permitidas)
                sql += f" AND unit_name IN ({marcadores})"
                params.extend(permitidas)
            else:
                sql += " AND 1 = 0"
    if client_key:
        sql += " AND client_key = ?"
        params.append(client_key)
    if status:
        sql += " AND status = ?"
        params.append(normalize_upper(status))
    sql += " ORDER BY date(COALESCE(NULLIF(occurred_at,''), scheduled_for, created_at)) DESC, id DESC LIMIT ?"
    params.append(int(limit))
    return [visit_row_to_dict(r) for r in conn.execute(sql, params).fetchall()]


def save_visit(
    conn: sqlite3.Connection, company_id: int, user: sqlite3.Row, payload: dict[str, Any]
) -> dict[str, Any]:
    if not user_can_manage_visits(conn, user):
        raise PermissionError("Apenas gestão registra visitas.")

    client_key = normalize_whitespace(payload.get("clientKey"))
    if not client_key:
        raise ValueError("Selecione o cliente.")
    tipo = normalize_upper(payload.get("visitType")) or "RELACIONAMENTO"
    if tipo not in VISIT_TYPE_IDS:
        raise ValueError("Tipo de visita inválido.")
    status = normalize_upper(payload.get("status")) or "PLANEJADA"
    ocorrida = normalize_whitespace(payload.get("occurredAt"))
    if status == "REALIZADA":
        if not ocorrida:
            ocorrida = today_in_brazil().isoformat()
        if not normalize_whitespace(payload.get("outcome")):
            raise ValueError("Descreva o que aconteceu na visita antes de marcar como realizada.")

    perfil = conn.execute(
        "SELECT client_name, city_name, neighborhood, address_line, address_number "
        "FROM crm_client_profiles WHERE company_id = ? AND client_code = ?",
        (company_id, client_key),
    ).fetchone()
    rua = normalize_whitespace(perfil["address_line"]) if perfil else ""
    numero = normalize_whitespace(perfil["address_number"]) if perfil else ""

    unidade = normalize_unit(payload.get("unitName")) or None
    permitidas = crm_allowed_units_for_user(conn, user)
    if permitidas is not None:
        if not permitidas:
            raise ValueError("Seu usuário não tem unidade vinculada. Peça ao administrador para vincular.")
        if unidade not in permitidas:
            unidade = permitidas[0]

    campos = (
        client_key,
        normalize_whitespace(payload.get("clientName")) or (perfil["client_name"] if perfil else client_key),
        unidade,
        normalize_upper(perfil["city_name"]) if perfil else normalize_upper(payload.get("cityName")),
        normalize_upper(perfil["neighborhood"]) if perfil else None,
        f"{rua}, {numero}" if rua and numero else rua,
        tipo, status,
        normalize_whitespace(payload.get("scheduledFor")) or None,
        ocorrida or None,
        normalize_whitespace(payload.get("managerName")) or meeting_person_identity(user),
        normalize_whitespace(payload.get("sellerName")) or None,
        normalize_whitespace(payload.get("objective")),
        normalize_whitespace(payload.get("outcome")),
        normalize_whitespace(payload.get("agreement")),
        normalize_whitespace(payload.get("nextAction")),
        normalize_whitespace(payload.get("nextActionDue")) or None,
    )

    visit_id = payload.get("id")
    if visit_id:
        conn.execute(
            """
            UPDATE visits SET client_key=?, client_name=?, unit_name=?, city_name=?, neighborhood=?,
                   address_line=?, visit_type=?, status=?, scheduled_for=?, occurred_at=?,
                   manager_name=?, seller_name=?, objective=?, outcome=?, agreement=?,
                   next_action=?, next_action_due=?, updated_at=?
            WHERE company_id = ? AND id = ?
            """,
            (*campos, now_iso(), company_id, int(visit_id)),
        )
        novo_id = int(visit_id)
    else:
        cursor = conn.execute(
            """
            INSERT INTO visits (client_key, client_name, unit_name, city_name, neighborhood,
                address_line, visit_type, status, scheduled_for, occurred_at, manager_name,
                seller_name, objective, outcome, agreement, next_action, next_action_due,
                company_id, request_id, created_by_user_id, created_at)
            VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)
            """,
            (*campos, company_id, payload.get("requestId"), user["id"], now_iso()),
        )
        novo_id = int(cursor.lastrowid)

    # O pedido do vendedor é encerrado quando a visita nasce dele.
    if payload.get("requestId"):
        conn.execute(
            "UPDATE visit_requests SET status='ACEITA', resolved_by_user_id=?, resolved_at=? "
            "WHERE company_id = ? AND id = ? AND status = 'PENDENTE'",
            (user["id"], now_iso(), company_id, int(payload["requestId"])),
        )

    # Ação direcionada vira tarefa do vendedor. É o elo que faz a visita
    # continuar depois que o gerente vai embora.
    tarefa_id = None
    proxima = normalize_whitespace(payload.get("nextAction"))
    vendedor = normalize_whitespace(payload.get("sellerName"))
    if status == "REALIZADA" and proxima and vendedor:
        aberta = conn.execute(
            "SELECT id FROM crm_tasks WHERE company_id = ? AND client_key = ? AND seller_name = ? "
            "AND status IN ('ABERTA','ATRASADA')",
            (company_id, client_key, vendedor),
        ).fetchone()
        if not aberta:
            cur = conn.execute(
                """
                INSERT INTO crm_tasks (company_id, client_key, client_name, seller_name, title,
                    description, due_at, status, origin, created_by_name, created_by_user_id, created_at)
                VALUES (?,?,?,?,?,?,?, 'ABERTA', 'VISITA', ?, ?, ?)
                """,
                (company_id, client_key, campos[1], vendedor,
                 f"Pós-visita: {campos[1]}", proxima,
                 normalize_whitespace(payload.get("nextActionDue")) or today_in_brazil().isoformat(),
                 meeting_person_identity(user), user["id"], now_iso()),
            )
            tarefa_id = int(cur.lastrowid)

    audit_log(conn, company_id, user["id"], "salvar", "visits", str(novo_id),
              {"cliente": client_key, "tipo": tipo, "status": status})
    conn.commit()
    invalidate_crm_cache(company_id)
    if status == "REALIZADA":
        measure_visit_effect(conn, company_id, novo_id)
    return {"visitId": novo_id, "taskId": tarefa_id}


def delete_visit(conn: sqlite3.Connection, company_id: int, user: sqlite3.Row, visit_id: int) -> None:
    if not user_can_manage_visits(conn, user):
        raise PermissionError("Apenas gestão exclui visitas.")
    conn.execute("DELETE FROM visits WHERE company_id = ? AND id = ?", (company_id, visit_id))
    audit_log(conn, company_id, user["id"], "excluir", "visits", str(visit_id), {})
    conn.commit()


def client_contact_effect(
    conn: sqlite3.Connection, company_id: int, client_key: str
) -> dict[str, Any]:
    """Histórico de contatos e visitas do cliente, com o efeito de cada visita.

    É o que responde a pergunta que o gerente faz: "adiantou ter ido lá?".
    """
    visitas = [
        visit_row_to_dict(r)
        for r in conn.execute(
            "SELECT * FROM visits WHERE company_id = ? AND client_key = ? "
            "ORDER BY date(COALESCE(NULLIF(occurred_at,''), scheduled_for)) DESC LIMIT 20",
            (company_id, client_key),
        ).fetchall()
    ]
    # Mede o que ainda não foi medido e já fechou a janela
    for v in visitas:
        if v["status"] == "REALIZADA" and v["revenueBefore"] is None:
            medido = measure_visit_effect(conn, company_id, v["id"])
            if medido:
                v["revenueBefore"] = medido["revenueBefore"]
                v["revenueAfter"] = medido["revenueAfter"]
                v["effectPct"] = (round(safe_div(medido["revenueAfter"] - medido["revenueBefore"],
                                                 medido["revenueBefore"]) * 100, 1)
                                  if medido["revenueBefore"] else None)

    ligacoes = conn.execute(
        """
        SELECT COUNT(*) total,
               SUM(CASE WHEN date(substr(replace(occurred_at,'T',' '),1,10)) >= date(?) THEN 1 ELSE 0 END) recentes,
               MAX(occurred_at) ultima
        FROM crm_interactions
        WHERE company_id = ? AND client_key = ? AND contact_type_code = 'LIGACAO' AND initiative = 'ATIVO'
        """,
        ((today_in_brazil() - timedelta(days=VISIT_CALL_WINDOW_DAYS)).isoformat(), company_id, client_key),
    ).fetchone()

    pedido = conn.execute(
        "SELECT * FROM visit_requests WHERE company_id = ? AND client_key = ? AND status = 'PENDENTE'",
        (company_id, client_key),
    ).fetchone()

    return {
        "visits": visitas,
        "callsTotal": int(ligacoes["total"] or 0),
        "callsRecent": int(ligacoes["recentes"] or 0),
        "lastCallAt": ligacoes["ultima"] or "",
        "pendingRequest": {
            "id": int(pedido["id"]), "reason": pedido["reason"], "sellerName": pedido["seller_name"],
            "createdAt": pedido["created_at"],
        } if pedido else None,
        "eligibleForVisit": int(ligacoes["recentes"] or 0) > 0,
        "callWindowDays": VISIT_CALL_WINDOW_DAYS,
        "canRequestVisit": (int(ligacoes["recentes"] or 0) > 0 and pedido is None),
    }


# ─────────────────────────────────────────────────────────────────────────────
# Tarefas
#
# A tabela nasceu só para follow-up de cliente. Agora carrega quatro origens
# diferentes, e distinguir isso importa: o vendedor precisa saber se a tarefa
# veio da própria ligação dele, de uma cobrança do gerente, de uma visita ou de
# um direcionamento da gestão — a conversa que ele terá é outra em cada caso.
#
# Tarefa "para a equipe" vira UMA tarefa POR PESSOA. Uma tarefa coletiva que
# qualquer um pode concluir não é de ninguém, e ninguém faz.
# ─────────────────────────────────────────────────────────────────────────────

TASK_ORIGINS = [
    {"id": "FOLLOWUP", "label": "Follow-up", "icon": "🔁", "color": "#1a5276", "bg": "#e8f0fe",
     "hint": "Nasceu do registro de contato do próprio vendedor."},
    {"id": "COBRANCA", "label": "Cobrança da gestão", "icon": "📣", "color": "#b06000", "bg": "#fef7e0",
     "hint": "O gestor pediu o contato com este cliente."},
    {"id": "VISITA", "label": "Pós-visita", "icon": "🗺️", "color": "#1e8e3e", "bg": "#e6f4ea",
     "hint": "Ação combinada durante uma visita gerencial."},
    {"id": "LIVRE", "label": "Direcionamento", "icon": "🎯", "color": "#6a1b9a", "bg": "#f3e5f5",
     "hint": "Tarefa de gestão, com ou sem cliente vinculado."},
    {"id": "APOIO", "label": "Atendido por outro", "icon": "🤝", "color": "#00695c", "bg": "#e0f2f1",
     "hint": "Outro vendedor atendeu este cliente da sua carteira — retome o contato."},
]
TASK_ORIGIN_IDS = {o["id"] for o in TASK_ORIGINS}

TASK_PRIORITIES = [
    {"id": "ALTA", "label": "Alta", "color": "#c5221f", "bg": "#fce8e6"},
    {"id": "NORMAL", "label": "Normal", "color": "#5f6368", "bg": "#f1f3f4"},
]

TASK_STATUS_FILTERS = [
    {"id": "ABERTAS", "label": "Em aberto"},
    {"id": "CONCLUIDAS", "label": "Concluídas"},
    {"id": "ATRASADAS", "label": "Atrasadas"},
    {"id": "TODAS", "label": "Todas"},
]


def task_assignable_people(
    conn: sqlite3.Connection, company_id: int, user: sqlite3.Row
) -> list[dict[str, Any]]:
    """Para quem este usuário pode criar tarefa.

    Gestor de unidade: a equipe dele (inclusive ele mesmo). Diretoria: todo
    mundo, o que inclui os gerentes — é assim que o direcionamento tático desce.
    """
    pessoas = list_meeting_people(conn, company_id, user)
    eu = meeting_person_identity(user)
    if eu and not any(person_key(p["personName"]) == person_key(eu) for p in pessoas):
        # O gestor pode não estar no cadastro de pessoas da unidade dele, mas
        # precisa poder criar tarefa para si mesmo.
        pessoas.insert(0, {"personName": eu, "personKey": person_key(eu),
                           "unitName": "", "role": "Gestão", "hasLogin": True})
    return pessoas


# ─── Histórico de contatos ──────────────────────────────────────────────────

CALL_GOAL_MONTH = 60  # piso do MEC: 60 ligações ativas no mês


def contact_history(
    conn: sqlite3.Connection, company_id: int, user: sqlite3.Row,
    filtros: dict[str, Any],
) -> dict[str, Any]:
    """Histórico de registros com KPIs por vendedor.

    Uma consulta só traz as linhas; os indicadores saem de uma agregação
    separada, para o KPI não mudar quando o gerente pagina ou filtra por tipo.
    Vendedor vê os dele; gerente vê os da equipe; diretor vê tudo.
    """
    inicio = normalize_whitespace(filtros.get("start"))
    fim = normalize_whitespace(filtros.get("end"))
    if not inicio or not fim:
        hoje = today_in_brazil()
        inicio = inicio or hoje.replace(day=1).isoformat()
        fim = fim or hoje.isoformat()

    condicoes = ["i.company_id = ?",
                 "date(substr(replace(i.occurred_at,'T',' '),1,10)) BETWEEN date(?) AND date(?)"]
    params: list[Any] = [company_id, inicio, fim]

    visiveis = task_visible_sellers(conn, company_id, user)
    if visiveis is not None:
        if not visiveis:
            return {"items": [], "sellers": [], "totals": {}, "start": inicio, "end": fim,
                    "sellerOptions": []}
        marcadores = ",".join("?" for _ in visiveis)
        condicoes.append(f"UPPER(i.seller_name) IN ({marcadores})")
        params.extend(normalize_upper(v) for v in visiveis)

    vendedor = normalize_whitespace(filtros.get("seller"))
    if vendedor:
        condicoes.append("UPPER(i.seller_name) = ?")
        params.append(normalize_upper(vendedor))
    tipo = normalize_upper(filtros.get("type"))
    if tipo:
        condicoes.append("i.contact_type_code = ?")
        params.append(tipo)
    resultado = normalize_upper(filtros.get("result"))
    if resultado:
        condicoes.append("i.result_code = ?")
        params.append(resultado)
    iniciativa = normalize_upper(filtros.get("initiative"))
    if iniciativa in {INITIATIVE_ACTIVE, INITIATIVE_RECEPTIVE, INITIATIVE_SUPPORT}:
        condicoes.append("i.initiative = ?")
        params.append(iniciativa)
    busca = normalize_whitespace(filtros.get("search"))
    if busca:
        condicoes.append("(UPPER(i.client_name) LIKE ? OR i.client_key LIKE ?)")
        params.extend([f"%{normalize_upper(busca)}%", f"%{busca}%"])

    onde = " AND ".join(condicoes)
    limite = min(int(filtros.get("limit") or 300), 1000)

    itens = [dict(r) for r in conn.execute(
        f"""
        SELECT i.id, i.client_key, i.client_name, i.seller_name, i.unit_name,
               i.contact_type_code, i.result_code, i.occurred_at, i.notes,
               i.next_action, i.followup_due_at, i.initiative, i.offer_title,
               t.label AS type_label, r.label AS result_label
        FROM crm_interactions i
        LEFT JOIN crm_contact_types t ON t.code = i.contact_type_code
        LEFT JOIN crm_contact_results r ON r.code = i.result_code
        WHERE {onde}
        ORDER BY i.occurred_at DESC, i.id DESC
        LIMIT ?
        """,
        (*params, limite),
    ).fetchall()]

    # ── KPIs por vendedor ─────────────────────────────────────────────────
    # Sem o LIMIT: o indicador tem de refletir o período inteiro, não a página.
    agregado = conn.execute(
        f"""
        SELECT
            COUNT(*) AS registros,
            SUM(CASE WHEN i.initiative = 'ATIVO' THEN 1 ELSE 0 END) AS ativos,
            SUM(CASE WHEN i.initiative = 'RECEPTIVO' THEN 1 ELSE 0 END) AS receptivos,
            SUM(CASE WHEN i.initiative = 'APOIO' THEN 1 ELSE 0 END) AS apoios,
            SUM(CASE WHEN i.initiative = 'ATIVO' AND i.contact_type_code = 'LIGACAO'
                     THEN 1 ELSE 0 END) AS ligacoes,
            SUM(CASE WHEN i.initiative = 'ATIVO' AND i.result_code = 'FALOU_CLIENTE'
                     THEN 1 ELSE 0 END) AS falou,
            SUM(CASE WHEN i.initiative = 'ATIVO'
                      AND i.result_code IN ('GEROU_ORCAMENTO','GEROU_PEDIDO')
                     THEN 1 ELSE 0 END) AS converteu,
            COUNT(DISTINCT CASE WHEN i.initiative = 'ATIVO' THEN i.client_key END) AS clientes
        FROM crm_interactions i
        WHERE {onde}
        """,
        params,
    ).fetchone()

    por_vendedor = [dict(r) for r in conn.execute(
        f"""
        SELECT i.seller_name,
            COUNT(*) AS registros,
            SUM(CASE WHEN i.initiative = 'ATIVO' THEN 1 ELSE 0 END) AS ativos,
            SUM(CASE WHEN i.initiative = 'RECEPTIVO' THEN 1 ELSE 0 END) AS receptivos,
            SUM(CASE WHEN i.initiative = 'APOIO' THEN 1 ELSE 0 END) AS apoios,
            SUM(CASE WHEN i.initiative = 'ATIVO' AND i.contact_type_code = 'LIGACAO'
                     THEN 1 ELSE 0 END) AS ligacoes,
            SUM(CASE WHEN i.initiative = 'ATIVO' AND i.result_code = 'FALOU_CLIENTE'
                     THEN 1 ELSE 0 END) AS falou,
            SUM(CASE WHEN i.initiative = 'ATIVO'
                      AND i.result_code IN ('GEROU_ORCAMENTO','GEROU_PEDIDO')
                     THEN 1 ELSE 0 END) AS converteu,
            COUNT(DISTINCT CASE WHEN i.initiative = 'ATIVO' THEN i.client_key END) AS clientes,
            MAX(i.occurred_at) AS ultimo
        FROM crm_interactions i
        WHERE {onde}
        GROUP BY i.seller_name
        ORDER BY ligacoes DESC, i.seller_name
        """,
        params,
    ).fetchall()]

    # Meta de ligações no RITMO do período, não a meta fechada do mês. Cobrar 60
    # no dia 6 marcaria todo mundo como irregular — é a mesma regra dos faróis.
    competencia = inicio[:7]
    calendario = get_business_calendar(conn, company_id, competencia)
    decorridos = int(calendario.get("elapsedWorkingDays") or 0)
    totais = int(calendario.get("totalWorkingDays") or 0)
    ritmo = safe_div(decorridos, totais) if totais else 1.0
    meta_ate_hoje = round(CALL_GOAL_MONTH * ritmo) if ritmo else CALL_GOAL_MONTH

    def enriquece(linha: dict[str, Any]) -> dict[str, Any]:
        ativos = int(linha.get("ativos") or 0)
        ligacoes = int(linha.get("ligacoes") or 0)
        return {
            **linha,
            "callsTargetToDate": meta_ate_hoje,
            "callsPacePct": round(safe_div(ligacoes, meta_ate_hoje) * 100, 1) if meta_ate_hoje else None,
            "talkRatePct": round(safe_div(int(linha.get("falou") or 0), ativos) * 100, 1) if ativos else 0.0,
            "conversionPct": round(safe_div(int(linha.get("converteu") or 0), ativos) * 100, 1) if ativos else 0.0,
        }

    escopo = data_scope_for_user(conn, user)
    return {
        "start": inicio,
        "end": fim,
        "items": itens,
        "truncated": len(itens) >= limite,
        "totals": enriquece(dict(agregado) if agregado else {}),
        "sellers": [enriquece(l) for l in por_vendedor],
        "sellerOptions": sorted({l["seller_name"] for l in por_vendedor if l["seller_name"]}),
        "contactTypes": [dict(r) for r in conn.execute(
            "SELECT code, label FROM crm_contact_types WHERE is_active = 1 ORDER BY code").fetchall()],
        "contactResults": [dict(r) for r in conn.execute(
            "SELECT code, label FROM crm_contact_results WHERE is_active = 1 ORDER BY code").fetchall()],
        "receptiveTypes": CRM_RECEPTIVE_TYPES,
        "callGoalMonth": CALL_GOAL_MONTH,
        "isManagerView": escopo != "proprio",
    }


def task_visible_sellers(
    conn: sqlite3.Connection, company_id: int, user: sqlite3.Row
) -> list[str] | None:
    """Nomes cujas tarefas este usuário pode ver. None = sem restrição."""
    scope = data_scope_for_user(conn, user)
    if scope == "todos":
        return None
    if scope == "proprio":
        return [seller_identity_for_user(user)]
    nomes = {p["personName"] for p in task_assignable_people(conn, company_id, user)}
    nomes.add(meeting_person_identity(user))
    return sorted(n for n in nomes if n)


def list_crm_tasks(
    conn: sqlite3.Connection, company_id: int, user: sqlite3.Row,
    status: str = "ABERTAS", seller: str = "", date_from: str = "", date_to: str = "",
    origin: str = "", search: str = "", limit: int = 400,
) -> list[dict[str, Any]]:
    sql = """
        SELECT id, client_key, client_name, seller_name, title, description, due_at, status,
               source_interaction_id, created_at, completed_at,
               COALESCE(origin, 'FOLLOWUP') AS origin,
               COALESCE(priority, 'NORMAL') AS priority,
               created_by_name
        FROM crm_tasks
        WHERE company_id = ?
    """
    params: list[Any] = [company_id]

    visiveis = task_visible_sellers(conn, company_id, user)
    if visiveis is not None:
        if not visiveis:
            return []
        marcadores = ",".join("?" for _ in visiveis)
        sql += f" AND UPPER(seller_name) IN ({marcadores})"
        params.extend(normalize_upper(n) for n in visiveis)

    if seller:
        sql += " AND UPPER(seller_name) = ?"
        params.append(normalize_upper(seller))

    filtro = normalize_upper(status) or "ABERTAS"
    if filtro == "ABERTAS":
        sql += " AND status NOT IN ('CONCLUIDA', 'CANCELADA')"
    elif filtro == "CONCLUIDAS":
        sql += " AND status = 'CONCLUIDA'"
    elif filtro == "ATRASADAS":
        sql += " AND status NOT IN ('CONCLUIDA','CANCELADA') AND date(due_at) < date(?)"
        params.append(today_in_brazil().isoformat())

    if date_from:
        sql += " AND date(due_at) >= date(?)"
        params.append(date_from)
    if date_to:
        sql += " AND date(due_at) <= date(?)"
        params.append(date_to)
    if origin in TASK_ORIGIN_IDS:
        sql += " AND COALESCE(origin,'FOLLOWUP') = ?"
        params.append(origin)
    termo = normalize_whitespace(search)
    if termo:
        alvo = f"%{termo.upper()}%"
        sql += (" AND (UPPER(title) LIKE ? OR UPPER(COALESCE(client_name,'')) LIKE ?"
                " OR UPPER(COALESCE(description,'')) LIKE ?)")
        params.extend([alvo, alvo, alvo])

    sql += (" ORDER BY CASE status WHEN 'ATRASADA' THEN 0 WHEN 'ABERTA' THEN 1"
            " WHEN 'REAGENDADA' THEN 2 ELSE 3 END, datetime(due_at) ASC LIMIT ?")
    params.append(int(limit))

    hoje = today_in_brazil().isoformat()
    linhas = []
    for r in conn.execute(sql, params).fetchall():
        item = dict(r)
        vencimento = (item.get("due_at") or "")[:10]
        item["overdue"] = bool(
            vencimento and vencimento < hoje and item["status"] not in ("CONCLUIDA", "CANCELADA")
        )
        linhas.append(item)
    return linhas


def crm_task_counters(
    conn: sqlite3.Connection, company_id: int, user: sqlite3.Row
) -> dict[str, int]:
    """Números do topo da tela. Contam sobre o mesmo recorte da lista."""
    visiveis = task_visible_sellers(conn, company_id, user)
    where = "company_id = ?"
    params: list[Any] = [company_id]
    if visiveis is not None:
        if not visiveis:
            return {"open": 0, "overdue": 0, "today": 0, "doneMonth": 0}
        marcadores = ",".join("?" for _ in visiveis)
        where += f" AND UPPER(seller_name) IN ({marcadores})"
        params.extend(normalize_upper(n) for n in visiveis)
    hoje = today_in_brazil().isoformat()
    mes = hoje[:7]

    def conta(extra: str, extras: list[Any]) -> int:
        row = conn.execute(
            f"SELECT COUNT(*) n FROM crm_tasks WHERE {where} AND {extra}", (*params, *extras)
        ).fetchone()
        return int(row["n"] or 0)

    return {
        "open": conta("status NOT IN ('CONCLUIDA','CANCELADA')", []),
        "overdue": conta("status NOT IN ('CONCLUIDA','CANCELADA') AND date(due_at) < date(?)", [hoje]),
        "today": conta("status NOT IN ('CONCLUIDA','CANCELADA') AND date(due_at) = date(?)", [hoje]),
        "doneMonth": conta("status = 'CONCLUIDA' AND substr(COALESCE(completed_at, ''),1,7) = ?", [mes]),
    }


def create_crm_tasks(
    conn: sqlite3.Connection, company_id: int, user: sqlite3.Row, payload: dict[str, Any]
) -> dict[str, Any]:
    """Cria tarefa de direcionamento — uma por pessoa escolhida.

    Cliente é OPCIONAL aqui: "revisar os orçamentos do dia e levantar os motivos
    de desistência" é tarefa legítima e não pertence a nenhum cliente.
    """
    if data_scope_for_user(conn, user) == "proprio":
        raise PermissionError("Apenas gestão cria direcionamento.")

    titulo = normalize_whitespace(payload.get("title"))
    if not titulo:
        raise ValueError("Escreva o que precisa ser feito.")
    vencimento = normalize_whitespace(payload.get("dueAt")) or today_in_brazil().isoformat()

    permitidas = {person_key(p["personName"]): p["personName"]
                  for p in task_assignable_people(conn, company_id, user)}
    escolhidos = payload.get("assignees") or []
    if isinstance(escolhidos, str):
        escolhidos = [escolhidos]

    destinos: list[str] = []
    for nome in escolhidos:
        chave = person_key(nome)
        if chave in permitidas:
            destinos.append(permitidas[chave])
        elif normalize_whitespace(nome):
            # Nome fora da equipe do gestor é recusado, não ignorado em silêncio.
            raise PermissionError(f"{nome} não está entre as pessoas que você pode direcionar.")
    destinos = list(dict.fromkeys(destinos))
    if not destinos:
        raise ValueError("Escolha ao menos uma pessoa para receber a tarefa.")

    prioridade = normalize_upper(payload.get("priority")) or "NORMAL"
    if prioridade not in {p["id"] for p in TASK_PRIORITIES}:
        prioridade = "NORMAL"
    client_key = normalize_whitespace(payload.get("clientKey"))
    client_name = normalize_whitespace(payload.get("clientName"))
    if client_key and not client_name:
        perfil = conn.execute(
            "SELECT client_name FROM crm_client_profiles WHERE company_id = ? AND client_code = ?",
            (company_id, client_key),
        ).fetchone()
        client_name = perfil["client_name"] if perfil else client_key

    criados = []
    for nome in destinos:
        cursor = conn.execute(
            """
            INSERT INTO crm_tasks (company_id, client_key, client_name, seller_name, title,
                description, due_at, status, origin, priority, created_by_name,
                created_by_user_id, created_at)
            VALUES (?,?,?,?,?,?,?, 'ABERTA', 'LIVRE', ?,?,?,?)
            """,
            (company_id, client_key or "", client_name or "", nome, titulo,
             normalize_whitespace(payload.get("description")), vencimento,
             prioridade, meeting_person_identity(user), user["id"], now_iso()),
        )
        criados.append({"taskId": int(cursor.lastrowid), "sellerName": nome})

    audit_log(conn, company_id, user["id"], "criar", "crm_tasks", client_key or "-",
              {"titulo": titulo, "destinos": len(criados), "origem": "direcionamento"})
    conn.commit()
    invalidate_crm_cache(company_id)
    return {"created": len(criados), "tasks": criados}


def delete_crm_task(conn: sqlite3.Connection, company_id: int, user: sqlite3.Row, task_id: int) -> None:
    if data_scope_for_user(conn, user) == "proprio":
        raise PermissionError("Apenas gestão exclui tarefa.")
    conn.execute("DELETE FROM crm_tasks WHERE company_id = ? AND id = ?", (company_id, task_id))
    audit_log(conn, company_id, user["id"], "excluir", "crm_tasks", str(task_id), {})
    conn.commit()



# ─────────────────────────────────────────────────────────────────────────────
# Assistente: tutorial, FAQ e dicas
#
# A caixa de perguntas NÃO é um chat inteligente — o servidor é local e não fala
# com nenhum serviço de IA. Ela busca na base de conhecimento por palavra. Isso
# fica dito na tela: prometer conversa e entregar busca destrói a confiança na
# primeira pergunta.
#
# Quando a busca não acha, a pergunta é registrada para o admin responder. É
# assim que o FAQ cresce com a dúvida real da equipe em vez da que imaginamos.
# ─────────────────────────────────────────────────────────────────────────────

def _help_content() -> Any:
    import help_content
    return help_content


def help_role_for_user(conn: sqlite3.Connection, user: sqlite3.Row) -> str:
    """Perfil do usuário na linguagem do conteúdo de ajuda."""
    scope = data_scope_for_user(conn, user)
    if scope == "proprio":
        return "VENDEDOR"
    if scope == "todos":
        return "DIRETOR"
    return "GERENTE"


def _matches_role(roles: str | None, papel: str) -> bool:
    """Campo `roles` vazio vale para todos os perfis."""
    texto = normalize_upper(roles)
    return not texto or papel in texto.split()


def seed_help_content(conn: sqlite3.Connection, company_id: int) -> None:
    """Popula FAQ e dicas na primeira execução e a cada conteúdo novo.

    INSERT OR IGNORE de propósito: o que o admin editou pela tela nunca é
    sobrescrito por uma atualização do arquivo.
    """
    try:
        conteudo = _help_content()
    except Exception as exc:
        print(f"[ajuda] help_content.py indisponível: {exc}", flush=True)
        return

    novos_faq = 0
    for i, item in enumerate(conteudo.FAQ_SEED):
        cur = conn.execute(
            """
            INSERT OR IGNORE INTO help_articles
                (company_id, category, question, answer, keywords, roles, source, sort_order, created_at)
            VALUES (?,?,?,?,?,?,'SEED',?,?)
            """,
            (company_id, item["category"], item["question"], item["answer"],
             item.get("keywords", ""), item.get("roles", ""), i, now_iso()),
        )
        novos_faq += cur.rowcount or 0

    novas_dicas = 0
    for item in conteudo.TIPS_SEED:
        cur = conn.execute(
            """
            INSERT OR IGNORE INTO assistant_tips
                (company_id, kind, title, body, roles, trigger_code, source, created_at)
            VALUES (?,?,?,?,?,?,'SEED',?)
            """,
            (company_id, item["kind"], item["title"], item["body"],
             item.get("roles", ""), item.get("trigger", ""), now_iso()),
        )
        novas_dicas += cur.rowcount or 0

    if novos_faq or novas_dicas:
        conn.commit()
        print(f"[ajuda] {novos_faq} pergunta(s) e {novas_dicas} dica(s) adicionadas", flush=True)


def list_help_articles(
    conn: sqlite3.Connection, company_id: int, papel: str, incluir_inativos: bool = False
) -> list[dict[str, Any]]:
    sql = "SELECT * FROM help_articles WHERE company_id = ?"
    if not incluir_inativos:
        sql += " AND is_active = 1"
    sql += " ORDER BY sort_order, id"
    return [
        {
            "id": int(r["id"]), "category": r["category"], "question": r["question"],
            "answer": r["answer"], "keywords": r["keywords"] or "", "roles": r["roles"] or "",
            "source": r["source"], "isActive": bool(r["is_active"]),
        }
        for r in conn.execute(sql, (company_id,)).fetchall()
        if _matches_role(r["roles"], papel)
    ]


def search_help(
    conn: sqlite3.Connection, company_id: int, papel: str, pergunta: str, limite: int = 5
) -> list[dict[str, Any]]:
    """Busca por palavras, com pontuação simples.

    Sem motor de busca e sem IA: cada palavra da pergunta com 3+ letras vale
    pontos conforme onde aparece. Pergunta vale mais que palavra-chave, que vale
    mais que o corpo da resposta — quem acerta o título costuma ser a resposta
    certa. É simples, previsível e explicável, que é o que importa aqui.
    """
    termo = normalize_upper(strip_accents(normalize_whitespace(pergunta)))
    # Palavras vazias de significado ficam de fora. Sem isso, "esqueci MINHA
    # senha" casava com qualquer resposta que tivesse "minha" no meio do texto.
    vazias = {
        "COMO", "ONDE", "QUANDO", "QUAL", "QUAIS", "QUE", "POR", "PARA", "COM",
        "SEM", "DOS", "DAS", "UMA", "UNS", "UMAS", "MEU", "MINHA", "MEUS",
        "MINHAS", "SEU", "SUA", "NAO", "SIM", "ISSO", "ESSE", "ESSA", "ESTA",
        "ESTE", "AQUI", "MAS", "TEM", "FAZ", "FOI", "SAO", "ERA", "DEU", "VOU",
        "PODE", "POSSO", "DEVO", "AGORA", "DEPOIS", "ANTES", "MAIS", "MENOS",
    }
    palavras = [p for p in re.split(r"[^A-Z0-9]+", termo) if len(p) >= 3 and p not in vazias]
    if not palavras:
        return []

    resultados = []
    for artigo in list_help_articles(conn, company_id, papel):
        alvo_q = normalize_upper(strip_accents(artigo["question"]))
        alvo_k = normalize_upper(strip_accents(artigo["keywords"]))
        alvo_a = normalize_upper(strip_accents(artigo["answer"]))
        pontos = 0
        acertos = 0
        for palavra in palavras:
            achou = False
            if palavra in alvo_q:
                pontos += 5; achou = True
            if palavra in alvo_k:
                pontos += 3; achou = True
            if palavra in alvo_a:
                pontos += 1; achou = True
            if achou:
                acertos += 1
        # Exige boa parte das palavras da pergunta. Metade era frouxo demais e
        # trazia resposta que só compartilhava uma palavra qualquer.
        # Até 2 palavras, uma boa já basta ("como funciona o farol" vira
        # FUNCIONA + FAROL, e só FAROL importa). De 3 em diante, exige 60%.
        minimo = 1 if len(palavras) <= 2 else math.ceil(len(palavras) * 0.6)
        if pontos and acertos >= minimo:
            resultados.append({**artigo, "score": pontos})

    resultados.sort(key=lambda x: x["score"], reverse=True)
    return resultados[:limite]


def register_help_question(
    conn: sqlite3.Connection, company_id: int, user: sqlite3.Row, pergunta: str
) -> int:
    texto = normalize_whitespace(pergunta)
    if not texto:
        return 0
    ja = conn.execute(
        "SELECT id FROM help_questions WHERE company_id = ? AND UPPER(question) = ? AND status = 'PENDENTE'",
        (company_id, normalize_upper(texto)),
    ).fetchone()
    if ja:
        return int(ja["id"])
    cur = conn.execute(
        """
        INSERT INTO help_questions (company_id, user_id, user_name, user_role, question, created_at)
        VALUES (?,?,?,?,?,?)
        """,
        (company_id, user["id"], meeting_person_identity(user),
         help_role_for_user(conn, user), texto, now_iso()),
    )
    conn.commit()
    return int(cur.lastrowid)


def list_help_questions(
    conn: sqlite3.Connection, company_id: int, status: str = "PENDENTE"
) -> list[dict[str, Any]]:
    sql = "SELECT * FROM help_questions WHERE company_id = ?"
    params: list[Any] = [company_id]
    if status:
        sql += " AND status = ?"
        params.append(normalize_upper(status))
    sql += " ORDER BY datetime(created_at) DESC LIMIT 200"
    return [
        {
            "id": int(r["id"]), "question": r["question"], "userName": r["user_name"] or "",
            "userRole": r["user_role"] or "", "status": r["status"],
            "answer": r["answer"] or "", "createdAt": r["created_at"],
        }
        for r in conn.execute(sql, params).fetchall()
    ]


def answer_help_question(
    conn: sqlite3.Connection, company_id: int, user: sqlite3.Row, payload: dict[str, Any]
) -> dict[str, Any]:
    """Responde a dúvida e, se pedido, publica no FAQ para os próximos."""
    if not user_can_manage_users(conn, user):
        raise PermissionError("Apenas a diretoria responde as dúvidas.")
    question_id = int(payload.get("questionId") or 0)
    linha = conn.execute(
        "SELECT * FROM help_questions WHERE company_id = ? AND id = ?", (company_id, question_id)
    ).fetchone()
    if not linha:
        raise ValueError("Pergunta não encontrada.")

    if payload.get("discard"):
        conn.execute("UPDATE help_questions SET status='DESCARTADA' WHERE id = ?", (question_id,))
        conn.commit()
        return {"discarded": True}

    resposta = normalize_whitespace(payload.get("answer"))
    if not resposta:
        raise ValueError("Escreva a resposta.")

    conn.execute(
        "UPDATE help_questions SET status='RESPONDIDA', answer=?, answered_by_user_id=?, answered_at=? "
        "WHERE id = ?",
        (resposta, user["id"], now_iso(), question_id),
    )

    publicado = False
    if payload.get("publish"):
        titulo = normalize_whitespace(payload.get("question")) or linha["question"]
        conn.execute(
            """
            INSERT INTO help_articles (company_id, category, question, answer, keywords, roles,
                source, sort_order, created_by_user_id, created_at)
            VALUES (?,?,?,?,?,?, 'MANUAL', 100, ?, ?)
            ON CONFLICT(company_id, question) DO UPDATE SET
                answer = excluded.answer, keywords = excluded.keywords,
                category = excluded.category, updated_at = excluded.created_at
            """,
            (company_id, normalize_whitespace(payload.get("category")) or "dia-a-dia",
             titulo, resposta, normalize_whitespace(payload.get("keywords")) or titulo,
             normalize_upper(payload.get("roles")), user["id"], now_iso()),
        )
        publicado = True

    conn.commit()
    return {"answered": True, "published": publicado}


def save_help_article(
    conn: sqlite3.Connection, company_id: int, user: sqlite3.Row, payload: dict[str, Any]
) -> dict[str, Any]:
    if not user_can_manage_users(conn, user):
        raise PermissionError("Apenas a diretoria edita o FAQ.")
    pergunta = normalize_whitespace(payload.get("question"))
    resposta = normalize_whitespace(payload.get("answer"))
    if not pergunta or not resposta:
        raise ValueError("Pergunta e resposta são obrigatórias.")
    artigo_id = payload.get("id")
    campos = (
        normalize_whitespace(payload.get("category")) or "dia-a-dia",
        pergunta, resposta,
        normalize_whitespace(payload.get("keywords")) or pergunta,
        normalize_upper(payload.get("roles")),
        0 if payload.get("isActive") is False else 1,
    )
    if artigo_id:
        conn.execute(
            "UPDATE help_articles SET category=?, question=?, answer=?, keywords=?, roles=?, "
            "is_active=?, updated_at=? WHERE company_id = ? AND id = ?",
            (*campos, now_iso(), company_id, int(artigo_id)),
        )
        novo = int(artigo_id)
    else:
        cur = conn.execute(
            "INSERT INTO help_articles (category, question, answer, keywords, roles, is_active, "
            "company_id, source, sort_order, created_by_user_id, created_at) "
            "VALUES (?,?,?,?,?,?,?, 'MANUAL', 100, ?, ?)",
            (*campos, company_id, user["id"], now_iso()),
        )
        novo = int(cur.lastrowid)
    conn.commit()
    return {"articleId": novo}


def delete_help_article(conn: sqlite3.Connection, company_id: int, user: sqlite3.Row, article_id: int) -> None:
    if not user_can_manage_users(conn, user):
        raise PermissionError("Apenas a diretoria edita o FAQ.")
    conn.execute("DELETE FROM help_articles WHERE company_id = ? AND id = ?", (company_id, article_id))
    conn.commit()


def save_assistant_tip(
    conn: sqlite3.Connection, company_id: int, user: sqlite3.Row, payload: dict[str, Any]
) -> dict[str, Any]:
    if not user_can_manage_users(conn, user):
        raise PermissionError("Apenas a diretoria edita as dicas.")
    titulo = normalize_whitespace(payload.get("title"))
    corpo = normalize_whitespace(payload.get("body"))
    if not titulo or not corpo:
        raise ValueError("Título e texto são obrigatórios.")
    kind = normalize_upper(payload.get("kind")) or "MENSAGEM"
    campos = (kind, titulo, corpo, normalize_upper(payload.get("roles")),
              normalize_upper(payload.get("trigger")), 0 if payload.get("isActive") is False else 1)
    tip_id = payload.get("id")
    if tip_id:
        conn.execute(
            "UPDATE assistant_tips SET kind=?, title=?, body=?, roles=?, trigger_code=?, "
            "is_active=?, updated_at=? WHERE company_id = ? AND id = ?",
            (*campos, now_iso(), company_id, int(tip_id)),
        )
        novo = int(tip_id)
    else:
        cur = conn.execute(
            "INSERT INTO assistant_tips (kind, title, body, roles, trigger_code, is_active, "
            "company_id, source, created_by_user_id, created_at) VALUES (?,?,?,?,?,?,?, 'MANUAL', ?, ?)",
            (*campos, company_id, user["id"], now_iso()),
        )
        novo = int(cur.lastrowid)
    conn.commit()
    return {"tipId": novo}


def delete_assistant_tip(conn: sqlite3.Connection, company_id: int, user: sqlite3.Row, tip_id: int) -> None:
    if not user_can_manage_users(conn, user):
        raise PermissionError("Apenas a diretoria edita as dicas.")
    conn.execute("DELETE FROM assistant_tips WHERE company_id = ? AND id = ?", (company_id, tip_id))
    conn.commit()


def user_situation_triggers(
    conn: sqlite3.Connection, company_id: int, user: sqlite3.Row
) -> list[str]:
    """Situações em que este usuário está agora.

    É o que faz a dica ser útil em vez de decorativa: só entra o que responde ao
    número real da pessoa hoje.
    """
    gatilhos: list[str] = []
    papel = help_role_for_user(conn, user)
    hoje = today_in_brazil().isoformat()

    if papel == "VENDEDOR":
        nome = seller_identity_for_user(user)
        atrasadas = conn.execute(
            "SELECT COUNT(*) n FROM crm_tasks WHERE company_id = ? AND UPPER(seller_name) = ? "
            "AND status NOT IN ('CONCLUIDA','CANCELADA') AND date(due_at) < date(?)",
            (company_id, normalize_upper(nome), hoje),
        ).fetchone()["n"]
        if atrasadas:
            gatilhos.append("TASKS_OVERDUE")
        try:
            competencia = crm_latest_competence(conn, company_id) or hoje[:7]
            ind = safe_feedback_indicators(conn, company_id, "VENDEDOR", nome, "", competencia)
            if ind.get("inDeployment"):
                gatilhos.append("DEPLOYMENT")
            if ind.get("found"):
                # Mesmo critério dos faróis: ritmo, não acumulado.
                meta_ligacoes = int(ind.get("callsTargetToDate") or ind.get("callsTarget") or 60)
                if int(ind.get("calls") or 0) < meta_ligacoes:
                    gatilhos.append("CALLS_LOW")
                if not ind.get("inDeployment"):
                    ritmo = ind.get("projectedGoalAttainmentPct")
                    if ritmo is None:
                        ritmo = ind.get("goalAttainmentPct")
                    if float(ritmo or 0) < 90:
                        gatilhos.append("GOAL_LOW")
                # Ligações continuam valendo em implantação — é o esforço que sobra.
                if float(ind.get("returnsPct") or 0) > 3:
                    gatilhos.append("RETURNS_HIGH")
                ticket_unidade = float(ind.get("ticketAverageUnit") or 0)
                if ticket_unidade and float(ind.get("ticketAverage") or 0) < ticket_unidade * 0.85:
                    gatilhos.append("TICKET_LOW")
                total = int(ind.get("portfolioTotal") or 0)
                if total:
                    parados = int(ind.get("portfolioInactive") or 0) + int(ind.get("portfolioPreInactive") or 0)
                    if safe_div(parados, total) > 0.35:
                        gatilhos.append("INACTIVE_HIGH")
        except Exception as exc:
            print(f"[ajuda] indicadores indisponíveis para dicas: {exc}", flush=True)
    else:
        pendentes = conn.execute(
            "SELECT COUNT(*) n FROM visit_requests WHERE company_id = ? AND status = 'PENDENTE'",
            (company_id,),
        ).fetchone()["n"]
        if pendentes:
            gatilhos.append("VISITS_PENDING")
        if papel == "GERENTE":
            competencia = crm_latest_competence(conn, company_id) or hoje[:7]
            feitos = conn.execute(
                "SELECT COUNT(*) n FROM feedbacks WHERE company_id = ? AND kind='VENDEDOR' "
                "AND competence = ? AND status='PUBLICADO'",
                (company_id, competencia),
            ).fetchone()["n"]
            equipe = len([p for p in list_meeting_people(conn, company_id, user)
                          if p["role"] == "Vendedor"])
            if equipe and feitos < equipe:
                gatilhos.append("FEEDBACK_PENDING")
    return gatilhos


def assistant_payload(
    conn: sqlite3.Connection, company_id: int, user: sqlite3.Row
) -> dict[str, Any]:
    conteudo = _help_content()
    papel = help_role_for_user(conn, user)
    gatilhos = user_situation_triggers(conn, company_id, user)

    dicas = []
    for r in conn.execute(
        "SELECT * FROM assistant_tips WHERE company_id = ? AND is_active = 1 ORDER BY id",
        (company_id,),
    ).fetchall():
        if not _matches_role(r["roles"], papel):
            continue
        gatilho = normalize_upper(r["trigger_code"])
        if gatilho and gatilho not in gatilhos:
            continue
        dicas.append({
            "id": int(r["id"]), "kind": r["kind"], "title": r["title"], "body": r["body"],
            "roles": r["roles"] or "", "trigger": gatilho, "source": r["source"],
            "isActive": bool(r["is_active"]),
        })

    # Mensagem do dia: gira pelo dia do ano, para toda a equipe ver a mesma e a
    # frase virar assunto. Aleatório faria cada um ver uma coisa.
    mensagens = [d for d in dicas if d["kind"] == "MENSAGEM"]
    do_dia = mensagens[today_in_brazil().toordinal() % len(mensagens)] if mensagens else None

    # Dica situacional tem prioridade sobre a mensagem: quem está com tarefa
    # atrasada precisa mais de um empurrão concreto que de uma frase.
    situacionais = [d for d in dicas if d["trigger"]]

    tour_key = f"tour-{papel.lower()}-{getattr(conteudo, 'TOUR_VERSION', 'v1')}"
    visto = conn.execute(
        "SELECT completed_at, skipped FROM user_onboarding WHERE user_id = ? AND tour_key = ?",
        (user["id"], tour_key),
    ).fetchone()

    pode_gerir = user_can_manage_users(conn, user)
    return {
        "role": papel,
        "myName": meeting_person_identity(user),
        "tour": conteudo.TOURS.get(papel, []),
        "tourKey": tour_key,
        "tourSeen": bool(visto),
        "messageOfDay": do_dia,
        "situationTips": situacionais,
        "tips": dicas,
        "tipKinds": conteudo.TIP_KINDS,
        "faq": list_help_articles(conn, company_id, papel),
        "faqCategories": conteudo.FAQ_CATEGORIES,
        "canManage": pode_gerir,
        "pendingQuestions": (list_help_questions(conn, company_id, "PENDENTE") if pode_gerir else []),
        "allTips": ([
            {"id": int(r["id"]), "kind": r["kind"], "title": r["title"], "body": r["body"],
             "roles": r["roles"] or "", "trigger": r["trigger_code"] or "",
             "source": r["source"], "isActive": bool(r["is_active"])}
            for r in conn.execute("SELECT * FROM assistant_tips WHERE company_id = ? ORDER BY kind, id",
                                  (company_id,)).fetchall()
        ] if pode_gerir else []),
        "allFaq": (list_help_articles(conn, company_id, papel, incluir_inativos=True)
                   if pode_gerir else []),
    }


def mark_tour_seen(conn: sqlite3.Connection, user: sqlite3.Row, tour_key: str, skipped: bool) -> None:
    conn.execute(
        """
        INSERT INTO user_onboarding (user_id, tour_key, completed_at, skipped)
        VALUES (?,?,?,?)
        ON CONFLICT(user_id, tour_key) DO UPDATE SET
            completed_at = excluded.completed_at, skipped = excluded.skipped
        """,
        (user["id"], normalize_whitespace(tour_key) or "tour", now_iso(), 1 if skipped else 0),
    )
    conn.commit()


# ─────────────────────────────────────────────────────────────────────────────
# Unidade em implantação e prospecção
#
# Uma unidade nova passa meses sem meta e sem carteira. A estrutura do sistema
# pressupõe as duas coisas, então em vez de criar um sistema paralelo — que
# duplicaria carteira, tarefas e placar — a fase da unidade muda a LEITURA do
# que já existe:
#
#   • sem meta de faturamento, o farol vira neutro em vez de crítico;
#   • no lugar da meta entram metas de ESFORÇO (ligações, prospects, cadastros);
#   • o contato com prospect usa a MESMA tabela de interações, com a chave
#     "P-<id>". Assim ele conta no placar, gera tarefa de retorno e aparece no
#     histórico. Quando o CNPJ é cadastrado no Alfa, as interações são
#     reapontadas para o código real e o histórico vai junto para a ficha.
#
# A ideia é que, quando a Zona Norte inaugurar, nada precise ser migrado: a
# unidade só muda de fase.
# ─────────────────────────────────────────────────────────────────────────────

PROSPECT_KEY_PREFIX = "P-"

PROSPECT_STATUSES = [
    {"id": "NOVO",       "label": "A contatar",  "icon": "○", "color": "#5f6368", "bg": "#f1f3f4",
     "hint": "Cadastrado no CRM, ainda sem contato."},
    {"id": "EM_CONTATO", "label": "Em contato",  "icon": "◐", "color": "#1a5276", "bg": "#e8f0fe",
     "hint": "Já houve ligação, ainda sem qualificação completa."},
    {"id": "QUALIFICADO","label": "Qualificado", "icon": "◆", "color": "#b06000", "bg": "#fef7e0",
     "hint": "Respondeu as 4 perguntas e aceitou um gatilho de fechamento."},
    {"id": "CADASTRADO", "label": "Cadastrado",  "icon": "●", "color": "#1e8e3e", "bg": "#e6f4ea",
     "hint": "Virou cliente no Alfa. Já está na carteira."},
    {"id": "PERDIDO",    "label": "Perdido",     "icon": "✕", "color": "#c5221f", "bg": "#fce8e6",
     "hint": "Recusou ou não é público-alvo."},
]
PROSPECT_STATUS_IDS = {s["id"] for s in PROSPECT_STATUSES}

PROSPECT_TRIGGERS = [
    {"id": "ORCAMENTO",  "label": "Participar do próximo orçamento"},
    {"id": "COTACAO",    "label": "Entrar nas cotações da linha"},
    {"id": "DIA_COMPRA", "label": "Dia da semana de compra combinado"},
]

ACTIVITY_METRICS = [
    {"id": "CALLS", "label": "Ligações", "icon": "📞",
     "hint": "Ligações registradas no mês, para clientes e prospects."},
    {"id": "PROSPECTS_NEW", "label": "Prospects novos", "icon": "🆕",
     "hint": "Oficinas novas cadastradas no CRM no mês."},
    {"id": "PROSPECTS_REGISTERED", "label": "Cadastros concluídos", "icon": "✅",
     "hint": "Prospects que viraram cliente no Alfa no mês."},
    {"id": "FIRST_PURCHASES", "label": "Primeiras compras", "icon": "💰",
     "hint": "Prospects convertidos que compraram pela primeira vez."},
]
ACTIVITY_METRIC_IDS = {m["id"] for m in ACTIVITY_METRICS}


def only_digits(value: Any) -> str:
    return re.sub(r"\D", "", str(value or ""))


def repair_prospect_keys(conn: sqlite3.Connection, company_id: int) -> None:
    """Conserta interações e tarefas gravadas com "P 1" em vez de "P-1".

    A normalização antiga trocava o hífen por espaço, então todo contato feito
    em prospect virava órfão: a oficina continuava "nunca contatada" e a fila da
    Missão do Dia nunca esvaziava. Roda uma vez, na subida.
    """
    total = 0
    for tabela in ("crm_interactions", "crm_tasks"):
        linhas = conn.execute(
            f"SELECT id, client_key FROM {tabela} "
            f"WHERE company_id = ? AND client_key GLOB 'P [0-9]*'",
            (company_id,),
        ).fetchall()
        for linha in linhas:
            corrigida = normalize_client_key(linha["client_key"])
            if corrigida != linha["client_key"]:
                conn.execute(f"UPDATE {tabela} SET client_key = ? WHERE id = ?",
                             (corrigida, linha["id"]))
                total += 1
    if total:
        conn.commit()
        print(f"[prospects] {total} registro(s) religado(s) ao prospect correto")
        for row in conn.execute(
            "SELECT id FROM prospects WHERE company_id = ? AND status NOT IN ('CADASTRADO','PERDIDO')",
            (company_id,),
        ).fetchall():
            _refresh_prospect_status(conn, company_id, int(row["id"]))
        conn.commit()


def prospect_client_key(prospect_id: int) -> str:
    """Chave usada nas interações e tarefas de um prospect.

    O prefixo garante que nenhuma consulta de carteira encontre esse registro
    por engano — código de cliente do Alfa é numérico.
    """
    return f"{PROSPECT_KEY_PREFIX}{int(prospect_id)}"


# ── Fase da unidade ──────────────────────────────────────────────────────────

def get_unit_phase(conn: sqlite3.Connection, company_id: int, unit_name: str) -> dict[str, Any]:
    unidade = normalize_unit(unit_name)
    row = conn.execute(
        "SELECT * FROM unit_phases WHERE company_id = ? AND unit_name = ?", (company_id, unidade)
    ).fetchone()
    if not row:
        return {"unitName": unidade, "phase": "OPERACAO", "openingDate": "",
                "goalExemptUntil": "", "notes": "", "isDeployment": False}
    return {
        "unitName": unidade,
        "phase": row["phase"],
        "openingDate": row["opening_date"] or "",
        "goalExemptUntil": row["goal_exempt_until"] or "",
        "notes": row["notes"] or "",
        "isDeployment": row["phase"] == "IMPLANTACAO",
    }


def list_unit_phases(conn: sqlite3.Connection, company_id: int) -> list[dict[str, Any]]:
    return [
        {
            "unitName": r["unit_name"], "phase": r["phase"],
            "openingDate": r["opening_date"] or "", "goalExemptUntil": r["goal_exempt_until"] or "",
            "notes": r["notes"] or "", "isDeployment": r["phase"] == "IMPLANTACAO",
        }
        for r in conn.execute(
            "SELECT * FROM unit_phases WHERE company_id = ? ORDER BY unit_name", (company_id,)
        ).fetchall()
    ]


def save_unit_phase(
    conn: sqlite3.Connection, company_id: int, user: sqlite3.Row, payload: dict[str, Any]
) -> dict[str, Any]:
    if not user_can_manage_users(conn, user):
        raise PermissionError("Apenas a diretoria define a fase da unidade.")
    unidade = normalize_unit(payload.get("unitName"))
    if not unidade:
        raise ValueError("Informe a unidade.")
    fase = normalize_upper(payload.get("phase")) or "OPERACAO"
    if fase not in {"IMPLANTACAO", "OPERACAO"}:
        raise ValueError("Fase inválida.")
    conn.execute(
        """
        INSERT INTO unit_phases (company_id, unit_name, phase, opening_date, goal_exempt_until,
            notes, updated_by_user_id, created_at, updated_at)
        VALUES (?,?,?,?,?,?,?,?,?)
        ON CONFLICT(company_id, unit_name) DO UPDATE SET
            phase = excluded.phase, opening_date = excluded.opening_date,
            goal_exempt_until = excluded.goal_exempt_until, notes = excluded.notes,
            updated_by_user_id = excluded.updated_by_user_id, updated_at = excluded.updated_at
        """,
        (company_id, unidade, fase, normalize_whitespace(payload.get("openingDate")) or None,
         normalize_whitespace(payload.get("goalExemptUntil")) or None,
         normalize_whitespace(payload.get("notes")), user["id"], now_iso(), now_iso()),
    )
    audit_log(conn, company_id, user["id"], "salvar", "unit_phases", unidade, {"fase": fase})
    conn.commit()
    invalidate_crm_cache(company_id)
    return {"saved": True}


def unit_is_in_deployment(conn: sqlite3.Connection, company_id: int, unit_name: str,
                          competence: str = "") -> bool:
    """Se a unidade está em implantação nesta competência.

    `goal_exempt_until` permite manter a isenção até uma competência específica
    mesmo depois de inaugurar — a loja abre, mas a meta só entra no ano seguinte.
    """
    fase = get_unit_phase(conn, company_id, unit_name)
    if fase["phase"] == "IMPLANTACAO":
        return True
    limite = fase.get("goalExemptUntil")
    return bool(limite and competence and competence <= limite)


def seller_unit_name(conn: sqlite3.Connection, company_id: int, seller_name: str,
                     competence: str = "") -> str:
    competence = competence or crm_latest_competence(conn, company_id) or date.today().strftime("%Y-%m")
    _, unidade = current_role_and_unit(conn, company_id, seller_name, competence)
    return normalize_unit(unidade)


# ── Metas de atividade ───────────────────────────────────────────────────────

def list_activity_goals(
    conn: sqlite3.Connection, company_id: int, competence: str, unit_name: str = ""
) -> list[dict[str, Any]]:
    sql = "SELECT * FROM activity_goals WHERE company_id = ? AND competence = ?"
    params: list[Any] = [company_id, competence]
    if unit_name:
        sql += " AND unit_name = ?"
        params.append(normalize_unit(unit_name))
    sql += " ORDER BY unit_name, seller_name, metric"
    return [
        {"id": int(r["id"]), "competence": r["competence"], "unitName": r["unit_name"],
         "sellerName": r["seller_name"] or "", "metric": r["metric"], "target": float(r["target"] or 0)}
        for r in conn.execute(sql, params).fetchall()
    ]


def save_activity_goal(
    conn: sqlite3.Connection, company_id: int, user: sqlite3.Row, payload: dict[str, Any]
) -> dict[str, Any]:
    if data_scope_for_user(conn, user) == "proprio":
        raise PermissionError("Apenas gestão define metas de atividade.")
    metric = normalize_upper(payload.get("metric"))
    if metric not in ACTIVITY_METRIC_IDS:
        raise ValueError("Indicador inválido.")
    competence = normalize_whitespace(payload.get("competence"))
    if not competence or len(competence) != 7:
        raise ValueError("Informe a competência no formato AAAA-MM.")
    unidade = normalize_unit(payload.get("unitName"))
    if not unidade:
        raise ValueError("Informe a unidade.")
    conn.execute(
        """
        INSERT INTO activity_goals (company_id, competence, unit_name, seller_name, metric, target, created_at, updated_at)
        VALUES (?,?,?,?,?,?,?,?)
        ON CONFLICT(company_id, competence, unit_name, seller_name, metric) DO UPDATE SET
            target = excluded.target, updated_at = excluded.updated_at
        """,
        (company_id, competence, unidade, normalize_whitespace(payload.get("sellerName")) or None,
         metric, float(payload.get("target") or 0), now_iso(), now_iso()),
    )
    conn.commit()
    return {"saved": True}


def activity_progress(
    conn: sqlite3.Connection, company_id: int, competence: str,
    unit_name: str = "", seller_name: str = "",
) -> list[dict[str, Any]]:
    """Realizado × meta de esforço, com o mesmo raciocínio de ritmo dos faróis."""
    inicio = first_day_of_competence(competence).isoformat()
    fim = last_day_of_competence(competence).isoformat()
    unidade = normalize_unit(unit_name)
    filtro_vendedor = normalize_upper(seller_name)

    def conta(sql: str, params: tuple) -> int:
        return int(conn.execute(sql, params).fetchone()["n"] or 0)

    cond_v = " AND UPPER(seller_name) = ?" if filtro_vendedor else ""
    p_v = (filtro_vendedor,) if filtro_vendedor else ()

    ligacoes = conta(
        f"""SELECT COUNT(*) n FROM crm_interactions
            WHERE company_id = ? AND contact_type_code = 'LIGACAO' AND initiative = 'ATIVO'
              AND date(substr(replace(occurred_at,'T',' '),1,10)) BETWEEN date(?) AND date(?){cond_v}""",
        (company_id, inicio, fim, *p_v),
    )
    cond_u = " AND unit_name = ?" if unidade else ""
    p_u = (unidade,) if unidade else ()
    novos = conta(
        f"""SELECT COUNT(*) n FROM prospects
            WHERE company_id = ? AND date(created_at) BETWEEN date(?) AND date(?){cond_u}{cond_v}""",
        (company_id, inicio, fim, *p_u, *p_v),
    )
    cadastrados = conta(
        f"""SELECT COUNT(*) n FROM prospects
            WHERE company_id = ? AND converted_at IS NOT NULL
              AND date(converted_at) BETWEEN date(?) AND date(?){cond_u}{cond_v}""",
        (company_id, inicio, fim, *p_u, *p_v),
    )
    primeiras = conta(
        f"""SELECT COUNT(*) n FROM prospects
            WHERE company_id = ? AND first_purchase_at IS NOT NULL
              AND date(first_purchase_at) BETWEEN date(?) AND date(?){cond_u}{cond_v}""",
        (company_id, inicio, fim, *p_u, *p_v),
    )
    realizado = {"CALLS": ligacoes, "PROSPECTS_NEW": novos,
                 "PROSPECTS_REGISTERED": cadastrados, "FIRST_PURCHASES": primeiras}

    metas = {
        g["metric"]: g["target"]
        for g in list_activity_goals(conn, company_id, competence, unidade)
        if not g["sellerName"] or normalize_upper(g["sellerName"]) == filtro_vendedor
    }

    # Ritmo do mês, para não cobrar 100% no dia 6 — mesma regra dos faróis.
    calendario = get_business_calendar(conn, company_id, competence)
    decorridos = int(calendario.get("elapsedWorkingDays") or 0)
    totais = int(calendario.get("totalWorkingDays") or 0)
    ritmo = safe_div(decorridos, totais) if totais else 1.0

    resultado = []
    for m in ACTIVITY_METRICS:
        alvo = float(metas.get(m["id"], 0))
        feito = realizado.get(m["id"], 0)
        esperado = round(alvo * ritmo) if alvo else 0
        pct_ritmo = safe_div(feito, esperado) * 100 if esperado else None
        resultado.append({
            **m,
            "target": alvo,
            "expectedToDate": esperado,
            "actual": feito,
            "pacePct": round(pct_ritmo, 1) if pct_ritmo is not None else None,
            "onTrack": None if pct_ritmo is None else pct_ritmo >= 90,
        })
    return resultado


# ── Prospects ────────────────────────────────────────────────────────────────

def prospect_row_to_dict(row: sqlite3.Row) -> dict[str, Any]:
    cfg = next((s for s in PROSPECT_STATUSES if s["id"] == row["status"]), None)
    return {
        "id": int(row["id"]),
        "clientKey": prospect_client_key(row["id"]),
        "unitName": row["unit_name"],
        "sellerName": row["seller_name"],
        "companyName": row["company_name"],
        "tradeName": row["trade_name"] or "",
        "documentNumber": row["document_number"] or "",
        "phone": row["phone"] or "",
        "contactName": row["contact_name"] or "",
        "email": row["email"] or "",
        "cityName": row["city_name"] or "",
        "neighborhood": row["neighborhood"] or "",
        "addressLine": row["address_line"] or "",
        "origin": row["origin"] or "",
        "status": row["status"],
        "statusLabel": cfg["label"] if cfg else row["status"],
        "statusIcon": cfg["icon"] if cfg else "•",
        "serviceType": row["q_service_type"] or "",
        "carsWeek": row["q_cars_week"],
        "mainLine": row["q_main_line"] or "",
        "payment": row["q_payment"] or "",
        "closingTrigger": row["closing_trigger"] or "",
        "notes": row["notes"] or "",
        "clientCode": row["client_code"] or "",
        "convertedAt": row["converted_at"] or "",
        "firstPurchaseAt": row["first_purchase_at"] or "",
        "lostReason": row["lost_reason"] or "",
        "createdAt": row["created_at"],
        # Qualificado = as 4 respostas + um gatilho, como manda o modelo Passini.
        "isQualified": bool(row["q_service_type"] and row["q_cars_week"]
                            and row["q_main_line"] and row["q_payment"] and row["closing_trigger"]),
    }


def list_prospects(
    conn: sqlite3.Connection, company_id: int, user: sqlite3.Row,
    status: str = "", search: str = "", seller: str = "", limit: int = 500,
) -> list[dict[str, Any]]:
    sql = "SELECT * FROM prospects WHERE company_id = ?"
    params: list[Any] = [company_id]

    if data_scope_for_user(conn, user) == "proprio":
        sql += " AND UPPER(seller_name) = ?"
        params.append(normalize_upper(seller_identity_for_user(user)))
    else:
        permitidas = crm_allowed_units_for_user(conn, user)
        if permitidas is not None:
            if permitidas:
                marcadores = ",".join("?" for _ in permitidas)
                sql += f" AND unit_name IN ({marcadores})"
                params.extend(permitidas)
            else:
                sql += " AND 1 = 0"
        if seller:
            sql += " AND UPPER(seller_name) = ?"
            params.append(normalize_upper(seller))

    if status in PROSPECT_STATUS_IDS:
        sql += " AND status = ?"
        params.append(status)
    termo = normalize_whitespace(search)
    if termo:
        alvo = f"%{termo.upper()}%"
        sql += (" AND (UPPER(company_name) LIKE ? OR UPPER(COALESCE(trade_name,'')) LIKE ?"
                " OR UPPER(COALESCE(contact_name,'')) LIKE ? OR COALESCE(document_digits,'') LIKE ?"
                " OR UPPER(COALESCE(city_name,'')) LIKE ?)")
        params.extend([alvo, alvo, alvo, f"%{only_digits(termo)}%" if only_digits(termo) else "%%", alvo])

    sql += """ ORDER BY CASE status WHEN 'QUALIFICADO' THEN 0 WHEN 'EM_CONTATO' THEN 1
               WHEN 'NOVO' THEN 2 WHEN 'CADASTRADO' THEN 3 ELSE 4 END, company_name LIMIT ?"""
    params.append(int(limit))

    linhas = [prospect_row_to_dict(r) for r in conn.execute(sql, params).fetchall()]

    # Último contato e retorno marcado vêm das mesmas tabelas usadas para cliente.
    for p in linhas:
        chave = p["clientKey"]
        ultimo = conn.execute(
            "SELECT MAX(occurred_at) u, COUNT(*) n FROM crm_interactions "
            "WHERE company_id = ? AND client_key = ?", (company_id, chave),
        ).fetchone()
        p["lastContactAt"] = ultimo["u"] or ""
        p["contactCount"] = int(ultimo["n"] or 0)
        tarefa = conn.execute(
            "SELECT MIN(due_at) d FROM crm_tasks WHERE company_id = ? AND client_key = ? "
            "AND status NOT IN ('CONCLUIDA','CANCELADA')", (company_id, chave),
        ).fetchone()
        p["nextTaskAt"] = tarefa["d"] or ""
        dias = None
        if p["lastContactAt"]:
            dt = parse_datetime_flexible(p["lastContactAt"][:10])
            if dt:
                dias = (today_in_brazil() - dt.date()).days
        p["daysSinceContact"] = dias
    return linhas


def save_prospect(
    conn: sqlite3.Connection, company_id: int, user: sqlite3.Row, payload: dict[str, Any]
) -> dict[str, Any]:
    nome = normalize_whitespace(payload.get("companyName"))
    if not nome:
        raise ValueError("Informe o nome da oficina.")

    escopo = data_scope_for_user(conn, user)
    if escopo == "proprio":
        vendedor = seller_identity_for_user(user)
    else:
        vendedor = normalize_whitespace(payload.get("sellerName")) or meeting_person_identity(user)

    # Quem é o dono do prospect: primeiro o que foi escolhido na tela, depois o
    # TERRITÓRIO (bairro/cidade) e só então a unidade do vendedor. O território
    # vem antes porque é ele que direciona a prospecção — um bairro da Zona
    # Norte prospectado por engano pela Zona Sul entra na unidade certa.
    # O território devolve None em cidade compartilhada (Canoas, Viamão) e aí
    # manda o vendedor, que é exatamente a regra combinada.
    unidade = (
        normalize_unit(payload.get("unitName"))
        or resolve_territory_unit(conn, company_id,
                                  payload.get("cityName"), payload.get("neighborhood"))
        or seller_unit_name(conn, company_id, vendedor)
    )
    permitidas = crm_allowed_units_for_user(conn, user)
    if permitidas is not None and permitidas and unidade not in permitidas:
        unidade = permitidas[0]

    documento = normalize_whitespace(payload.get("documentNumber"))
    digitos = only_digits(documento)

    prospect_id = payload.get("id")
    campos = (
        unidade, vendedor, nome,
        normalize_whitespace(payload.get("tradeName")),
        documento or None, digitos or None,
        normalize_whitespace(payload.get("phone")),
        normalize_whitespace(payload.get("contactName")),
        normalize_whitespace(payload.get("email")),
        normalize_upper(payload.get("cityName")),
        normalize_upper(payload.get("neighborhood")),
        normalize_whitespace(payload.get("addressLine")),
        normalize_whitespace(payload.get("origin")),
        normalize_whitespace(payload.get("serviceType")),
        int(payload["carsWeek"]) if str(payload.get("carsWeek") or "").strip().isdigit() else None,
        normalize_whitespace(payload.get("mainLine")),
        normalize_whitespace(payload.get("payment")),
        normalize_upper(payload.get("closingTrigger")) or None,
        normalize_whitespace(payload.get("notes")),
    )

    if prospect_id:
        conn.execute(
            """
            UPDATE prospects SET unit_name=?, seller_name=?, company_name=?, trade_name=?,
                   document_number=?, document_digits=?, phone=?, contact_name=?, email=?,
                   city_name=?, neighborhood=?, address_line=?, origin=?,
                   q_service_type=?, q_cars_week=?, q_main_line=?, q_payment=?,
                   closing_trigger=?, notes=?, updated_at=?
            WHERE company_id = ? AND id = ?
            """,
            (*campos, now_iso(), company_id, int(prospect_id)),
        )
        novo_id = int(prospect_id)
    else:
        duplicado = None
        if digitos:
            duplicado = conn.execute(
                "SELECT id, company_name, seller_name FROM prospects "
                "WHERE company_id = ? AND document_digits = ?", (company_id, digitos),
            ).fetchone()
        if duplicado:
            return {"prospectId": int(duplicado["id"]), "duplicated": True,
                    "message": f"Este CNPJ já está com {duplicado['seller_name']} "
                               f"como {duplicado['company_name']}."}
        cursor = conn.execute(
            """
            INSERT INTO prospects (unit_name, seller_name, company_name, trade_name,
                document_number, document_digits, phone, contact_name, email, city_name,
                neighborhood, address_line, origin, q_service_type, q_cars_week, q_main_line,
                q_payment, closing_trigger, notes, company_id, status, created_by_user_id, created_at)
            VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?, 'NOVO',?,?)
            """,
            (*campos, company_id, user["id"], now_iso()),
        )
        novo_id = int(cursor.lastrowid)

    _refresh_prospect_status(conn, company_id, novo_id)
    audit_log(conn, company_id, user["id"], "salvar", "prospects", str(novo_id), {"nome": nome})
    conn.commit()

    # Já testa se este CNPJ existe no cadastro — pode ser cliente antigo da casa.
    achado = try_convert_prospect(conn, company_id, novo_id)
    return {"prospectId": novo_id, "duplicated": False, "converted": bool(achado), **(achado or {})}


def _refresh_prospect_status(conn: sqlite3.Connection, company_id: int, prospect_id: int) -> None:
    """Status derivado do que já aconteceu, para ninguém precisar mantê-lo na mão."""
    row = conn.execute(
        "SELECT * FROM prospects WHERE company_id = ? AND id = ?", (company_id, prospect_id)
    ).fetchone()
    if not row or row["status"] in ("CADASTRADO", "PERDIDO"):
        return
    contatos = conn.execute(
        "SELECT COUNT(*) n FROM crm_interactions WHERE company_id = ? AND client_key = ?",
        (company_id, prospect_client_key(prospect_id)),
    ).fetchone()["n"]
    qualificado = all([row["q_service_type"], row["q_cars_week"], row["q_main_line"],
                       row["q_payment"], row["closing_trigger"]])
    novo = "QUALIFICADO" if qualificado else ("EM_CONTATO" if contatos else "NOVO")
    if novo != row["status"]:
        conn.execute("UPDATE prospects SET status = ?, updated_at = ? WHERE id = ?",
                     (novo, now_iso(), prospect_id))


def try_convert_prospect(
    conn: sqlite3.Connection, company_id: int, prospect_id: int
) -> dict[str, Any] | None:
    """Procura o CNPJ do prospect no cadastro de clientes e vincula.

    Achando, o histórico de contatos e as tarefas passam a apontar para o código
    real. É o que faz o esforço de prospecção aparecer na ficha do cliente em vez
    de desaparecer no dia em que ele é cadastrado.
    """
    row = conn.execute(
        "SELECT * FROM prospects WHERE company_id = ? AND id = ?", (company_id, prospect_id)
    ).fetchone()
    if not row or row["client_code"] or not row["document_digits"]:
        return None

    cliente = conn.execute(
        """
        SELECT client_code, client_name FROM crm_client_profiles
        WHERE company_id = ? AND REPLACE(REPLACE(REPLACE(COALESCE(document_number,''),'.',''),'/',''),'-','') = ?
        LIMIT 1
        """,
        (company_id, row["document_digits"]),
    ).fetchone()
    if not cliente:
        return None

    return link_prospect_to_client(conn, company_id, prospect_id, cliente["client_code"],
                                   automatic=True)


def link_prospect_to_client(
    conn: sqlite3.Connection, company_id: int, prospect_id: int, client_code: str,
    automatic: bool = False,
) -> dict[str, Any]:
    codigo = normalize_whitespace(client_code)
    if not codigo:
        raise ValueError("Informe o código do cliente.")
    chave_prospect = prospect_client_key(prospect_id)
    perfil = conn.execute(
        "SELECT client_name FROM crm_client_profiles WHERE company_id = ? AND client_code = ?",
        (company_id, codigo),
    ).fetchone()

    movidas = conn.execute(
        "UPDATE crm_interactions SET client_key = ?, client_name = COALESCE(?, client_name) "
        "WHERE company_id = ? AND client_key = ?",
        (codigo, perfil["client_name"] if perfil else None, company_id, chave_prospect),
    ).rowcount
    conn.execute(
        "UPDATE crm_tasks SET client_key = ?, client_name = COALESCE(?, client_name) "
        "WHERE company_id = ? AND client_key = ?",
        (codigo, perfil["client_name"] if perfil else None, company_id, chave_prospect),
    )
    conn.execute(
        "UPDATE prospects SET status = 'CADASTRADO', client_code = ?, converted_at = ?, updated_at = ? "
        "WHERE company_id = ? AND id = ?",
        (codigo, now_iso(), now_iso(), company_id, prospect_id),
    )
    conn.commit()
    invalidate_crm_cache(company_id)
    print(f"[prospeccao] prospect {prospect_id} vinculado ao cliente {codigo} "
          f"({'automático' if automatic else 'manual'}) · {movidas} contato(s) migrado(s)", flush=True)
    return {"clientCode": codigo, "clientName": perfil["client_name"] if perfil else codigo,
            "movedInteractions": int(movidas or 0), "automatic": automatic}


def reconcile_prospects(conn: sqlite3.Connection, company_id: int) -> int:
    """Passa por todos os prospects pendentes e tenta casar com o cadastro.

    Roda depois de cada importação de clientes: é o momento em que o cadastro do
    Alfa acabou de mudar e o vínculo passa a ser possível.
    """
    pendentes = conn.execute(
        "SELECT id FROM prospects WHERE company_id = ? AND client_code IS NULL "
        "AND document_digits IS NOT NULL AND document_digits <> '' AND status <> 'PERDIDO'",
        (company_id,),
    ).fetchall()
    vinculados = 0
    for r in pendentes:
        if try_convert_prospect(conn, company_id, int(r["id"])):
            vinculados += 1
    if vinculados:
        print(f"[prospeccao] {vinculados} prospect(s) viraram cliente nesta importação", flush=True)
    return vinculados


def refresh_prospect_first_purchases(conn: sqlite3.Connection, company_id: int) -> int:
    """Marca a primeira compra de quem já foi convertido.

    É o indicador que fecha o funil: prospectado → cadastrado → comprou.
    """
    convertidos = conn.execute(
        "SELECT id, client_code FROM prospects WHERE company_id = ? AND client_code IS NOT NULL "
        "AND first_purchase_at IS NULL", (company_id,),
    ).fetchall()
    marcados = 0
    for r in convertidos:
        nomes = client_sales_names(conn, company_id, r["client_code"])
        if not nomes:
            continue
        marcadores = ",".join("?" for _ in nomes)
        linha = conn.execute(
            f"""SELECT MIN(COALESCE(NULLIF(issue_date,''), competence || '-01')) d
                FROM fact_sales_detail
                WHERE company_id = ? AND net_value > 0 AND UPPER(client_name) IN ({marcadores})""",
            (company_id, *[normalize_upper(n) for n in nomes]),
        ).fetchone()
        if linha and linha["d"]:
            conn.execute("UPDATE prospects SET first_purchase_at = ? WHERE id = ?",
                         (linha["d"][:10], r["id"]))
            marcados += 1
    if marcados:
        conn.commit()
    return marcados


def mark_prospect_lost(
    conn: sqlite3.Connection, company_id: int, user: sqlite3.Row, prospect_id: int, reason: str
) -> None:
    conn.execute(
        "UPDATE prospects SET status='PERDIDO', lost_reason=?, updated_at=? WHERE company_id=? AND id=?",
        (normalize_whitespace(reason), now_iso(), company_id, prospect_id),
    )
    conn.commit()


def delete_prospect(conn: sqlite3.Connection, company_id: int, user: sqlite3.Row, prospect_id: int) -> None:
    if data_scope_for_user(conn, user) == "proprio":
        dono = conn.execute(
            "SELECT seller_name FROM prospects WHERE company_id = ? AND id = ?", (company_id, prospect_id)
        ).fetchone()
        if not dono or normalize_upper(dono["seller_name"]) != normalize_upper(seller_identity_for_user(user)):
            raise PermissionError("Você só pode excluir os seus prospects.")
    conn.execute("DELETE FROM prospects WHERE company_id = ? AND id = ?", (company_id, prospect_id))
    conn.commit()


def prospect_funnel(
    conn: sqlite3.Connection, company_id: int, user: sqlite3.Row
) -> dict[str, Any]:
    linhas = list_prospects(conn, company_id, user, limit=5000)
    por_status = {s["id"]: 0 for s in PROSPECT_STATUSES}
    for p in linhas:
        por_status[p["status"]] = por_status.get(p["status"], 0) + 1
    total = len(linhas)
    convertidos = por_status.get("CADASTRADO", 0)
    return {
        "total": total,
        "byStatus": por_status,
        "conversionPct": round(safe_div(convertidos, total) * 100, 1) if total else 0.0,
        "withoutContact": sum(1 for p in linhas if not p["contactCount"]),
        "stale": sum(1 for p in linhas
                     if p["status"] in ("NOVO", "EM_CONTATO", "QUALIFICADO")
                     and (p["daysSinceContact"] is None or p["daysSinceContact"] >= 7)),
    }


def data_scope_for_user(conn: sqlite3.Connection, user: sqlite3.Row) -> str:
    """Escopo de dados vindo do perfil, com fallback pelo nome do papel (legado)."""
    profile = get_access_profile_for_user(conn, user)
    if profile:
        return profile["dataScope"]
    if user["role"] == "Vendedor":
        return "proprio"
    if user["role"] in {"Gerente", "Analista"}:
        return "unidade"
    return "todos"


def scoped_filters_for_user(conn: sqlite3.Connection, company_id: int, user: sqlite3.Row, filters: dict[str, str | None]) -> dict[str, str | None]:
    scoped = dict(filters)
    data_scope = data_scope_for_user(conn, user)

    if data_scope == "todos":
        return scoped

    if data_scope == "proprio":
        seller_name = seller_identity_for_user(user)
        _, base_unit = current_role_and_unit(conn, company_id, seller_name, selected_primary_competence(scoped, query_competences(conn, company_id)) or date.today().strftime("%Y-%m"))
        scoped["seller_name"] = seller_name
        # Sempre sobrescreve a unidade pedida: o vendedor só enxerga a própria base.
        # Sem unidade cadastrada, zera o filtro (o recorte por vendedor já delimita).
        scoped["unit_name"] = normalize_unit(base_unit) if base_unit else None
        return scoped

    # unidade / unidade_consolidado
    linked_units = linked_units_for_user(user)
    if not linked_units:
        scoped["unit_name"] = "__NO_ACCESS__"
        scoped["city_name"] = None
        scoped["seller_name"] = None
        return scoped

    requested_unit = normalize_unit(scoped.get("unit_name"))
    requested_city = normalize_upper(scoped.get("city_name"))
    requested_seller = normalize_whitespace(scoped.get("seller_name"))
    competence = selected_primary_competence(scoped, query_competences(conn, company_id)) or date.today().strftime("%Y-%m")

    # "unidade + consolidado": pode escolher "Todas" para ver o total da empresa,
    # mas ao filtrar por uma unidade específica só valem as vinculadas a ele.
    if data_scope == "unidade_consolidado":
        scoped["allowed_units"] = []
        if requested_unit:
            scoped["unit_name"] = requested_unit if requested_unit in linked_units else linked_units[0]
    else:
        scoped["allowed_units"] = linked_units
        scoped["unit_name"] = (
            requested_unit if requested_unit in linked_units else (requested_unit and linked_units[0]) or None
        )

    # Cidade e vendedor pedidos precisam pertencer ao escopo permitido
    scope_units = [scoped["unit_name"]] if scoped.get("unit_name") else linked_units
    if requested_city:
        valid_cities = set(active_mapped_cities_for_units(conn, company_id, scope_units))
        if requested_city not in valid_cities:
            scoped["city_name"] = None
    if requested_seller:
        _, seller_base_unit = current_role_and_unit(conn, company_id, requested_seller, competence)
        if normalize_unit(seller_base_unit) not in linked_units:
            scoped["seller_name"] = None
    return scoped


def crm_allowed_units_for_user(conn: sqlite3.Connection, user: sqlite3.Row) -> list[str] | None:
    """Unidades que o usuário pode ver NO CRM.

    Diferente do dashboard de resultados, a carteira nunca é consolidada: mesmo o
    perfil "unidade + consolidado" enxerga só as unidades vinculadas. O consolidado
    serve para comparar números, não para acessar a carteira de outras equipes.
    Retorna None quando não há restrição.
    """
    scope = data_scope_for_user(conn, user)
    if scope in {"unidade", "unidade_consolidado"}:
        return linked_units_for_user(user)
    return None


def crm_unit_filter_for_user(
    conn: sqlite3.Connection, user: sqlite3.Row, requested_unit: str | None
) -> list[str] | None:
    """Resolve o filtro de unidade do CRM respeitando o vínculo do usuário."""
    allowed = crm_allowed_units_for_user(conn, user)
    requested = normalize_unit(requested_unit) if requested_unit else None
    if allowed is None:
        return [requested] if requested else None
    if not allowed:
        return ["__NO_ACCESS__"]
    if requested and requested in allowed:
        return [requested]
    return allowed


def crm_scoped_filters_for_user(
    conn: sqlite3.Connection, company_id: int, user: sqlite3.Row, filters: dict[str, str | None]
) -> dict[str, str | None]:
    """Filtros do CRM. Igual ao dashboard, exceto que NUNCA consolida:
    o perfil "unidade + consolidado" fica restrito às unidades vinculadas aqui."""
    scoped = scoped_filters_for_user(conn, company_id, user, filters)

    # Cobertura de carteira: o vendedor pode pedir a carteira de um colega
    # DESDE QUE o gerente tenha autorizado e a autorização esteja vigente.
    # Sem isso o escopo "proprio" travaria o nome dele e a cobertura seria
    # apenas decorativa.
    pedido = normalize_whitespace(filters.get("coverage_of"))
    if pedido and data_scope_for_user(conn, user) == "proprio":
        if seller_can_see_portfolio(conn, company_id, user, pedido):
            scoped["seller_name"] = pedido
            scoped["coverage_of"] = pedido
        else:
            scoped["seller_name"] = seller_identity_for_user(user)

    allowed = crm_allowed_units_for_user(conn, user)
    if allowed is None:
        return scoped
    if not allowed:
        scoped["unit_name"] = "__NO_ACCESS__"
        scoped["allowed_units"] = []
        return scoped
    scoped["allowed_units"] = allowed
    current_unit = normalize_unit(scoped.get("unit_name"))
    if not current_unit or current_unit not in allowed:
        # Sem unidade escolhida (ou fora do vínculo): usa a primeira unidade do usuário
        scoped["unit_name"] = allowed[0]
    return scoped


def crm_base_client_scope_query(
    conn: sqlite3.Connection, company_id: int, filters: dict[str, str | None]
) -> tuple[str | None, list[Any], str | None]:
    current_competence = crm_summary_latest_competence(conn, company_id)
    if not current_competence:
        return None, [], None
    seller_name = normalize_upper(filters.get("seller_name"))
    unit_name = normalize_unit(filters.get("unit_name"))
    city_name = normalize_upper(filters.get("city_name"))
    where_clauses = ["base.client_code IS NOT NULL", "base.client_code <> ''"]
    base_params: list[Any] = [company_id, current_competence, company_id, company_id, current_competence, company_id]
    filter_params: list[Any] = []
    if seller_name:
        where_clauses.append("UPPER(TRIM(COALESCE(p.internal_seller_name, ''))) = ?")
        filter_params.append(seller_name)
    if city_name:
        where_clauses.append("COALESCE(p.city_name, s.summary_city_name) = ?")
        filter_params.append(city_name)
    # O filtro de UNIDADE não entra mais aqui. Ele existia como um EXISTS sobre
    # city_mappings, ou seja, só conhecia a cidade — e desde que a unidade passou
    # a vir do vendedor (e do bairro, quando não há vendedor), esse filtro
    # devolvia zero para a Zona Norte: nenhuma CIDADE aponta para ela, o vínculo
    # é pelo vendedor e pelo bairro. Agora a filtragem acontece em Python, logo
    # depois de a unidade ser resolvida, usando exatamente a mesma regra que a
    # tela exibe. Ver unit_for_client_row().
    scope_query = f"""
        WITH current_summary AS (
            SELECT
                client_code,
                MAX(client_name) AS summary_client_name,
                MAX(city_name) AS summary_city_name,
                MAX(seller_name) AS summary_seller_name,
                ROUND(SUM(net_value), 2) AS current_revenue,
                MAX(last_purchase_at) AS last_purchase_at
            FROM crm_client_summary
            WHERE company_id = ? AND competence = ?
            GROUP BY client_code
        ),
        base_clients AS (
            SELECT client_code
            FROM crm_client_profiles
            WHERE company_id = ?
            UNION
            SELECT client_code
            FROM crm_client_summary
            WHERE company_id = ? AND competence = ?
        )
        SELECT
            base.client_code,
            s.summary_client_name,
            COALESCE(NULLIF(p.client_name, ''), NULLIF(s.summary_client_name, ''), base.client_code) AS client_name,
            p.trade_name,
            COALESCE(NULLIF(p.city_name, ''), NULLIF(s.summary_city_name, '')) AS city_name,
            p.phone,
            p.updated_phone,
            p.primary_contact_name,
            p.contact_notes,
            p.document_number,
            p.credit_limit,
            p.economic_group,
            p.neighborhood,
            -- CARTEIRA = vendedor INTERNO do cadastro. Só ele.
            --
            -- Aqui havia uma cascata interno → externo → quem faturou, e por
            -- isso a mesma vendedora aparecia com 177 clientes na lista e 132
            -- no painel. Vendedor externo e quem emitiu a nota não definem
            -- carteira: representante atende cliente de vários vendedores, e o
            -- faturamento de um mês não transfere a propriedade do cliente.
            NULLIF(TRIM(p.internal_seller_name), '') AS assigned_seller,
            COALESCE(s.current_revenue, 0) AS current_revenue,
            COALESCE(s.last_purchase_at, p.last_sale_at) AS last_purchase_at
        FROM base_clients base
        LEFT JOIN crm_client_profiles p
          ON p.company_id = ? AND p.client_code = base.client_code
        LEFT JOIN current_summary s
          ON s.client_code = base.client_code
        WHERE {" AND ".join(where_clauses)}
    """
    return scope_query, [*base_params, *filter_params], current_competence


def crm_base_client_rows(
    conn: sqlite3.Connection, company_id: int, filters: dict[str, str | None]
) -> list[dict[str, Any]]:
    scope_query, params, current_competence = crm_base_client_scope_query(conn, company_id, filters)
    if not scope_query or not current_competence:
        return []
    c0 = current_competence
    c1 = shift_competence(c0, -1)
    c2 = shift_competence(c0, -2)
    c3 = shift_competence(c0, -3)
    previous_competences = [c1, c2, c3]

    aggregate_rows = conn.execute(scope_query, params).fetchall()

    # Precarrega os três mapas de unidade em uma query cada (evita N+1)
    city_unit_map = build_city_unit_map(conn, company_id, c0)
    seller_unit_map = build_seller_unit_map(conn, company_id, c0)
    territory_map = build_territory_map(conn, company_id, c0)
    unit_filter = normalize_unit(filters.get("unit_name"))

    # Agrega fact_sales_detail por cliente no SQL — evita iterar milhares de linhas no Python
    detail_rows = conn.execute(
        """
        SELECT
            client_name,
            SUM(CASE WHEN competence = ? THEN net_value ELSE 0 END) AS revenue_c1,
            SUM(CASE WHEN competence = ? THEN net_value ELSE 0 END) AS revenue_c2,
            SUM(CASE WHEN competence = ? THEN net_value ELSE 0 END) AS revenue_c3,
            MAX(issue_date) AS last_purchase_at,
            COUNT(DISTINCT CASE WHEN competence = ? AND net_value > 0 THEN
                COALESCE(NULLIF(manufacturer_sku, ''), NULLIF(sku_key, ''), NULLIF(gtin_value, ''), 'ITEM')
            END) AS current_sku_count
        FROM fact_sales_detail
        WHERE company_id = ? AND competence IN (?, ?, ?, ?)
        GROUP BY client_name
        """,
        [c1, c2, c3, c0, company_id, c0, c1, c2, c3],
    ).fetchall()

    # Receita c0 scoped por vendedor — query separada para não contaminar c1/c2/c3 (sem filtro de vendedor)
    seller_name_for_c0 = normalize_upper(filters.get("seller_name"))
    c0_seller_revenue: dict[str, float] = {}
    if seller_name_for_c0:
        c0_rows = conn.execute(
            """
            SELECT client_name, SUM(net_value) AS revenue_c0
            FROM fact_sales_detail
            WHERE company_id = ? AND competence = ? AND UPPER(seller_name) = ?
            GROUP BY client_name
            """,
            [company_id, c0, seller_name_for_c0],
        ).fetchall()
        for c0row in c0_rows:
            key = normalize_client_key(c0row["client_name"])
            if key:
                c0_seller_revenue[key] = c0_seller_revenue.get(key, 0.0) + float(c0row["revenue_c0"] or 0.0)

    interaction_rows = conn.execute(
        """
        SELECT client_key, MAX(occurred_at) AS last_interaction_at
        FROM crm_interactions
        WHERE company_id = ?
        GROUP BY client_key
        """,
        (company_id,),
    ).fetchall()
    interaction_map = {row["client_key"]: row["last_interaction_at"] for row in interaction_rows}

    # Constrói mapa de métricas por cliente já agregado (uma entrada por nome único)
    detail_metrics: dict[str, dict[str, Any]] = {}
    for row in detail_rows:
        client_name_value = normalize_whitespace(row["client_name"])
        if not client_name_value:
            continue
        name_key = normalize_client_key(client_name_value)
        if not name_key:
            continue
        detail_metrics[name_key] = {
            "revenues_c123": [
                float(row["revenue_c1"] or 0.0),
                float(row["revenue_c2"] or 0.0),
                float(row["revenue_c3"] or 0.0),
            ],
            "currentSkuCount": int(row["current_sku_count"] or 0),
            "lastPurchaseAt": parse_datetime_flexible(row["last_purchase_at"]),
        }

    # ── Cadastros duplicados: mesmo cliente com mais de um código ──────────────
    # O faturamento detalhado é indexado por NOME; a carteira, por CÓDIGO. Quando a
    # mesma empresa tem código antigo e novo (recadastro), os dois códigos casariam
    # com o mesmo faturamento e o valor apareceria em dobro nos totais.
    # Regra: o faturamento fica com um único código — o que tem movimento no mês
    # corrente; sem movimento em nenhum, fica com o de compra mais recente.
    revenue_owner_by_name: dict[str, str] = {}
    _candidates_by_name: dict[str, list[tuple[float, str, str, str]]] = defaultdict(list)
    for row in aggregate_rows:
        code = normalize_whitespace(row["client_code"])
        if not code:
            continue
        # NOME FANTASIA NÃO ENTRA AQUI.
        #
        # Ele entrava, e criava duplicata falsa: fantasia costuma ser curta e
        # repetida ("RAFAEL", "AUTO CENTER", "OFICINA"), então dois clientes sem
        # nenhuma relação caíam na mesma chave e um era marcado como cópia do
        # outro. Caso real: RAFAEL MOSCARELLI (Pelotas) apontando para RAFAEL
        # SPANIOL E CIA (Matriz). O faturamento vem pela RAZÃO SOCIAL, então a
        # fantasia não ajudaria nem se fosse confiável.
        for key in {
            normalize_client_key(row["client_name"]),
            normalize_client_key(row["summary_client_name"]),
        }:
            if key:
                _candidates_by_name[key].append((
                    float(row["current_revenue"] or 0.0),
                    normalize_whitespace(row["last_purchase_at"]) or "",
                    code,
                    only_digits(row["document_number"]),
                ))
    for name_key, candidates in _candidates_by_name.items():
        if len(candidates) < 2:
            revenue_owner_by_name[name_key] = candidates[0][2]
            continue
        # Segunda trava: documentos diferentes são clientes diferentes, por mais
        # parecido que o nome seja. Homônimo existe — duas empresas do mesmo dono,
        # matriz e filial, pessoa física e jurídica com o mesmo nome. Sem isso, o
        # faturamento de uma seria creditado à outra.
        documentos = {c[3] for c in candidates if c[3]}
        if len(documentos) > 1:
            continue
        # Maior receita no mês corrente; empate resolve pela compra mais recente
        candidates.sort(key=lambda c: (c[0], c[1]), reverse=True)
        revenue_owner_by_name[name_key] = candidates[0][2]

    client_rows: list[dict[str, Any]] = []
    for row in aggregate_rows:
        current_revenue = float(row["current_revenue"] or 0.0)
        client_key = normalize_whitespace(row["client_code"])
        # O cliente pode ser encontrado por razão social, nome do resumo ou nome
        # fantasia. DEDUPLICAR é essencial: quando dois desses campos normalizam
        # para a mesma chave (o caso comum — razão social igual em ambos), o loop
        # abaixo somava o MESMO faturamento duas ou três vezes, dobrando a média.
        # Também sem fantasia, pela mesma razão: o faturamento é indexado pela
        # razão social, e casar por fantasia trazia receita de outro cliente.
        candidate_keys = list(dict.fromkeys(
            key for key in (
                normalize_client_key(row["client_name"]),
                normalize_client_key(row["summary_client_name"]),
            ) if key
        ))
        merged_revenues = [0.0, 0.0, 0.0]
        merged_current_sku_count = 0
        merged_last_purchase_at: datetime | None = None
        merged_detail_c0_revenue = 0.0
        duplicated_codes: list[str] = []
        for name_key in candidate_keys:
            # Só o código eleito dono recebe o faturamento daquele nome
            owner = revenue_owner_by_name.get(name_key)
            if owner and owner != client_key:
                duplicated_codes.append(owner)
                continue
            if name_key in detail_metrics:
                metrics = detail_metrics[name_key]
                for i, v in enumerate(metrics["revenues_c123"]):
                    merged_revenues[i] += v
                merged_current_sku_count = max(merged_current_sku_count, metrics["currentSkuCount"])
                m_last = metrics["lastPurchaseAt"]
                if m_last and (merged_last_purchase_at is None or m_last > merged_last_purchase_at):
                    merged_last_purchase_at = m_last
            if name_key in c0_seller_revenue:
                merged_detail_c0_revenue += c0_seller_revenue[name_key]

        # Usa o maior entre CRM summary e faturamento detalhado scoped por vendedor
        current_revenue = max(current_revenue, merged_detail_c0_revenue)
        previous_revenues = merged_revenues
        # MÉDIA = soma dos 3 meses ANTERIORES a c0, dividida por 3 (sempre por 3,
        # inclusive meses sem compra — mede volume médio mensal, não ticket médio).
        # Base: fact_sales_detail, faturamento total do cliente (todos os vendedores).
        average_revenue = sum(previous_revenues) / 3
        months_with_purchase = sum(1 for v in previous_revenues if v > 0)
        summary_last_purchase = parse_datetime_flexible(row["last_purchase_at"])
        last_purchase_at = summary_last_purchase
        if merged_last_purchase_at and (last_purchase_at is None or merged_last_purchase_at > last_purchase_at):
            last_purchase_at = merged_last_purchase_at
        days_without_purchase = (date.today() - last_purchase_at.date()).days if last_purchase_at else None
        status_code = crm_status_from_days(days_without_purchase)
        class_code = crm_class_from_average(average_revenue)
        drop_pct = safe_div(current_revenue - average_revenue, average_revenue) if average_revenue else 0.0
        has_mix_opportunity = current_revenue > 0 and merged_current_sku_count <= 2
        resolved_unit_name = unit_for_client_row(
            row["assigned_seller"], row["city_name"], row["neighborhood"],
            seller_unit_map, territory_map, city_unit_map)
        if unit_filter and normalize_unit(resolved_unit_name) != unit_filter:
            continue
        priorities: list[str] = []
        if status_code == "INATIVO":
            priorities.append("REATIVACAO_INATIVO")
        if status_code == "PRE_INATIVO":
            priorities.append("PRE_INATIVO")
        if current_revenue <= 0:
            priorities.append("SEM_COMPRA_MES")
        if average_revenue > 0 and drop_pct <= -0.1:
            priorities.append("QUEDA_FATURAMENTO")
        if class_code in {"DIAMANTE", "OURO"}:
            priorities.append("CLIENTE_CLASSE_ALTA")
        if has_mix_opportunity:
            priorities.append("OPORTUNIDADE_MIX")
        primary_reason_code = priorities[0] if priorities else "PROSPECCAO_NOVA"
        secondary_reasons = priorities[1:]
        # Determina tipo de pessoa pelo documento; fallback por heurística de nome
        doc_person_type, _ = person_type_from_document(row["document_number"])
        if doc_person_type:
            person_type = doc_person_type
        else:
            person_type, _, _ = infer_person_type_from_name(row["client_name"])
        row_summary = {"daysWithoutPurchase": days_without_purchase, "dropPct": drop_pct, "classCode": class_code}
        client_rows.append(
            {
                "clientKey": client_key,
                "clientCode": client_key,
                "clientName": row["client_name"],
                "summaryClientName": row["summary_client_name"],
                "tradeName": row["trade_name"],
                "cityName": row["city_name"],
                "phone": row["updated_phone"] or row["phone"],
                "phoneRaw": row["phone"],
                "updatedPhone": row["updated_phone"],
                "primaryContactName": row["primary_contact_name"],
                "contactNotes": row["contact_notes"],
                "documentNumber": row["document_number"],
                "personType": person_type,
                "creditLimit": float(row["credit_limit"] or 0.0),
                "economicGroup": row["economic_group"],
                "assignedSeller": row["assigned_seller"],
                "unitName": resolved_unit_name,
                "currentRevenue": round(current_revenue, 2),
                "averageRevenue": round(average_revenue, 2),
                "growthPct": round(drop_pct, 4),
                "trimesterRevenue1": round(previous_revenues[0], 2),
                "trimesterRevenue2": round(previous_revenues[1], 2),
                "trimesterRevenue3": round(previous_revenues[2], 2),
                # Sinaliza cadastro duplicado: o faturamento deste nome está em outro código
                "duplicateOfCode": duplicated_codes[0] if duplicated_codes else None,
                # Memória de cálculo da média — permite auditar o número na tela
                "averageBasis": {
                    "currentCompetence": c0,
                    "months": [
                        {"competence": c1, "revenue": round(previous_revenues[0], 2)},
                        {"competence": c2, "revenue": round(previous_revenues[1], 2)},
                        {"competence": c3, "revenue": round(previous_revenues[2], 2)},
                    ],
                    "total": round(sum(previous_revenues), 2),
                    "divisor": 3,
                    "monthsWithPurchase": months_with_purchase,
                    "formula": "soma dos 3 meses anteriores ÷ 3",
                },
                "neighborhood": normalize_upper(row["neighborhood"]),
                "dropPct": round(drop_pct, 4),
                "classCode": class_code,
                "statusCode": status_code,
                "daysWithoutPurchase": days_without_purchase,
                "lastPurchaseAt": last_purchase_at.isoformat(timespec="seconds") if last_purchase_at else None,
                "lastInteractionAt": interaction_map.get(client_key),
                "currentSkuCount": merged_current_sku_count,
                "primaryReasonCode": primary_reason_code,
                "primaryReason": crm_reason_message(row_summary, primary_reason_code),
                "secondaryReasonCodes": secondary_reasons,
            }
        )
    return client_rows


# --- Cache TTL para crm_base_client_rows (90s) ---
_crm_base_cache: dict[tuple, tuple[float, list]] = {}
_crm_base_cache_lock = threading.Lock()
_CRM_CACHE_TTL = 90  # segundos


def _crm_cache_key(company_id: int, filters: dict) -> tuple:
    return (
        company_id,
        normalize_whitespace(filters.get("seller_name")) or "",
        normalize_whitespace(filters.get("unit_name")) or "",
        normalize_whitespace(filters.get("competenceEnd") or filters.get("competence")) or "",
    )


def crm_base_client_rows_cached(conn: sqlite3.Connection, company_id: int, filters: dict) -> list:
    key = _crm_cache_key(company_id, filters)
    now = time.monotonic()
    with _crm_base_cache_lock:
        entry = _crm_base_cache.get(key)
    if entry and now - entry[0] < _CRM_CACHE_TTL:
        return entry[1]
    result = crm_base_client_rows(conn, company_id, filters)
    with _crm_base_cache_lock:
        _crm_base_cache[key] = (now, result)
    return result


def invalidate_crm_cache(company_id: int | None = None) -> None:
    with _crm_base_cache_lock:
        if company_id is None:
            _crm_base_cache.clear()
        else:
            for k in list(_crm_base_cache.keys()):
                if k[0] == company_id:
                    del _crm_base_cache[k]
    invalidate_dashboard_cache(company_id)
    invalidate_calendar_cache(company_id)


# ─────────────────────────────────────────────────────────────────────────────
# Cache do dashboard
#
# Entre importações os dados são estáticos, então o resultado de get_dashboard_data
# pode ser reaproveitado integralmente. Sem isso, cada troca de competência
# reprocessava dezenas de milhares de linhas do faturamento detalhado em Python.
# O cache é invalidado em toda importação e pré-aquecido em segundo plano.
# ─────────────────────────────────────────────────────────────────────────────
_dashboard_cache: dict[tuple, dict[str, Any]] = {}
_dashboard_cache_lock = threading.Lock()
_DASHBOARD_CACHE_MAX = 400


def _dashboard_cache_key(company_id: int, filters: dict[str, str | None]) -> tuple:
    # A competência efetiva vem de competence_end/competence_start (ver
    # selected_primary_competence); competence_start vazio significa "mais recente".
    return (
        company_id,
        normalize_whitespace(filters.get("competence_start")),
        normalize_whitespace(filters.get("competence_end")),
        normalize_whitespace(filters.get("unit_name")),
        normalize_whitespace(filters.get("seller_name")),
        normalize_whitespace(filters.get("city_name")),
        tuple(sorted(normalize_unit_list(filters.get("allowed_units")))),
    )


def invalidate_dashboard_cache(company_id: int | None = None) -> None:
    with _dashboard_cache_lock:
        if company_id is None:
            _dashboard_cache.clear()
        else:
            for k in list(_dashboard_cache.keys()):
                if k[0] == company_id:
                    del _dashboard_cache[k]


def get_dashboard_data_cached(
    conn: sqlite3.Connection, company_id: int, filters: dict[str, str | None]
) -> dict[str, Any]:
    key = _dashboard_cache_key(company_id, filters)
    with _dashboard_cache_lock:
        hit = _dashboard_cache.get(key)
    if hit is not None:
        print(f"[dashboard] CACHE HIT {key[1:5]}", flush=True)
        return hit
    started = time.time()
    data = get_dashboard_data(conn, company_id, filters)
    elapsed = time.time() - started
    with _dashboard_cache_lock:
        if len(_dashboard_cache) >= _DASHBOARD_CACHE_MAX:
            _dashboard_cache.clear()
        _dashboard_cache[key] = data
    print(f"[dashboard] CALCULADO {key[1:5]} em {elapsed:.1f}s", flush=True)
    return data


def warm_dashboard_cache(company_id: int | None = None) -> None:
    """Pré-calcula o dashboard de cada competência para que a troca seja instantânea."""
    try:
        with closing(get_connection()) as conn:
            if company_id is None:
                company_ids = [r["id"] for r in conn.execute("SELECT id FROM companies").fetchall()]
            else:
                company_ids = [company_id]
            started = time.time()
            for cid in company_ids:
                competences = query_competences(conn, cid)
                units = [
                    normalize_unit(r["unit_name"])
                    for r in conn.execute(
                        "SELECT DISTINCT unit_name FROM fact_unit_summary WHERE company_id = ?", (cid,)
                    ).fetchall()
                    if r["unit_name"]
                ]
                # Visão padrão (sem competência escolhida = mais recente)
                try:
                    get_dashboard_data_cached(conn, cid, build_filters_from_query({}))
                except Exception:
                    traceback.print_exc()

                # A tela envia competenceStart e competenceEnd com o mesmo valor.
                # Competências recentes sem filtro de unidade.
                for competence in competences[:8]:
                    filters = build_filters_from_query({})
                    filters["competence_start"] = competence
                    filters["competence_end"] = competence
                    try:
                        get_dashboard_data_cached(conn, cid, filters)
                    except Exception:
                        traceback.print_exc()

                # Combinação competência × unidade para as 3 competências mais recentes,
                # que é o recorte usado no dia a dia.
                for competence in competences[:3]:
                    for unit in units:
                        filters = build_filters_from_query({})
                        filters["competence_start"] = competence
                        filters["competence_end"] = competence
                        filters["unit_name"] = unit
                        try:
                            get_dashboard_data_cached(conn, cid, filters)
                        except Exception:
                            traceback.print_exc()
            print(f"[dashboard] cache pré-aquecido: {len(_dashboard_cache)} combinação(ões) "
                  f"em {time.time() - started:.0f}s")
    except Exception:
        traceback.print_exc()


def crm_base_client_count(conn: sqlite3.Connection, company_id: int, filters: dict[str, str | None]) -> int:
    scope_query, params, _ = crm_base_client_scope_query(conn, company_id, filters)
    if not scope_query:
        return 0
    row = conn.execute(
        f"""
        SELECT COUNT(*) AS total
        FROM (
            {scope_query}
        ) scoped_clients
        """,
        params,
    ).fetchone()
    return int(row["total"] or 0)


def crm_reason_message(summary: dict[str, Any], reason_code: str) -> str:
    if reason_code == "REATIVACAO_INATIVO":
        days = summary.get("daysWithoutPurchase")
        return f"Inativo ha {days} dias" if days is not None else "Inativo sem data recente de compra"
    if reason_code == "PRE_INATIVO":
        days = summary.get("daysWithoutPurchase")
        return f"Pre-inativo ha {days} dias sem compra" if days is not None else "Pre-inativo sem data recente de compra"
    if reason_code == "SEM_COMPRA_MES":
        return "Sem compra no mes atual"
    if reason_code == "QUEDA_FATURAMENTO":
        return f"Queda de {abs(summary['dropPct']) * 100:.0f}% vs media do trimestre"
    if reason_code == "CLIENTE_CLASSE_ALTA":
        return f"Cliente {summary['classCode']}"
    if reason_code == "OPORTUNIDADE_MIX":
        return "Baixa variedade recente de itens para o perfil"
    return "Prospeccao nova"


def crm_generate_questions(summary: dict[str, Any]) -> dict[str, Any]:
    if summary["statusCode"] == "INATIVO":
        return {
            "primary": "O que mudou na sua rotina de compra desde o ultimo pedido?",
            "secondary": [
                "Qual linha esta com mais giro hoje na oficina/loja?",
                "Tem alguma ruptura ou fornecedor que deixou de atender bem?",
                "Qual dia da semana faz mais sentido eu te acompanhar?",
            ],
        }
    if summary["statusCode"] == "PRE_INATIVO":
        return {
            "primary": "Qual item voce esta precisando girar nesta semana para eu te ajudar antes da proxima compra?",
            "secondary": [
                "O mix atual esta cobrindo bem freio, suspensao e direcao?",
                "Tem alguma marca ou linha com falta recente?",
                "Qual proximo pedido voce imagina fazer?",
            ],
        }
    if summary["primaryReasonCode"] == "QUEDA_FATURAMENTO":
        return {
            "primary": "O volume deste mes caiu frente ao seu ritmo normal; onde voce sentiu mais essa queda?",
            "secondary": [
                "Qual linha perdeu mais giro recentemente?",
                "Teve mudanca de demanda, estoque ou fornecedor?",
                "Que oferta faria sentido para retomar compra agora?",
            ],
        }
    return {
        "primary": "Qual linha esta girando mais no momento para eu te apoiar com oferta certa?",
        "secondary": [
            "Tem algum item que voce costumava comprar e precisa voltar a girar?",
            "Existe alguma necessidade de reposicao imediata?",
            "Qual dia voce costuma comprar para eu te acompanhar melhor?",
        ],
    }


# ── Parâmetros das sugestões de oferta ──────────────────────────────────────
# Ficam juntos e nomeados de propósito: são regras comerciais que mudam com o
# tempo e precisam ser ajustáveis sem caçar número solto no meio da query.

# Janela em que uma parada de compra ainda é "recompra". Item que o cliente
# largou há mais tempo que isso vira histórico morto e não gera gancho.
OFFER_REPURCHASE_WINDOW_MONTHS = 12
# Janela usada para medir o giro do item na praça (competência atual + 3 anteriores).
OFFER_PEER_WINDOW_MONTHS = 4
# Mínimo de oficinas distintas comprando o item na cidade para virar oportunidade.
OFFER_PEER_MIN_CLIENTS = 3
# Quantas sugestões de cada tipo devolver.
OFFER_MAX_SUGGESTIONS = 5
# Teto de itens no histórico exibido. NÃO usar esse recorte para decidir o que o
# cliente já comprou — foi exatamente esse erro que fez um item comprado ser
# anunciado como "nunca pediu".
OFFER_HISTORY_DISPLAY_LIMIT = 25


def crm_get_offer_suggestions(
    conn: sqlite3.Connection, company_id: int, client_name: str, city_name: str | None = None
) -> dict[str, Any]:
    """Sugestões de oferta com contexto para o vendedor conseguir abordar.

    Dois tipos, sempre rotulados:
      - RECOMPRA: item que o cliente comprava e parou. Traz meses de compra,
        valor médio e há quanto tempo não pede — dá o gancho da conversa.
      - OPORTUNIDADE: item que oficinas da mesma praça compram bastante e este
        cliente nunca comprou. Abre mix novo.
    """
    latest = crm_latest_competence(conn, company_id)
    if not latest:
        return {"primary": None, "secondary": [], "repurchase": [], "opportunity": []}

    ITEM_EXPR = "COALESCE(NULLIF(manufacturer_sku, ''), NULLIF(sku_key, ''), NULLIF(gtin_value, ''), 'ITEM')"

    def item_label(codigo: str | None, gtin: str | None) -> str:
        """Código do fabricante com o código interno entre parênteses.

        O vendedor procura a peça pelo código interno no sistema, mas conversa
        com o mecânico usando a referência do fabricante — os dois precisam
        aparecer. O interno vem da coluna sem cabeçalho do relatório (coluna E).
        """
        cod = normalize_whitespace(codigo)
        g = normalize_whitespace(gtin)
        if g and g != cod:
            return f"{cod} ({g})"
        return cod or g or "ITEM"

    # Histórico do cliente por item dentro da janela de recompra (fora do mês corrente).
    # Este recorte serve para MONTAR A LISTA DE RECOMPRA — nunca para decidir se o
    # cliente já comprou um item.
    inicio_recompra = shift_competence(latest, -(OFFER_REPURCHASE_WINDOW_MONTHS - 1))
    history = conn.execute(
        f"""
        SELECT {ITEM_EXPR} AS item_code,
               MAX(NULLIF(gtin_value, '')) AS gtin,
               COUNT(DISTINCT competence) AS meses,
               SUM(net_value) AS total,
               SUM(quantity) AS qtd,
               MAX(competence) AS ultima_competencia
        FROM fact_sales_detail
        WHERE company_id = ? AND client_name = ? AND competence <> ?
          AND competence >= ? AND net_value > 0
        GROUP BY item_code
        ORDER BY meses DESC, total DESC
        LIMIT {OFFER_HISTORY_DISPLAY_LIMIT}
        """,
        (company_id, client_name, latest, inicio_recompra),
    ).fetchall()

    current_items = {
        r["item_code"]
        for r in conn.execute(
            f"SELECT DISTINCT {ITEM_EXPR} AS item_code FROM fact_sales_detail "
            "WHERE company_id = ? AND client_name = ? AND competence = ?",
            (company_id, client_name, latest),
        ).fetchall()
    }

    # Tudo que o cliente já comprou, sem janela e SEM LIMITE. É esta lista — e não
    # o histórico exibido — que impede anunciar como novidade um item que ele já
    # levou. O bug anterior: o histórico vinha com LIMIT 25 ordenado por valor, e
    # um item barato comprado uma vez ficava de fora, virando "nunca pediu".
    ever_bought_rows = conn.execute(
        f"""
        SELECT DISTINCT {ITEM_EXPR} AS item_code
        FROM fact_sales_detail
        WHERE company_id = ? AND client_name = ? AND net_value > 0
        """,
        (company_id, client_name),
    ).fetchall()
    all_client_items = {r["item_code"] for r in ever_bought_rows} | current_items
    client_has_history = bool(all_client_items)

    def months_since(competence: str | None) -> int | None:
        if not competence or len(competence) < 7:
            return None
        try:
            y1, m1 = int(latest[:4]), int(latest[5:7])
            y2, m2 = int(competence[:4]), int(competence[5:7])
            return (y1 - y2) * 12 + (m1 - m2)
        except ValueError:
            return None

    repurchase: list[dict[str, Any]] = []
    for r in history:
        if r["item_code"] in current_items:
            continue
        meses = int(r["meses"] or 0)
        gap = months_since(r["ultima_competencia"])
        media = float(r["total"] or 0) / max(meses, 1)
        if meses >= 2:
            freq = f"comprava em {meses} meses"
        else:
            freq = "comprou uma vez"
        gap_txt = "sem pedido no mês atual" if not gap else f"sem pedido há {gap} {'mês' if gap == 1 else 'meses'}"
        repurchase.append({
            "itemCode": r["item_code"],
            "gtin": normalize_whitespace(r["gtin"]),
            "type": "RECOMPRA",
            "typeLabel": "Recompra",
            "title": item_label(r["item_code"], r["gtin"]),
            "reason": f"{freq} · {gap_txt} · média {brl(media)}/mês",
            "months": meses,
            "avgValue": round(media, 2),
            "monthsSincePurchase": gap,
        })

    # Oportunidade: itens girando na praça que este cliente nunca comprou
    opportunity: list[dict[str, Any]] = []
    if city_name:
        peer_rows = conn.execute(
            f"""
            SELECT {ITEM_EXPR} AS item_code,
                   MAX(NULLIF(gtin_value, '')) AS gtin,
                   COUNT(DISTINCT client_name) AS clientes,
                   SUM(net_value) AS total
            FROM fact_sales_detail
            WHERE company_id = ? AND city_name = ? AND client_name <> ? AND net_value > 0
              AND competence >= ?
            GROUP BY item_code
            HAVING clientes >= ?
            ORDER BY clientes DESC, total DESC
            LIMIT 40
            """,
            (company_id, normalize_upper(city_name), client_name,
             shift_competence(latest, -(OFFER_PEER_WINDOW_MONTHS - 1)), OFFER_PEER_MIN_CLIENTS),
        ).fetchall()
        for r in peer_rows:
            if r["item_code"] in all_client_items:
                continue
            # Só afirma "nunca pediu" quando existe histórico de compra do cliente
            # para comparar. Sem histórico, a frase seria um chute com cara de dado.
            complemento = ("ele nunca pediu" if client_has_history
                           else "não há registro de compra dele")
            opportunity.append({
                "itemCode": r["item_code"],
                "gtin": normalize_whitespace(r["gtin"]),
                "type": "OPORTUNIDADE",
                "typeLabel": "Oportunidade",
                "title": item_label(r["item_code"], r["gtin"]),
                "reason": f"{int(r['clientes'])} oficinas da região compram e {complemento}",
                "peerClients": int(r["clientes"]),
            })
            if len(opportunity) >= OFFER_MAX_SUGGESTIONS:
                break

    # Prioriza recompra: gancho mais forte e conversão mais provável
    primary = repurchase[0] if repurchase else (opportunity[0] if opportunity else None)
    secondary = ([o for o in repurchase[1:3]] + [o for o in opportunity[:2]])[:4]
    return {
        "primary": primary,
        "secondary": secondary,
        "repurchase": repurchase[:OFFER_MAX_SUGGESTIONS],
        "opportunity": opportunity[:OFFER_MAX_SUGGESTIONS],
    }


def crm_situation_for_client(summary: dict[str, Any]) -> str:
    """Traduz o estado do cliente na situação usada pela biblioteca de conteúdo."""
    status = summary.get("statusCode")
    if status == "INATIVO":
        return "INATIVO"
    if status == "PRE_INATIVO":
        return "PRE_INATIVO"
    if float(summary.get("currentRevenue") or 0) <= 0:
        return "SEM_COMPRA_MES"
    if float(summary.get("dropPct") or 0) <= -0.1:
        return "QUEDA"
    if int(summary.get("currentSkuCount") or 0) <= 2:
        return "MIX"
    return "GERAL"


def crm_next_action(summary: dict[str, Any]) -> str:
    """Próxima ação concreta, com número e prazo — não frase genérica."""
    status = summary.get("statusCode")
    dias = summary.get("daysWithoutPurchase")
    offer = (summary.get("offerPrimary") or {}).get("title")
    media = float(summary.get("averageRevenue") or 0)

    if status == "INATIVO":
        base = f"Ligar hoje e perguntar por que parou de comprar há {dias} dias" if dias else "Ligar hoje e entender o motivo da parada"
        if offer:
            base += f". Oferecer {offer} como gancho"
        if media > 0:
            base += f". Cliente valia {brl(media)}/mês"
        return base + ". Agendar retorno na própria ligação."
    if status == "PRE_INATIVO":
        base = "Contato preventivo ainda esta semana"
        if offer:
            base += f", puxando {offer} que ele deixou de pedir"
        return base + ". Fechar pedido pequeno para não perder a frequência."
    if float(summary.get("currentRevenue") or 0) <= 0:
        base = "Ligar para reposição do mês"
        if offer:
            base += f" — confirmar se {offer} ainda tem em estoque"
        return base + ". Fechar antes do corte de entrega."
    if float(summary.get("dropPct") or 0) <= -0.1:
        queda = abs(float(summary.get("dropPct") or 0)) * 100
        return (f"Investigar queda de {queda:.0f}% na compra. Perguntar que linha migrou para "
                "outro fornecedor e cotar 2 ou 3 códigos hoje.")
    if int(summary.get("currentSkuCount") or 0) <= 2:
        base = "Cliente compra pouca variedade. Perguntar que serviço a oficina mais faz"
        if offer:
            base += f" e oferecer {offer} para teste de giro"
        return base + "."
    return "Manter frequência: confirmar reposição e explorar uma linha que ele ainda não compra."


def crm_attach_context(
    conn: sqlite3.Connection, company_id: int, summaries: list[dict[str, Any]],
    seller_name: str | None = None, with_scripts: bool = False,
) -> list[dict[str, Any]]:
    enriched: list[dict[str, Any]] = []
    # Carrega a biblioteca uma vez só quando os scripts forem necessários
    library = list_content_library(conn, company_id) if with_scripts else []
    for summary in summaries:
        offers = crm_get_offer_suggestions(
            conn, company_id, summary["clientName"], summary.get("cityName")
        )
        questions = crm_generate_questions(summary)
        summary["primaryReason"] = crm_reason_message(summary, summary["primaryReasonCode"])
        summary["secondaryReasons"] = [crm_reason_message(summary, code) for code in summary["secondaryReasonCodes"]]
        summary["offerPrimary"] = offers["primary"]
        summary["offerSecondary"] = offers["secondary"]
        summary["offerRepurchase"] = offers.get("repurchase", [])
        summary["offerOpportunity"] = offers.get("opportunity", [])
        summary["questionPrimary"] = questions["primary"]
        summary["questionSecondary"] = questions["secondary"]
        summary["situationCode"] = crm_situation_for_client(summary)
        summary["nextAction"] = crm_next_action(summary)
        if with_scripts:
            ctx = content_context_for_client(summary, seller_name)
            situation = summary["situationCode"]
            summary["scripts"] = [
                {**item, "body": render_content_text(item["body"], ctx)}
                for item in library
                if item["situation"] in {situation, "GERAL"} and item["category"] in {"ligacao", "whatsapp"}
            ]
        enriched.append(summary)
    return enriched


def crm_agenda_action_map(
    conn: sqlite3.Connection, company_id: int, seller_name: str
) -> dict[str, dict[str, Any]]:
    rows = conn.execute(
        """
        SELECT a1.*
        FROM crm_agenda_actions a1
        JOIN (
            SELECT client_key, MAX(id) AS max_id
            FROM crm_agenda_actions
            WHERE company_id = ? AND seller_name = ?
            GROUP BY client_key
        ) latest ON latest.max_id = a1.id
        WHERE a1.company_id = ? AND a1.seller_name = ?
        """,
        (company_id, seller_name, company_id, seller_name),
    ).fetchall()
    return {row["client_key"]: dict(row) for row in rows}


# Rodízio da fila: ritmo diário assumido e limites da janela de descanso.
ROTATION_DAILY_TARGET = 5      # os 5 contatos prioritários do dia
ROTATION_MIN_DAYS = 7          # nunca voltar no mesmo dia nem na mesma semana
ROTATION_MAX_DAYS = 90         # carteira gigante não pode congelar o cliente


def last_active_contact_map(
    conn: sqlite3.Connection, company_id: int, seller_name: str = ""
) -> dict[str, str]:
    """Data do último contato ATIVO por cliente.

    Só o ativo conta: registro receptivo é o cliente que ligou, e isso não
    significa que o vendedor trabalhou a carteira. Sem essa distinção bastaria
    anotar o que chegou sozinho para o cliente sumir da fila.
    """
    if seller_name:
        linhas = conn.execute(
            """SELECT client_key, MAX(occurred_at) AS u FROM crm_interactions
               WHERE company_id = ? AND UPPER(seller_name) = ? AND initiative = 'ATIVO'
               GROUP BY client_key""",
            (company_id, normalize_upper(seller_name)),
        ).fetchall()
    else:
        linhas = conn.execute(
            """SELECT client_key, MAX(occurred_at) AS u FROM crm_interactions
               WHERE company_id = ? AND initiative = 'ATIVO'
               GROUP BY client_key""",
            (company_id,),
        ).fetchall()
    return {normalize_client_key(r["client_key"]): r["u"] for r in linhas if r["u"]}


def rotation_window_days(rows: list[dict[str, Any]]) -> int:
    """Quantos dias um cliente descansa depois de contatado.

    A janela não é um número fixo: é o tempo que a carteira PRIORITÁRIA leva
    para girar inteira no ritmo de 5 contatos por dia. Carteira de 100 clientes
    prioritários gira em 20 dias; de 400, em 80. Assim "só volta depois que
    todos rodarem" vira consequência da conta, não uma regra arbitrária que
    envelhece quando a carteira muda de tamanho.
    """
    prioritarios = sum(1 for r in rows
                       if r.get("primaryReasonCode") in CRM_PRIORITY_ORDER)
    base = prioritarios or len(rows)
    if not base:
        return ROTATION_MIN_DAYS
    dias = math.ceil(base / ROTATION_DAILY_TARGET)
    return max(ROTATION_MIN_DAYS, min(dias, ROTATION_MAX_DAYS))


def days_since(valor: str | None, referencia: date | None = None) -> int | None:
    dt = parse_datetime_flexible((valor or "")[:10])
    if not dt:
        return None
    return ((referencia or today_in_brazil()) - dt.date()).days


def crm_priority_sort_key(summary: dict[str, Any]) -> tuple[Any, ...]:
    code = summary["primaryReasonCode"]
    priority_index = CRM_PRIORITY_ORDER.index(code) if code in CRM_PRIORITY_ORDER else len(CRM_PRIORITY_ORDER)
    return (
        priority_index,
        crm_class_rank(summary["classCode"]),
        -(summary["daysWithoutPurchase"] or 0),
        summary["dropPct"],
        -summary["currentRevenue"],
        summary["clientName"],
    )


def list_crm_clients(
    conn: sqlite3.Connection,
    company_id: int,
    filters: dict[str, str | None],
    limit: int | None = None,
    attach_context: bool = True,
    exclude_contacted_today: bool = False,
    stats: dict[str, Any] | None = None,
) -> list[dict[str, Any]]:
    seller_name = normalize_whitespace(filters.get("seller_name"))
    rows = crm_base_client_rows_cached(conn, company_id, filters)
    action_map = crm_agenda_action_map(conn, company_id, seller_name) if seller_name else {}
    # ── Clientes já contactados hoje ──────────────────────────────────────────
    # Três correções aplicadas aqui, todas causa de o cliente não sair da fila:
    #  1. Nome do vendedor comparado sem diferenciar maiúsculas — o cadastro e o
    #     registro da interação nem sempre gravam a mesma capitalização.
    #  2. Data avaliada no fuso de São Paulo. O navegador enviava horário UTC:
    #     um contato feito às 21h30 virava "amanhã" e nunca casava com hoje.
    #  3. Sem vendedor no filtro (visão do gerente) a exclusão simplesmente não
    #     rodava. Agora considera o contato de QUALQUER vendedor do escopo.
    ultimo_contato: dict[str, str] = {}
    janela = 0
    if exclude_contacted_today:
        ultimo_contato = last_active_contact_map(conn, company_id, seller_name)
        janela = rotation_window_days(rows)

    visible_rows: list[dict[str, Any]] = []
    adiados: list[dict[str, Any]] = []
    now_dt = datetime.now()
    hoje = today_in_brazil()
    for row in rows:
        action = action_map.get(row["clientKey"])
        row["agendaAction"] = action
        if action and action.get("action_type") == "ADIAR":
            next_visible_at = parse_datetime_flexible(action.get("next_visible_at"))
            if next_visible_at and next_visible_at > now_dt:
                continue
        chave = normalize_client_key(row["clientKey"])
        row["lastActiveContactAt"] = ultimo_contato.get(chave, "")
        if exclude_contacted_today and row["lastActiveContactAt"]:
            dias = days_since(row["lastActiveContactAt"], hoje)
            row["daysSinceActiveContact"] = dias
            if dias is not None and dias < janela:
                # Já foi trabalhado neste ciclo. Guarda de lado: se a fila
                # inteira estiver nesta situação, ele volta como reciclado.
                adiados.append(row)
                continue
        visible_rows.append(row)

    # Fila esgotada = todo mundo já rodou. Em vez de deixar a tela vazia (e o
    # vendedor sem o que fazer), recomeça o ciclo pelos contatados há mais tempo.
    reciclado = False
    if exclude_contacted_today and not visible_rows and adiados:
        adiados.sort(key=lambda r: r.get("lastActiveContactAt") or "")
        visible_rows = adiados
        reciclado = True
        for row in visible_rows:
            row["recycled"] = True

    if exclude_contacted_today:
        # Dentro da mesma prioridade e classe, quem está há mais tempo sem
        # contato vem primeiro — nunca contatado antes de todos. É o que faz a
        # carteira girar inteira em vez de orbitar os mesmos nomes.
        #
        # A data entra ANTES do resto do critério de prioridade: deixá-la no
        # fim fazia o nome do cliente decidir a ordem, e a fila voltava a ser
        # sempre a mesma em ordem alfabética.
        def chave_rodizio(r: dict[str, Any]) -> tuple[Any, ...]:
            base = crm_priority_sort_key(r)
            return (base[0], base[1], r.get("lastActiveContactAt") or "", *base[2:])

        visible_rows.sort(key=chave_rodizio)
    else:
        visible_rows.sort(key=crm_priority_sort_key)

    if stats is not None:
        stats.update({"cycleDays": janela, "recycled": reciclado,
                      "restingCount": len(adiados) if not reciclado else 0,
                      "eligibleCount": len(visible_rows)})
    if limit is not None:
        visible_rows = visible_rows[:limit]
    # Os sinais de relacionamento NÃO são calculados aqui. Eles montam um IN com
    # as chaves recebidas e esta função devolve a carteira inteira — depois que a
    # base passou de 100 mil clientes, virou um IN de 100 mil parâmetros por
    # consulta e a tela levava minutos. Quem chama aplica em cima da PÁGINA.
    if not attach_context:
        return visible_rows
    attach_engagement_markers(conn, company_id, visible_rows)
    return crm_attach_context(conn, company_id, visible_rows)


ENGAGEMENT_RECENT_DAYS = 30


def attach_engagement_markers(
    conn: sqlite3.Connection, company_id: int, rows: list[dict[str, Any]]
) -> None:
    """Marca cada cliente da lista com o que já aconteceu com ele.

    Quatro sinais, decididos com o Felipe: contato ativo nos últimos 30 dias,
    visita registrada, retorno pendente e nunca contatado. O último é o que
    mais importa — mostra o vazio, não só o cheio.

    Tudo em três consultas em lote sobre as chaves JÁ VISÍVEIS. Uma consulta por
    linha derrubaria a carteira, que tem centenas de clientes por vendedor.
    """
    if not rows:
        return
    chaves = [normalize_client_key(r.get("clientKey")) for r in rows if r.get("clientKey")]
    if not chaves:
        return
    # Lote pequeno por segurança: SQLite tem teto de parâmetros por consulta e,
    # mesmo onde não estoura, um IN enorme fica lento.
    if len(chaves) > 400:
        for inicio in range(0, len(rows), 400):
            attach_engagement_markers(conn, company_id, rows[inicio : inicio + 400])
        return
    marcadores = ",".join("?" for _ in chaves)
    limite = (today_in_brazil() - timedelta(days=ENGAGEMENT_RECENT_DAYS)).isoformat()

    contatos: dict[str, dict[str, Any]] = {}
    for r in conn.execute(
        f"""
        SELECT client_key,
               MAX(CASE WHEN initiative = 'ATIVO' THEN occurred_at END) AS ultimo_ativo,
               MAX(occurred_at) AS ultimo_qualquer,
               COUNT(*) AS total
        FROM crm_interactions
        WHERE company_id = ? AND client_key IN ({marcadores})
        GROUP BY client_key
        """,
        (company_id, *chaves),
    ).fetchall():
        contatos[normalize_client_key(r["client_key"])] = dict(r)

    visitas: dict[str, str] = {}
    for r in conn.execute(
        f"""
        SELECT client_key, MAX(COALESCE(occurred_at, scheduled_for)) AS ultima
        FROM visits
        WHERE company_id = ? AND client_key IN ({marcadores}) AND status <> 'CANCELADA'
        GROUP BY client_key
        """,
        (company_id, *chaves),
    ).fetchall():
        if r["ultima"]:
            visitas[normalize_client_key(r["client_key"])] = r["ultima"]

    pendentes: dict[str, str] = {}
    for r in conn.execute(
        f"""
        SELECT client_key, MIN(due_at) AS proxima
        FROM crm_tasks
        WHERE company_id = ? AND client_key IN ({marcadores})
          AND status IN ('ABERTA','REAGENDADA','ATRASADA')
        GROUP BY client_key
        """,
        (company_id, *chaves),
    ).fetchall():
        pendentes[normalize_client_key(r["client_key"])] = r["proxima"]

    for row in rows:
        chave = normalize_client_key(row.get("clientKey"))
        info = contatos.get(chave) or {}
        ultimo_ativo = info.get("ultimo_ativo")
        row["engagement"] = {
            "lastActiveContactAt": ultimo_ativo,
            "activeRecent": bool(ultimo_ativo and ultimo_ativo[:10] >= limite),
            "lastVisitAt": visitas.get(chave),
            "pendingFollowupAt": pendentes.get(chave),
            # "Nunca contatado" é ausência de QUALQUER toque: nenhum registro
            # (ativo ou receptivo) E nenhuma visita. Marcar de vermelho um
            # cliente que o gerente visitou queimaria a confiança no sinal.
            "neverContacted": not info.get("total") and chave not in visitas,
        }


def crm_matches_search(row: dict[str, Any], search_value: str) -> bool:
    haystack = " ".join(
        [
            normalize_whitespace(row.get("clientKey")),
            normalize_whitespace(row.get("clientCode")),
            normalize_whitespace(row.get("clientName")),
            normalize_whitespace(row.get("cityName")),
            normalize_whitespace(row.get("phone")),
            normalize_whitespace(row.get("primaryContactName")),
        ]
    ).lower()
    return search_value.lower() in haystack


def clients_who_bought_item(
    conn: sqlite3.Connection, company_id: int, item_code: str, months_window: int = 12
) -> set[str] | None:
    """Nomes normalizados dos clientes que compraram determinado item.

    Busca no código do fabricante E no código interno (a coluna sem cabeçalho do
    relatório) — o vendedor nem sempre tem em mãos o mesmo código que digitaria.
    Aceita correspondência parcial para funcionar com código digitado incompleto.
    Devolve None quando não há filtro a aplicar.
    """
    termo = normalize_whitespace(item_code)
    if not termo:
        return None
    detalhes = item_purchase_details(conn, company_id, item_code, months_window)
    return None if detalhes is None else set(detalhes.keys())


def item_purchase_details(
    conn: sqlite3.Connection, company_id: int, item_code: str, months_window: int = 12
) -> dict[str, dict[str, Any]] | None:
    """Quando, quanto e por quanto cada cliente comprou o item pesquisado.

    Devolve, por cliente: a data da ÚLTIMA compra do item, a quantidade e o preço
    unitário praticado nessa compra, mais o acumulado da janela (total de peças e
    quantas compras). O preço unitário sai do líquido dividido pela quantidade —
    é o que o cliente efetivamente pagou, já com desconto e devolução.

    Devolve None quando não há filtro de item a aplicar.
    """
    termo = normalize_whitespace(item_code)
    if not termo:
        return None
    latest = crm_latest_competence(conn, company_id) or date.today().strftime("%Y-%m")
    inicio = shift_competence(latest, -(months_window - 1))
    padrao = f"%{termo.upper()}%"
    rows = conn.execute(
        """
        SELECT client_name,
               COALESCE(NULLIF(issue_date, ''), competence) AS purchase_at,
               COALESCE(NULLIF(manufacturer_sku, ''), NULLIF(sku_key, ''), NULLIF(gtin_value, '')) AS item_code,
               SUM(quantity) AS quantity,
               SUM(net_value) AS net_value
        FROM fact_sales_detail
        WHERE company_id = ? AND competence >= ? AND net_value > 0
          AND (UPPER(COALESCE(manufacturer_sku, '')) LIKE ?
            OR UPPER(COALESCE(gtin_value, '')) LIKE ?
            OR UPPER(COALESCE(sku_key, '')) LIKE ?)
        GROUP BY client_name, purchase_at, item_code
        ORDER BY purchase_at ASC
        """,
        (company_id, inicio, padrao, padrao, padrao),
    ).fetchall()

    detalhes: dict[str, dict[str, Any]] = {}
    for row in rows:
        chave = normalize_client_key(row["client_name"])
        if not chave:
            continue
        qtd = float(row["quantity"] or 0.0)
        valor = float(row["net_value"] or 0.0)
        atual = detalhes.setdefault(chave, {
            "lastPurchaseAt": "", "lastQuantity": 0.0, "lastUnitPrice": None,
            "totalQuantity": 0.0, "totalValue": 0.0, "purchaseCount": 0, "itemCode": "",
        })
        atual["totalQuantity"] += qtd
        atual["totalValue"] += valor
        atual["purchaseCount"] += 1
        # Rows vêm em ordem crescente de data, então a última iteração é a compra mais recente.
        atual["lastPurchaseAt"] = normalize_whitespace(row["purchase_at"])
        atual["lastQuantity"] = qtd
        atual["lastUnitPrice"] = finite_or_none(valor / qtd) if qtd > 0 else None
        atual["itemCode"] = normalize_whitespace(row["item_code"]) or atual["itemCode"]

    for valores in detalhes.values():
        total_qtd = valores["totalQuantity"]
        valores["avgUnitPrice"] = finite_or_none(valores["totalValue"] / total_qtd) if total_qtd > 0 else None
    return detalhes


def sort_by_search_relevance(
    rows: list[dict[str, Any]], termo: str
) -> list[dict[str, Any]]:
    """Põe na frente quem casou pelo CÓDIGO exato.

    A busca varre código, nome, cidade, telefone e contato — então digitar
    "99856" também traz quem tem 99856 no telefone. Sem ordenar por relevância,
    o cliente procurado aparecia no fim de uma lista de homônimos acidentais, o
    que na prática equivale a não ter encontrado.

    A ordenação é ESTÁVEL: dentro de cada faixa, a ordem de prioridade da
    carteira (inativo primeiro, classe, etc.) continua valendo.
    """
    alvo = normalize_upper(termo)
    if not alvo:
        return rows

    def faixa(row: dict[str, Any]) -> int:
        codigo = normalize_upper(row.get("clientKey") or row.get("clientCode"))
        nome = normalize_upper(row.get("clientName"))
        if codigo == alvo:
            return 0                      # o código que a pessoa digitou
        if codigo.startswith(alvo):
            return 1                      # código que começa com o termo
        if nome.startswith(alvo):
            return 2                      # nome que começa com o termo
        if alvo in codigo:
            return 3                      # termo em algum lugar do código
        return 4                          # casou por nome, cidade, telefone…

    return sorted(rows, key=faixa)


def filter_crm_client_rows(
    rows: list[dict[str, Any]], filters: dict[str, str | None],
    item_buyers: set[str] | None = None,
) -> list[dict[str, Any]]:
    status_filter = normalize_upper(filters.get("status"))
    purchase_filter = normalize_upper(filters.get("purchaseMonth"))
    growth_filter = normalize_upper(filters.get("growth"))
    class_filter = normalize_upper(filters.get("classCode"))
    person_type_filter = normalize_upper(filters.get("personType"))
    search_filter = normalize_whitespace(filters.get("search"))
    filtered: list[dict[str, Any]] = []
    for row in rows:
        if status_filter and row.get("statusCode") != status_filter:
            continue
        current_revenue = float(row.get("currentRevenue") or 0.0)
        if purchase_filter == "COM_COMPRA" and current_revenue <= 0:
            continue
        if purchase_filter == "SEM_COMPRA" and current_revenue > 0:
            continue
        growth_pct = float(row.get("growthPct") or 0.0)
        if growth_filter == "ACIMA" and growth_pct <= 0.03:
            continue
        if growth_filter == "ESTAVEL" and not (-0.03 <= growth_pct <= 0.03):
            continue
        if growth_filter == "ABAIXO" and growth_pct >= -0.03:
            continue
        class_code = normalize_upper(row.get("classCode"))
        if class_filter == "SEM_CLASSE" and class_code:
            continue
        if class_filter and class_filter != "SEM_CLASSE" and class_code != class_filter:
            continue
        if person_type_filter and normalize_upper(row.get("personType")) != person_type_filter:
            continue
        if search_filter and not crm_matches_search(row, search_filter):
            continue
        if item_buyers is not None:
            chaves = {
                normalize_client_key(row.get("clientName")),
                normalize_client_key(row.get("summaryClientName")),
                normalize_client_key(row.get("tradeName")),
            }
            if not (chaves & item_buyers):
                continue
        filtered.append(row)
    return filtered


def query_crm_clients_page(
    conn: sqlite3.Connection,
    company_id: int,
    filters: dict[str, str | None],
    page: int,
    page_size: int,
) -> dict[str, Any]:
    all_rows = list_crm_clients(conn, company_id, filters, attach_context=False)
    item_details = item_purchase_details(conn, company_id, filters.get("itemCode"))
    item_buyers = None if item_details is None else set(item_details.keys())
    filtered_rows = filter_crm_client_rows(all_rows, filters, item_buyers)
    filtered_rows = sort_by_search_relevance(filtered_rows, filters.get("search") or "")
    total = len(filtered_rows)
    safe_page_size = min(max(int(page_size or 50), 1), 100)
    total_pages = max(math.ceil(total / safe_page_size), 1) if total else 1
    safe_page = min(max(int(page or 1), 1), total_pages)
    offset = (safe_page - 1) * safe_page_size
    visible_rows = filtered_rows[offset : offset + safe_page_size]
    attach_engagement_markers(conn, company_id, visible_rows)
    rows = crm_attach_context(conn, company_id, visible_rows)
    # Só nos clientes da página visível — não faz sentido carregar o detalhe de quem não aparece.
    if item_details:
        for row in rows:
            for candidata in (row.get("clientName"), row.get("summaryClientName"), row.get("tradeName")):
                detalhe = item_details.get(normalize_client_key(candidata))
                if detalhe:
                    row["itemPurchase"] = detalhe
                    break
    print(
        "[CRM CLIENTS PAGE DEBUG]",
        {
            "baseRows": len(all_rows),
            "filteredRows": len(filtered_rows),
            "total": total,
            "page": safe_page,
            "pageSize": safe_page_size,
            "totalPages": total_pages,
            "rowsReturned": len(rows),
            "filters": filters,
        },
    )
    return {
        "rows": rows,
        "total": total,
        "page": safe_page,
        "pageSize": safe_page_size,
        "totalPages": total_pages,
    }


def count_crm_clients(
    conn: sqlite3.Connection,
    company_id: int,
    filters: dict[str, str | None],
) -> int:
    has_client_filters = any(
        normalize_whitespace(filters.get(key))
        for key in ("status", "purchaseMonth", "growth", "classCode", "search")
    )
    if not has_client_filters:
        return crm_base_client_count(conn, company_id, filters)
    filtered_rows = filter_crm_client_rows(list_crm_clients(conn, company_id, filters, attach_context=False), filters)
    return len(filtered_rows)


def get_crm_client_summary(
    conn: sqlite3.Connection, company_id: int, filters: dict[str, str | None], client_key: str,
    seller_name: str | None = None, allow_outside: bool = False,
) -> dict[str, Any] | None:
    base_summary = next(
        (row for row in list_crm_clients(conn, company_id, filters, attach_context=False) if row["clientKey"] == client_key),
        None,
    )
    if not base_summary and allow_outside:
        # Atendimento de apoio: o vendedor chegou pelo CÓDIGO EXATO, informado
        # pelo cliente na linha. A ficha é completa — histórico e vendas — para
        # ele conseguir negociar com o mesmo contexto de quem é dono.
        livres = dict(filters)
        livres["seller_name"] = ""
        livres["unit_name"] = ""
        livres["allowed_units"] = None
        base_summary = next(
            (row for row in list_crm_clients(conn, company_id, livres, attach_context=False)
             if row["clientKey"] == client_key), None)
    if not base_summary:
        return None
    # A ficha traz os scripts prontos, já com o nome do cliente preenchido
    summary = crm_attach_context(
        conn, company_id, [base_summary], seller_name=seller_name, with_scripts=True
    )[0]
    profile = conn.execute(
        """
        SELECT client_code, client_name, city_name, phone, updated_phone, primary_contact_name, contact_notes
        FROM crm_client_profiles
        WHERE company_id = ? AND client_code = ?
        """,
        (company_id, client_key),
    ).fetchone()
    profile_payload = {
        "clientKey": client_key,
        "clientCode": summary.get("clientCode") or client_key,
        "clientName": summary.get("clientName"),
        "cityName": summary.get("cityName"),
        "city_name": summary.get("cityName"),
        "phone": None,
        "updatedPhone": None,
        "primaryContactName": None,
        "contactNotes": None,
    }
    if profile:
        profile_payload.update(
            {
                "clientCode": profile["client_code"],
                "clientName": profile["client_name"] or profile_payload["clientName"],
                "cityName": profile["city_name"] or profile_payload["cityName"],
                "city_name": profile["city_name"] or profile_payload["city_name"],
                "phone": profile["phone"],
                "updatedPhone": profile["updated_phone"],
                "primaryContactName": profile["primary_contact_name"],
                "contactNotes": profile["contact_notes"],
            }
        )
    summary_payload = {
        **summary,
        "clientCode": summary.get("clientCode") or summary.get("clientKey"),
        "phone": summary.get("phone") or (profile["phone"] if profile else None),
        "updatedPhone": summary.get("updatedPhone") or (profile["updated_phone"] if profile else None),
        "primaryContactName": summary.get("primaryContactName") or (profile["primary_contact_name"] if profile else None),
        "contactNotes": summary.get("contactNotes") or (profile["contact_notes"] if profile else None),
    }
    return {"summary": summary_payload, "profile": profile_payload}


def crm_client_detail_name(
    conn: sqlite3.Connection, company_id: int, filters: dict[str, str | None], client_key: str
) -> tuple[dict[str, Any] | None, str | None]:
    summary_data = get_crm_client_summary(conn, company_id, filters, client_key)
    if not summary_data:
        return None, None
    summary = summary_data["summary"]
    detail_name = summary.get("summaryClientName") or summary.get("clientName")
    return summary_data, detail_name


def get_crm_client_purchases(
    conn: sqlite3.Connection, company_id: int, filters: dict[str, str | None], client_key: str
) -> list[dict[str, Any]] | None:
    summary_data, detail_name = crm_client_detail_name(conn, company_id, filters, client_key)
    if not summary_data:
        return None
    current_month_rows = conn.execute(
        """
        SELECT competence, ROUND(SUM(net_value), 2) AS revenue
        FROM crm_client_summary
        WHERE company_id = ? AND client_code = ?
        GROUP BY competence
        ORDER BY competence DESC
        LIMIT 6
        """,
        (company_id, client_key),
    ).fetchall()
    current_competence = crm_latest_competence(conn, company_id)
    history_competences = [shift_competence(current_competence, -offset) for offset in range(1, 6)] if current_competence else []
    candidate_names = {
        normalize_whitespace(summary_data["summary"].get("clientName")),
        normalize_whitespace(summary_data["summary"].get("summaryClientName")),
        normalize_whitespace(detail_name),
    }
    candidate_names = {name for name in candidate_names if name}
    detail_history: dict[str, float] = {}
    if candidate_names and history_competences:
        placeholders_names = ", ".join("?" for _ in candidate_names)
        placeholders_competences = ", ".join("?" for _ in history_competences)
        detail_history_rows = conn.execute(
            f"""
            SELECT competence, ROUND(SUM(net_value), 2) AS revenue
            FROM fact_sales_detail
            WHERE company_id = ? AND client_name IN ({placeholders_names}) AND competence IN ({placeholders_competences})
            GROUP BY competence
            ORDER BY competence DESC
            """,
            [company_id] + list(candidate_names) + history_competences,
        ).fetchall()
        detail_history = {row["competence"]: float(row["revenue"] or 0.0) for row in detail_history_rows}
    monthly_map = {row["competence"]: float(row["revenue"] or 0.0) for row in current_month_rows}
    for competence, revenue in detail_history.items():
        monthly_map.setdefault(competence, revenue)
    return [
        {"competence": competence, "revenue": round(monthly_map[competence], 2)}
        for competence in sorted(monthly_map.keys(), reverse=True)[:6]
    ]


def get_crm_client_items(
    conn: sqlite3.Connection, company_id: int, filters: dict[str, str | None], client_key: str, page: int, page_size: int
) -> dict[str, Any] | None:
    _, detail_name = crm_client_detail_name(conn, company_id, filters, client_key)
    if not detail_name:
        return None
    safe_page_size = min(max(int(page_size or 20), 1), 100)
    safe_page = max(int(page or 1), 1)
    offset = (safe_page - 1) * safe_page_size
    total_row = conn.execute(
        """
        SELECT COUNT(*) AS total
        FROM (
            SELECT issue_date, COALESCE(NULLIF(manufacturer_sku, ''), NULLIF(sku_key, ''), NULLIF(gtin_value, ''), 'ITEM') AS item_code
            FROM fact_sales_detail
            WHERE company_id = ? AND client_name = ?
            GROUP BY issue_date, item_code
        )
        """,
        (company_id, detail_name),
    ).fetchone()
    total = int(total_row["total"] or 0)
    total_pages = max(math.ceil(total / safe_page_size), 1) if total else 1
    safe_page = min(safe_page, total_pages)
    rows = conn.execute(
        """
        SELECT
            issue_date,
            COALESCE(NULLIF(manufacturer_sku, ''), NULLIF(sku_key, ''), NULLIF(gtin_value, ''), 'ITEM') AS item_code,
            SUM(quantity) AS quantity,
            ROUND(SUM(net_value), 2) AS net_value
        FROM fact_sales_detail
        WHERE company_id = ? AND client_name = ?
        GROUP BY issue_date, item_code
        ORDER BY datetime(issue_date) DESC
        LIMIT ? OFFSET ?
        """,
        (company_id, detail_name, safe_page_size, (safe_page - 1) * safe_page_size),
    ).fetchall()
    return {"rows": [dict(row) for row in rows], "total": total, "page": safe_page, "pageSize": safe_page_size, "totalPages": total_pages}


def get_crm_client_interactions(
    conn: sqlite3.Connection, company_id: int, filters: dict[str, str | None], client_key: str, page: int, page_size: int
) -> dict[str, Any] | None:
    summary_data = get_crm_client_summary(conn, company_id, filters, client_key)
    if not summary_data:
        return None
    safe_page_size = min(max(int(page_size or 20), 1), 100)
    total_row = conn.execute("SELECT COUNT(*) AS total FROM crm_interactions WHERE company_id = ? AND client_key = ?", (company_id, client_key)).fetchone()
    total = int(total_row["total"] or 0)
    total_pages = max(math.ceil(total / safe_page_size), 1) if total else 1
    safe_page = min(max(int(page or 1), 1), total_pages)
    rows = conn.execute(
        """
        SELECT id, contact_type_code, result_code, occurred_at, notes, question_used, had_progress,
               offer_title, next_action, followup_due_at, contact_phone, contact_name
        FROM crm_interactions
        WHERE company_id = ? AND client_key = ?
        ORDER BY datetime(occurred_at) DESC
        LIMIT ? OFFSET ?
        """,
        (company_id, client_key, safe_page_size, (safe_page - 1) * safe_page_size),
    ).fetchall()
    return {"rows": [dict(row) for row in rows], "total": total, "page": safe_page, "pageSize": safe_page_size, "totalPages": total_pages}


def get_crm_client_tasks(
    conn: sqlite3.Connection, company_id: int, filters: dict[str, str | None], client_key: str
) -> list[dict[str, Any]] | None:
    summary_data = get_crm_client_summary(conn, company_id, filters, client_key)
    if not summary_data:
        return None
    rows = conn.execute(
        """
        SELECT id, title, description, due_at, status, created_at, completed_at
        FROM crm_tasks
        WHERE company_id = ? AND client_key = ?
        ORDER BY
            CASE status WHEN 'ATRASADA' THEN 0 WHEN 'ABERTA' THEN 1 WHEN 'REAGENDADA' THEN 2 ELSE 3 END,
            datetime(due_at) ASC
        LIMIT 20
        """,
        (company_id, client_key),
    ).fetchall()
    return [dict(row) for row in rows]


def get_crm_client_360(
    conn: sqlite3.Connection, company_id: int, filters: dict[str, str | None], client_key: str,
    allow_outside: bool = False,
) -> dict[str, Any] | None:
    summaries = {row["clientKey"]: row for row in list_crm_clients(conn, company_id, filters)}
    summary = summaries.get(client_key)
    if not summary and allow_outside:
        # Atendimento de apoio: o vendedor chegou pelo CÓDIGO EXATO, informado
        # pelo próprio cliente na linha. Aqui a ficha é completa, incluindo
        # histórico e vendas — decisão do Felipe: quem atende precisa negociar
        # com o mesmo contexto de quem é dono.
        livres = dict(filters)
        livres["seller_name"] = ""
        livres["unit_name"] = ""
        livres["allowed_units"] = None
        summary = {row["clientKey"]: row for row in list_crm_clients(
            conn, company_id, livres)}.get(client_key)
    if not summary:
        return None
    detail_name = summary.get("summaryClientName") or summary["clientName"]
    profile = conn.execute(
        """
        SELECT client_code, client_name, trade_name, document_number, state_registration, address_line, address_number,
               neighborhood, city_name, state_name, phone, updated_phone, primary_contact_name, contact_notes,
               contact_updated_at, postal_code, first_sale_at, last_sale_at,
               credit_limit, economic_group, internal_seller_name, external_seller_name, email
        FROM crm_client_profiles
        WHERE company_id = ? AND client_code = ?
        """,
        (company_id, client_key),
    ).fetchone()
    current_month_rows = conn.execute(
        """
        SELECT competence, ROUND(SUM(net_value), 2) AS revenue
        FROM crm_client_summary
        WHERE company_id = ? AND client_code = ?
        GROUP BY competence
        ORDER BY competence DESC
        LIMIT 6
        """,
        (company_id, client_key),
    ).fetchall()
    current_competence = crm_latest_competence(conn, company_id)
    history_competences = [shift_competence(current_competence, -offset) for offset in range(1, 6)] if current_competence else []
    candidate_names = {
        normalize_whitespace(summary.get("clientName")),
        normalize_whitespace(summary.get("summaryClientName")),
        normalize_whitespace(profile["client_name"]) if profile else "",
        normalize_whitespace(profile["trade_name"]) if profile else "",
    }
    candidate_names = {name for name in candidate_names if name}
    detail_history: dict[str, float] = {}
    if candidate_names and history_competences:
        placeholders_names = ", ".join("?" for _ in candidate_names)
        placeholders_competences = ", ".join("?" for _ in history_competences)
        detail_history_rows = conn.execute(
            f"""
            SELECT competence, ROUND(SUM(net_value), 2) AS revenue
            FROM fact_sales_detail
            WHERE company_id = ? AND client_name IN ({placeholders_names}) AND competence IN ({placeholders_competences})
            GROUP BY competence
            ORDER BY competence DESC
            """,
            [company_id] + list(candidate_names) + history_competences,
        ).fetchall()
        detail_history = {row["competence"]: float(row["revenue"] or 0.0) for row in detail_history_rows}
    monthly_map = {row["competence"]: float(row["revenue"] or 0.0) for row in current_month_rows}
    for competence, revenue in detail_history.items():
        monthly_map.setdefault(competence, revenue)
    monthly_rows = [
        {"competence": competence, "revenue": round(monthly_map[competence], 2)}
        for competence in sorted(monthly_map.keys(), reverse=True)[:6]
    ]
    recent_items = conn.execute(
        """
        SELECT
            issue_date,
            COALESCE(NULLIF(manufacturer_sku, ''), NULLIF(sku_key, ''), NULLIF(gtin_value, ''), 'ITEM') AS item_code,
            SUM(quantity) AS quantity,
            ROUND(SUM(net_value), 2) AS net_value
        FROM fact_sales_detail
        WHERE company_id = ? AND client_name = ?
        GROUP BY issue_date, item_code
        ORDER BY datetime(issue_date) DESC
        LIMIT 15
        """,
        (company_id, detail_name),
    ).fetchall()
    interaction_rows = conn.execute(
        """
        SELECT id, contact_type_code, result_code, occurred_at, notes, question_used, had_progress,
               offer_title, next_action, followup_due_at, contact_phone, contact_name
        FROM crm_interactions
        WHERE company_id = ? AND client_key = ?
        ORDER BY datetime(occurred_at) DESC
        LIMIT 20
        """,
        (company_id, client_key),
    ).fetchall()
    task_rows = conn.execute(
        """
        SELECT id, title, description, due_at, status, created_at, completed_at
        FROM crm_tasks
        WHERE company_id = ? AND client_key = ?
        ORDER BY
            CASE status WHEN 'ATRASADA' THEN 0 WHEN 'ABERTA' THEN 1 WHEN 'REAGENDADA' THEN 2 ELSE 3 END,
            datetime(due_at) ASC
        LIMIT 20
        """,
        (company_id, client_key),
    ).fetchall()
    profile_payload = None
    if profile:
        profile_payload = {
            **dict(profile),
            "clientKey": client_key,
            "clientCode": profile["client_code"],
            "clientName": profile["client_name"],
            "cityName": profile["city_name"],
            "unitName": summary.get("unitName"),
            "classCode": summary.get("classCode"),
            "statusCode": summary.get("statusCode"),
            "currentRevenue": summary.get("currentRevenue"),
            "averageRevenue": summary.get("averageRevenue"),
            "growthPct": summary.get("growthPct"),
            "lastPurchaseAt": summary.get("lastPurchaseAt"),
            "daysWithoutPurchase": summary.get("daysWithoutPurchase"),
            "primaryReason": summary.get("primaryReason"),
            "updatedPhone": profile["updated_phone"],
            "primaryContactName": profile["primary_contact_name"],
            "contactNotes": profile["contact_notes"],
            # Território é informativo na ficha: mostra de quem o bairro é,
            # mesmo quando o vendedor que atende é de outra unidade. A carteira
            # NÃO muda por causa disso — serve para o gerente enxergar a
            # sobreposição e decidir se troca o atendimento.
            "territory": territory_for_client(conn, company_id, profile["client_code"]),
        }
    summary_payload = {**summary, "clientCode": summary.get("clientCode") or summary.get("clientKey")}
    return {
        "summary": summary_payload,
        "profile": profile_payload,
        "monthlyRevenue": monthly_rows,
        "recentItems": [dict(row) for row in recent_items],
        "interactions": [dict(row) for row in interaction_rows],
        "tasks": [dict(row) for row in task_rows],
    }


def crm_summary_for_user(
    conn: sqlite3.Connection, company_id: int, user: sqlite3.Row, filters: dict[str, str | None]
) -> dict[str, Any]:
    seller_name = normalize_whitespace(filters.get("seller_name")) or seller_identity_for_user(user)
    base_clients = crm_base_client_rows_cached(conn, company_id, filters)
    today_str = date.today().isoformat()  # "2026-06-02" — funciona com ambos separadores T e espaço
    contacts_today = conn.execute(
        """
        SELECT
            SUM(CASE WHEN initiative = 'ATIVO' THEN 1 ELSE 0 END) AS total_contacts,
            SUM(CASE WHEN initiative = 'RECEPTIVO' THEN 1 ELSE 0 END) AS receptive_contacts,
            SUM(CASE WHEN result_code NOT IN ('NAO_ATENDEU','PEDIU_RETORNO')
                      AND initiative = 'ATIVO' THEN 1 ELSE 0 END) AS active_contacts,
            SUM(CASE WHEN result_code = 'FALOU_CLIENTE' THEN 1 ELSE 0 END) AS success_contacts,
            SUM(CASE WHEN result_code = 'GEROU_ORCAMENTO' THEN 1 ELSE 0 END) AS generated_quotes,
            SUM(CASE WHEN result_code = 'GEROU_PEDIDO' THEN 1 ELSE 0 END) AS generated_orders
        FROM crm_interactions
        WHERE company_id = ? AND seller_name = ? AND substr(occurred_at, 1, 10) = ?
        """,
        (company_id, seller_name, today_str),
    ).fetchone()
    open_tasks = conn.execute(
        """
        SELECT
            SUM(CASE WHEN status IN ('ABERTA', 'REAGENDADA') THEN 1 ELSE 0 END) AS open_tasks,
            SUM(CASE WHEN status = 'ATRASADA' THEN 1 ELSE 0 END) AS overdue_tasks
        FROM crm_tasks
        WHERE company_id = ? AND seller_name = ?
        """,
        (company_id, seller_name),
    ).fetchone()
    return {
        "portfolioSize": len(base_clients),
        "top5Count": min(len(base_clients), 5),
        "contactsToday": int(contacts_today["active_contacts"] or 0),
        "successContactsToday": int(contacts_today["success_contacts"] or 0),
        "quotesToday": int(contacts_today["generated_quotes"] or 0),
        "ordersToday": int(contacts_today["generated_orders"] or 0),
        "inactiveClients": sum(1 for client in base_clients if client["statusCode"] == "INATIVO"),
        "preInactiveClients": sum(1 for client in base_clients if client["statusCode"] == "PRE_INATIVO"),
        "openTasks": int(open_tasks["open_tasks"] or 0),
        "overdueTasks": int(open_tasks["overdue_tasks"] or 0),
    }


def save_crm_client_contact(
    conn: sqlite3.Connection,
    company_id: int,
    user: sqlite3.Row,
    payload: dict[str, Any],
) -> dict[str, Any]:
    client_key = normalize_client_key(payload.get("clientKey"))
    client_name = normalize_whitespace(payload.get("clientName"))
    updated_phone = normalize_whitespace(payload.get("updatedPhone"))
    primary_contact_name = normalize_whitespace(payload.get("primaryContactName"))
    contact_notes = normalize_whitespace(payload.get("notes") or payload.get("contactNotes"))
    if not client_key or not client_name:
        raise ValueError("Cliente invalido para atualizacao de contato")

    existing = conn.execute(
        """
        SELECT id, phone, updated_phone, primary_contact_name, contact_notes
        FROM crm_client_profiles
        WHERE company_id = ? AND client_code = ?
        """,
        (company_id, client_key),
    ).fetchone()

    if existing:
        conn.execute(
            """
            UPDATE crm_client_profiles
            SET client_name = ?,
                updated_phone = ?,
                primary_contact_name = ?,
                contact_notes = ?,
                contact_updated_at = ?,
                contact_updated_by_user_id = ?,
                updated_at = ?
            WHERE company_id = ? AND client_code = ?
            """,
            (
                client_name,
                updated_phone or None,
                primary_contact_name or None,
                contact_notes or None,
                now_iso(),
                user["id"],
                now_iso(),
                company_id,
                client_key,
            ),
        )
        return {
            "clientKey": client_key,
            "updatedPhone": updated_phone or existing["updated_phone"] or existing["phone"],
            "primaryContactName": primary_contact_name or existing["primary_contact_name"],
        }

    conn.execute(
        """
        INSERT INTO crm_client_profiles (
            company_id, client_code, client_name, updated_phone, primary_contact_name,
            contact_notes, contact_updated_at, contact_updated_by_user_id, updated_at, created_at
        ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
        """,
        (
            company_id,
            client_key,
            client_name,
            updated_phone or None,
            primary_contact_name or None,
            contact_notes or None,
            now_iso(),
            user["id"],
            now_iso(),
            now_iso(),
        ),
    )
    return {
        "clientKey": client_key,
        "updatedPhone": updated_phone,
        "primaryContactName": primary_contact_name,
    }


def active_coverages_for_seller(
    conn: sqlite3.Connection, company_id: int, seller_name: str, quando: str = ""
) -> list[dict[str, Any]]:
    """Carteiras que este vendedor está autorizado a cobrir hoje.

    A cobertura é nominal e tem prazo: o gerente diz quem cobre quem e até
    quando. Sem prazo, "cobertura de férias" viraria acesso permanente à
    carteira do colega — que é exatamente o que não queremos.
    """
    alvo = person_key(seller_name)
    if not alvo:
        return []
    hoje = quando or today_in_brazil().isoformat()
    linhas = []
    for r in conn.execute(
        "SELECT * FROM portfolio_coverage WHERE company_id = ? "
        "AND date(start_date) <= date(?) "
        "AND (end_date IS NULL OR end_date = '' OR date(end_date) >= date(?))",
        (company_id, hoje, hoje),
    ).fetchall():
        if person_key(r["covering_seller"]) == alvo:
            linhas.append(dict(r))
    return linhas


def seller_can_see_portfolio(
    conn: sqlite3.Connection, company_id: int, user: sqlite3.Row, seller_name: str
) -> bool:
    """O usuário pode abrir a carteira deste vendedor?"""
    if data_scope_for_user(conn, user) != "proprio":
        return True
    alvo = person_key(seller_name)
    if alvo == person_key(seller_identity_for_user(user)):
        return True
    return any(person_key(c["covered_seller"]) == alvo
               for c in active_coverages_for_seller(
                   conn, company_id, seller_identity_for_user(user)))


def save_coverage(
    conn: sqlite3.Connection, company_id: int, user_id: int, payload: dict[str, Any]
) -> dict[str, Any]:
    cobre = normalize_whitespace(payload.get("coveringSeller"))
    coberto = normalize_whitespace(payload.get("coveredSeller"))
    if not cobre or not coberto:
        raise ValueError("Informe quem cobre e qual carteira será coberta.")
    if person_key(cobre) == person_key(coberto):
        raise ValueError("O vendedor não precisa de autorização para a própria carteira.")
    inicio = normalize_whitespace(payload.get("startDate")) or today_in_brazil().isoformat()
    fim = normalize_whitespace(payload.get("endDate")) or None
    if fim and fim < inicio:
        raise ValueError("A data final não pode ser anterior à inicial.")
    registro_id = payload.get("id")
    if registro_id:
        conn.execute(
            "UPDATE portfolio_coverage SET covering_seller=?, covered_seller=?, start_date=?, "
            "end_date=?, reason=? WHERE company_id = ? AND id = ?",
            (cobre, coberto, inicio, fim, normalize_whitespace(payload.get("reason")) or None,
             company_id, int(registro_id)),
        )
    else:
        conn.execute(
            """
            INSERT INTO portfolio_coverage
                (company_id, covering_seller, covered_seller, unit_name, start_date, end_date,
                 reason, created_by_user_id, created_at)
            VALUES (?,?,?,?,?,?,?,?,?)
            """,
            (company_id, cobre, coberto, normalize_unit(payload.get("unitName")) or None,
             inicio, fim, normalize_whitespace(payload.get("reason")) or None, user_id, now_iso()),
        )
    audit_log(conn, company_id, user_id, "salvar", "portfolio_coverage",
              f"{cobre} cobre {coberto}", {"inicio": inicio, "fim": fim})
    conn.commit()
    prazo = f"até {fim}" if fim else "sem prazo definido"
    return {"message": f"{cobre} passa a enxergar a carteira de {coberto} ({prazo})."}


def delete_coverage(
    conn: sqlite3.Connection, company_id: int, user_id: int, registro_id: Any
) -> dict[str, Any]:
    conn.execute("DELETE FROM portfolio_coverage WHERE company_id = ? AND id = ?",
                 (company_id, int(registro_id)))
    audit_log(conn, company_id, user_id, "excluir", "portfolio_coverage", str(registro_id), {})
    conn.commit()
    return {"message": "Cobertura encerrada."}


def list_coverages(
    conn: sqlite3.Connection, company_id: int, user: sqlite3.Row
) -> list[dict[str, Any]]:
    permitidas = crm_allowed_units_for_user(conn, user)
    linhas = [dict(r) for r in conn.execute(
        "SELECT * FROM portfolio_coverage WHERE company_id = ? ORDER BY start_date DESC",
        (company_id,),
    ).fetchall()]
    if permitidas is None:
        return linhas
    return [l for l in linhas if not l.get("unit_name") or l["unit_name"] in permitidas]


def client_is_outside_own_portfolio(
    conn: sqlite3.Connection, company_id: int, user: sqlite3.Row, client_key: str
) -> bool:
    """O cliente pertence a outro vendedor?"""
    if data_scope_for_user(conn, user) != "proprio":
        return False
    dono = conn.execute(
        "SELECT NULLIF(TRIM(internal_seller_name), '') AS vendedor "
        "FROM crm_client_profiles WHERE company_id = ? AND client_code = ?",
        (company_id, client_key),
    ).fetchone()
    if not dono or not normalize_whitespace(dono["vendedor"]):
        return False
    return person_key(dono["vendedor"]) != person_key(seller_identity_for_user(user))


def support_client_view(
    conn: sqlite3.Connection, company_id: int, user: sqlite3.Row, codigo: str
) -> dict[str, Any] | None:
    """Ficha REDUZIDA de um cliente de outra carteira, por código exato.

    Duas escolhas de desenho, ambas deliberadas:

    1. Só código EXATO. Aceitar nome ou trecho deixaria qualquer vendedor varrer
       a carteira do colega em minutos. Pelo código, ele só chega em quem já
       está falando com ele — é o cliente que informa o código.
    2. Sem valores. Vai o necessário para ATENDER (telefone, contato, endereço,
       última compra, retornos em aberto, histórico de contatos). Não vai
       faturamento, média, classe nem margem: isso é resultado do colega e
       alimenta comparação e comissão.
    """
    code = normalize_whitespace(codigo)
    if not code:
        return None
    perfil = conn.execute(
        "SELECT client_code, client_name, trade_name, document_number, phone, updated_phone, "
        "       primary_contact_name, contact_notes, address_line, address_number, neighborhood, "
        "       city_name, state_name, postal_code, last_sale_at, "
        "       NULLIF(TRIM(internal_seller_name), '') AS owner_name "
        "FROM crm_client_profiles WHERE company_id = ? AND TRIM(client_code) = ?",
        (company_id, code),
    ).fetchone()
    if not perfil:
        return None

    dados = dict(perfil)
    dados["isOwnClient"] = not client_is_outside_own_portfolio(conn, company_id, user, code)
    dados["interactions"] = [dict(r) for r in conn.execute(
        "SELECT i.occurred_at, i.seller_name, i.notes, i.initiative, "
        "       t.label AS type_label, r.label AS result_label "
        "FROM crm_interactions i "
        "LEFT JOIN crm_contact_types t ON t.code = i.contact_type_code "
        "LEFT JOIN crm_contact_results r ON r.code = i.result_code "
        "WHERE i.company_id = ? AND i.client_key = ? "
        "ORDER BY i.occurred_at DESC LIMIT 10",
        (company_id, code),
    ).fetchall()]
    dados["openTasks"] = [dict(r) for r in conn.execute(
        "SELECT title, due_at, seller_name FROM crm_tasks "
        "WHERE company_id = ? AND client_key = ? AND status IN ('ABERTA','ATRASADA','REAGENDADA') "
        "ORDER BY due_at LIMIT 5",
        (company_id, code),
    ).fetchall()]
    return dados


def create_crm_interaction(
    conn: sqlite3.Connection, company_id: int, user: sqlite3.Row, payload: dict[str, Any]
) -> dict[str, Any]:
    client_key = normalize_client_key(payload.get("clientKey"))
    client_name = normalize_whitespace(payload.get("clientName"))
    contact_phone = normalize_whitespace(payload.get("updatedPhone") or payload.get("contactPhone"))
    contact_name = normalize_whitespace(payload.get("primaryContactName") or payload.get("contactName"))
    contact_type_code = normalize_upper(payload.get("contactTypeCode"))
    result_code = normalize_upper(payload.get("resultCode"))
    occurred_at = normalize_whitespace(payload.get("occurredAt")) or now_iso()
    notes = normalize_whitespace(payload.get("notes"))
    next_action = normalize_whitespace(payload.get("nextAction"))
    followup_due_at = normalize_whitespace(payload.get("followupDueAt"))
    if not client_key or not client_name:
        raise ValueError("Cliente invalido para registro de interacao")
    if not contact_type_code:
        raise ValueError("Tipo de contato obrigatorio")
    # RECEPTIVO: o cliente procurou, ou é uma anotação sobre ele. Entra no
    # histórico e na ficha, mas fica fora da meta de ligações ativas — senão
    # bastaria anotar o que chegou sozinho para "bater" a meta sem prospectar.
    iniciativa = (normalize_upper(payload.get("initiative")) or INITIATIVE_ACTIVE)
    if iniciativa not in {INITIATIVE_ACTIVE, INITIATIVE_RECEPTIVE, INITIATIVE_SUPPORT}:
        iniciativa = INITIATIVE_ACTIVE
    if iniciativa == INITIATIVE_SUPPORT and not client_is_outside_own_portfolio(
            conn, company_id, user, client_key):
        # Cliente é da carteira dele: apoio não se aplica, é atendimento normal.
        iniciativa = INITIATIVE_ACTIVE
    if contact_type_code in CRM_RECEPTIVE_TYPES:
        iniciativa = INITIATIVE_RECEPTIVE
    if iniciativa == INITIATIVE_RECEPTIVE and not result_code:
        # No registro receptivo o resultado não é a pergunta certa: ninguém
        # "não atendeu" uma mensagem que chegou. O padrão evita um campo a mais
        # numa tela que precisa ser de dois cliques.
        result_code = "OUTRO"
    if not result_code:
        raise ValueError("Resultado do contato obrigatorio")
    if not notes:
        raise ValueError("Observacao obrigatoria")
    contact_type = conn.execute("SELECT code FROM crm_contact_types WHERE code = ? AND is_active = 1", (contact_type_code,)).fetchone()
    if not contact_type:
        raise ValueError("Tipo de contato invalido")
    result = conn.execute("SELECT * FROM crm_contact_results WHERE code = ? AND is_active = 1", (result_code,)).fetchone()
    if not result:
        raise ValueError("Resultado do contato invalido")
    if result["requires_followup_date"] and not followup_due_at:
        raise ValueError("Este resultado exige data de retorno")
    seller_name = seller_identity_for_user(user)
    unit_name = normalize_unit(payload.get("unitName"))
    if contact_phone or contact_name:
        save_crm_client_contact(
            conn,
            company_id,
            user,
            {
                "clientKey": client_key,
                "clientName": client_name,
                "updatedPhone": contact_phone,
                "primaryContactName": contact_name,
                "notes": normalize_whitespace(payload.get("contactNotes")),
            },
        )
    cursor = conn.execute(
        """
        INSERT INTO crm_interactions (
            company_id, client_key, client_name, seller_name, unit_name,
            contact_phone, contact_name,
            contact_type_code, result_code, occurred_at, notes, question_used,
            had_progress, offer_title, next_action, followup_due_at, initiative,
            created_at, created_by_user_id
        ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
        """,
        (
            company_id,
            client_key,
            client_name,
            seller_name,
            unit_name,
            contact_phone or None,
            contact_name or None,
            contact_type_code,
            result_code,
            occurred_at,
            notes,
            normalize_whitespace(payload.get("questionUsed")),
            1 if payload.get("hadProgress") else 0,
            normalize_whitespace(payload.get("offerTitle")),
            next_action,
            followup_due_at or None,
            iniciativa,
            now_iso(),
            user["id"],
        ),
    )
    interaction_id = cursor.lastrowid

    # Contato em prospect: o status é derivado do que aconteceu, então precisa
    # ser recalculado aqui. Antes só era recalculado ao SALVAR o prospect — por
    # isso a oficina seguia "a contatar" mesmo depois da ligação registrada.
    if client_key.startswith(PROSPECT_KEY_PREFIX):
        try:
            _refresh_prospect_status(conn, company_id, int(client_key[len(PROSPECT_KEY_PREFIX):]))
        except (ValueError, sqlite3.Error) as exc:
            print(f"[prospects] status não atualizado para {client_key}: {exc}", flush=True)

    # Registrar o contato encerra as tarefas abertas daquele cliente para aquele
    # vendedor. Sem isso o retorno continuava aparecendo em "Retornos de hoje"
    # depois do vendedor já ter feito a ligação — mesma queixa do TOP 5.
    conn.execute(
        """
        UPDATE crm_tasks SET status = 'CONCLUIDA', completed_at = ?
        WHERE company_id = ? AND client_key = ? AND UPPER(seller_name) = ?
          AND status IN ('ABERTA', 'ATRASADA', 'REAGENDADA')
        """,
        (now_iso(), company_id, client_key, normalize_upper(seller_name)),
    )

    # Atendimento de apoio: o DONO precisa saber que alguém falou com o cliente
    # dele, senão o atendimento morre ali e o cliente fica sem continuidade.
    # A tarefa é o canal que ele já usa todo dia — não inventa notificação nova.
    if iniciativa == INITIATIVE_SUPPORT:
        dono = conn.execute(
            "SELECT NULLIF(TRIM(internal_seller_name), '') AS vendedor "
            "FROM crm_client_profiles WHERE company_id = ? AND client_code = ?",
            (company_id, client_key),
        ).fetchone()
        nome_dono = normalize_whitespace(dono["vendedor"]) if dono else ""
        if nome_dono:
            conn.execute(
                """
                INSERT INTO crm_tasks
                    (company_id, client_key, client_name, seller_name, title, description,
                     due_at, status, origin, priority, created_by_name, created_at)
                VALUES (?,?,?,?,?,?,?, 'ABERTA', 'APOIO', 'ALTA', ?, ?)
                """,
                (company_id, client_key, client_name, nome_dono,
                 f"Cliente atendido por {seller_name}",
                 f"{seller_name} atendeu este cliente da sua carteira em "
                 f"{occurred_at[:10]}. Observação: {notes}",
                 today_in_brazil().isoformat(), seller_name, now_iso()),
            )

    task_id = None
    if result["generates_followup"] and followup_due_at:
        task_cursor = conn.execute(
            """
            INSERT INTO crm_tasks (
                company_id, client_key, client_name, seller_name, title, description,
                due_at, status, origin, source_interaction_id, created_at
            ) VALUES (?, ?, ?, ?, ?, ?, ?, 'ABERTA', 'FOLLOWUP', ?, ?)
            """,
            (
                company_id,
                client_key,
                client_name,
                seller_name,
                next_action or result["label"],
                notes,
                followup_due_at,
                interaction_id,
                now_iso(),
            ),
        )
        task_id = task_cursor.lastrowid
    conn.commit()
    return {"interactionId": interaction_id, "taskId": task_id}


def save_crm_agenda_action(
    conn: sqlite3.Connection, company_id: int, user: sqlite3.Row, payload: dict[str, Any]
) -> dict[str, Any]:
    client_key = normalize_client_key(payload.get("clientKey"))
    client_name = normalize_whitespace(payload.get("clientName"))
    action_type = normalize_upper(payload.get("actionType"))
    justification = normalize_whitespace(payload.get("justification"))
    next_visible_at = normalize_whitespace(payload.get("nextVisibleAt"))
    if not client_key or not client_name:
        raise ValueError("Cliente invalido para acao de agenda")
    if action_type not in {"ADIAR", "REORDENAR"}:
        raise ValueError("Acao de agenda invalida")
    if not justification:
        raise ValueError("Justificativa obrigatoria")
    seller_name = seller_identity_for_user(user)
    cursor = conn.execute(
        """
        INSERT INTO crm_agenda_actions (
            company_id, seller_name, client_key, client_name, action_type,
            justification, next_visible_at, created_at, created_by_user_id
        ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)
        """,
        (
            company_id,
            seller_name,
            client_key,
            client_name,
            action_type,
            justification,
            next_visible_at or None,
            now_iso(),
            user["id"],
        ),
    )
    conn.commit()
    return {"actionId": cursor.lastrowid}


def complete_crm_task(
    conn: sqlite3.Connection, company_id: int, user: sqlite3.Row, task_id: int
) -> None:
    seller_name = seller_identity_for_user(user)
    row = conn.execute(
        """
        SELECT id FROM crm_tasks
        WHERE id = ? AND company_id = ? AND (? != 'Vendedor' OR seller_name = ?)
        """,
        (task_id, company_id, user["role"], seller_name),
    ).fetchone()
    if not row:
        raise ValueError("Tarefa nao encontrada")
    conn.execute(
        "UPDATE crm_tasks SET status = 'CONCLUIDA', completed_at = ? WHERE id = ?",
        (now_iso(), task_id),
    )
    conn.commit()


def reschedule_crm_task(
    conn: sqlite3.Connection, company_id: int, user: sqlite3.Row, task_id: int, due_at: str
) -> None:
    seller_name = seller_identity_for_user(user)
    row = conn.execute(
        """
        SELECT id FROM crm_tasks
        WHERE id = ? AND company_id = ? AND (? != 'Vendedor' OR seller_name = ?)
        """,
        (task_id, company_id, user["role"], seller_name),
    ).fetchone()
    if not row:
        raise ValueError("Tarefa nao encontrada")
    if not normalize_whitespace(due_at):
        raise ValueError("Nova data obrigatoria")
    conn.execute(
        "UPDATE crm_tasks SET status = 'REAGENDADA', due_at = ?, completed_at = NULL WHERE id = ?",
        (due_at, task_id),
    )
    conn.commit()


CORPORATE_HINTS = {
    "LTDA",
    "ME",
    "MEI",
    "EIRELI",
    "EPP",
    "SA",
    "S A",
    "S/A",
    "AUTO PECAS",
    "AUTOPECAS",
    "DISTRIBUIDORA",
    "COMERCIO",
    "COMERCIAL",
    "INDUSTRIA",
    "INDUSTRIAL",
    "OFICINA",
    "MECANICA",
    "CENTRO AUTOMOTIVO",
    "SERVICOS",
    "SERVICOS AUTOMOTIVOS",
    "PECAS",
    "TRANSPORTES",
    "BORRACHARIA",
    "POSTO",
    "RESTAURANTE",
    "MERCADO",
    "SUPERMERCADO",
    "FERRAGENS",
}


@functools.lru_cache(maxsize=4096)
def infer_person_type_from_name(client_name: str | None) -> tuple[str, float, str]:
    normalized = normalize_client_key(client_name)
    if not normalized:
        return "PF", 0.3, "nome_vazio"
    if any(hint in normalized for hint in CORPORATE_HINTS):
        return "PJ", 0.9, "palavra_corporativa"
    if re.search(r"\d", normalized):
        return "PJ", 0.85, "nome_com_numero"
    tokens = [token for token in normalized.split() if token]
    if len(tokens) >= 2 and all(token.isalpha() for token in tokens):
        return "PF", 0.7, "nome_pessoal"
    if len(tokens) == 1 and tokens[0].isalpha():
        return "PF", 0.55, "nome_curto"
    return "PJ", 0.6, "heuristica_empresa"


def person_type_from_document(document_value: str | None) -> tuple[str | None, str]:
    digits = re.sub(r"\D", "", document_value or "")
    if len(digits) == 11:
        return "PF", digits
    if len(digits) == 14:
        return "PJ", digits
    return None, digits


def upsert_client_registry_row(
    conn: sqlite3.Connection,
    company_id: int,
    client_name: str,
    document_number: str | None,
    person_type: str,
    source: str,
    confidence_score: float,
    notes: str | None = None,
) -> None:
    normalized_client_name = normalize_client_key(client_name)
    _, document_digits = person_type_from_document(document_number)
    now = now_iso()
    conn.execute(
        """
        INSERT INTO client_registry
            (company_id, client_name, normalized_client_name, document_number, document_digits, person_type, source, confidence_score, notes, created_at, updated_at)
        VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
        ON CONFLICT(company_id, normalized_client_name) DO UPDATE SET
            client_name = excluded.client_name,
            document_number = excluded.document_number,
            document_digits = excluded.document_digits,
            person_type = excluded.person_type,
            source = excluded.source,
            confidence_score = excluded.confidence_score,
            notes = excluded.notes,
            updated_at = excluded.updated_at
        """,
        (
            company_id,
            normalize_whitespace(client_name),
            normalized_client_name,
            normalize_whitespace(document_number),
            document_digits,
            person_type,
            source,
            confidence_score,
            notes,
            now,
            now,
        ),
    )


def ensure_client_registry_for_sales(conn: sqlite3.Connection, company_id: int) -> None:
    existing = {
        row["normalized_client_name"]
        for row in conn.execute("SELECT normalized_client_name FROM client_registry WHERE company_id = ?", (company_id,)).fetchall()
    }
    missing_rows = conn.execute(
        """
        SELECT DISTINCT client_name
        FROM fact_sales_detail
        WHERE company_id = ? AND client_name IS NOT NULL AND TRIM(client_name) <> ''
        """,
        (company_id,),
    ).fetchall()
    created = 0
    for row in missing_rows:
        client_name = normalize_whitespace(row["client_name"])
        normalized = normalize_client_key(client_name)
        if not normalized or normalized in existing:
            continue
        person_type, confidence, reason = infer_person_type_from_name(client_name)
        upsert_client_registry_row(conn, company_id, client_name, None, person_type, "heuristica", confidence, reason)
        existing.add(normalized)
        created += 1
    if created:
        conn.commit()


def get_dashboard_data(conn: sqlite3.Connection, company_id: int, filters: dict[str, str | None]) -> dict[str, Any]:
    competences = query_competences(conn, company_id)
    primary_competence = selected_primary_competence(filters, competences)
    primary_competence = primary_competence or date.today().strftime("%Y-%m")
    competence_state = dashboard_competence_state(primary_competence)
    cutoff_date = competence_state["cutoffDate"]
    allowed_units = normalize_unit_list(filters.get("allowed_units"))
    scoped_units = [normalize_unit(filters["unit_name"])] if filters.get("unit_name") else allowed_units

    detail_conditions = ["company_id = ?", "competence = ?"]
    detail_params: list[Any] = [company_id, primary_competence]
    summary_vendor_conditions = ["company_id = ?", "competence = ?"]
    summary_vendor_params: list[Any] = [company_id, primary_competence]
    summary_unit_conditions = ["company_id = ?", "competence = ?"]
    summary_unit_params: list[Any] = [company_id, primary_competence]

    if filters["seller_name"]:
        detail_conditions.append("seller_name = ?")
        detail_params.append(filters["seller_name"])
        summary_vendor_conditions.append("seller_name = ?")
        summary_vendor_params.append(filters["seller_name"])

    detail_source_rows = conn.execute(
        f"""
        SELECT seller_name, client_name, city_name, sku_key, gross_value, discount_value, return_value, net_value, competence, sale_share, issue_date
        FROM fact_sales_detail
        WHERE {' AND '.join(detail_conditions)}
        """,
        detail_params,
    ).fetchall()

    # ── PRÉ-CARREGAR LOOKUPS (elimina N+1 queries) ──────────────────────────
    _comp_target = first_day_of_competence(primary_competence).isoformat()

    _city_map: dict[str, str | None] = {}
    for _r in conn.execute(
        """
        SELECT city_name, principal_unit
        FROM city_mappings
        WHERE company_id = ? AND date(valid_from) <= date(?)
          AND (valid_to IS NULL OR date(valid_to) >= date(?))
        ORDER BY date(valid_from) DESC
        """,
        (company_id, _comp_target, _comp_target),
    ).fetchall():
        _city_name_key = normalize_upper(_r["city_name"])
        if _city_name_key and _city_name_key not in _city_map:
            _city_map[_city_name_key] = normalize_unit(_r["principal_unit"])

    _seller_unit_map: dict[str, str | None] = {}
    _seller_role_map: dict[str, str | None] = {}
    for _r in conn.execute(
        """
        SELECT person_name, base_unit, role_classification
        FROM people_records
        WHERE company_id = ? AND date(valid_from) <= date(?)
          AND (valid_to IS NULL OR date(valid_to) >= date(?))
        ORDER BY date(valid_from) DESC
        """,
        (company_id, _comp_target, _comp_target),
    ).fetchall():
        _pname_key = normalize_whitespace(_r["person_name"])
        if _pname_key and _pname_key not in _seller_unit_map:
            _seller_unit_map[_pname_key] = normalize_unit(_r["base_unit"])
            _seller_role_map[_pname_key] = _r["role_classification"]
    # ────────────────────────────────────────────────────────────────────────

    detail_rows_all: list[dict[str, Any]] = []
    for row in detail_source_rows:
        issue_dt = parse_datetime_flexible(row["issue_date"])
        if competence_state["isFutureCompetence"]:
            continue
        if competence_state["isCurrentCompetence"]:
            # Exclui só se a data é conhecida E está no futuro; issue_date nulo = inclui (confia na competência)
            if issue_dt and issue_dt.date() > cutoff_date:
                continue
        seller_name = normalize_whitespace(row["seller_name"])
        city_name = normalize_upper(row["city_name"])
        resolved_unit = _city_map.get(city_name)
        seller_base_unit = _seller_unit_map.get(seller_name)
        enriched = dict(row)
        enriched["seller_name"] = seller_name
        enriched["client_name"] = normalize_whitespace(row["client_name"])
        enriched["city_name"] = city_name
        enriched["resolved_unit"] = resolved_unit
        enriched["seller_base_unit"] = normalize_unit(seller_base_unit)
        detail_rows_all.append(enriched)

    if filters["city_name"]:
        detail_rows_all = [row for row in detail_rows_all if row["city_name"] == filters["city_name"]]

    detail_rows_scope = detail_rows_all
    if scoped_units:
        detail_rows_scope = [row for row in detail_rows_scope if row["resolved_unit"] in scoped_units]
    if filters["unit_name"]:
        detail_rows_scope = [row for row in detail_rows_scope if row["resolved_unit"] == filters["unit_name"]]

    seller_detail_rows = detail_rows_all
    if scoped_units:
        seller_detail_rows = [row for row in seller_detail_rows if row["seller_base_unit"] in scoped_units]
    if filters["city_name"]:
        seller_detail_rows = [row for row in seller_detail_rows if row["city_name"] == filters["city_name"]]
    if filters["unit_name"]:
        seller_detail_rows = [row for row in seller_detail_rows if row["seller_base_unit"] == filters["unit_name"]]

    if scoped_units:
        placeholders = ", ".join("?" for _ in scoped_units)
        summary_unit_conditions.append(f"unit_name IN ({placeholders})")
        summary_unit_params.extend(scoped_units)
    if filters["unit_name"]:
        summary_unit_conditions.append("unit_name = ?")
        summary_unit_params.append(filters["unit_name"])

    vendor_summary_rows = [dict(row) for row in conn.execute(
        f"""
        SELECT
            seller_name,
            SUM(qty_sold) AS qty_sold,
            SUM(cost_value) AS cost_value,
            SUM(sale_value) AS sale_value,
            SUM(profit_value) AS profit_value,
            SUM(net_profit_value) AS net_profit_value,
            AVG(profit_pct) AS profit_pct,
            SUM(return_cost) AS return_cost,
            SUM(return_value) AS return_value,
            SUM(net_value) AS net_value,
            AVG(margin_value) AS margin_value
        FROM fact_vendor_summary
        WHERE {' AND '.join(summary_vendor_conditions)}
        GROUP BY seller_name
        ORDER BY net_value DESC
        """,
        summary_vendor_params,
    ).fetchall()]
    unit_summary_rows = [dict(row) for row in conn.execute(
        f"""
        SELECT
            unit_name,
            SUM(qty_sold) AS qty_sold,
            SUM(cost_value) AS cost_value,
            SUM(sale_value) AS sale_value,
            SUM(profit_value) AS profit_value,
            SUM(net_profit_value) AS net_profit_value,
            AVG(profit_pct) AS profit_pct,
            SUM(return_cost) AS return_cost,
            SUM(return_value) AS return_value,
            SUM(net_value) AS net_value,
            AVG(margin_value) AS margin_value
        FROM fact_unit_summary
        WHERE {' AND '.join(summary_unit_conditions)}
        GROUP BY unit_name
        ORDER BY net_value DESC
        """,
        summary_unit_params,
    ).fetchall()]

    vendor_summary_by_seller = {normalize_whitespace(row["seller_name"]): row for row in vendor_summary_rows}
    unit_summary_by_unit = {normalize_unit(row["unit_name"]): row for row in unit_summary_rows}

    goal_by_seller, duplicate_seller_goals = load_goal_maps(
        conn,
        company_id,
        primary_competence,
        "goals_seller",
        "seller_name",
        normalize_whitespace,
    )
    goal_by_unit, duplicate_unit_goals = load_goal_maps(
        conn,
        company_id,
        primary_competence,
        "goals_unit",
        "unit_name",
        normalize_unit,
    )
    if scoped_units:
        goal_by_unit = {unit_name: values for unit_name, values in goal_by_unit.items() if normalize_unit(unit_name) in scoped_units}
    total_unit_goal = float(sum(item["revenueGoal"] or 0 for item in goal_by_unit.values()))
    total_unit_returns_goal = float(sum(item["returnsGoal"] or 0 for item in goal_by_unit.values()))

    client_registry = {
        row["normalized_client_name"]: dict(row)
        for row in conn.execute(
            "SELECT client_name, normalized_client_name, document_number, person_type, source, confidence_score, notes FROM client_registry WHERE company_id = ?",
            (company_id,),
        ).fetchall()
    }

    # ── Código do cliente e histórico trimestral ──────────────────────────────
    # O ranking nasce do faturamento, que só tem o NOME. O código vem do cadastro
    # e é o que permite abrir a ficha. O trimestre anterior serve de base para
    # dizer se o cliente está crescendo, estável ou caindo.
    client_code_map: dict[str, str] = {}
    for _c in conn.execute(
        "SELECT client_code, client_name FROM crm_client_profiles WHERE company_id = ?",
        (company_id,),
    ).fetchall():
        _k = normalize_client_key(_c["client_name"])
        _code = normalize_whitespace(_c["client_code"])
        if _k and _code and _k not in client_code_map:
            client_code_map[_k] = _code

    _pc1 = shift_competence(primary_competence, -1)
    _pc2 = shift_competence(primary_competence, -2)
    _pc3 = shift_competence(primary_competence, -3)
    client_prev_quarter: dict[str, float] = {}
    for _r in conn.execute(
        """
        SELECT client_name, SUM(net_value) AS total
        FROM fact_sales_detail
        WHERE company_id = ? AND competence IN (?, ?, ?)
        GROUP BY client_name
        """,
        (company_id, _pc1, _pc2, _pc3),
    ).fetchall():
        _k = normalize_client_key(_r["client_name"])
        if _k:
            client_prev_quarter[_k] = float(_r["total"] or 0.0)

    # ── Devoluções em garantia ────────────────────────────────────────────────
    # Vêm dentro do total de devoluções do custo/venda, mas são defeito de peça,
    # não erro de venda. Ficam separadas para não penalizar o resultado comercial.
    warranty_by_seller: dict[str, float] = {}
    for _w in conn.execute(
        "SELECT seller_name, SUM(total_value) AS total FROM fact_warranty_returns "
        "WHERE company_id = ? AND competence = ? GROUP BY seller_name",
        (company_id, primary_competence),
    ).fetchall():
        warranty_by_seller[normalize_whitespace(_w["seller_name"])] = float(_w["total"] or 0.0)

    warranty_by_unit: dict[str, float] = {}
    for _w in conn.execute(
        "SELECT unit_name, SUM(total_value) AS total FROM fact_warranty_returns "
        "WHERE company_id = ? AND competence = ? GROUP BY unit_name",
        (company_id, primary_competence),
    ).fetchall():
        warranty_by_unit[normalize_unit(_w["unit_name"])] = float(_w["total"] or 0.0)

    # ── Dono do cliente no cadastro CRM ───────────────────────────────────────
    # Permite separar o que o vendedor faturou da PRÓPRIA carteira do que faturou
    # de clientes de outros (ou sem dono). Chave é o nome normalizado, porque o
    # faturamento não traz o código do cliente.
    client_owner_map: dict[str, str] = {}
    for _o in conn.execute(
        "SELECT client_name, internal_seller_name, external_seller_name "
        "FROM crm_client_profiles WHERE company_id = ?",
        (company_id,),
    ).fetchall():
        _key = normalize_client_key(_o["client_name"])
        if not _key:
            continue
        # Só o vendedor interno define dono, como no resto do sistema.
        _owner = normalize_whitespace(_o["internal_seller_name"])
        if _owner:
            client_owner_map[_key] = _owner

    city_metrics: dict[str, dict[str, Any]] = defaultdict(lambda: {"revenueNet": 0.0, "grossSales": 0.0, "grossSalesPct": 0.0, "discountValue": 0.0, "returnValue": 0.0, "clients": set()})
    detail_by_seller: dict[str, dict[str, Any]] = defaultdict(lambda: {
        "revenueNet": 0.0, "grossSales": 0.0, "grossSalesPct": 0.0, "discountValue": 0.0,
        "returnValue": 0.0, "clients": set(), "sku": set(), "baseUnit": None,
        # Carteira própria x fora da carteira
        "ownClients": set(), "otherClients": set(),
        "ownRevenue": 0.0, "otherRevenue": 0.0,
    })
    detail_by_client: dict[str, dict[str, Any]] = defaultdict(lambda: {"revenueNet": 0.0, "grossSales": 0.0, "grossSalesPct": 0.0, "discountValue": 0.0, "returnValue": 0.0, "cities": set(), "personType": None, "typeSource": None, "typeConfidence": 0.0})
    detail_by_unit: dict[str, dict[str, Any]] = defaultdict(lambda: {"revenueNet": 0.0, "grossSales": 0.0, "grossSalesPct": 0.0, "discountValue": 0.0, "returnValue": 0.0, "clients": set()})
    client_top_by_unit_source: dict[str, dict[str, dict[str, Any]]] = defaultdict(lambda: defaultdict(lambda: {"revenueNet": 0.0, "grossSales": 0.0, "grossSalesPct": 0.0, "discountValue": 0.0, "returnValue": 0.0}))

    for row in detail_rows_scope:
        city_name = row["city_name"]
        seller_name = row["seller_name"]
        client_name = row["client_name"]
        resolved_unit = row["resolved_unit"] or "NAO_MAPEADO"
        net_value = float(row["net_value"] or 0)
        gross_value = float(row["gross_value"] or 0)
        gross_pct_base = max(gross_value, 0.0)
        discount_value = float(row["discount_value"] or 0)
        return_value = float(row["return_value"] or 0)
        if city_name:
            city_metrics[city_name]["revenueNet"] += net_value
            city_metrics[city_name]["grossSales"] += gross_value
            city_metrics[city_name]["grossSalesPct"] += gross_pct_base
            city_metrics[city_name]["discountValue"] += discount_value
            city_metrics[city_name]["returnValue"] += return_value
            if client_name:
                city_metrics[city_name]["clients"].add(client_name)
        if client_name:
            client_key = normalize_client_key(client_name)
            registry_row = client_registry.get(client_key)
            if registry_row:
                person_type = registry_row["person_type"]
                type_source = registry_row["source"]
                type_confidence = float(registry_row["confidence_score"] or 0)
            else:
                person_type, type_confidence, type_source = infer_person_type_from_name(client_name)
            detail_by_client[client_name]["revenueNet"] += net_value
            detail_by_client[client_name]["grossSales"] += gross_value
            detail_by_client[client_name]["grossSalesPct"] += gross_pct_base
            detail_by_client[client_name]["discountValue"] += discount_value
            detail_by_client[client_name]["returnValue"] += return_value
            detail_by_client[client_name]["personType"] = person_type
            detail_by_client[client_name]["typeSource"] = type_source
            detail_by_client[client_name]["typeConfidence"] = type_confidence
            if city_name:
                detail_by_client[client_name]["cities"].add(city_name)
            client_top_by_unit_source[resolved_unit][client_name]["revenueNet"] += net_value
            client_top_by_unit_source[resolved_unit][client_name]["grossSales"] += gross_value
            client_top_by_unit_source[resolved_unit][client_name]["grossSalesPct"] += gross_pct_base
            client_top_by_unit_source[resolved_unit][client_name]["discountValue"] += discount_value
            client_top_by_unit_source[resolved_unit][client_name]["returnValue"] += return_value
        detail_by_unit[resolved_unit]["revenueNet"] += net_value
        detail_by_unit[resolved_unit]["grossSales"] += gross_value
        detail_by_unit[resolved_unit]["grossSalesPct"] += gross_pct_base
        detail_by_unit[resolved_unit]["discountValue"] += discount_value
        detail_by_unit[resolved_unit]["returnValue"] += return_value
        if client_name:
            detail_by_unit[resolved_unit]["clients"].add(client_name)

    for row in seller_detail_rows:
        seller_name = row["seller_name"]
        client_name = row["client_name"]
        net_value = float(row["net_value"] or 0)
        gross_value = float(row["gross_value"] or 0)
        gross_pct_base = max(gross_value, 0.0)
        discount_value = float(row["discount_value"] or 0)
        return_value = float(row["return_value"] or 0)
        detail_by_seller[seller_name]["revenueNet"] += net_value
        detail_by_seller[seller_name]["grossSales"] += gross_value
        detail_by_seller[seller_name]["grossSalesPct"] += gross_pct_base
        detail_by_seller[seller_name]["discountValue"] += discount_value
        detail_by_seller[seller_name]["returnValue"] += return_value
        detail_by_seller[seller_name]["baseUnit"] = row["seller_base_unit"]
        if client_name:
            detail_by_seller[seller_name]["clients"].add(client_name)
            # Cliente é da carteira do vendedor se o cadastro CRM aponta para ele
            owner = client_owner_map.get(normalize_client_key(client_name))
            if owner and normalize_upper(owner) == normalize_upper(seller_name):
                detail_by_seller[seller_name]["ownClients"].add(client_name)
                detail_by_seller[seller_name]["ownRevenue"] += net_value
            else:
                detail_by_seller[seller_name]["otherClients"].add(client_name)
                detail_by_seller[seller_name]["otherRevenue"] += net_value
        if row["sku_key"]:
            detail_by_seller[seller_name]["sku"].add(row["sku_key"])

    city_ranking = []
    for city_name, metrics in city_metrics.items():
        distinct_clients = len(metrics["clients"])
        city_ranking.append(
            {
                "cityName": city_name,
                "revenueNet": round(metrics["revenueNet"], 2),
                "ticketAverage": round(safe_div(metrics["revenueNet"], distinct_clients), 2),
                "distinctClients": distinct_clients,
                "discountValue": round(metrics["discountValue"], 2),
                "discountPct": round(safe_div(metrics["discountValue"], metrics["grossSalesPct"]) * 100, 2),
                "returnsValue": round(metrics["returnValue"], 2),
            }
        )
    city_ranking.sort(key=lambda item: item["revenueNet"], reverse=True)

    client_ranking = []
    client_type_summary: dict[str, dict[str, Any]] = defaultdict(lambda: {"revenueNet": 0.0, "clients": 0})
    for client_name, metrics in detail_by_client.items():
        person_type = metrics["personType"] or "Nao classificado"
        client_type_summary[person_type]["revenueNet"] += metrics["revenueNet"]
        client_type_summary[person_type]["clients"] += 1
        # Comparativo com a média mensal do trimestre anterior
        _ckey = normalize_client_key(client_name)
        _prev_total = client_prev_quarter.get(_ckey, 0.0)
        _prev_avg = _prev_total / 3
        _current = float(metrics["revenueNet"] or 0.0)
        _variation = safe_div(_current - _prev_avg, _prev_avg) * 100 if _prev_avg > 0 else None
        client_ranking.append(
            {
                "clientName": client_name,
                # Código do cadastro — sem ele não dá para abrir a ficha
                "clientKey": client_code_map.get(_ckey),
                "personType": person_type,
                "typeSource": metrics["typeSource"],
                "typeConfidence": round(float(metrics["typeConfidence"] or 0), 2),
                "revenueNet": round(_current, 2),
                "discountValue": round(metrics["discountValue"], 2),
                "discountPct": round(safe_div(metrics["discountValue"], metrics["grossSalesPct"]) * 100, 2),
                "returnsValue": round(metrics["returnValue"], 2),
                "citiesCount": len(metrics["cities"]),
                # Tendência contra a média mensal dos 3 meses anteriores
                "quarterAverage": round(_prev_avg, 2),
                "quarterTotal": round(_prev_total, 2),
                "quarterVariationPct": round(_variation, 1) if _variation is not None else None,
                "quarterMonths": [_pc3, _pc2, _pc1],
            }
        )
    client_ranking.sort(key=lambda item: item["revenueNet"], reverse=True)

    client_top_by_unit = []
    for unit_name, clients in client_top_by_unit_source.items():
        if scoped_units and unit_name not in scoped_units:
            continue
        top_clients = []
        for client_name, metrics in clients.items():
            client_registry_row = client_registry.get(normalize_client_key(client_name))
            if client_registry_row:
                person_type = client_registry_row["person_type"]
            else:
                person_type, _, _ = infer_person_type_from_name(client_name)
            top_clients.append(
                {
                    "clientName": client_name,
                    "personType": person_type,
                    "revenueNet": round(metrics["revenueNet"], 2),
                    "discountValue": round(metrics["discountValue"], 2),
                    "discountPct": round(safe_div(metrics["discountValue"], metrics["grossSalesPct"]) * 100, 2),
                    "returnsValue": round(metrics["returnValue"], 2),
                }
            )
        top_clients.sort(key=lambda item: item["revenueNet"], reverse=True)
        client_top_by_unit.append({"unitName": unit_name, "clients": top_clients[:10]})
    client_top_by_unit.sort(key=lambda item: item["unitName"])

    summary_calendar = get_business_calendar(
        conn,
        company_id,
        primary_competence,
        reference_today=competence_state["today"],
        include_current_day=False,
    )
    # ── Faróis ────────────────────────────────────────────────────────────────
    # Percentual do mês já decorrido: base para avaliar ritmo em vez de valor cheio.
    farol_thresholds = load_kpi_thresholds(conn, company_id)
    # Unidades em implantação: carregadas de uma vez para não consultar por linha.
    # Sem meta, o farol precisa ficar neutro — pintar de vermelho quem ainda nem
    # inaugurou ensina a equipe a ignorar o painel.
    competencia_atual = filters.get("competence_end") or filters.get("competence_start") or ""
    unidades_implantacao = set()
    for _linha_fase in conn.execute(
        "SELECT unit_name, phase, goal_exempt_until FROM unit_phases WHERE company_id = ?",
        (company_id,),
    ).fetchall():
        _isento = _linha_fase["goal_exempt_until"]
        if _linha_fase["phase"] == "IMPLANTACAO" or (
            _isento and competencia_atual and competencia_atual <= _isento
        ):
            unidades_implantacao.add(normalize_unit(_linha_fase["unit_name"]))
    pace_pct = (
        safe_div(summary_calendar["elapsedWorkingDays"], summary_calendar["totalWorkingDays"]) * 100
        if summary_calendar["totalWorkingDays"] else 0.0
    )

    score_config = get_score_config(conn, company_id, primary_competence)
    max_ticket = 1.0
    max_clients = 1.0
    max_mix = 1.0
    candidate_sellers = []
    for seller_name, official_row in vendor_summary_by_seller.items():
        metrics = detail_by_seller.get(
            seller_name,
            {"revenueNet": 0.0, "grossSales": 0.0, "grossSalesPct": 0.0, "discountValue": 0.0, "returnValue": 0.0, "clients": set(), "sku": set(), "baseUnit": None},
        )
        role = _seller_role_map.get(seller_name)
        base_unit = _seller_unit_map.get(seller_name)
        resolved_base_unit = normalize_unit(base_unit or metrics.get("baseUnit"))
        if role not in (None, "Vendedor"):
            continue
        if scoped_units and resolved_base_unit not in scoped_units:
            continue
        if filters["unit_name"] and resolved_base_unit != filters["unit_name"]:
            continue
        official_revenue_net = float(official_row.get("net_value") or 0.0)
        ticket = safe_div(official_revenue_net, len(metrics["clients"]))
        max_ticket = max(max_ticket, ticket)
        max_clients = max(max_clients, len(metrics["clients"]))
        max_mix = max(max_mix, len(metrics["sku"]))
        candidate_sellers.append((seller_name, metrics, official_row, role, resolved_base_unit))

    seller_rows = []
    total_company_seller_goal = 0.0
    total_company_seller_returns_goal = 0.0
    for seller_name, metrics, official_row, role, base_unit in candidate_sellers:
        gross_sales = float(official_row.get("sale_value") or 0.0)
        revenue_net_raw = float(official_row.get("net_value") or 0.0)
        gross_sales_pct = float(metrics.get("grossSalesPct") or 0.0)
        returns_total = float(official_row.get("return_value") or 0.0)
        # Devolução em garantia é defeito de peça, não erro de venda: sai do
        # resultado comercial e é devolvida ao líquido do vendedor.
        warranty_value = min(warranty_by_seller.get(seller_name, 0.0), returns_total)
        returns_value = max(returns_total - warranty_value, 0.0)
        revenue_net = revenue_net_raw + warranty_value
        qty_sold = float(official_row.get("qty_sold") or 0)
        return_cost = float(official_row.get("return_cost") or 0.0)
        cost_value = float(official_row.get("cost_value") or 0.0)
        profit_value = float(official_row.get("profit_value") or 0.0)
        net_profit_value = float(official_row.get("net_profit_value") or 0.0)
        distinct_clients = len(metrics["clients"])
        mix_count = len(metrics["sku"])
        ticket = safe_div(revenue_net, distinct_clients)
        ticket_per_piece = safe_div(revenue_net, qty_sold)
        seller_goal = goal_by_seller.get(seller_name, {"revenueGoal": 0.0, "returnsGoal": 0.0})
        revenue_goal = float(seller_goal["revenueGoal"] or 0)
        returns_goal = float(seller_goal["returnsGoal"] or 0)
        total_company_seller_goal += revenue_goal
        total_company_seller_returns_goal += returns_goal
        goal_attainment = safe_div(revenue_net, revenue_goal) * 100 if revenue_goal else 0.0
        # Calendário individual do vendedor (ajusta por férias)
        seller_calendar = get_business_calendar(
            conn, company_id, primary_competence,
            reference_today=competence_state["today"],
            include_current_day=False,
            seller_name=seller_name,
        )
        seller_total_days = seller_calendar["sellerWorkingDays"] or summary_calendar["totalWorkingDays"]
        seller_elapsed_days = seller_calendar["sellerElapsedWorkingDays"] or summary_calendar["elapsedWorkingDays"]
        daily_revenue_actual, projected_revenue_raw = dashboard_metric_projection(
            revenue_net,
            seller_elapsed_days,
            seller_total_days,
        )
        projected_revenue = round(projected_revenue_raw, 2)
        projected_goal_attainment = safe_div(projected_revenue, revenue_goal) * 100 if revenue_goal else 0.0
        return_ratio = safe_div(returns_value, revenue_net) * 100 if revenue_net else 0.0
        discount_pct = safe_div(metrics["discountValue"], gross_sales_pct) * 100
        margin_value = float(official_row.get("margin_value") or 0) if official_row and not filters["city_name"] else None
        goal_component = min(goal_attainment, 120) / 120 * 100
        ticket_component = safe_div(ticket, max_ticket) * 100
        client_component = safe_div(distinct_clients, max_clients) * 100
        mix_component = safe_div(mix_count, max_mix) * 100
        returns_component = max(0.0, 100 - min(return_ratio, 8) / 8 * 100)
        score = (
            goal_component * float(score_config["weight_goal"])
            + ticket_component * float(score_config["weight_ticket"])
            + client_component * float(score_config["weight_clients"])
            + mix_component * float(score_config["weight_mix"])
            + returns_component * float(score_config["weight_returns"])
        ) / 100
        daily_goal_value = safe_div(revenue_goal, seller_total_days) if seller_total_days else 0.0
        seller_rows.append(
            {
                "sellerName": seller_name,
                "baseUnit": base_unit,
                "role": role or "Pendente",
                "revenueNet": round(revenue_net, 2),
                "revenueGoal": round(revenue_goal, 2),
                "projectedRevenue": projected_revenue,
                "returnsValue": round(returns_value, 2),
                "warrantyReturnsValue": round(warranty_value, 2),
                "returnsTotalValue": round(returns_total, 2),
                "revenueNetWithWarranty": round(revenue_net_raw, 2),
                "returnCost": round(return_cost, 2),
                "revenueGross": round(gross_sales, 2),
                "costValue": round(cost_value, 2),
                "profitValue": round(profit_value, 2),
                "netProfitValue": round(net_profit_value, 2),
                "returnsGoal": round(returns_goal, 2),
                "goalAttainmentPct": round(goal_attainment, 2),
                "projectedGoalAttainmentPct": round(projected_goal_attainment, 2),
                "dailyRevenueActual": round(daily_revenue_actual, 2),
                "dailyGoal": round(daily_goal_value, 2),
                "ticketAverage": round(ticket, 2),
                "qtySold": round(qty_sold, 2),
                "ticketPerPiece": round(ticket_per_piece, 2),
                "distinctClients": distinct_clients,
                # Clientes da própria carteira x atendidos fora dela
                "ownClients": len(metrics.get("ownClients") or set()),
                "otherClients": len(metrics.get("otherClients") or set()),
                "ownRevenue": round(float(metrics.get("ownRevenue") or 0.0), 2),
                "otherRevenue": round(float(metrics.get("otherRevenue") or 0.0), 2),
                "mixSku": mix_count,
                "discountValue": round(metrics["discountValue"], 2),
                "discountPct": round(discount_pct, 2),
                "returnRatioPct": round(return_ratio, 2),
                "farol": {
                    "goalAttainment": evaluate_farol("goal_attainment", goal_attainment if revenue_goal else None, farol_thresholds, pace_pct),
                    "projectedAttainment": evaluate_farol("projected_attainment", projected_goal_attainment if revenue_goal else None, farol_thresholds),
                    "returnRatio": evaluate_farol("return_ratio", return_ratio, farol_thresholds),
                    "discountPct": evaluate_farol("discount_pct", discount_pct, farol_thresholds),
                },
                "warrantyRatioPct": round(safe_div(warranty_value, revenue_net) * 100 if revenue_net else 0.0, 2),
                "returnsTotalRatioPct": round(safe_div(returns_total, revenue_net) * 100 if revenue_net else 0.0, 2),
                "marginValue": (lambda _m: round(_m, 2) if _m is not None else None)(finite_or_none(margin_value)),
                "score": round(score, 2),
                "pendingMapping": role is None,
                "missingGoal": revenue_goal <= 0,
                "inDeployment": normalize_unit(base_unit) in unidades_implantacao,
                "metaDiaria": round(daily_goal_value, 2),
                "sellerWorkingDays": seller_total_days,
                "sellerElapsedWorkingDays": seller_elapsed_days,
                "vacationDays": summary_calendar["totalWorkingDays"] - seller_total_days,
            }
        )
    seller_rows.sort(key=lambda item: item["score"], reverse=True)

    unit_calendar = summary_calendar
    unit_rows = []
    unit_names = sorted(set(unit_summary_by_unit) | set(goal_by_unit))
    for unit_name in unit_names:
        official_row = unit_summary_by_unit.get(unit_name, {})
        unit_goal = goal_by_unit.get(unit_name, {"revenueGoal": 0.0, "returnsGoal": 0.0})
        revenue_goal = float(unit_goal["revenueGoal"] or 0)
        returns_goal = float(unit_goal["returnsGoal"] or 0)
        revenue_net_raw = float(official_row.get("net_value") or 0.0)
        gross_sales = float(official_row.get("sale_value") or 0.0)
        returns_total = float(official_row.get("return_value") or 0.0)
        # Garantia sai do resultado comercial: devolução comercial = total − garantia,
        # e o líquido volta a somar a garantia (que o custo/venda já havia descontado).
        warranty_value = min(warranty_by_unit.get(unit_name, 0.0), returns_total)
        returns_value = max(returns_total - warranty_value, 0.0)
        revenue_net = revenue_net_raw + warranty_value
        qty_sold = float(official_row.get("qty_sold") or 0.0)
        return_cost = float(official_row.get("return_cost") or 0.0)
        cost_value = float(official_row.get("cost_value") or 0.0)
        profit_value = float(official_row.get("profit_value") or 0.0)
        net_profit_value = float(official_row.get("net_profit_value") or 0.0)
        daily_revenue_actual, projected_revenue_raw = dashboard_metric_projection(
            revenue_net,
            unit_calendar["elapsedWorkingDays"],
            unit_calendar["totalWorkingDays"],
        )
        projected_revenue = round(projected_revenue_raw, 2)
        daily_goal_value = safe_div(revenue_goal, unit_calendar["totalWorkingDays"]) if unit_calendar["totalWorkingDays"] else 0.0
        unit_rows.append(
            {
                "unitName": unit_name,
                "revenueNet": round(revenue_net, 2),
                "revenueGoal": round(revenue_goal, 2),
                "projectedRevenue": projected_revenue,
                "dailyRevenueActual": round(daily_revenue_actual, 2),
                "dailyGoal": round(daily_goal_value, 2),
                "returnsValue": round(returns_value, 2),
                "warrantyReturnsValue": round(warranty_value, 2),
                "returnsTotalValue": round(returns_total, 2),
                "revenueNetWithWarranty": round(revenue_net_raw, 2),
                "returnCost": round(return_cost, 2),
                "revenueGross": round(gross_sales, 2),
                "costValue": round(cost_value, 2),
                "profitValue": round(profit_value, 2),
                "netProfitValue": round(net_profit_value, 2),
                "returnsGoal": round(returns_goal, 2),
                "returnRatioPct": round(safe_div(returns_value, revenue_net) * 100 if revenue_net else 0.0, 2),
                "farol": {
                    "goalAttainment": evaluate_farol(
                        "goal_attainment",
                        safe_div(revenue_net, revenue_goal) * 100 if revenue_goal else None,
                        farol_thresholds, pace_pct),
                    "projectedAttainment": evaluate_farol(
                        "projected_attainment",
                        safe_div(projected_revenue, revenue_goal) * 100 if revenue_goal else None,
                        farol_thresholds),
                    "returnRatio": evaluate_farol(
                        "return_ratio",
                        safe_div(returns_value, revenue_net) * 100 if revenue_net else None,
                        farol_thresholds),
                    "marginValue": evaluate_farol(
                        "margin_value", finite_or_none(official_row.get("margin_value")), farol_thresholds),
                },
                "warrantyRatioPct": round(safe_div(warranty_value, revenue_net) * 100 if revenue_net else 0.0, 2),
                "returnsTotalRatioPct": round(safe_div(returns_total, revenue_net) * 100 if revenue_net else 0.0, 2),
                "goalAttainmentPct": round(safe_div(revenue_net, revenue_goal) * 100 if revenue_goal else 0.0, 2),
                "projectedGoalAttainmentPct": round(safe_div(projected_revenue, revenue_goal) * 100 if revenue_goal else 0.0, 2),
                "marginValue": round(finite_or_none(official_row.get("margin_value")) or 0, 2) if official_row else None,
                "qtySold": round(qty_sold, 2),
                "ticketPerPiece": round(safe_div(revenue_net, qty_sold), 2),
                "metaDiaria": round(daily_goal_value, 2),
            }
        )
    unit_rows.sort(key=lambda item: item["revenueNet"], reverse=True)

    detail_totals = {
        "revenueNet": float(sum(row["net_value"] or 0 for row in detail_rows_scope)),
        "grossSales": float(sum(row["gross_value"] or 0 for row in detail_rows_scope)),
        "grossSalesPct": float(sum(max(float(row["gross_value"] or 0), 0.0) for row in detail_rows_scope)),
        "discountValue": float(sum(row["discount_value"] or 0 for row in detail_rows_scope)),
        "returnsValue": float(sum(row["return_value"] or 0 for row in detail_rows_scope)),
        "distinctClients": len({row["client_name"] for row in detail_rows_scope if row["client_name"]}),
    }
    detail_totals["ticketAverage"] = safe_div(detail_totals["revenueNet"], detail_totals["distinctClients"])
    detail_totals["discountPct"] = safe_div(detail_totals["discountValue"], detail_totals["grossSalesPct"]) * 100
    detail_totals["returnRatioPct"] = safe_div(detail_totals["returnsValue"], detail_totals["revenueNet"]) * 100 if detail_totals["revenueNet"] else 0.0

    # ── Carteira x fora da carteira e quebra PJ/PF no consolidado ─────────────
    _own_clients: set[str] = set()
    _other_clients: set[str] = set()
    _own_revenue = 0.0
    _other_revenue = 0.0
    _pj_clients: set[str] = set()
    _pf_clients: set[str] = set()
    _pj_revenue = 0.0
    _pf_revenue = 0.0
    for row in detail_rows_scope:
        _cname = row["client_name"]
        if not _cname:
            continue
        _net = float(row["net_value"] or 0)
        _ckey = normalize_client_key(_cname)
        _owner = client_owner_map.get(_ckey)
        if _owner and normalize_upper(_owner) == normalize_upper(row["seller_name"]):
            _own_clients.add(_cname)
            _own_revenue += _net
        else:
            _other_clients.add(_cname)
            _other_revenue += _net
        _reg = client_registry.get(_ckey)
        _ptype = _reg["person_type"] if _reg else (infer_person_type_from_name(_cname)[0])
        if normalize_upper(_ptype) == "PJ":
            _pj_clients.add(_cname)
            _pj_revenue += _net
        elif normalize_upper(_ptype) == "PF":
            _pf_clients.add(_cname)
            _pf_revenue += _net

    detail_totals["ownClients"] = len(_own_clients)
    detail_totals["otherClients"] = len(_other_clients)
    detail_totals["ownRevenue"] = round(_own_revenue, 2)
    detail_totals["otherRevenue"] = round(_other_revenue, 2)
    detail_totals["pjClients"] = len(_pj_clients)
    detail_totals["pfClients"] = len(_pf_clients)
    detail_totals["pjRevenue"] = round(_pj_revenue, 2)
    detail_totals["pfRevenue"] = round(_pf_revenue, 2)
    detail_totals["ticketAveragePj"] = round(safe_div(_pj_revenue, len(_pj_clients)), 2)
    detail_totals["ticketAveragePf"] = round(safe_div(_pf_revenue, len(_pf_clients)), 2)
    detail_totals["ticketAverageOwn"] = round(safe_div(_own_revenue, len(_own_clients)), 2)
    detail_totals["ticketAverageOther"] = round(safe_div(_other_revenue, len(_other_clients)), 2)

    official_totals_vendor = aggregate_official_summary_rows(vendor_summary_rows)
    official_totals_unit = aggregate_official_summary_rows(unit_summary_rows)

    use_detail_summary = bool(filters["city_name"])
    if filters["seller_name"] and not filters["city_name"]:
        seller_metrics = detail_by_seller.get(filters["seller_name"], {})
        official_seller = vendor_summary_by_seller.get(filters["seller_name"], {})
        seller_goal = goal_by_seller.get(filters["seller_name"], {"revenueGoal": 0.0, "returnsGoal": 0.0})
        summary_revenue = float(official_seller.get("net_value") or 0.0)
        summary_goal = float(seller_goal["revenueGoal"] or 0)
        summary_returns = float(official_seller.get("return_value") or 0.0)
        summary_returns_goal = float(seller_goal["returnsGoal"] or 0)
        summary_margin = float(official_seller.get("margin_value") or 0) if official_seller else None
        summary_qty_sold = float(official_seller.get("qty_sold") or 0)
        summary_gross = float(official_seller.get("sale_value") or 0.0)
        summary_return_cost = float(official_seller.get("return_cost") or 0.0)
        summary_cost_value = float(official_seller.get("cost_value") or 0.0)
        summary_profit_value = float(official_seller.get("profit_value") or 0.0)
        summary_net_profit_value = float(official_seller.get("net_profit_value") or 0.0)
    elif use_detail_summary:
        summary_revenue = detail_totals["revenueNet"]
        summary_goal = total_unit_goal
        summary_returns = detail_totals["returnsValue"]
        summary_returns_goal = total_unit_returns_goal
        summary_margin = None
        summary_qty_sold = 0.0
        summary_gross = detail_totals["grossSales"]
        summary_return_cost = 0.0
        summary_cost_value = 0.0
        summary_profit_value = 0.0
        summary_net_profit_value = 0.0
    else:
        summary_revenue = float(official_totals_unit["revenueNet"] or 0.0)
        summary_goal = total_unit_goal
        summary_returns = float(official_totals_unit["returnsValue"] or 0.0)
        summary_returns_goal = total_unit_returns_goal
        summary_margin = official_totals_unit["marginAverage"]
        summary_qty_sold = float(official_totals_unit["qtySold"] or 0.0)
        summary_gross = float(official_totals_unit["revenueGross"] or 0.0)
        summary_return_cost = float(official_totals_unit["returnCost"] or 0.0)
        summary_cost_value = float(official_totals_unit["costValue"] or 0.0)
        summary_profit_value = float(official_totals_unit["profitValue"] or 0.0)
        summary_net_profit_value = float(official_totals_unit["netProfitValue"] or 0.0)

    # ── Dedução da garantia no consolidado ────────────────────────────────────
    # Aplica o mesmo tratamento do vendedor e da unidade no total do grupo,
    # respeitando o recorte de vendedor/unidade que estiver ativo.
    _warranty_conditions = ["company_id = ?", "competence = ?"]
    _warranty_params: list[Any] = [company_id, primary_competence]
    if filters.get("seller_name"):
        _warranty_conditions.append("seller_name = ?")
        _warranty_params.append(filters["seller_name"])
    if filters.get("unit_name"):
        _warranty_conditions.append("unit_name = ?")
        _warranty_params.append(normalize_unit(filters["unit_name"]))
    elif scoped_units:
        _ph = ", ".join("?" for _ in scoped_units)
        _warranty_conditions.append(f"unit_name IN ({_ph})")
        _warranty_params.extend(scoped_units)
    summary_warranty = float(conn.execute(
        f"SELECT COALESCE(SUM(total_value), 0) AS total FROM fact_warranty_returns "
        f"WHERE {' AND '.join(_warranty_conditions)}",
        _warranty_params,
    ).fetchone()["total"] or 0.0)

    summary_returns_total = summary_returns
    summary_warranty = min(summary_warranty, summary_returns_total)
    summary_returns = max(summary_returns_total - summary_warranty, 0.0)
    summary_revenue_with_warranty = summary_revenue
    summary_revenue = summary_revenue + summary_warranty

    elapsed_days_current = summary_calendar["elapsedWorkingDays"]
    total_days_current = summary_calendar["totalWorkingDays"]
    daily_revenue_actual_raw, projection_revenue_raw = dashboard_metric_projection(summary_revenue, elapsed_days_current, total_days_current)
    projection_revenue = round(projection_revenue_raw, 2)
    score_average = round(sum(item["score"] for item in seller_rows) / len(seller_rows), 2) if seller_rows else 0.0
    daily_revenue_actual = round(daily_revenue_actual_raw, 2)
    daily_goal = round(safe_div(summary_goal, total_days_current), 2) if total_days_current else 0.0
    revenue_pf_pct = round(safe_div(client_type_summary.get("PF", {}).get("revenueNet", 0.0), detail_totals["revenueNet"]) * 100, 2) if detail_totals["revenueNet"] else 0.0
    revenue_pj_pct = round(safe_div(client_type_summary.get("PJ", {}).get("revenueNet", 0.0), detail_totals["revenueNet"]) * 100, 2) if detail_totals["revenueNet"] else 0.0
    summary_distinct_clients = detail_totals["distinctClients"]
    summary_ticket_average = round(safe_div(summary_revenue, summary_distinct_clients), 2)
    summary_ticket_per_piece = round(safe_div(summary_revenue, summary_qty_sold), 2)
    summary_discount_value = round(detail_totals["discountValue"], 2)
    # Mesma base do ranking de vendedores: desconto sobre a venda bruta POSITIVA
    # do faturamento detalhado. Antes o denominador vinha do arquivo de
    # custo/venda (`summary_gross`), que é outra fonte e outra magnitude — o
    # Executivo mostrava 59% onde o ranking mostrava 26% para a mesma pessoa.
    # Numerador e denominador precisam sair do mesmo lugar.
    summary_discount_pct = round(detail_totals["discountPct"], 2)
    unidade_filtrada = normalize_unit(filters.get("unit_name"))
    resumo_em_implantacao = bool(unidade_filtrada and unidade_filtrada in unidades_implantacao)

    comparison_previous = {}
    comparison_yoy = {}
    if primary_competence:
        year, month = map(int, primary_competence.split("-"))
        prev_date = first_day_of_competence(primary_competence)
        previous_date = (prev_date.replace(day=1) - timedelta(days=1)).replace(day=1)
        comparison_previous = _cached_single_competence_summary(
            conn,
            company_id,
            previous_date.strftime("%Y-%m"),
            filters,
        )
        comparison_yoy = _cached_single_competence_summary(
            conn,
            company_id,
            f"{year - 1:04d}-{month:02d}",
            filters,
        )

    data_quality = {
        "pendingSellers": conn.execute(
            "SELECT COUNT(*) AS total FROM import_issues WHERE company_id = ? AND status = 'pendente' AND issue_type = 'vendedor_sem_vinculo'",
            (company_id,),
        ).fetchone()["total"],
        "pendingCities": conn.execute(
            "SELECT COUNT(*) AS total FROM import_issues WHERE company_id = ? AND status = 'pendente' AND issue_type = 'cidade_sem_correspondencia'",
            (company_id,),
        ).fetchone()["total"],
        "duplicateRowsSkipped": conn.execute(
            "SELECT COALESCE(SUM(duplicate_rows_skipped), 0) AS total FROM imports WHERE company_id = ?",
            (company_id,),
        ).fetchone()["total"],
    }

    debug_projection = {
        "timezone": "America/Sao_Paulo",
        "today": competence_state["today"].isoformat(),
        "cutoffDate": cutoff_date.isoformat(),
        "competenceUsed": primary_competence,
        "isCurrentCompetence": competence_state["isCurrentCompetence"],
        "isPastCompetence": competence_state["isPastCompetence"],
        "isFutureCompetence": competence_state["isFutureCompetence"],
        "revenueNet": round(summary_revenue, 2),
        "revenueNetUntilCutoff": round(summary_revenue, 2),
        "revenueGoal": round(summary_goal, 2),
        "expectedGroupGoal": 6477000,
        "totalWorkingDays": total_days_current,
        "elapsedWorkingDays": elapsed_days_current,
        "remainingWorkingDays": summary_calendar["remainingWorkingDays"],
        "dailyRevenueActual": daily_revenue_actual,
        "dailyGoal": daily_goal,
        "projectedRevenue": projection_revenue,
        "goalAttainmentPct": round(safe_div(summary_revenue, summary_goal) * 100 if summary_goal else 0.0, 2),
        "projectedGoalAttainmentPct": round(safe_div(projection_revenue, summary_goal) * 100 if summary_goal else 0.0, 2),
        "formula": {
            "cutoffDate": "today - 1 day",
            "dailyRevenueActual": "revenueNet / elapsedWorkingDays",
            "projectedRevenue": "dailyRevenueActual * totalWorkingDays",
            "dailyGoal": "revenueGoal / totalWorkingDays",
            "goalAttainmentPct": "revenueNet / revenueGoal * 100",
            "projectedGoalAttainmentPct": "projectedRevenue / revenueGoal * 100",
        },
        "unitGoalsBreakdown": [
            {"unitName": unit_name, "revenueGoal": round(values["revenueGoal"], 2)}
            for unit_name, values in sorted(goal_by_unit.items())
        ],
        "unitGoalsSum": round(total_unit_goal, 2),
        "sellerGoalsSum": round(total_company_seller_goal, 2),
        "duplicateUnitGoals": duplicate_unit_goals,
        "duplicateSellerGoals": duplicate_seller_goals,
    }
    if DASHBOARD_DEBUG_LOG:
        print("[DASHBOARD D-1 PROJECTION DEBUG]", debug_projection)
    debug_comparisons = {
        "current": {
            "competence": primary_competence,
            "unit": filters.get("unit_name"),
            "seller": filters.get("seller_name"),
            "city": filters.get("city_name"),
            "revenueNet": round(summary_revenue, 2),
        },
        "previousCompetence": {
            "competence": comparison_previous.get("competence"),
            "unit": comparison_previous.get("unit"),
            "seller": comparison_previous.get("seller"),
            "city": comparison_previous.get("city"),
            "revenueNet": comparison_previous.get("revenueNet"),
        },
        "yearOverYear": {
            "competence": comparison_yoy.get("competence"),
            "unit": comparison_yoy.get("unit"),
            "seller": comparison_yoy.get("seller"),
            "city": comparison_yoy.get("city"),
            "revenueNet": comparison_yoy.get("revenueNet"),
        },
    }
    if DASHBOARD_DEBUG_LOG:
        print("[debugComparisons]", debug_comparisons)

    return {
        "filters": filters,
        "revenueSourcePolicy": {
            "executiveSummary": "fact_unit_summary" if not filters.get("seller_name") and not filters.get("city_name") else ("fact_vendor_summary" if filters.get("seller_name") and not filters.get("city_name") else "fact_sales_detail"),
            "unitPerformance": "fact_unit_summary",
            "sellerRanking": "fact_vendor_summary",
            "cityRanking": "fact_sales_detail",
            "clientRanking": "fact_sales_detail",
            "crm": "crm_client_summary + fact_sales_detail",
        },
        "availableCompetences": competences,
        "primaryCompetence": primary_competence,
        "calendar": summary_calendar,
        "summary": {
            "revenueNet": round(summary_revenue, 2),
            "revenueGoal": round(summary_goal, 2),
            "projectedRevenue": projection_revenue,
            "returnsValue": round(summary_returns, 2),
            # Garantia separada do resultado comercial
            "warrantyReturnsValue": round(summary_warranty, 2),
            "returnsTotalValue": round(summary_returns_total, 2),
            "revenueNetWithWarranty": round(summary_revenue_with_warranty, 2),
            "returnCost": round(summary_return_cost, 2),
            "revenueGross": round(summary_gross, 2),
            "costValue": round(summary_cost_value, 2),
            "profitValue": round(summary_profit_value, 2),
            "netProfitValue": round(summary_net_profit_value, 2),
            "returnsGoal": round(summary_returns_goal, 2),
            "returnRatioPct": round(safe_div(summary_returns, summary_revenue) * 100 if summary_revenue else 0.0, 2),
            "warrantyRatioPct": round(safe_div(summary_warranty, summary_revenue) * 100 if summary_revenue else 0.0, 2),
            "returnsTotalRatioPct": round(safe_div(summary_returns_total, summary_revenue) * 100 if summary_revenue else 0.0, 2),
            "goalAttainmentPct": round(safe_div(summary_revenue, summary_goal) * 100 if summary_goal else 0.0, 2),
            "projectedGoalAttainmentPct": round(safe_div(projection_revenue, summary_goal) * 100 if summary_goal else 0.0, 2),
            "dailyRevenueActual": daily_revenue_actual,
            "dailyGoal": daily_goal,
            "revenuePfPct": revenue_pf_pct,
            "revenuePjPct": revenue_pj_pct,
            "ticketAverage": summary_ticket_average,
            "qtySold": round(summary_qty_sold, 2),
            "ticketPerPiece": summary_ticket_per_piece,
            "distinctClients": summary_distinct_clients,
            # Clientes atendidos: quantos são de carteira e quantos vieram de fora
            "ownClients": detail_totals["ownClients"],
            "otherClients": detail_totals["otherClients"],
            "ownRevenue": detail_totals["ownRevenue"],
            "otherRevenue": detail_totals["otherRevenue"],
            "ticketAverageOwn": detail_totals["ticketAverageOwn"],
            "ticketAverageOther": detail_totals["ticketAverageOther"],
            # Quebra por tipo de pessoa
            "pjClients": detail_totals["pjClients"],
            "pfClients": detail_totals["pfClients"],
            "pjRevenue": detail_totals["pjRevenue"],
            "pfRevenue": detail_totals["pfRevenue"],
            "ticketAveragePj": detail_totals["ticketAveragePj"],
            "ticketAveragePf": detail_totals["ticketAveragePf"],
            "marginAverage": round(summary_margin, 2) if summary_margin is not None else None,
            "discountValue": summary_discount_value,
            "discountPct": summary_discount_pct,
            "inDeployment": resumo_em_implantacao,
            "scoreAverage": score_average,
            "workingDaysTotal": total_days_current,
            "workingDaysElapsed": elapsed_days_current,
            # Ritmo do mês e faróis prontos para a tela
            "paceExpectedPct": round(pace_pct, 2),
            "farol": {
                "goalAttainment": evaluate_farol(
                    "goal_attainment",
                    safe_div(summary_revenue, summary_goal) * 100 if summary_goal else None,
                    farol_thresholds, pace_pct),
                "projectedAttainment": evaluate_farol(
                    "projected_attainment",
                    safe_div(projection_revenue, summary_goal) * 100 if summary_goal else None,
                    farol_thresholds),
                "returnRatio": evaluate_farol(
                    "return_ratio",
                    safe_div(summary_returns, summary_revenue) * 100 if summary_revenue else None,
                    farol_thresholds),
                "discountPct": evaluate_farol(
                    "discount_pct", detail_totals["discountPct"], farol_thresholds),
            },
            "dailyRevenueTarget": daily_goal,
        },
        "debugProjection": debug_projection,
        "debugComparisons": debug_comparisons,
        "comparisons": {
            "previousCompetence": comparison_previous,
            "yearOverYear": comparison_yoy,
        },
        "sellerRanking": seller_rows,
        "sellerTop10": seller_rows[:10],
        "unitPerformance": unit_rows,
        "cityRanking": city_ranking,
        "clientRanking": client_ranking,
        "clientTopByUnit": client_top_by_unit,
        "clientTypeSummary": {key: {"revenueNet": round(value["revenueNet"], 2), "clients": value["clients"]} for key, value in client_type_summary.items()},
        "scoreWeights": {
            "goal": score_config["weight_goal"],
            "ticket": score_config["weight_ticket"],
            "clients": score_config["weight_clients"],
            "mix": score_config["weight_mix"],
            "returns": score_config["weight_returns"],
        },
        "quadrant": {
            "xReference": 100,
            "yReference": score_average,
            "points": [{"sellerName": row["sellerName"], "x": row["goalAttainmentPct"], "y": row["score"]} for row in seller_rows],
        },
        "dataQuality": data_quality,
    }


def single_competence_summary(
    conn: sqlite3.Connection,
    company_id: int,
    competence: str | None,
    filters: dict[str, str | None] | None = None,
) -> dict[str, Any]:
    if not competence:
        return {}

    scoped_filters = dict(filters or {})
    scoped_filters.setdefault("unit_name", None)
    scoped_filters.setdefault("seller_name", None)
    scoped_filters.setdefault("city_name", None)
    scoped_filters.setdefault("allowed_units", None)
    allowed_units = normalize_unit_list(scoped_filters.get("allowed_units"))
    scoped_units = [normalize_unit(scoped_filters["unit_name"])] if scoped_filters.get("unit_name") else allowed_units

    detail_conditions = ["company_id = ?", "competence = ?"]
    detail_params: list[Any] = [company_id, competence]
    summary_vendor_conditions = ["company_id = ?", "competence = ?"]
    summary_vendor_params: list[Any] = [company_id, competence]
    summary_unit_conditions = ["company_id = ?", "competence = ?"]
    summary_unit_params: list[Any] = [company_id, competence]

    if scoped_filters["seller_name"]:
        detail_conditions.append("seller_name = ?")
        detail_params.append(scoped_filters["seller_name"])
        summary_vendor_conditions.append("seller_name = ?")
        summary_vendor_params.append(scoped_filters["seller_name"])

    detail_source_rows = conn.execute(
        f"""
        SELECT seller_name, client_name, city_name, gross_value, discount_value, return_value, net_value, competence, issue_date
        FROM fact_sales_detail
        WHERE {' AND '.join(detail_conditions)}
        """,
        detail_params,
    ).fetchall()

    competence_state = dashboard_competence_state(competence)
    cutoff_date = competence_state["cutoffDate"]

    # ── PRÉ-CARREGAR LOOKUPS (elimina N+1 queries) ──────────────────────────
    _comp_target = first_day_of_competence(competence).isoformat()

    _city_map: dict[str, str | None] = {}
    for _r in conn.execute(
        """
        SELECT city_name, principal_unit FROM city_mappings
        WHERE company_id = ? AND date(valid_from) <= date(?)
          AND (valid_to IS NULL OR date(valid_to) >= date(?))
        ORDER BY date(valid_from) DESC
        """,
        (company_id, _comp_target, _comp_target),
    ).fetchall():
        _k = normalize_upper(_r["city_name"])
        if _k and _k not in _city_map:
            _city_map[_k] = normalize_unit(_r["principal_unit"])

    _seller_unit_map: dict[str, str | None] = {}
    for _r in conn.execute(
        """
        SELECT person_name, base_unit FROM people_records
        WHERE company_id = ? AND date(valid_from) <= date(?)
          AND (valid_to IS NULL OR date(valid_to) >= date(?))
        ORDER BY date(valid_from) DESC
        """,
        (company_id, _comp_target, _comp_target),
    ).fetchall():
        _k = normalize_whitespace(_r["person_name"])
        if _k and _k not in _seller_unit_map:
            _seller_unit_map[_k] = normalize_unit(_r["base_unit"])
    # ────────────────────────────────────────────────────────────────────────

    detail_rows_all: list[dict[str, Any]] = []
    for row in detail_source_rows:
        issue_dt = parse_datetime_flexible(row["issue_date"])
        if competence_state["isFutureCompetence"]:
            continue
        if competence_state["isCurrentCompetence"] and issue_dt and issue_dt.date() > cutoff_date:
            continue
        seller_name = normalize_whitespace(row["seller_name"])
        city_name = normalize_upper(row["city_name"])
        resolved_unit = _city_map.get(city_name)
        seller_base_unit = _seller_unit_map.get(seller_name)
        enriched = dict(row)
        enriched["seller_name"] = seller_name
        enriched["client_name"] = normalize_whitespace(row["client_name"])
        enriched["city_name"] = city_name
        enriched["resolved_unit"] = resolved_unit
        enriched["seller_base_unit"] = normalize_unit(seller_base_unit)
        detail_rows_all.append(enriched)

    if scoped_filters["city_name"]:
        detail_rows_all = [row for row in detail_rows_all if row["city_name"] == scoped_filters["city_name"]]

    detail_rows_scope = detail_rows_all
    if scoped_units:
        detail_rows_scope = [row for row in detail_rows_scope if row["resolved_unit"] in scoped_units]
    if scoped_filters["unit_name"]:
        detail_rows_scope = [row for row in detail_rows_scope if row["resolved_unit"] == scoped_filters["unit_name"]]

    seller_detail_rows = detail_rows_all
    if scoped_units:
        seller_detail_rows = [row for row in seller_detail_rows if row["seller_base_unit"] in scoped_units]
    if scoped_filters["city_name"]:
        seller_detail_rows = [row for row in seller_detail_rows if row["city_name"] == scoped_filters["city_name"]]
    if scoped_filters["unit_name"]:
        seller_detail_rows = [row for row in seller_detail_rows if row["seller_base_unit"] == scoped_filters["unit_name"]]

    if scoped_units:
        placeholders = ", ".join("?" for _ in scoped_units)
        summary_unit_conditions.append(f"unit_name IN ({placeholders})")
        summary_unit_params.extend(scoped_units)
    if scoped_filters["unit_name"]:
        summary_unit_conditions.append("unit_name = ?")
        summary_unit_params.append(scoped_filters["unit_name"])

    vendor_summary_rows = [
        dict(row)
        for row in conn.execute(
            f"""
        SELECT
            seller_name,
            SUM(qty_sold) AS qty_sold,
            SUM(cost_value) AS cost_value,
            SUM(sale_value) AS sale_value,
            SUM(profit_value) AS profit_value,
            SUM(net_profit_value) AS net_profit_value,
            SUM(return_cost) AS return_cost,
            SUM(return_value) AS return_value,
            SUM(net_value) AS net_value,
            AVG(margin_value) AS margin_value
        FROM fact_vendor_summary
            WHERE {' AND '.join(summary_vendor_conditions)}
            GROUP BY seller_name
            """,
            summary_vendor_params,
        ).fetchall()
    ]
    unit_summary_rows = [
        dict(row)
        for row in conn.execute(
            f"""
        SELECT
            unit_name,
            SUM(qty_sold) AS qty_sold,
            SUM(cost_value) AS cost_value,
            SUM(sale_value) AS sale_value,
            SUM(profit_value) AS profit_value,
            SUM(net_profit_value) AS net_profit_value,
            SUM(return_cost) AS return_cost,
            SUM(return_value) AS return_value,
            SUM(net_value) AS net_value,
            AVG(margin_value) AS margin_value
        FROM fact_unit_summary
            WHERE {' AND '.join(summary_unit_conditions)}
            GROUP BY unit_name
            """,
            summary_unit_params,
        ).fetchall()
    ]
    vendor_summary_by_seller = {normalize_whitespace(row["seller_name"]): row for row in vendor_summary_rows}

    goal_by_seller, _ = load_goal_maps(
        conn,
        company_id,
        competence,
        "goals_seller",
        "seller_name",
        normalize_whitespace,
    )
    goal_by_unit, _ = load_goal_maps(
        conn,
        company_id,
        competence,
        "goals_unit",
        "unit_name",
        normalize_unit,
    )
    if scoped_units:
        goal_by_unit = {
            unit_name: values
            for unit_name, values in goal_by_unit.items()
            if normalize_unit(unit_name) in scoped_units
        }
    total_unit_goal = float(sum(item["revenueGoal"] or 0 for item in goal_by_unit.values()))
    total_unit_returns_goal = float(sum(item["returnsGoal"] or 0 for item in goal_by_unit.values()))

    detail_by_seller: dict[str, dict[str, Any]] = defaultdict(
        lambda: {"revenueNet": 0.0, "grossSales": 0.0, "returnValue": 0.0, "clients": set()}
    )
    detail_by_unit: dict[str, dict[str, Any]] = defaultdict(
        lambda: {"revenueNet": 0.0, "grossSales": 0.0, "returnValue": 0.0, "clients": set()}
    )

    for row in detail_rows_scope:
        resolved_unit = row["resolved_unit"] or "NAO_MAPEADO"
        net_value = float(row["net_value"] or 0.0)
        gross_value = float(row["gross_value"] or 0.0)
        return_value = float(row["return_value"] or 0.0)
        client_name = row["client_name"]
        detail_by_unit[resolved_unit]["revenueNet"] += net_value
        detail_by_unit[resolved_unit]["grossSales"] += gross_value
        detail_by_unit[resolved_unit]["returnValue"] += return_value
        if client_name:
            detail_by_unit[resolved_unit]["clients"].add(client_name)

    for row in seller_detail_rows:
        seller_name = row["seller_name"]
        net_value = float(row["net_value"] or 0.0)
        gross_value = float(row["gross_value"] or 0.0)
        return_value = float(row["return_value"] or 0.0)
        client_name = row["client_name"]
        detail_by_seller[seller_name]["revenueNet"] += net_value
        detail_by_seller[seller_name]["grossSales"] += gross_value
        detail_by_seller[seller_name]["returnValue"] += return_value
        if client_name:
            detail_by_seller[seller_name]["clients"].add(client_name)

    detail_totals = {
        "revenueNet": float(sum(float(row["net_value"] or 0.0) for row in detail_rows_scope)),
        "grossSales": float(sum(float(row["gross_value"] or 0.0) for row in detail_rows_scope)),
        "returnsValue": float(sum(float(row["return_value"] or 0.0) for row in detail_rows_scope)),
        "discountValue": float(sum(float(row["discount_value"] or 0.0) for row in detail_rows_scope)),
        "distinctClients": len({row["client_name"] for row in detail_rows_scope if row["client_name"]}),
    }
    official_totals_unit = aggregate_official_summary_rows(unit_summary_rows)

    if scoped_filters["seller_name"] and not scoped_filters["city_name"]:
        seller_metrics = detail_by_seller.get(scoped_filters["seller_name"], {})
        official_seller = vendor_summary_by_seller.get(scoped_filters["seller_name"], {})
        seller_goal = goal_by_seller.get(scoped_filters["seller_name"], {"revenueGoal": 0.0, "returnsGoal": 0.0})
        summary_revenue = float(official_seller.get("net_value") or 0.0)
        summary_goal = float(seller_goal["revenueGoal"] or 0)
        summary_returns = float(official_seller.get("return_value") or 0.0)
        summary_returns_goal = float(seller_goal["returnsGoal"] or 0)
        summary_margin = float(official_seller.get("margin_value") or 0) if official_seller else None
        summary_qty_sold = float(official_seller.get("qty_sold") or 0.0)
        summary_gross = float(official_seller.get("sale_value") or 0.0)
    elif scoped_filters["city_name"]:
        summary_revenue = detail_totals["revenueNet"]
        summary_goal = total_unit_goal
        summary_returns = detail_totals["returnsValue"]
        summary_returns_goal = total_unit_returns_goal
        summary_margin = None
        summary_qty_sold = 0.0
        summary_gross = detail_totals["grossSales"]
    else:
        summary_revenue = float(official_totals_unit["revenueNet"] or 0.0)
        summary_goal = total_unit_goal
        summary_returns = float(official_totals_unit["returnsValue"] or 0.0)
        summary_returns_goal = total_unit_returns_goal
        summary_margin = official_totals_unit["marginAverage"]
        summary_qty_sold = float(official_totals_unit["qtySold"] or 0.0)
        summary_gross = float(official_totals_unit["revenueGross"] or 0.0)

    calendar = get_business_calendar(conn, company_id, competence)
    return {
        "competence": competence,
        "unit": scoped_filters.get("unit_name"),
        "seller": scoped_filters.get("seller_name"),
        "city": scoped_filters.get("city_name"),
        "revenueNet": round(summary_revenue, 2),
        "revenueGoal": round(summary_goal, 2),
        "returnsValue": round(summary_returns, 2),
        "returnsGoal": round(summary_returns_goal, 2),
        "returnRatioPct": round(safe_div(summary_returns, summary_revenue) * 100 if summary_revenue else 0.0, 2),
        "dailyRevenue": round(safe_div(summary_revenue, max(calendar["totalWorkingDays"], 1)), 2),
        "marginAverage": round(summary_margin, 2) if summary_margin is not None else None,
        "distinctClients": detail_totals["distinctClients"],
        "discountValue": round(detail_totals["discountValue"], 2),
        "qtySold": round(summary_qty_sold, 2),
    }


AUDIT_TOLERANCE = 0.05


def audit_filters_for_competence(competence: str, **overrides: Any) -> dict[str, str | None]:
    filters: dict[str, str | None] = {
        "competence_start": competence,
        "competence_end": competence,
        "unit_name": None,
        "seller_name": None,
        "city_name": None,
        "status": None,
        "purchaseMonth": None,
        "growth": None,
        "classCode": None,
        "search": None,
        "allowed_units": None,
    }
    filters.update(overrides)
    return filters


def audit_round(value: float | int | None) -> float:
    return round(float(value or 0.0), 2)


def audit_same(a: float | int | None, b: float | int | None, tolerance: float = AUDIT_TOLERANCE) -> bool:
    return abs(float(a or 0.0) - float(b or 0.0)) <= tolerance


def append_audit_issue(
    issues: list[dict[str, Any]],
    severity: str,
    area: str,
    message: str,
    expected: Any,
    actual: Any,
) -> None:
    issues.append(
        {
            "severity": severity,
            "area": area,
            "message": message,
            "expected": expected,
            "actual": actual,
        }
    )


def count_rows_for_import_file(conn: sqlite3.Connection, import_id: int, file_type: str) -> int:
    table_map = {
        "faturamento_detalhado": "fact_sales_detail",
        "custo_vendedor": "fact_vendor_summary",
        "custo_unidade": "fact_unit_summary",
        "cadastro_clientes": "crm_client_profiles",
        "faturamento_cliente_consolidado": "crm_client_summary",
    }
    table_name = table_map.get(file_type)
    if not table_name:
        return 0
    if file_type == "cadastro_clientes":
        row = conn.execute(
            f"SELECT COUNT(*) AS total FROM {table_name} WHERE source_import_id = ?",
            (import_id,),
        ).fetchone()
    else:
        row = conn.execute(
            f"SELECT COUNT(*) AS total FROM {table_name} WHERE import_id = ?",
            (import_id,),
        ).fetchone()
    return int(row["total"] or 0)


def summarize_imports_for_competence(conn: sqlite3.Connection, company_id: int, competence: str) -> dict[str, Any]:
    import_rows = conn.execute(
        """
        SELECT imports.*, users.username AS imported_by_username
        FROM imports
        LEFT JOIN users ON users.id = imports.imported_by
        WHERE imports.company_id = ? AND imports.competence = ?
        ORDER BY datetime(imports.imported_at) DESC, imports.id DESC
        """,
        (company_id, competence),
    ).fetchall()
    summaries: list[dict[str, Any]] = []
    totals = {
        "imports": 0,
        "files": 0,
        "rowsRead": 0,
        "rowsWritten": 0,
        "duplicateRowsSkipped": 0,
        "errors": 0,
        "pendingIssues": 0,
    }
    for import_row in import_rows:
        file_rows = conn.execute(
            """
            SELECT file_type, original_name, row_count, file_hash
            FROM import_files
            WHERE import_id = ?
            ORDER BY id
            """,
            (import_row["id"],),
        ).fetchall()
        pending_issues = int(
            conn.execute(
                "SELECT COUNT(*) AS total FROM import_issues WHERE import_id = ?",
                (import_row["id"],),
            ).fetchone()["total"]
            or 0
        )
        files_summary = []
        rows_read = 0
        rows_written = 0
        for file_row in file_rows:
            read_count = int(file_row["row_count"] or 0)
            written_count = count_rows_for_import_file(conn, import_row["id"], file_row["file_type"])
            rows_read += read_count
            rows_written += written_count
            files_summary.append(
                {
                    "fileType": file_row["file_type"],
                    "fileName": file_row["original_name"],
                    "rowsRead": read_count,
                    "rowsWritten": written_count,
                    "fileHash": file_row["file_hash"],
                }
            )
        summaries.append(
            {
                "importId": import_row["id"],
                "competence": import_row["competence"],
                "importedAt": import_row["imported_at"],
                "importedBy": import_row["imported_by_username"],
                "action": import_row["import_action"],
                "suggestedCompetence": import_row["suggested_competence"],
                "files": files_summary,
                "rowsRead": rows_read,
                "rowsWritten": rows_written,
                "duplicateRowsSkipped": int(import_row["duplicate_rows_skipped"] or 0),
                "errors": 0,
                "pendingIssues": pending_issues,
            }
        )
        totals["imports"] += 1
        totals["files"] += len(files_summary)
        totals["rowsRead"] += rows_read
        totals["rowsWritten"] += rows_written
        totals["duplicateRowsSkipped"] += int(import_row["duplicate_rows_skipped"] or 0)
        totals["pendingIssues"] += pending_issues
    return {"competence": competence, "imports": summaries, "totals": totals}


def audit_revenue_gap_detail(
    conn: sqlite3.Connection,
    company_id: int,
    competence: str,
    dashboard: dict[str, Any],
    filters: dict[str, Any] | None = None,
) -> dict[str, Any]:
    scoped_filters = dict(filters or {})
    scoped_filters.setdefault("unit_name", None)
    scoped_filters.setdefault("seller_name", None)
    scoped_filters.setdefault("city_name", None)
    scoped_filters.setdefault("allowed_units", None)
    allowed_units = normalize_unit_list(scoped_filters.get("allowed_units"))
    scoped_units = [normalize_unit(scoped_filters["unit_name"])] if scoped_filters.get("unit_name") else allowed_units

    detail_rows = conn.execute(
        """
        SELECT seller_name, city_name, net_value, competence, issue_date
        FROM fact_sales_detail
        WHERE company_id = ? AND competence = ?
        """,
        (company_id, competence),
    ).fetchall()
    competence_state = dashboard_competence_state(competence)
    cutoff_date = competence_state["cutoffDate"]

    raw_detail_revenue = 0.0
    raw_detail_rows = 0
    scope_detail_revenue = 0.0
    scope_detail_rows = 0
    unit_revenue_map: dict[str, float] = defaultdict(float)
    seller_revenue_map: dict[str, float] = defaultdict(float)
    missing_unit_revenue_map: dict[str, float] = defaultdict(float)
    missing_seller_revenue_map: dict[str, float] = defaultdict(float)
    missing_city_revenue_map: dict[str, float] = defaultdict(float)

    discarded = {
        "unitWithoutMapping": {"rows": 0, "revenue": 0.0},
        "sellerWithoutMapping": {"rows": 0, "revenue": 0.0},
        "cityWithoutMapping": {"rows": 0, "revenue": 0.0},
        "invalidCompetence": {"rows": 0, "revenue": 0.0},
        "userScope": {"rows": 0, "revenue": 0.0},
        "nonSellerRole": {"rows": 0, "revenue": 0.0},
    }

    # ── PRÉ-CARREGAR LOOKUPS ────────────────────────────────────────────────
    _comp_target_a = first_day_of_competence(competence).isoformat()
    _city_map_a: dict[str, str | None] = {}
    for _r in conn.execute(
        "SELECT city_name, principal_unit FROM city_mappings WHERE company_id = ? AND date(valid_from) <= date(?) AND (valid_to IS NULL OR date(valid_to) >= date(?)) ORDER BY date(valid_from) DESC",
        (company_id, _comp_target_a, _comp_target_a),
    ).fetchall():
        _k = normalize_upper(_r["city_name"])
        if _k and _k not in _city_map_a:
            _city_map_a[_k] = normalize_unit(_r["principal_unit"])
    _seller_role_map_a: dict[str, tuple[str | None, str | None]] = {}
    for _r in conn.execute(
        "SELECT person_name, role_classification, base_unit FROM people_records WHERE company_id = ? AND date(valid_from) <= date(?) AND (valid_to IS NULL OR date(valid_to) >= date(?)) ORDER BY date(valid_from) DESC",
        (company_id, _comp_target_a, _comp_target_a),
    ).fetchall():
        _k = normalize_whitespace(_r["person_name"])
        if _k and _k not in _seller_role_map_a:
            _seller_role_map_a[_k] = (_r["role_classification"], normalize_unit(_r["base_unit"]))
    # ────────────────────────────────────────────────────────────────────────

    for row in detail_rows:
        issue_dt = parse_datetime_flexible(row["issue_date"])
        net_value = float(row["net_value"] or 0.0)
        if competence_state["isFutureCompetence"]:
            discarded["invalidCompetence"]["rows"] += 1
            discarded["invalidCompetence"]["revenue"] += net_value
            continue
        if competence_state["isCurrentCompetence"] and issue_dt and issue_dt.date() > cutoff_date:
            discarded["invalidCompetence"]["rows"] += 1
            discarded["invalidCompetence"]["revenue"] += net_value
            continue

        raw_detail_rows += 1
        raw_detail_revenue += net_value

        seller_name = normalize_whitespace(row["seller_name"])
        city_name = normalize_upper(row["city_name"])
        resolved_unit = _city_map_a.get(city_name)
        role, seller_base_unit = _seller_role_map_a.get(seller_name, (None, None))
        seller_base_unit = normalize_unit(seller_base_unit)

        if not city_name:
            discarded["cityWithoutMapping"]["rows"] += 1
            discarded["cityWithoutMapping"]["revenue"] += net_value
            missing_city_revenue_map["SEM_CIDADE"] += net_value
        if not resolved_unit:
            discarded["unitWithoutMapping"]["rows"] += 1
            discarded["unitWithoutMapping"]["revenue"] += net_value
            missing_unit_revenue_map[city_name or "SEM_CIDADE"] += net_value
        if role is None:
            discarded["sellerWithoutMapping"]["rows"] += 1
            discarded["sellerWithoutMapping"]["revenue"] += net_value
            missing_seller_revenue_map[seller_name or "SEM_VENDEDOR"] += net_value
        elif role != "Vendedor":
            discarded["nonSellerRole"]["rows"] += 1
            discarded["nonSellerRole"]["revenue"] += net_value

        in_scope = True
        if scoped_filters.get("seller_name") and seller_name != scoped_filters["seller_name"]:
            in_scope = False
        if scoped_filters.get("city_name") and city_name != scoped_filters["city_name"]:
            in_scope = False
        if scoped_units and resolved_unit not in scoped_units:
            in_scope = False
        if scoped_filters.get("unit_name") and resolved_unit != scoped_filters["unit_name"]:
            in_scope = False
        if scoped_units and seller_base_unit and scoped_filters.get("seller_name") and seller_base_unit not in scoped_units:
            in_scope = False

        if not in_scope:
            discarded["userScope"]["rows"] += 1
            discarded["userScope"]["revenue"] += net_value
            continue

        scope_detail_rows += 1
        scope_detail_revenue += net_value
        unit_revenue_map[resolved_unit or "SEM_MAPEAMENTO"] += net_value
        seller_revenue_map[seller_name or "SEM_VENDEDOR"] += net_value

    official_unit_rows = conn.execute(
        """
        SELECT unit_name, SUM(sale_value) AS sale_value, SUM(return_value) AS return_value, SUM(net_value) AS net_value
        FROM fact_unit_summary
        WHERE company_id = ? AND competence = ?
        GROUP BY unit_name
        """,
        (company_id, competence),
    ).fetchall()
    official_vendor_rows = conn.execute(
        """
        SELECT seller_name, SUM(sale_value) AS sale_value, SUM(return_value) AS return_value, SUM(net_value) AS net_value
        FROM fact_vendor_summary
        WHERE company_id = ? AND competence = ?
        GROUP BY seller_name
        """,
        (company_id, competence),
    ).fetchall()
    official_unit_map = {
        normalize_unit(row["unit_name"]) or "SEM_MAPEAMENTO": round(float(row["net_value"] or 0.0), 2)
        for row in official_unit_rows
    }
    official_vendor_map = {
        normalize_whitespace(row["seller_name"]) or "SEM_VENDEDOR": round(float(row["net_value"] or 0.0), 2)
        for row in official_vendor_rows
    }

    missing_units = []
    for unit_name in sorted(set(official_unit_map) | set(unit_revenue_map), key=lambda value: value or ""):
        official_value = round(float(official_unit_map.get(unit_name, 0.0)), 2)
        detail_value = round(float(unit_revenue_map.get(unit_name, 0.0)), 2)
        delta_value = round(official_value - detail_value, 2)
        if abs(delta_value) > 0.01:
            missing_units.append(
                {
                    "unitName": unit_name,
                    "officialRevenue": official_value,
                    "detailRevenue": detail_value,
                    "deltaRevenue": delta_value,
                }
            )

    missing_sellers = []
    for seller_name in sorted(set(official_vendor_map) | set(seller_revenue_map), key=lambda value: value or ""):
        official_value = round(float(official_vendor_map.get(seller_name, 0.0)), 2)
        detail_value = round(float(seller_revenue_map.get(seller_name, 0.0)), 2)
        delta_value = round(official_value - detail_value, 2)
        if abs(delta_value) > 0.01:
            missing_sellers.append(
                {
                    "sellerName": seller_name,
                    "officialRevenue": official_value,
                    "detailRevenue": detail_value,
                    "deltaRevenue": delta_value,
                }
            )

    missing_cities = [
        {"cityName": city_name, "revenue": round(revenue, 2)}
        for city_name, revenue in sorted(missing_city_revenue_map.items(), key=lambda item: item[1], reverse=True)
        if abs(revenue) > 0.01
    ]

    official_unit_revenue = round(sum(official_unit_map.values()), 2)
    official_vendor_revenue = round(sum(official_vendor_map.values()), 2)
    dashboard_revenue = round(float(dashboard["summary"]["revenueNet"] or 0.0), 2)
    unit_revenue_sum = round(sum(float(row["revenueNet"] or 0.0) for row in dashboard["unitPerformance"]), 2)
    seller_revenue_sum = round(sum(float(row["revenueNet"] or 0.0) for row in dashboard["sellerRanking"]), 2)

    return {
        "competence": competence,
        "officialUnitRevenue": official_unit_revenue,
        "officialVendorRevenue": official_vendor_revenue,
        "importedRevenue": round(raw_detail_revenue, 2),
        "detailRevenue": round(raw_detail_revenue, 2),
        "importedRows": raw_detail_rows,
        "dashboardRevenue": dashboard_revenue,
        "unitRevenueSum": unit_revenue_sum,
        "sellerRevenueSum": seller_revenue_sum,
        "discardedRevenue": round(official_unit_revenue - dashboard_revenue, 2),
        "discardedRows": {
            "total": sum(int(item["rows"]) for item in discarded.values()),
            **{key: {"rows": int(value["rows"]), "revenue": round(float(value["revenue"]), 2)} for key, value in discarded.items()},
        },
        "missingUnits": missing_units,
        "missingSellers": missing_sellers,
        "missingCities": missing_cities,
        "sourceGapRevenue": round(official_unit_revenue - raw_detail_revenue, 2),
        "gapUnitVsDetail": round(official_unit_revenue - raw_detail_revenue, 2),
        "gapSellerVsDetail": round(official_vendor_revenue - raw_detail_revenue, 2),
        "gapUnitVsSeller": round(official_unit_revenue - official_vendor_revenue, 2),
        "sourceGapRows": max(raw_detail_rows - scope_detail_rows, 0),
        "scopeRevenue": round(scope_detail_revenue, 2),
        "scopeRows": scope_detail_rows,
        "mappingRevenue": {
            "missingUnitMapping": round(sum(item["revenue"] for item in [discarded["unitWithoutMapping"]]), 2),
            "missingSellerMapping": round(sum(item["revenue"] for item in [discarded["sellerWithoutMapping"]]), 2),
            "missingCityMapping": round(sum(item["revenue"] for item in [discarded["cityWithoutMapping"]]), 2),
            "nonSellerRole": round(sum(item["revenue"] for item in [discarded["nonSellerRole"]]), 2),
        },
        "classification": {
            "executiveSummary": "OK" if audit_same(dashboard_revenue, official_unit_revenue) else ("ERRO" if audit_same(dashboard_revenue, raw_detail_revenue) else "ATENCAO"),
            "sellerSummary": "OK" if audit_same(seller_revenue_sum, official_vendor_revenue) else "ATENCAO",
        },
    }


def audit_revenue_integrity(
    conn: sqlite3.Connection,
    company_id: int,
    competence: str,
    dashboard: dict[str, Any],
    issues: list[dict[str, Any]],
) -> dict[str, Any]:
    revenue_gap = audit_revenue_gap_detail(conn, company_id, competence, dashboard, dashboard.get("filters"))
    summary = dashboard["summary"]
    unit_rows = dashboard["unitPerformance"]
    seller_rows = dashboard["sellerRanking"]
    city_rows = dashboard["cityRanking"]
    client_rows = dashboard["clientRanking"]
    official_unit = conn.execute(
        """
        SELECT
            ROUND(COALESCE(SUM(net_value), 0), 2) AS revenue_net,
            ROUND(COALESCE(SUM(return_value), 0), 2) AS returns_value,
            ROUND(COALESCE(SUM(sale_value), 0), 2) AS gross_sales
        FROM fact_unit_summary
        WHERE company_id = ? AND competence = ?
        """,
        (company_id, competence),
    ).fetchone()
    official_vendor = conn.execute(
        """
        SELECT ROUND(COALESCE(SUM(net_value), 0), 2) AS revenue_net
        FROM fact_vendor_summary
        WHERE company_id = ? AND competence = ?
        """,
        (company_id, competence),
    ).fetchone()
    detail_scope = conn.execute(
        """
        SELECT
            ROUND(COALESCE(SUM(net_value), 0), 2) AS revenue_net,
            ROUND(COALESCE(SUM(return_value), 0), 2) AS returns_value,
            ROUND(COALESCE(SUM(gross_value), 0), 2) AS gross_sales,
            ROUND(COALESCE(SUM(CASE WHEN city_name IS NULL OR city_name = '' THEN net_value ELSE 0 END), 0), 2) AS revenue_without_city
        FROM fact_sales_detail
        WHERE company_id = ? AND competence = ?
        """,
        (company_id, competence),
    ).fetchone()
    summary_revenue = audit_round(summary["revenueNet"])
    unit_sum = audit_round(sum(row["revenueNet"] for row in unit_rows))
    seller_sum = audit_round(sum(row["revenueNet"] for row in seller_rows))
    city_sum = audit_round(sum(row["revenueNet"] for row in city_rows))
    client_sum = audit_round(sum(row["revenueNet"] for row in client_rows))
    official_unit_revenue = audit_round(official_unit["revenue_net"])
    official_vendor_revenue = audit_round(official_vendor["revenue_net"])
    detail_revenue = audit_round(detail_scope["revenue_net"])
    missing_city_revenue = audit_round(detail_scope["revenue_without_city"])
    if not audit_same(revenue_gap["officialUnitRevenue"], revenue_gap["importedRevenue"]):
        append_audit_issue(
            issues,
            "CRITICO",
            "FATURAMENTO",
            "A base detalhada importada não bate com o valor oficial do custo por unidade.",
            revenue_gap["officialUnitRevenue"],
            revenue_gap["importedRevenue"],
        )
    if not audit_same(summary_revenue, unit_sum):
        append_audit_issue(issues, "CRITICO", "FATURAMENTO", "Resumo do grupo não bate com a soma das unidades.", unit_sum, summary_revenue)
    if not audit_same(summary_revenue, official_unit_revenue):
        append_audit_issue(issues, "CRITICO", "FATURAMENTO", "Resumo do grupo não bate com o valor oficial do custo por unidade.", official_unit_revenue, summary_revenue)
    if not audit_same(seller_sum, official_vendor_revenue):
        append_audit_issue(issues, "ATENCAO", "FATURAMENTO", "Soma do ranking de vendedores diverge do valor oficial do custo por vendedor.", official_vendor_revenue, seller_sum)
    expected_city_sum = audit_round(detail_revenue - missing_city_revenue)
    if not audit_same(city_sum, expected_city_sum):
        append_audit_issue(issues, "ATENCAO", "FATURAMENTO", "Soma das cidades diverge do faturamento detalhado esperado para cidades válidas.", expected_city_sum, city_sum)
    if not audit_same(client_sum, detail_revenue):
        append_audit_issue(issues, "ATENCAO", "FATURAMENTO", "Soma do ranking de clientes diverge do faturamento detalhado.", detail_revenue, client_sum)
    filter_samples: dict[str, Any] = {}
    if unit_rows:
        sample_unit = unit_rows[0]["unitName"]
        sample_unit_summary = single_competence_summary(conn, company_id, competence, audit_filters_for_competence(competence, unit_name=sample_unit))
        filter_samples["unit"] = {"unitName": sample_unit, "summaryRevenueNet": sample_unit_summary.get("revenueNet"), "rankingRevenueNet": unit_rows[0]["revenueNet"]}
        if not audit_same(sample_unit_summary.get("revenueNet"), unit_rows[0]["revenueNet"]):
            append_audit_issue(issues, "CRITICO", "FATURAMENTO", "Filtro por unidade não retorna o mesmo faturamento da linha da unidade no ranking.", unit_rows[0]["revenueNet"], sample_unit_summary.get("revenueNet"))
    if seller_rows:
        sample_seller = seller_rows[0]["sellerName"]
        sample_seller_summary = single_competence_summary(conn, company_id, competence, audit_filters_for_competence(competence, seller_name=sample_seller))
        filter_samples["seller"] = {"sellerName": sample_seller, "summaryRevenueNet": sample_seller_summary.get("revenueNet"), "rankingRevenueNet": seller_rows[0]["revenueNet"]}
        if not audit_same(sample_seller_summary.get("revenueNet"), seller_rows[0]["revenueNet"]):
            append_audit_issue(issues, "CRITICO", "FATURAMENTO", "Filtro por vendedor não retorna o mesmo faturamento da linha do vendedor no ranking.", seller_rows[0]["revenueNet"], sample_seller_summary.get("revenueNet"))
    return {
        "summaryRevenueNet": summary_revenue,
        "officialUnitRevenueNet": official_unit_revenue,
        "officialVendorRevenueNet": official_vendor_revenue,
        "detailRevenueNet": detail_revenue,
        "sumUnitsRevenueNet": unit_sum,
        "sumSellersRevenueNet": seller_sum,
        "sumCitiesRevenueNet": city_sum,
        "sumClientsRevenueNet": client_sum,
        "revenueWithoutCity": missing_city_revenue,
        "filterSamples": filter_samples,
        "gapDetail": revenue_gap,
    }


def audit_goals_integrity(
    conn: sqlite3.Connection,
    company_id: int,
    competence: str,
    dashboard: dict[str, Any],
    issues: list[dict[str, Any]],
) -> dict[str, Any]:
    summary_goal = audit_round(dashboard["summary"]["revenueGoal"])
    unit_sum = audit_round(sum(row["revenueGoal"] for row in dashboard["unitPerformance"]))
    seller_sum = audit_round(sum(row["revenueGoal"] for row in dashboard["sellerRanking"]))
    goal_by_seller, duplicate_seller = load_goal_maps(conn, company_id, competence, "goals_seller", "seller_name", normalize_whitespace)
    goal_by_unit, duplicate_unit = load_goal_maps(conn, company_id, competence, "goals_unit", "unit_name", normalize_unit)
    duplicate_unit_normalized = normalized_goal_duplicates(conn, company_id, "goals_unit", "unit_name", normalize_unit)
    duplicate_seller_normalized = normalized_goal_duplicates(conn, company_id, "goals_seller", "seller_name", normalize_whitespace)
    official_unit_goal = audit_round(sum(item["revenueGoal"] for item in goal_by_unit.values()))
    official_seller_goal = audit_round(sum(item["revenueGoal"] for item in goal_by_seller.values()))
    if not audit_same(summary_goal, official_unit_goal):
        append_audit_issue(issues, "CRITICO", "META", "Meta do grupo não bate com a soma das metas das unidades.", official_unit_goal, summary_goal)
    if not audit_same(summary_goal, unit_sum):
        append_audit_issue(issues, "CRITICO", "META", "Resumo da meta do grupo não bate com a soma das metas exibidas por unidade.", unit_sum, summary_goal)
    if duplicate_unit:
        append_audit_issue(issues, "ATENCAO", "META", "Foram encontradas metas duplicadas de unidade na competência.", "Sem duplicidades", duplicate_unit)
    if duplicate_seller:
        append_audit_issue(issues, "ATENCAO", "META", "Foram encontradas metas duplicadas de vendedor na competência.", "Sem duplicidades", duplicate_seller)
    for duplicate in duplicate_unit_normalized:
        append_audit_issue(
            issues,
            "CRITICO" if duplicate["competence"] == competence else "ATENCAO",
            "META",
            "Foram encontradas metas duplicadas por unidade normalizada.",
            "Uma meta por unidade normalizada",
            duplicate,
        )
    for duplicate in duplicate_seller_normalized:
        append_audit_issue(
            issues,
            "CRITICO" if duplicate["competence"] == competence else "ATENCAO",
            "META",
            "Foram encontradas metas duplicadas por vendedor normalizado.",
            "Uma meta por vendedor normalizado",
            duplicate,
        )
    sample_checks: dict[str, Any] = {}
    if dashboard["unitPerformance"]:
        sample_unit = dashboard["unitPerformance"][0]
        expected = audit_round(goal_by_unit.get(normalize_unit(sample_unit["unitName"]), {}).get("revenueGoal"))
        actual = audit_round(sample_unit["revenueGoal"])
        sample_checks["unit"] = {"unitName": sample_unit["unitName"], "expectedGoal": expected, "actualGoal": actual}
        if not audit_same(expected, actual):
            append_audit_issue(issues, "CRITICO", "META", "Meta exibida na unidade diverge da meta cadastrada.", expected, actual)
    if dashboard["sellerRanking"]:
        sample_seller = dashboard["sellerRanking"][0]
        expected = audit_round(goal_by_seller.get(normalize_whitespace(sample_seller["sellerName"]), {}).get("revenueGoal"))
        actual = audit_round(sample_seller["revenueGoal"])
        sample_checks["seller"] = {"sellerName": sample_seller["sellerName"], "expectedGoal": expected, "actualGoal": actual}
        if not audit_same(expected, actual):
            append_audit_issue(issues, "CRITICO", "META", "Meta exibida no vendedor diverge da meta cadastrada.", expected, actual)
    return {
        "summaryRevenueGoal": summary_goal,
        "officialUnitRevenueGoal": official_unit_goal,
        "dashboardUnitGoalsSum": unit_sum,
        "officialSellerGoalsSum": official_seller_goal,
        "dashboardSellerGoalsSum": seller_sum,
        "duplicateUnitGoals": duplicate_unit,
        "duplicateSellerGoals": duplicate_seller,
        "duplicateUnitGoalsNormalized": [
            {
                "competence": item["competence"],
                "normalizedUnit": item["normalizedKey"],
                "rows": [
                    {
                        "id": row["id"],
                        "unitName": row["goalKey"],
                        "revenueGoal": row["revenueGoal"],
                    }
                    for row in item["rows"]
                ],
                "recommendation": "Manter unidade normalizada e excluir duplicadas",
            }
            for item in duplicate_unit_normalized
        ],
        "duplicateSellerGoalsNormalized": [
            {
                "competence": item["competence"],
                "normalizedSeller": item["normalizedKey"],
                "rows": [
                    {
                        "id": row["id"],
                        "sellerName": row["goalKey"],
                        "revenueGoal": row["revenueGoal"],
                    }
                    for row in item["rows"]
                ],
            }
            for item in duplicate_seller_normalized
        ],
        "sampleChecks": sample_checks,
    }


def audit_projection_integrity(
    conn: sqlite3.Connection,
    company_id: int,
    competence: str,
    dashboard: dict[str, Any],
    issues: list[dict[str, Any]],
) -> dict[str, Any]:
    summary = dashboard["summary"]
    calendar = dashboard["calendar"]
    debug_projection = dashboard["debugProjection"]
    competence_state = dashboard_competence_state(competence)
    expected_daily = audit_round(safe_div(summary["revenueNet"], calendar["elapsedWorkingDays"]))
    expected_projected = audit_round(expected_daily * calendar["totalWorkingDays"]) if calendar["elapsedWorkingDays"] and calendar["totalWorkingDays"] else 0.0
    expected_daily_goal = audit_round(safe_div(summary["revenueGoal"], calendar["totalWorkingDays"]))
    expected_goal_attainment = audit_round(safe_div(summary["revenueNet"], summary["revenueGoal"]) * 100 if summary["revenueGoal"] else 0.0)
    expected_projected_goal_attainment = audit_round(safe_div(expected_projected, summary["revenueGoal"]) * 100 if summary["revenueGoal"] else 0.0)
    expected_cutoff = dashboard_cutoff_date(competence_state["today"]).isoformat()
    if debug_projection.get("cutoffDate") != expected_cutoff:
        append_audit_issue(issues, "CRITICO", "PROJECAO", "Cutoff D-1 está incorreto.", expected_cutoff, debug_projection.get("cutoffDate"))
    if competence_state["isCurrentCompetence"] and calendar["effectiveToday"] != expected_cutoff:
        append_audit_issue(issues, "CRITICO", "PROJECAO", "Calendário da competência atual está incluindo o dia atual.", expected_cutoff, calendar["effectiveToday"])
    if competence_state["isPastCompetence"] and calendar["elapsedWorkingDays"] != calendar["totalWorkingDays"]:
        append_audit_issue(issues, "CRITICO", "PROJECAO", "Competência passada não está fechando com mês completo.", calendar["totalWorkingDays"], calendar["elapsedWorkingDays"])
    future_competence = shift_competence(competence, 1)
    future_summary = single_competence_summary(conn, company_id, future_competence, audit_filters_for_competence(future_competence))
    if future_summary and audit_round(future_summary.get("revenueNet")) == 0.0 and audit_round(future_summary.get("dailyRevenue")) != 0.0:
        append_audit_issue(issues, "ATENCAO", "PROJECAO", "Competência futura tem média diária diferente de zero.", 0.0, future_summary.get("dailyRevenue"))
    for label, expected, actual in [
        ("dailyRevenueActual", expected_daily, summary["dailyRevenueActual"]),
        ("dailyGoal", expected_daily_goal, summary["dailyGoal"]),
        ("projectedRevenue", expected_projected, summary["projectedRevenue"]),
        ("goalAttainmentPct", expected_goal_attainment, summary["goalAttainmentPct"]),
        ("projectedGoalAttainmentPct", expected_projected_goal_attainment, summary["projectedGoalAttainmentPct"]),
    ]:
        if not audit_same(expected, actual):
            append_audit_issue(issues, "CRITICO", "PROJECAO", f"Cálculo de {label} divergente.", expected, actual)
    return {
        "calendar": calendar,
        "cutoffDate": debug_projection.get("cutoffDate"),
        "expectedCutoffDate": expected_cutoff,
        "dailyRevenueActual": {"expected": expected_daily, "actual": audit_round(summary["dailyRevenueActual"])},
        "dailyGoal": {"expected": expected_daily_goal, "actual": audit_round(summary["dailyGoal"])},
        "projectedRevenue": {"expected": expected_projected, "actual": audit_round(summary["projectedRevenue"])},
        "goalAttainmentPct": {"expected": expected_goal_attainment, "actual": audit_round(summary["goalAttainmentPct"])},
        "projectedGoalAttainmentPct": {"expected": expected_projected_goal_attainment, "actual": audit_round(summary["projectedGoalAttainmentPct"])},
        "futureCompetenceCheck": {"competence": future_competence, "dailyRevenue": future_summary.get("dailyRevenue") if future_summary else None},
    }


def audit_comparisons_integrity(
    conn: sqlite3.Connection,
    company_id: int,
    competence: str,
    dashboard: dict[str, Any],
    issues: list[dict[str, Any]],
) -> dict[str, Any]:
    previous = dashboard["comparisons"]["previousCompetence"]
    yoy = dashboard["comparisons"]["yearOverYear"]
    previous_expected = shift_competence(competence, -1)
    yoy_expected = shift_competence(competence, -12)
    if previous.get("competence") and previous.get("competence") != previous_expected:
        append_audit_issue(issues, "CRITICO", "COMPARATIVO", "Comparativo do mês anterior está apontando competência incorreta.", previous_expected, previous.get("competence"))
    if yoy.get("competence") and yoy.get("competence") != yoy_expected:
        append_audit_issue(issues, "CRITICO", "COMPARATIVO", "Comparativo ano contra ano está apontando competência incorreta.", yoy_expected, yoy.get("competence"))
    scope_samples: dict[str, Any] = {}
    if dashboard["unitPerformance"]:
        sample_unit = dashboard["unitPerformance"][0]["unitName"]
        sample_previous = single_competence_summary(conn, company_id, previous_expected, audit_filters_for_competence(previous_expected, unit_name=sample_unit))
        sample_yoy = single_competence_summary(conn, company_id, yoy_expected, audit_filters_for_competence(yoy_expected, unit_name=sample_unit))
        scope_samples["unit"] = {"unitName": sample_unit, "previous": sample_previous, "yearOverYear": sample_yoy}
        if sample_previous.get("unit") != sample_unit or sample_yoy.get("unit") != sample_unit:
            append_audit_issue(issues, "CRITICO", "COMPARATIVO", "Comparativos filtrados por unidade não mantêm o mesmo escopo.", sample_unit, {"previous": sample_previous.get("unit"), "yearOverYear": sample_yoy.get("unit")})
    if dashboard["sellerRanking"]:
        sample_seller = dashboard["sellerRanking"][0]["sellerName"]
        sample_previous = single_competence_summary(conn, company_id, previous_expected, audit_filters_for_competence(previous_expected, seller_name=sample_seller))
        sample_yoy = single_competence_summary(conn, company_id, yoy_expected, audit_filters_for_competence(yoy_expected, seller_name=sample_seller))
        scope_samples["seller"] = {"sellerName": sample_seller, "previous": sample_previous, "yearOverYear": sample_yoy}
        if sample_previous.get("seller") != sample_seller or sample_yoy.get("seller") != sample_seller:
            append_audit_issue(issues, "CRITICO", "COMPARATIVO", "Comparativos filtrados por vendedor não mantêm o mesmo escopo.", sample_seller, {"previous": sample_previous.get("seller"), "yearOverYear": sample_yoy.get("seller")})
    return {
        "group": {"current": competence, "previousExpected": previous_expected, "yearOverYearExpected": yoy_expected, "previousActual": previous, "yearOverYearActual": yoy},
        "scopeSamples": scope_samples,
    }


def audit_crm_integrity(
    conn: sqlite3.Connection,
    company_id: int,
    competence: str,
    issues: list[dict[str, Any]],
) -> dict[str, Any]:
    filters = audit_filters_for_competence(competence)
    all_rows = list_crm_clients(conn, company_id, filters, attach_context=False)
    base_count = len(all_rows)
    filtered_base = filter_crm_client_rows(all_rows, filters)
    page_50_rows_raw = filtered_base[:50]
    page_100_rows_raw = filtered_base[:100]
    page_50_rows = crm_attach_context(conn, company_id, page_50_rows_raw)
    page_100_rows = crm_attach_context(conn, company_id, page_100_rows_raw)
    page_50 = {
        "total": len(filtered_base),
        "page": 1,
        "pageSize": 50,
        "totalPages": max(math.ceil(len(filtered_base) / 50), 1) if filtered_base else 1,
        "rows": page_50_rows,
    }
    page_100 = {
        "total": len(filtered_base),
        "page": 1,
        "pageSize": 100,
        "totalPages": max(math.ceil(len(filtered_base) / 100), 1) if filtered_base else 1,
        "rows": page_100_rows,
    }
    first_row = page_50_rows[0] if page_50_rows else None
    sample_status = first_row["statusCode"] if first_row else None
    sample_class = normalize_upper(first_row["classCode"]) if first_row else None
    status_page_total = len(filter_crm_client_rows(all_rows, {**filters, "status": sample_status})) if sample_status else 0
    class_page_total = len(filter_crm_client_rows(all_rows, {**filters, "classCode": sample_class})) if sample_class else 0
    purchase_with_total = len(filter_crm_client_rows(all_rows, {**filters, "purchaseMonth": "COM_COMPRA"}))
    purchase_without_total = len(filter_crm_client_rows(all_rows, {**filters, "purchaseMonth": "SEM_COMPRA"}))
    growth_above_total = len(filter_crm_client_rows(all_rows, {**filters, "growth": "ACIMA"}))
    growth_stable_total = len(filter_crm_client_rows(all_rows, {**filters, "growth": "ESTAVEL"}))
    growth_below_total = len(filter_crm_client_rows(all_rows, {**filters, "growth": "ABAIXO"}))
    profile_count = int(conn.execute("SELECT COUNT(*) AS total FROM crm_client_profiles WHERE company_id = ?", (company_id,)).fetchone()["total"] or 0)
    summary_distinct_count = int(
        conn.execute(
            "SELECT COUNT(DISTINCT client_code) AS total FROM crm_client_summary WHERE company_id = ? AND competence = ?",
            (company_id, competence),
        ).fetchone()["total"]
        or 0
    )
    detail_checks: dict[str, Any] = {}
    expected_pages_50 = max(math.ceil(base_count / 50), 1) if base_count else 1
    expected_pages_100 = max(math.ceil(base_count / 100), 1) if base_count else 1
    if page_50["total"] != base_count:
        append_audit_issue(issues, "CRITICO", "CRM", "Total do CRM Clientes diverge da base real.", base_count, page_50["total"])
    if page_50["totalPages"] != expected_pages_50:
        append_audit_issue(issues, "CRITICO", "CRM", "Paginação do CRM Clientes com pageSize 50 está incorreta.", expected_pages_50, page_50["totalPages"])
    if page_100["totalPages"] != expected_pages_100:
        append_audit_issue(issues, "CRITICO", "CRM", "Paginação do CRM Clientes com pageSize 100 está incorreta.", expected_pages_100, page_100["totalPages"])
    search_sample_total = 0
    if first_row:
        search_sample_total = len(filter_crm_client_rows(all_rows, {**filters, "search": first_row["clientKey"]}))
        summary_data = get_crm_client_summary(conn, company_id, filters, first_row["clientKey"])
        purchases_data = get_crm_client_purchases(conn, company_id, filters, first_row["clientKey"])
        items_data = get_crm_client_items(conn, company_id, filters, first_row["clientKey"], 1, 20)
        interactions_data = get_crm_client_interactions(conn, company_id, filters, first_row["clientKey"], 1, 20)
        tasks_data = get_crm_client_tasks(conn, company_id, filters, first_row["clientKey"])
        detail_checks = {
            "clientKey": first_row["clientKey"],
            "summaryFound": summary_data is not None,
            "purchasesFound": purchases_data is not None,
            "itemsFound": items_data is not None,
            "interactionsFound": interactions_data is not None,
            "tasksFound": tasks_data is not None,
            "searchRows": search_sample_total,
        }
        if not summary_data:
            append_audit_issue(issues, "CRITICO", "CRM", "Ficha 360 não encontrou o cliente de teste no summary.", True, False)
        if items_data is None:
            append_audit_issue(issues, "CRITICO", "CRM", "Ficha 360 não encontrou itens do cliente de teste.", "rows ou lista vazia", None)
        if interactions_data is None:
            append_audit_issue(issues, "CRITICO", "CRM", "Ficha 360 não encontrou interações do cliente de teste.", "rows ou lista vazia", None)
        if search_sample_total <= 0:
            append_audit_issue(issues, "ATENCAO", "CRM", "Busca por clientKey não retornou o cliente de teste.", first_row["clientKey"], 0)
    return {
        "crmClientProfiles": profile_count,
        "crmClientSummaryDistinct": summary_distinct_count,
        "baseCount": int(base_count or 0),
        "page50": {"total": page_50["total"], "page": page_50["page"], "pageSize": page_50["pageSize"], "totalPages": page_50["totalPages"], "rowsReturned": len(page_50["rows"])},
        "page100": {"total": page_100["total"], "page": page_100["page"], "pageSize": page_100["pageSize"], "totalPages": page_100["totalPages"], "rowsReturned": len(page_100["rows"])},
        "filters": {
            "status": {"value": sample_status, "total": status_page_total},
            "purchaseWith": {"total": purchase_with_total},
            "purchaseWithout": {"total": purchase_without_total},
            "growthAbove": {"total": growth_above_total},
            "growthStable": {"total": growth_stable_total},
            "growthBelow": {"total": growth_below_total},
            "classCode": {"value": sample_class, "total": class_page_total},
            "search": {"query": first_row["clientKey"] if first_row else None, "total": search_sample_total},
        },
        "detailChecks": detail_checks,
    }


def audit_permissions_integrity(
    conn: sqlite3.Connection,
    company_id: int,
    competence: str,
    issues: list[dict[str, Any]],
) -> dict[str, Any]:
    all_users = [dict(row) for row in conn.execute("SELECT * FROM users WHERE company_id = ? AND is_active = 1 ORDER BY id", (company_id,)).fetchall()]
    users: list[dict[str, Any]] = []
    seen_roles: set[str] = set()
    for user in all_users:
        role = user["role"]
        if role in seen_roles:
            continue
        users.append(user)
        seen_roles.add(role)
    checks: list[dict[str, Any]] = []
    for user in users:
        scoped = scoped_filters_for_user(conn, company_id, user, audit_filters_for_competence(competence))
        if user["role"] == "Vendedor":
            expected_seller = seller_identity_for_user(user)
            if scoped.get("seller_name") != expected_seller:
                append_audit_issue(issues, "CRITICO", "PERMISSAO", "Escopo do vendedor não foi travado no vendedor vinculado.", expected_seller, scoped.get("seller_name"))
            allowed_units = []
        elif user["role"] in {"Gerente", "Analista"}:
            expected_units = linked_units_for_user(user)
            allowed_units = normalize_unit_list(scoped.get("allowed_units"))
            if allowed_units != expected_units:
                append_audit_issue(issues, "CRITICO", "PERMISSAO", "Escopo do gerente/analista não corresponde às unidades vinculadas.", expected_units, allowed_units)
        else:
            allowed_units = []
        dashboard = get_dashboard_data(conn, company_id, scoped)
        visible_units = sorted({normalize_unit(row["unitName"]) for row in dashboard["unitPerformance"]})
        if user["role"] in {"Gerente", "Analista"} and allowed_units:
            unauthorized = [unit for unit in visible_units if unit not in allowed_units]
            if unauthorized:
                append_audit_issue(issues, "CRITICO", "PERMISSAO", "Usuário de unidade está vendo unidades fora do vínculo.", allowed_units, unauthorized)
        if user["role"] == "Vendedor":
            expected_seller = seller_identity_for_user(user)
            visible_sellers = {normalize_whitespace(row["sellerName"]) for row in dashboard["sellerRanking"]}
            if visible_sellers and visible_sellers != {expected_seller}:
                append_audit_issue(issues, "CRITICO", "PERMISSAO", "Vendedor está vendo ranking fora do próprio escopo.", [expected_seller], sorted(visible_sellers))
        checks.append(
            {
                "username": user["username"],
                "role": user["role"],
                "linkedPersonName": user.get("linked_person_name"),
                "linkedUnits": linked_units_for_user(user),
                "scopedFilters": {
                    "unit": scoped.get("unit_name"),
                    "seller": scoped.get("seller_name"),
                    "city": scoped.get("city_name"),
                    "allowedUnits": normalize_unit_list(scoped.get("allowed_units")),
                },
                "visibleUnits": visible_units,
                "sellerCount": len(dashboard["sellerRanking"]),
            }
        )
    return {"users": checks}


def build_integrity_audit(conn: sqlite3.Connection, company_id: int, competence: str) -> dict[str, Any]:
    issues: list[dict[str, Any]] = []
    dashboard = get_dashboard_data(conn, company_id, audit_filters_for_competence(competence))
    imports_check = summarize_imports_for_competence(conn, company_id, competence)
    revenue_check = audit_revenue_integrity(conn, company_id, competence, dashboard, issues)
    goals_check = audit_goals_integrity(conn, company_id, competence, dashboard, issues)
    projection_check = audit_projection_integrity(conn, company_id, competence, dashboard, issues)
    comparisons_check = audit_comparisons_integrity(conn, company_id, competence, dashboard, issues)
    crm_check = audit_crm_integrity(conn, company_id, competence, issues)
    permission_check = audit_permissions_integrity(conn, company_id, competence, issues)
    return {
        "competence": competence,
        "revenueSourcePolicy": {
            "executiveSummary": "fact_unit_summary",
            "unitPerformance": "fact_unit_summary",
            "sellerRanking": "fact_vendor_summary",
            "cityRanking": "fact_sales_detail",
            "clientRanking": "fact_sales_detail",
            "crm": "crm_client_summary + fact_sales_detail",
        },
        "imports": imports_check,
        "revenueCheck": revenue_check,
        "revenueGapReport": revenue_check.get("gapDetail", {}),
        "goalsCheck": goals_check,
        "projectionCheck": projection_check,
        "comparisonsCheck": comparisons_check,
        "crmCheck": crm_check,
        "permissionCheck": permission_check,
        "issues": issues,
    }


def filter_admin_data_for_user(
    conn: sqlite3.Connection, company_id: int, user: sqlite3.Row, data: dict[str, Any]
) -> dict[str, Any]:
    """Aplica o recorte por unidade nos cadastros administrativos.

    Um gerente só administra a própria unidade: metas de vendedor da equipe dele,
    meta da unidade dele e férias das pessoas da unidade dele. Quem tem escopo
    "todos" não sofre restrição.
    """
    scope = data_scope_for_user(conn, user)
    if scope not in {"unidade", "unidade_consolidado"}:
        return data
    allowed = set(linked_units_for_user(user))
    if not allowed:
        for key in ("goalsSeller", "goalsUnit", "vacations", "people"):
            data[key] = []
        return data

    # Mapa pessoa -> unidade base, para filtrar férias e metas sem base_unit gravada
    person_unit: dict[str, str] = {}
    for r in conn.execute(
        "SELECT person_name, base_unit FROM people_records WHERE company_id = ? ORDER BY valid_from DESC",
        (company_id,),
    ).fetchall():
        key = normalize_whitespace(r["person_name"])
        if key and key not in person_unit:
            person_unit[key] = normalize_unit(r["base_unit"])

    def unit_of_person(name: str | None) -> str:
        return person_unit.get(normalize_whitespace(name), "")

    data["goalsSeller"] = [
        g for g in data.get("goalsSeller", [])
        if (normalize_unit(g.get("base_unit")) or unit_of_person(g.get("seller_name"))) in allowed
    ]
    data["goalsUnit"] = [
        g for g in data.get("goalsUnit", []) if normalize_unit(g.get("unit_name")) in allowed
    ]
    data["vacations"] = [
        v for v in data.get("vacations", []) if unit_of_person(v.get("person_name")) in allowed
    ]
    data["people"] = [
        p for p in data.get("people", []) if normalize_unit(p.get("base_unit")) in allowed
    ]
    # Configuração de score é exclusiva de quem administra o sistema
    data["scoreConfigs"] = []
    return data


def list_admin_data(conn: sqlite3.Connection, company_id: int) -> dict[str, Any]:
    sanitize_unit_goals(conn, company_id)
    conn.commit()
    users: list[dict[str, Any]] = []
    for row in conn.execute(
        """
        SELECT id, username, full_name, linked_person_name, linked_units_json, role, profile_id,
               is_active, created_at
        FROM users
        WHERE company_id = ?
        ORDER BY username
        """,
        (company_id,),
    ).fetchall():
        item = dict(row)
        linked_units = normalize_unit_list(item.get("linked_units_json"))
        item["linked_units"] = linked_units
        item["linked_units_display"] = ", ".join(linked_units)
        # Unidade do VENDEDOR. Ela não fica na conta: mora no cadastro de
        # pessoas, e é de lá que sai o filtro da carteira, a meta e a leitura
        # de unidade em implantação. Sem devolver aqui, a tela de Acessos não
        # tinha como mostrar que um vendedor recém-criado está sem unidade.
        item["person_unit"] = person_base_unit(conn, company_id, item.get("linked_person_name"))
        users.append(item)
    profiles = list_access_profiles(conn, company_id)
    profile_name_by_id = {p["id"]: p["name"] for p in profiles}
    for item in users:
        item["profile_name"] = profile_name_by_id.get(item.get("profile_id")) or item.get("role") or ""
    # A tela de Acessos precisa da lista de unidades para marcar o vínculo do
    # gerente. Antes ela dependia dos filtros do Dashboard: quem entrava direto
    # em Acessos via a seção "Unidades vinculadas" sem nenhuma caixa para
    # marcar. Aqui a lista é montada de todas as fontes que conhecem unidade,
    # inclusive a que ainda não faturou (unit_phases) — é o caso da Zona Norte.
    unidades: set[str] = set(CANONICAL_UNITS)
    for sql in (
        "SELECT DISTINCT unit_name AS u FROM fact_unit_summary WHERE company_id = ?",
        "SELECT DISTINCT base_unit AS u FROM people_records WHERE company_id = ?",
        "SELECT DISTINCT unit_name AS u FROM goals_unit WHERE company_id = ?",
        "SELECT DISTINCT unit_name AS u FROM unit_phases WHERE company_id = ?",
    ):
        try:
            for row in conn.execute(sql, (company_id,)).fetchall():
                nome = normalize_unit(row["u"])
                if nome:
                    unidades.add(nome)
        except sqlite3.OperationalError:
            continue

    return {
        "users": users,
        "profiles": profiles,
        "accessModules": ACCESS_MODULES,
        "dataScopes": DATA_SCOPES,
        "units": sorted(unidades),
        "clients": [dict(row) for row in conn.execute("SELECT * FROM client_registry WHERE company_id = ? ORDER BY updated_at DESC, client_name LIMIT 300", (company_id,)).fetchall()],
        "people": [dict(row) for row in conn.execute("SELECT * FROM people_records WHERE company_id = ? ORDER BY person_name, valid_from DESC", (company_id,)).fetchall()],
        "salesSellers": [row["seller_name"] for row in conn.execute("SELECT DISTINCT seller_name FROM fact_sales_detail WHERE company_id = ? AND seller_name IS NOT NULL AND TRIM(seller_name) <> '' ORDER BY seller_name", (company_id,)).fetchall()],
        "salesCities": [row["city_name"] for row in conn.execute("SELECT DISTINCT city_name FROM fact_sales_detail WHERE company_id = ? AND city_name IS NOT NULL AND TRIM(city_name) <> '' ORDER BY city_name", (company_id,)).fetchall()],
        # Vendedor interno e externo do cadastro de clientes. É a fonte que
        # enxerga quem AINDA NÃO VENDEU — vendedor recém-contratado e gerente
        # não aparecem no faturamento, mas estão no cadastro desde o primeiro dia.
        "clientSellers": [
            row["nome"] for row in conn.execute(
                """
                SELECT DISTINCT nome FROM (
                    SELECT TRIM(internal_seller_name) AS nome FROM crm_client_profiles
                     WHERE company_id = ? AND internal_seller_name IS NOT NULL AND TRIM(internal_seller_name) <> ''
                    UNION
                    SELECT TRIM(external_seller_name) AS nome FROM crm_client_profiles
                     WHERE company_id = ? AND external_seller_name IS NOT NULL AND TRIM(external_seller_name) <> ''
                ) ORDER BY nome
                """,
                (company_id, company_id),
            ).fetchall()
        ],
        "salesCoverage": dict(zip(("min", "max", "total"), conn.execute("SELECT MIN(issue_date), MAX(issue_date), COUNT(*) FROM fact_sales_detail WHERE company_id = ? AND issue_date IS NOT NULL AND TRIM(issue_date) <> ''", (company_id,)).fetchone())),
        "cityMappings": [dict(row) for row in conn.execute("SELECT * FROM city_mappings WHERE company_id = ? ORDER BY city_name, valid_from DESC", (company_id,)).fetchall()],
        "vacations": [dict(row) for row in conn.execute("SELECT * FROM vacations WHERE company_id = ? ORDER BY start_date DESC", (company_id,)).fetchall()],
        "holidays": [dict(row) for row in conn.execute("SELECT * FROM holidays WHERE company_id = ? ORDER BY holiday_date DESC", (company_id,)).fetchall()],
        "goalsSeller": [dict(row) for row in conn.execute("SELECT * FROM goals_seller WHERE company_id = ? ORDER BY competence DESC, seller_name ASC", (company_id,)).fetchall()],
        "goalsUnit": [dict(row) for row in conn.execute("SELECT * FROM goals_unit WHERE company_id = ? ORDER BY competence DESC, unit_name ASC", (company_id,)).fetchall()],
        "issues": [dict(row) for row in conn.execute("SELECT * FROM import_issues WHERE company_id = ? ORDER BY created_at DESC LIMIT 200", (company_id,)).fetchall()],
        "imports": [dict(row) for row in conn.execute("SELECT * FROM imports WHERE company_id = ? ORDER BY imported_at DESC LIMIT 100", (company_id,)).fetchall()],
        "audit": [dict(row) for row in conn.execute("SELECT * FROM audit_logs WHERE company_id = ? ORDER BY created_at DESC LIMIT 200", (company_id,)).fetchall()],
        "scoreConfigs": [dict(row) for row in conn.execute("SELECT * FROM score_configs WHERE company_id = ? ORDER BY valid_from_competence DESC", (company_id,)).fetchall()],
        "salesDetailSummary": [
            dict(row) for row in conn.execute(
                """
                SELECT competence,
                       MAX(issue_date) AS last_issue_date,
                       COUNT(*) AS row_count
                FROM fact_sales_detail
                WHERE company_id = ?
                GROUP BY competence
                ORDER BY competence DESC
                LIMIT 6
                """,
                (company_id,),
            ).fetchall()
        ],
    }


def resolve_import_issue(conn: sqlite3.Connection, company_id: int, user_id: int, payload: dict[str, Any]) -> dict[str, Any]:
    issue_id = int(payload.get("issueId") or 0)
    action = normalize_whitespace(payload.get("action")).lower() or "resolve"
    issue = conn.execute("SELECT * FROM import_issues WHERE id = ? AND company_id = ?", (issue_id, company_id)).fetchone()
    if not issue:
        raise ValueError("Pendência não encontrada")
    issue = dict(issue)
    if issue["status"] != "pendente":
        raise ValueError("Essa pendência já foi tratada")

    updates = {
        "resolvedBy": user_id,
        "resolvedAt": now_iso(),
        "action": action,
    }

    if action == "ignore":
        new_status = "ignorada"
    elif issue["issue_type"] == "vendedor_sem_vinculo":
        person_name = normalize_whitespace(payload.get("person_name") or issue["reference_value"])
        role_classification = normalize_whitespace(payload.get("role_classification") or "Vendedor") or "Vendedor"
        base_unit = normalize_unit(payload.get("base_unit"))
        valid_from = payload.get("valid_from") or first_day_of_competence(issue["competence"]).isoformat()
        valid_to = payload.get("valid_to")
        if not person_name or not base_unit:
            raise ValueError("Informe nome e unidade base para resolver o vendedor")
        conn.execute(
            """
            INSERT OR REPLACE INTO people_records
                (company_id, person_name, role_classification, base_unit, valid_from, valid_to, source, is_active, created_at)
            VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)
            """,
            (company_id, person_name, role_classification, base_unit, valid_from, valid_to, "resolucao_pendencia", 1, now_iso()),
        )
        updates.update({"person_name": person_name, "role_classification": role_classification, "base_unit": base_unit, "valid_from": valid_from})
        new_status = "resolvida"
    elif issue["issue_type"] == "cidade_sem_correspondencia":
        city_name = normalize_upper(payload.get("city_name") or issue["reference_value"])
        principal_unit = normalize_unit(payload.get("principal_unit"))
        if not city_name or not principal_unit:
            raise ValueError("Informe cidade e unidade principal para resolver a cidade")
        map_city_to_unit(conn, company_id, city_name, principal_unit,
                         payload.get("state_name"), payload.get("country_name"))
        updates.update({"city_name": city_name, "principal_unit": principal_unit,
                        "valid_from": CITY_MAPPING_EPOCH})
        new_status = "resolvida"
    else:
        raise ValueError("Tipo de pendência ainda não suportado")

    conn.execute(
        """
        UPDATE import_issues
        SET status = ?, details_json = ?
        WHERE company_id = ? AND issue_type = ? AND reference_value = ? AND status = 'pendente'
        """,
        (
            new_status,
            json.dumps(updates, ensure_ascii=False),
            company_id,
            issue["issue_type"],
            issue["reference_value"],
        ),
    )
    audit_log(conn, company_id, user_id, "resolver_pendencia", "import_issue", str(issue_id), updates | {"status": new_status})
    conn.commit()
    return {"issueId": issue_id, "status": new_status, "issueType": issue["issue_type"], "referenceValue": issue["reference_value"]}


def save_json_payload(conn: sqlite3.Connection, company_id: int, user_id: int, table_name: str, rows: list[dict[str, Any]]) -> int:
    created = 0
    if table_name == "people_records":
        for row in rows:
            conn.execute(
                """
                INSERT OR REPLACE INTO people_records
                    (company_id, person_name, role_classification, base_unit, valid_from, valid_to, source, is_active, created_at)
                VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)
                """,
                (
                    company_id,
                    normalize_whitespace(row["person_name"]),
                    row["role_classification"],
                    normalize_unit(row.get("base_unit")),
                    row["valid_from"],
                    row.get("valid_to"),
                    row.get("source", "manual"),
                    1 if row.get("is_active", True) else 0,
                    now_iso(),
                ),
            )
            created += 1
    elif table_name == "vacations":
        for row in rows:
            _sd = parse_datetime_flexible(row.get("start_date"))
            _ed = parse_datetime_flexible(row.get("end_date"))
            conn.execute(
                """
                INSERT OR REPLACE INTO vacations
                    (company_id, person_name, start_date, end_date, notes, created_at)
                VALUES (?, ?, ?, ?, ?, ?)
                """,
                (
                    company_id,
                    normalize_whitespace(row["person_name"]),
                    _sd.strftime("%Y-%m-%d") if _sd else (row.get("start_date") or ""),
                    _ed.strftime("%Y-%m-%d") if _ed else (row.get("end_date") or ""),
                    row.get("notes"),
                    now_iso(),
                ),
            )
            created += 1
    elif table_name == "holidays":
        for row in rows:
            conn.execute(
                """
                INSERT OR REPLACE INTO holidays
                    (company_id, holiday_date, holiday_name, scope, created_at)
                VALUES (?, ?, ?, ?, ?)
                """,
                (company_id, row["holiday_date"], row["holiday_name"], row.get("scope", "NACIONAL_RS"), now_iso()),
            )
            created += 1
    elif table_name == "goals_seller":
        for row in rows:
            seller_name = normalize_whitespace(row["seller_name"])
            conn.execute(
                """
                INSERT INTO goals_seller
                    (company_id, competence, seller_name, base_unit, revenue_goal, returns_goal, created_at)
                VALUES (?, ?, ?, ?, ?, ?, ?)
                ON CONFLICT(company_id, competence, seller_name)
                DO UPDATE SET
                    base_unit = excluded.base_unit,
                    revenue_goal = excluded.revenue_goal,
                    returns_goal = excluded.returns_goal,
                    created_at = excluded.created_at
                """,
                (
                    company_id,
                    row["competence"],
                    seller_name,
                    normalize_unit(row.get("base_unit")),
                    float(row.get("revenue_goal") or 0),
                    float(row.get("returns_goal") or 0),
                    now_iso(),
                ),
            )
            created += 1
    elif table_name == "goals_unit":
        for row in rows:
            normalized_unit = normalize_unit(row["unit_name"])
            conn.execute(
                """
                INSERT INTO goals_unit
                    (company_id, competence, unit_name, revenue_goal, returns_goal, created_at)
                VALUES (?, ?, ?, ?, ?, ?)
                ON CONFLICT(company_id, competence, unit_name)
                DO UPDATE SET
                    revenue_goal = excluded.revenue_goal,
                    returns_goal = excluded.returns_goal,
                    created_at = excluded.created_at
                """,
                (
                    company_id,
                    row["competence"],
                    normalized_unit,
                    float(row.get("revenue_goal") or 0),
                    float(row.get("returns_goal") or 0),
                    now_iso(),
                ),
            )
            created += 1
    elif table_name == "score_configs":
        for row in rows:
            conn.execute(
                """
                INSERT OR REPLACE INTO score_configs
                    (company_id, valid_from_competence, valid_to_competence, weight_goal, weight_ticket, weight_clients, weight_mix, weight_returns, created_at)
                VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)
                """,
                (
                    company_id,
                    row["valid_from_competence"],
                    row.get("valid_to_competence"),
                    float(row["weight_goal"]),
                    float(row["weight_ticket"]),
                    float(row["weight_clients"]),
                    float(row["weight_mix"]),
                    float(row["weight_returns"]),
                    now_iso(),
                ),
            )
            created += 1
    elif table_name == "users":
        for row in rows:
            username = normalize_whitespace(row.get("username")).lower()
            if not username:
                continue
            linked_units = normalize_unit_list(row.get("linked_units") or row.get("linked_units_json") or row.get("linked_units_csv"))
            password = row.get("password")
            existing = None
            row_id = row.get("id")
            if row_id:
                existing = conn.execute("SELECT * FROM users WHERE company_id = ? AND id = ?", (company_id, int(row_id))).fetchone()
            if not existing:
                existing = conn.execute("SELECT * FROM users WHERE company_id = ? AND username = ?", (company_id, username)).fetchone()
            duplicate = conn.execute("SELECT id FROM users WHERE company_id = ? AND username = ?", (company_id, username)).fetchone()
            if duplicate and (not existing or duplicate["id"] != existing["id"]):
                raise ValueError(f"Já existe um usuário cadastrado com o login {username}")
            if existing:
                pwd_hash = existing["password_hash"]
                salt = existing["password_salt"]
                if password:
                    pwd_hash, salt = pbkdf2_hash(str(password))
                conn.execute(
                    """
                    UPDATE users
                    SET username = ?, full_name = ?, linked_person_name = ?, linked_units_json = ?, role = ?, is_active = ?, password_hash = ?, password_salt = ?
                    WHERE company_id = ? AND id = ?
                    """,
                    (
                        username,
                        normalize_whitespace(row.get("full_name")),
                        normalize_whitespace(row.get("linked_person_name")),
                        json.dumps(linked_units, ensure_ascii=False),
                        row["role"],
                        1 if row.get("is_active", True) not in {False, "0", 0, "false", "False"} else 0,
                        pwd_hash,
                        salt,
                        company_id,
                        existing["id"],
                    ),
                )
                created += 1
                continue
            if not password:
                continue
            pwd_hash, salt = pbkdf2_hash(str(password))
            conn.execute(
                """
                INSERT INTO users
                    (company_id, username, full_name, linked_person_name, linked_units_json, password_hash, password_salt, role, is_active, created_at)
                VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                """,
                (
                    company_id,
                    username,
                    normalize_whitespace(row.get("full_name")),
                    normalize_whitespace(row.get("linked_person_name")),
                    json.dumps(linked_units, ensure_ascii=False),
                    pwd_hash,
                    salt,
                    row["role"],
                    1 if row.get("is_active", True) not in {False, "0", 0, "false", "False"} else 0,
                    now_iso(),
                ),
            )
            created += 1
    elif table_name == "client_registry":
        for row in rows:
            raw_name = row.get("client_name") or row.get("Razao Social/Nome") or row.get("razao_social_nome")
            client_name = normalize_whitespace(raw_name)
            if not client_name:
                continue
            raw_document = row.get("document_number") or row.get("CNPJ/CPF") or row.get("cnpj_cpf")
            doc_person_type, _ = person_type_from_document(raw_document)
            if doc_person_type:
                person_type = doc_person_type
                confidence_score = 1.0
                notes = "documento"
            else:
                person_type = row.get("person_type") or row.get("tipo_pessoa")
                if not person_type:
                    person_type, confidence_score, notes = infer_person_type_from_name(client_name)
                else:
                    confidence_score = float(row.get("confidence_score") or 0.9)
                    notes = row.get("notes")
            upsert_client_registry_row(
                conn,
                company_id,
                client_name,
                raw_document,
                person_type,
                row.get("source", "importacao_clientes"),
                confidence_score,
                notes,
            )
            created += 1
    if table_name == "goals_unit":
        sanitize_unit_goals(conn, company_id, user_id)
    audit_log(conn, company_id, user_id, "salvar", table_name, "batch", {"rows": created})
    conn.commit()
    # Metas, férias, feriados e cadastros alteram o dashboard — derruba os caches
    invalidate_calendar_cache(company_id)
    invalidate_dashboard_cache(company_id)
    return created


def csv_template(kind: str) -> bytes:
    buffer = io.StringIO()
    writer = csv.writer(buffer, delimiter=";")
    if kind == "people":
        writer.writerow(["person_name", "role_classification", "base_unit", "valid_from", "valid_to", "source"])
        writer.writerow(["NOME DO VENDEDOR", "Vendedor", "MATRIZ", "2026-04-01", "", "importacao"])
    elif kind == "vacations":
        writer.writerow(["person_name", "start_date", "end_date", "notes"])
        writer.writerow(["NOME DO VENDEDOR", "2026-04-10", "2026-04-20", "Ferias abril"])
    elif kind == "holidays":
        writer.writerow(["holiday_date", "holiday_name", "scope"])
        writer.writerow(["2026-09-20", "Revolução Farroupilha", "NACIONAL_RS"])
    elif kind == "goals_seller":
        writer.writerow(["competence", "seller_name", "base_unit", "revenue_goal"])
        writer.writerow(["2026-04", "NOME DO VENDEDOR", "MATRIZ", "150000"])
    elif kind == "goals_unit":
        writer.writerow(["competence", "unit_name", "revenue_goal"])
        writer.writerow(["2026-04", "MATRIZ", "1100000"])
    elif kind == "users":
        writer.writerow(["username", "full_name", "linked_person_name", "linked_units_csv", "role", "password"])
        writer.writerow(["gerente.matriz", "Gerente Matriz", "", "MATRIZ;LAJEADO", "Gerente", "Senha@123"])
    elif kind == "clients":
        writer.writerow(["client_name", "document_number", "person_type", "source", "confidence_score", "notes"])
        writer.writerow(["CLIENTE EXEMPLO LTDA", "12.345.678/0001-90", "PJ", "importacao_clientes", "1", "documento"])
    return buffer.getvalue().encode("utf-8-sig")


def delete_user_record(conn: sqlite3.Connection, company_id: int, actor_user_id: int, target_user_id: int) -> None:
    target = conn.execute("SELECT * FROM users WHERE company_id = ? AND id = ?", (company_id, target_user_id)).fetchone()
    if not target:
        raise ValueError("Usuário não encontrado")
    if target["id"] == actor_user_id:
        raise ValueError("Não é permitido excluir o próprio usuário logado")
    if target["role"] == "Administrador":
        admin_count = conn.execute(
            "SELECT COUNT(*) AS total FROM users WHERE company_id = ? AND role = 'Administrador' AND is_active = 1",
            (company_id,),
        ).fetchone()["total"]
        if admin_count <= 1:
            raise ValueError("Mantenha pelo menos um administrador ativo no sistema")
    conn.execute("DELETE FROM sessions WHERE user_id = ?", (target_user_id,))
    conn.execute("DELETE FROM users WHERE company_id = ? AND id = ?", (company_id, target_user_id))
    audit_log(conn, company_id, actor_user_id, "excluir", "users", str(target_user_id), {"username": target["username"]})


ALLOWED_USER_ROLES = {"Administrador", "Gerente", "Analista", "Vendedor"}


def upsert_user(conn: sqlite3.Connection, company_id: int, actor_user_id: int, payload: dict[str, Any]) -> dict[str, Any]:
    """Cria ou atualiza um usuário. Faz o hash da senha quando informada."""
    username = normalize_whitespace(payload.get("username"))
    full_name = normalize_whitespace(payload.get("full_name"))
    if not username or not full_name:
        raise ValueError("Informe o usuário e o nome completo.")

    # O perfil é a fonte de verdade das permissões. A coluna `role` é mantida em
    # sincronia com o NOME do perfil para compatibilidade com o código legado.
    profile_id = payload.get("profile_id") or None
    profile_row = None
    if profile_id:
        profile_row = conn.execute(
            "SELECT * FROM access_profiles WHERE id = ? AND company_id = ?", (profile_id, company_id)
        ).fetchone()
        if not profile_row:
            raise ValueError("Perfil selecionado não existe.")
        role = profile_row["name"]
    else:
        role = normalize_whitespace(payload.get("role")) or "Administrador"
        profile_row = conn.execute(
            "SELECT * FROM access_profiles WHERE company_id = ? AND name = ?", (company_id, role)
        ).fetchone()
        if not profile_row:
            raise ValueError("Selecione um perfil de acesso.")
        profile_id = profile_row["id"]

    data_scope = profile_row["data_scope"] or "todos"
    linked_person = normalize_whitespace(payload.get("linked_person_name")) or None
    linked_units_raw = payload.get("linked_units") or []
    if not isinstance(linked_units_raw, list):
        linked_units_raw = []
    linked_units = [normalize_unit(u) for u in linked_units_raw if u]

    # Escopo "próprio" exige vínculo com a pessoa; escopos por unidade exigem unidades.
    #
    # O vínculo com a pessoa é GRAVADO EM TODOS OS PERFIS. Antes ele era apagado
    # fora do escopo "próprio", partindo da ideia de que só serviria para filtrar
    # a carteira. Não serve: é por ele que o gerente recebe a ciência de uma ata,
    # aparece como participante de reunião e recebe feedback do diretor. Apagar
    # aqui fazia o campo "não salvar" sem nenhum aviso na tela.
    if data_scope == "proprio":
        linked_units = []
        if not linked_person:
            raise ValueError("Perfis com escopo 'própria carteira' exigem a pessoa vinculada.")
    elif data_scope in {"unidade", "unidade_consolidado"} and not linked_units:
        raise ValueError("Este perfil exige ao menos uma unidade vinculada.")

    linked_units_json = json.dumps(linked_units, ensure_ascii=False) if linked_units else None
    password = (payload.get("password") or "").strip()
    user_id = payload.get("id")

    # Unidade do vendedor. A conta não guarda esse dado — quem guarda é o
    # cadastro de pessoas, que alimenta carteira, meta, feedback e visita. A
    # tela de Acessos só serve de atalho para preencher lá, evitando que um
    # vendedor recém-criado (Zona Norte, por exemplo) fique sem unidade e
    # desapareça de todas as listas de equipe.
    base_unit = normalize_unit(payload.get("base_unit"))

    def aplicar_unidade_da_pessoa() -> None:
        # Só depois que a conta foi gravada — assim um login duplicado não
        # deixa um cadastro de pessoa criado pela metade.
        if base_unit and linked_person:
            save_person_record(conn, company_id, actor_user_id, {
                "personName": linked_person,
                "baseUnit": base_unit,
                "roleClassification": "Vendedor" if data_scope == "proprio" else "Gerente",
            })

    if user_id:
        existing = conn.execute("SELECT * FROM users WHERE company_id = ? AND id = ?", (company_id, user_id)).fetchone()
        if not existing:
            raise ValueError("Usuário não encontrado.")
        dup = conn.execute("SELECT id FROM users WHERE username = ? AND id <> ?", (username, user_id)).fetchone()
        if dup:
            raise ValueError("Já existe um usuário com esse login.")
        conn.execute(
            "UPDATE users SET username = ?, full_name = ?, role = ?, profile_id = ?, linked_person_name = ?, linked_units_json = ? WHERE company_id = ? AND id = ?",
            (username, full_name, role, profile_id, linked_person, linked_units_json, company_id, user_id),
        )
        if password:
            pwd_hash, salt = pbkdf2_hash(password)
            conn.execute("UPDATE users SET password_hash = ?, password_salt = ? WHERE company_id = ? AND id = ?", (pwd_hash, salt, company_id, user_id))
        audit_log(conn, company_id, actor_user_id, "atualizar", "users", str(user_id), {"username": username, "role": role})
        aplicar_unidade_da_pessoa()
        return {"id": user_id, "created": False}

    if not password:
        raise ValueError("Defina a senha inicial.")
    dup = conn.execute("SELECT id FROM users WHERE username = ?", (username,)).fetchone()
    if dup:
        raise ValueError("Já existe um usuário com esse login.")
    pwd_hash, salt = pbkdf2_hash(password)
    cur = conn.execute(
        "INSERT INTO users (company_id, username, full_name, linked_person_name, linked_units_json, password_hash, password_salt, role, profile_id, is_active, created_at) VALUES (?,?,?,?,?,?,?,?,?,1,?)",
        (company_id, username, full_name, linked_person, linked_units_json, pwd_hash, salt, role, profile_id, now_iso()),
    )
    audit_log(conn, company_id, actor_user_id, "criar", "users", str(cur.lastrowid), {"username": username, "role": role})
    aplicar_unidade_da_pessoa()
    return {"id": cur.lastrowid, "created": True}


def set_user_password(conn: sqlite3.Connection, company_id: int, actor_user_id: int, target_user_id: Any, new_password: str | None) -> None:
    password = (new_password or "").strip()
    if not password:
        raise ValueError("Informe a nova senha.")
    target = conn.execute("SELECT id FROM users WHERE company_id = ? AND id = ?", (company_id, target_user_id)).fetchone()
    if not target:
        raise ValueError("Usuário não encontrado.")
    pwd_hash, salt = pbkdf2_hash(password)
    conn.execute("UPDATE users SET password_hash = ?, password_salt = ? WHERE company_id = ? AND id = ?", (pwd_hash, salt, company_id, target_user_id))
    audit_log(conn, company_id, actor_user_id, "trocar_senha", "users", str(target_user_id), {})


def import_admin_csv(conn: sqlite3.Connection, company_id: int, user_id: int, table_name: str, content: bytes) -> int:
    text = decode_text_content(content)
    reader = csv.DictReader(io.StringIO(text, newline=""), delimiter=";")
    rows = [dict(row) for row in reader]
    total = save_json_payload(conn, company_id, user_id, table_name, rows)
    if table_name == "client_registry":
        ensure_client_registry_for_sales(conn, company_id)
    return total


def import_city_mappings_csv(conn: sqlite3.Connection, company_id: int, user_id: int, content: bytes) -> dict[str, Any]:
    """Importa em lote o vinculo cidade->unidade a partir de um CSV e resolve as
    pendencias de 'cidade_sem_correspondencia' correspondentes.
    Colunas aceitas (cabecalho, sem diferenciar maiusculas/acentos):
      CIDADE (obrigatoria), PRINCIPAL ou UNIDADE (obrigatoria), ESTADO (opcional), PAIS (opcional).
    Separador ; ou , detectado automaticamente. O CSV e autoritativo: substitui o
    vinculo existente de cada cidade."""
    text = decode_text_content(content)
    first_line = next((ln for ln in text.splitlines() if ln.strip()), "")
    delimiter = ";" if first_line.count(";") >= first_line.count(",") else ","
    reader = csv.DictReader(io.StringIO(text, newline=""), delimiter=delimiter)

    field_map: dict[str, str] = {}
    for header in (reader.fieldnames or []):
        key = normalize_upper(header)
        if key in ("CIDADE", "CITY", "MUNICIPIO"):
            field_map["city"] = header
        elif key in ("PRINCIPAL", "UNIDADE", "UNIDADE PRINCIPAL", "UNIT"):
            field_map["unit"] = header
        elif key in ("ESTADO", "UF", "STATE"):
            field_map["state"] = header
        elif key in ("PAIS", "PAÍS", "COUNTRY"):
            field_map["country"] = header
    if "city" not in field_map or "unit" not in field_map:
        raise ValueError("O CSV precisa ter as colunas CIDADE e PRINCIPAL (ou UNIDADE).")

    now = now_iso()
    updated = 0
    for row in reader:
        city = normalize_upper(row.get(field_map["city"]))
        unit = normalize_unit(row.get(field_map["unit"]))
        if not city or not unit:
            continue
        state = normalize_upper(row.get(field_map["state"])) if field_map.get("state") else None
        country = normalize_upper(row.get(field_map["country"])) if field_map.get("country") else None
        conn.execute("DELETE FROM city_mappings WHERE company_id = ? AND city_name = ?", (company_id, city))
        conn.execute(
            """
            INSERT INTO city_mappings
                (company_id, city_name, principal_unit, state_name, country_name, valid_from, valid_to, source, created_at)
            VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)
            """,
            (company_id, city, unit, state or None, country or None, "2025-01-01", None, "csv_lote", now),
        )
        updated += 1

    # Resolve as pendencias de cidade que agora tem vinculo
    cur = conn.execute(
        """
        UPDATE import_issues
        SET status = 'resolvida'
        WHERE company_id = ? AND issue_type = 'cidade_sem_correspondencia' AND status = 'pendente'
          AND reference_value IN (SELECT city_name FROM city_mappings WHERE company_id = ?)
        """,
        (company_id, company_id),
    )
    resolved = cur.rowcount if (cur.rowcount and cur.rowcount > 0) else 0
    return {"updated": updated, "resolved": resolved}


def person_base_unit(
    conn: sqlite3.Connection, company_id: int, person_name: str | None
) -> str:
    """Unidade base da pessoa no cadastro, casando o nome pela chave normalizada."""
    nome = normalize_whitespace(person_name)
    if not nome:
        return ""
    row = conn.execute(
        "SELECT base_unit FROM people_records WHERE company_id = ? AND person_name = ? "
        "AND base_unit IS NOT NULL AND TRIM(base_unit) <> '' "
        "ORDER BY valid_from DESC LIMIT 1",
        (company_id, nome),
    ).fetchone()
    if row:
        return normalize_unit(row["base_unit"])
    alvo = person_key(nome)
    for r in conn.execute(
        "SELECT person_name, base_unit FROM people_records WHERE company_id = ? "
        "AND base_unit IS NOT NULL AND TRIM(base_unit) <> '' ORDER BY valid_from DESC",
        (company_id,),
    ).fetchall():
        if person_key(r["person_name"]) == alvo:
            return normalize_unit(r["base_unit"])
    return ""


def save_person_record(
    conn: sqlite3.Connection, company_id: int, user_id: int, payload: dict[str, Any]
) -> dict[str, Any]:
    """Cadastra ou atualiza uma pessoa com função e unidade explícitas.

    O ajuste antigo só mexia na unidade e ADIVINHAVA a função pelo sufixo do
    nome — "(VENDAS)" virava Vendedor, o resto virava "Outro". Funciona para
    quem já apareceu no faturamento, mas não para o vendedor de uma unidade que
    ainda não vendeu nada: ele entrava classificado errado e sumia das listas
    de equipe, meta e feedback.
    """
    nome = normalize_whitespace(payload.get("personName"))
    if not nome:
        raise ValueError("Informe o nome da pessoa.")
    unidade = normalize_unit(payload.get("baseUnit"))
    if not unidade:
        raise ValueError("Informe a unidade.")
    funcao = normalize_whitespace(payload.get("roleClassification")) or "Vendedor"
    if funcao not in {"Vendedor", "Gerente", "Outro"}:
        funcao = "Vendedor"
    inicio = normalize_whitespace(payload.get("validFrom")) or today_in_brazil().isoformat()

    existente = conn.execute(
        "SELECT COUNT(*) n FROM people_records WHERE company_id = ? AND person_name = ?",
        (company_id, nome),
    ).fetchone()["n"]

    if existente:
        conn.execute(
            "UPDATE people_records SET base_unit = ?, role_classification = ? "
            "WHERE company_id = ? AND person_name = ?",
            (unidade, funcao, company_id, nome),
        )
        acao = "atualizada"
    else:
        conn.execute(
            """
            INSERT INTO people_records
                (company_id, person_name, role_classification, base_unit, valid_from, valid_to,
                 source, is_active, created_at)
            VALUES (?,?,?,?,?,?, 'cadastro_manual', 1, ?)
            """,
            (company_id, nome, funcao, unidade, inicio, None, now_iso()),
        )
        acao = "cadastrada"

    audit_log(conn, company_id, user_id, "salvar", "people_records", nome,
              {"funcao": funcao, "unidade": unidade})
    conn.commit()
    invalidate_crm_cache(company_id)
    return {"message": f"{nome} {acao} como {funcao} na unidade {unidade}.",
            "personName": nome, "roleClassification": funcao, "baseUnit": unidade}


def search_person_candidates(
    conn: sqlite3.Connection, company_id: int, termo: str, limite: int = 30
) -> list[dict[str, Any]]:
    """Busca a pessoa para vincular a uma conta, em todas as fontes que existem.

    A digitação livre foi removida de propósito: um caractere trocado quebra o
    vínculo em silêncio, e o efeito só aparece semanas depois — a pessoa não
    recebe ciência de reunião nem feedback. Aqui ela só pode ESCOLHER algo que
    existe de verdade.

    Quatro fontes, da mais confiável para a menos:
      1. cadastro de pessoas (já classificado, com unidade)
      2. vendedor no faturamento (quem já vendeu)
      3. vendedor interno/externo do cadastro de clientes (quem ainda não vendeu)
      4. cliente PESSOA FÍSICA — todo funcionário tem cadastro próprio de PF
    """
    texto = normalize_whitespace(termo)
    if len(texto) < 3:
        return []
    alvo = f"%{normalize_upper(strip_accents(texto))}%"

    # Sem isto, procurar "jose" não acha "JOSÉ" — o SQLite não conhece acento.
    # A função vive só nesta conexão; não muda nada no banco.
    conn.create_function(
        "sem_acento", 1,
        lambda v: normalize_upper(strip_accents(v)) if v is not None else None)

    encontrados: dict[str, dict[str, Any]] = {}

    def adiciona(nome: str, fonte: str, detalhe: str = "") -> None:
        nome = normalize_whitespace(nome)
        chave = person_key(nome)
        if not nome or not chave:
            return
        if chave in encontrados:
            # Mantém a fonte mais confiável, mas acumula o detalhe.
            if detalhe and detalhe not in encontrados[chave]["detail"]:
                encontrados[chave]["detail"] += f" · {detalhe}"
            return
        encontrados[chave] = {"personName": nome, "source": fonte, "detail": detalhe}

    hoje = today_in_brazil().isoformat()
    for r in conn.execute(
        """
        SELECT DISTINCT person_name, base_unit, role_classification, valid_to
        FROM people_records
        WHERE company_id = ? AND sem_acento(person_name) LIKE ?
        ORDER BY person_name LIMIT ?
        """,
        (company_id, alvo, limite),
    ).fetchall():
        desligado = bool(r["valid_to"] and r["valid_to"] < hoje)
        adiciona(r["person_name"], "Cadastro de pessoas",
                 " · ".join(filter(None, [r["base_unit"], r["role_classification"],
                                          "DESLIGADO" if desligado else ""])))

    for r in conn.execute(
        "SELECT DISTINCT seller_name FROM fact_sales_detail "
        "WHERE company_id = ? AND sem_acento(seller_name) LIKE ? ORDER BY seller_name LIMIT ?",
        (company_id, alvo, limite),
    ).fetchall():
        adiciona(r["seller_name"], "Faturamento", "já emitiu venda")

    for r in conn.execute(
        """
        SELECT DISTINCT nome FROM (
            SELECT TRIM(internal_seller_name) AS nome FROM crm_client_profiles
             WHERE company_id = ? AND sem_acento(internal_seller_name) LIKE ?
            UNION
            SELECT TRIM(external_seller_name) AS nome FROM crm_client_profiles
             WHERE company_id = ? AND sem_acento(external_seller_name) LIKE ?
        ) WHERE nome <> '' ORDER BY nome LIMIT ?
        """,
        (company_id, alvo, company_id, alvo, limite),
    ).fetchall():
        adiciona(r["nome"], "Vendedor no cadastro de clientes", "")

    # Cliente pessoa física: é onde todo funcionário aparece, mesmo sem venda.
    for r in conn.execute(
        """
        SELECT p.client_code, p.client_name, p.document_number, p.city_name
        FROM crm_client_profiles p
        LEFT JOIN client_registry r
          ON r.company_id = p.company_id AND r.normalized_client_name = UPPER(TRIM(p.client_name))
        WHERE p.company_id = ? AND sem_acento(p.client_name) LIKE ?
          AND (r.person_type = 'PF'
               OR LENGTH(REPLACE(REPLACE(REPLACE(COALESCE(p.document_number,''),'.',''),'/',''),'-','')) = 11)
        ORDER BY p.client_name LIMIT ?
        """,
        (company_id, alvo, limite),
    ).fetchall():
        adiciona(r["client_name"], "Cliente pessoa física",
                 " · ".join(filter(None, [f"cód. {r['client_code']}", r["city_name"] or ""])))

    ordem = {"Cadastro de pessoas": 0, "Faturamento": 1,
             "Vendedor no cadastro de clientes": 2, "Cliente pessoa física": 3}
    resultado = sorted(encontrados.values(),
                       key=lambda x: (ordem.get(x["source"], 9), x["personName"]))
    return resultado[:limite]


def terminate_person(
    conn: sqlite3.Connection, company_id: int, user_id: int, payload: dict[str, Any]
) -> dict[str, Any]:
    """Registra o desligamento de uma pessoa fechando a vigência do cadastro.

    Não apaga nada: o histórico de venda, reunião e feedback continua íntegro.
    Só define até quando ela é considerada ativa — assim ela para de aparecer
    em lista de equipe, meta e presença, sem sumir dos meses em que trabalhou.
    """
    nome = normalize_whitespace(payload.get("personName"))
    if not nome:
        raise ValueError("Informe a pessoa.")

    if payload.get("reactivate"):
        conn.execute(
            "UPDATE people_records SET valid_to = NULL, is_active = 1 "
            "WHERE company_id = ? AND person_name = ?",
            (company_id, nome),
        )
        audit_log(conn, company_id, user_id, "reativar", "people_records", nome, {})
        conn.commit()
        invalidate_crm_cache(company_id)
        return {"message": f"{nome} reativado."}

    # Aceita "AAAA-MM" (mês de desligamento) ou data completa.
    bruto = normalize_whitespace(payload.get("terminationMonth")) or normalize_whitespace(payload.get("validTo"))
    if not bruto:
        raise ValueError("Informe o mês de desligamento.")
    if len(bruto) == 7:
        fim = last_day_of_competence(bruto).isoformat()
    else:
        fim = bruto[:10]

    cur = conn.execute(
        "UPDATE people_records SET valid_to = ?, is_active = 0 WHERE company_id = ? AND person_name = ?",
        (fim, company_id, nome),
    )
    if not cur.rowcount:
        # Quem saiu antes do cadastro de pessoas existir só aparece no
        # faturamento e no cadastro de clientes. Antes o desligamento era
        # recusado ("não está no cadastro") e esses nomes ficavam para sempre
        # nas listas de vendedor. Aqui o registro nasce JÁ FECHADO: o histórico
        # dos meses trabalhados continua válido e ela some dos meses seguintes.
        conhecida = conn.execute(
            """
            SELECT 1 FROM fact_sales_detail
             WHERE company_id = ? AND UPPER(TRIM(seller_name)) = ? LIMIT 1
            """,
            (company_id, normalize_upper(nome)),
        ).fetchone() or conn.execute(
            """
            SELECT 1 FROM crm_client_profiles
             WHERE company_id = ? AND (UPPER(TRIM(internal_seller_name)) = ?
                                    OR UPPER(TRIM(external_seller_name)) = ?) LIMIT 1
            """,
            (company_id, normalize_upper(nome), normalize_upper(nome)),
        ).fetchone()
        if not conhecida:
            raise ValueError(f"{nome} não aparece no cadastro nem no faturamento. "
                             "Confira a grafia do nome.")
        primeira_venda = conn.execute(
            "SELECT MIN(issue_date) d FROM fact_sales_detail "
            "WHERE company_id = ? AND UPPER(TRIM(seller_name)) = ?",
            (company_id, normalize_upper(nome)),
        ).fetchone()["d"]
        inicio = (primeira_venda or "2020-01-01")[:10]
        conn.execute(
            """
            INSERT INTO people_records
                (company_id, person_name, role_classification, base_unit, valid_from, valid_to,
                 source, is_active, created_at)
            VALUES (?,?,?,?,?,?, 'desligamento_retroativo', 0, ?)
            """,
            (company_id, nome, "Vendedor",
             normalize_unit(payload.get("baseUnit")) or "", inicio, fim, now_iso()),
        )

    # Conta de acesso também é desativada: quem saiu não deve continuar entrando.
    contas = conn.execute(
        "UPDATE users SET is_active = 0 WHERE company_id = ? AND UPPER(COALESCE(linked_person_name,'')) = ?",
        (company_id, normalize_upper(nome)),
    ).rowcount

    audit_log(conn, company_id, user_id, "desligar", "people_records", nome,
              {"validTo": fim, "contasDesativadas": contas})
    conn.commit()
    invalidate_crm_cache(company_id)
    return {"message": f"{nome} desligado em {fim}."
                       + (f" {contas} conta(s) de acesso desativada(s)." if contas else ""),
            "validTo": fim, "deactivatedUsers": int(contas or 0)}


def update_person_unit(conn: sqlite3.Connection, company_id: int, user_id: int, person_name: str | None, base_unit: str | None) -> str:
    """Corrige a unidade de um vendedor/pessoa. Aplica a todos os registros da pessoa
    (fica consistente entre competencias). Cria o registro se ainda nao existir."""
    person = normalize_whitespace(person_name)
    unit = normalize_unit(base_unit)
    if not person or not unit:
        raise ValueError("Informe o vendedor e a unidade.")
    cur = conn.execute(
        "UPDATE people_records SET base_unit = ? WHERE company_id = ? AND person_name = ?",
        (unit, company_id, person),
    )
    if not cur.rowcount:
        conn.execute(
            """
            INSERT INTO people_records
                (company_id, person_name, role_classification, base_unit, valid_from, valid_to, source, is_active, created_at)
            VALUES (?, ?, ?, ?, ?, ?, ?, 1, ?)
            """,
            (company_id, person, infer_role_from_name(person), unit, "2025-01-01", None, "edicao_manual", now_iso()),
        )
    conn.execute(
        """
        UPDATE import_issues SET status = 'resolvida'
        WHERE company_id = ? AND issue_type = 'vendedor_sem_vinculo' AND status = 'pendente' AND reference_value = ?
        """,
        (company_id, person),
    )
    audit_log(conn, company_id, user_id, "ajustar_unidade", "people_records", person, {"base_unit": unit})
    return unit


def update_city_unit(conn: sqlite3.Connection, company_id: int, user_id: int, city_name: str | None, principal_unit: str | None) -> str:
    """Corrige a unidade principal de uma cidade e resolve a pendencia correspondente."""
    city = normalize_upper(city_name)
    unit = normalize_unit(principal_unit)
    if not city or not unit:
        raise ValueError("Informe a cidade e a unidade.")
    cur = conn.execute(
        "UPDATE city_mappings SET principal_unit = ? WHERE company_id = ? AND city_name = ?",
        (unit, company_id, city),
    )
    if not cur.rowcount:
        conn.execute(
            """
            INSERT INTO city_mappings
                (company_id, city_name, principal_unit, state_name, country_name, valid_from, valid_to, source, created_at)
            VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)
            """,
            (company_id, city, unit, None, None, "2025-01-01", None, "edicao_manual", now_iso()),
        )
    conn.execute(
        """
        UPDATE import_issues SET status = 'resolvida'
        WHERE company_id = ? AND issue_type = 'cidade_sem_correspondencia' AND status = 'pendente' AND reference_value = ?
        """,
        (company_id, city),
    )
    audit_log(conn, company_id, user_id, "ajustar_unidade", "city_mappings", city, {"principal_unit": unit})
    return unit


def export_dashboard_xlsx(data: dict[str, Any]) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "Resumo"
    ws.append(["Indicador", "Valor"])
    for key, value in data["summary"].items():
        ws.append([key, value])

    ws_sellers = wb.create_sheet("Vendedores")
    ws_sellers.append(["Vendedor", "Unidade Base", "Faturamento Líquido", "Meta", "% Meta", "Ticket",
                       "Clientes", "Mix", "Devolução Comercial", "Devolução Garantia", "Devolução Total", "Score"])
    for row in data["sellerRanking"]:
        ws_sellers.append([
            row["sellerName"],
            row.get("baseUnit"),
            row["revenueNet"],
            row["revenueGoal"],
            row["goalAttainmentPct"],
            row["ticketAverage"],
            row["distinctClients"],
            row["mixSku"],
            row["returnsValue"],
            row.get("warrantyReturnsValue", 0),
            row.get("returnsTotalValue", row["returnsValue"]),
            row["score"],
        ])

    ws_units = wb.create_sheet("Unidades")
    ws_units.append(["Unidade", "Faturamento Líquido", "Meta", "% Meta",
                     "Devolução Comercial", "Devolução Garantia", "Devolução Total", "Margem"])
    for row in data["unitPerformance"]:
        ws_units.append([row["unitName"], row["revenueNet"], row["revenueGoal"], row["goalAttainmentPct"],
                         row["returnsValue"], row.get("warrantyReturnsValue", 0),
                         row.get("returnsTotalValue", row["returnsValue"]), row["marginValue"]])

    ws_cities = wb.create_sheet("Cidades")
    ws_cities.append(["Cidade", "Faturamento Líquido", "Ticket Médio", "Clientes Distintos", "Desconto"])
    for row in data["cityRanking"]:
        ws_cities.append([row["cityName"], row["revenueNet"], row["ticketAverage"], row["distinctClients"], row["discountValue"]])

    output = io.BytesIO()
    wb.save(output)
    return output.getvalue()


def export_dashboard_pdf(data: dict[str, Any]) -> bytes:
    output = io.BytesIO()
    doc = SimpleDocTemplate(output, pagesize=landscape(A4), leftMargin=24, rightMargin=24, topMargin=24, bottomMargin=24)
    styles = getSampleStyleSheet()
    elements = [
        Paragraph("Dashboard Comercial Passini", styles["Title"]),
        Spacer(1, 12),
        Paragraph(f"Competência principal: {data.get('primaryCompetence') or 'Não definida'}", styles["Normal"]),
        Spacer(1, 12),
    ]

    summary_table = [["Indicador", "Valor"]]
    for key, value in data["summary"].items():
        summary_table.append([key, str(value)])
    table = Table(summary_table, repeatRows=1)
    table.setStyle(
        TableStyle(
            [
                ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#12324a")),
                ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
                ("GRID", (0, 0), (-1, -1), 0.5, colors.HexColor("#bfd7ea")),
                ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
            ]
        )
    )
    elements.append(table)
    elements.append(Spacer(1, 18))

    top10 = [["Vendedor", "% Meta", "Score", "Faturamento Líquido", "Clientes", "Mix"]]
    for row in data["sellerTop10"]:
        top10.append([row["sellerName"], row["goalAttainmentPct"], row["score"], row["revenueNet"], row["distinctClients"], row["mixSku"]])
    table_top = Table(top10, repeatRows=1)
    table_top.setStyle(
        TableStyle(
            [
                ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#1e5f74")),
                ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
                ("GRID", (0, 0), (-1, -1), 0.5, colors.HexColor("#dce6ef")),
                ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
            ]
        )
    )
    elements.append(Paragraph("Top 10 vendedores", styles["Heading2"]))
    elements.append(table_top)

    doc.build(elements)
    return output.getvalue()


def compute_seller_score(conn: sqlite3.Connection, company_id: int, user: sqlite3.Row, competence: str | None = None) -> dict[str, Any]:
    """Calcula os 9 indicadores de premiação para o vendedor logado."""
    seller_name = seller_identity_for_user(user)
    today = date.today()
    if not competence:
        competence = today.strftime("%Y-%m")
    comp_start = first_day_of_competence(competence).isoformat()
    comp_end = (first_day_of_competence(competence).replace(day=28) + timedelta(days=4)).replace(day=1) - timedelta(days=1)
    comp_end_iso = comp_end.isoformat()

    # ── 1. META DE VENDAS ────────────────────────────────────────────────────
    goal_row = conn.execute(
        "SELECT revenue_goal FROM goals_seller WHERE company_id = ? AND competence = ? AND seller_name = ?",
        (company_id, competence, seller_name),
    ).fetchone()
    revenue_goal = float(goal_row["revenue_goal"]) if goal_row else 0.0

    vendor_row = conn.execute(
        "SELECT COALESCE(SUM(net_value),0) AS net_value FROM fact_vendor_summary WHERE company_id = ? AND competence = ? AND seller_name = ?",
        (company_id, competence, seller_name),
    ).fetchone()
    revenue_actual = float(vendor_row["net_value"]) if vendor_row else 0.0
    goal_pct = (revenue_actual / revenue_goal * 100) if revenue_goal > 0 else 0.0

    if goal_pct >= 110:
        goal_pts = 50
    elif goal_pct >= 100:
        goal_pts = 30
    elif goal_pct >= 95:
        goal_pts = 15
    elif goal_pct >= 90:
        goal_pts = 5
    else:
        goal_pts = 0

    # ── 2. MARGEM DE VENDA ───────────────────────────────────────────────────
    margin_row = conn.execute(
        "SELECT COALESCE(AVG(margin_value),0) AS margin FROM fact_vendor_summary WHERE company_id = ? AND competence = ? AND seller_name = ?",
        (company_id, competence, seller_name),
    ).fetchone()
    margin_actual = float(margin_row["margin"]) if margin_row else 0.0
    if margin_actual >= 1.59:
        margin_pts = 20
    elif margin_actual >= 1.52:
        margin_pts = 10
    elif margin_actual >= 1.50:
        margin_pts = 5
    else:
        margin_pts = 0

    # ── 3. MIX DE ITENS (SKU distintos) ──────────────────────────────────────
    base_unit_row = conn.execute(
        "SELECT base_unit FROM people_records WHERE company_id = ? AND person_name = ? AND (valid_to IS NULL OR valid_to >= ?) ORDER BY valid_from DESC LIMIT 1",
        (company_id, seller_name, comp_start),
    ).fetchone()
    base_unit = normalize_unit(base_unit_row["base_unit"]) if base_unit_row else ""
    unit_item_goals = {
        "MATRIZ": 650, "LAJEADO": 1000, "PELOTAS": 850,
        "PORTO ALEGRE": 900, "XANGRI-LA": 750,
    }
    item_goal = unit_item_goals.get(base_unit, 800)

    sku_row = conn.execute(
        """SELECT COUNT(DISTINCT sku_key) AS sku_count
           FROM fact_sales_detail
           WHERE company_id = ? AND competence = ? AND seller_name = ? AND sku_key IS NOT NULL AND sku_key != ''""",
        (company_id, competence, seller_name),
    ).fetchone()
    sku_actual = int(sku_row["sku_count"]) if sku_row else 0
    item_pts = 10 if sku_actual >= item_goal else 0

    # ── GATILHO DA UNIDADE ───────────────────────────────────────────────────
    # Elegível se: unidade >= 95% da meta OU vendedor individualmente >= 105%
    unit_goal_row = conn.execute(
        "SELECT revenue_goal FROM goals_unit WHERE company_id = ? AND competence = ? AND unit_name = ?",
        (company_id, competence, base_unit),
    ).fetchone()
    unit_revenue_goal = float(unit_goal_row["revenue_goal"]) if unit_goal_row else 0.0
    unit_revenue_row = conn.execute(
        "SELECT COALESCE(SUM(net_value),0) AS net FROM fact_unit_summary WHERE company_id = ? AND competence = ? AND unit_name = ?",
        (company_id, competence, base_unit),
    ).fetchone()
    unit_revenue_actual = float(unit_revenue_row["net"]) if unit_revenue_row else 0.0
    unit_goal_pct = (unit_revenue_actual / unit_revenue_goal * 100) if unit_revenue_goal > 0 else 0.0
    unit_gate_met = unit_goal_pct >= 95.0
    seller_overrides_gate = goal_pct >= 105.0  # vendedor dispensa o gatilho

    # ── 4. POSITIVAÇÃO DA CARTEIRA ────────────────────────────────────────────
    total_clients_row = conn.execute(
        """SELECT COUNT(DISTINCT client_code) AS total
           FROM crm_client_summary WHERE company_id = ? AND seller_name = ?""",
        (company_id, seller_name),
    ).fetchone()
    total_clients = int(total_clients_row["total"]) if total_clients_row else 0

    active_clients_row = conn.execute(
        """SELECT COUNT(DISTINCT client_code) AS active
           FROM crm_client_summary WHERE company_id = ? AND competence = ? AND seller_name = ? AND net_value > 0""",
        (company_id, competence, seller_name),
    ).fetchone()
    active_clients = int(active_clients_row["active"]) if active_clients_row else 0
    pos_pct = (active_clients / total_clients * 100) if total_clients > 0 else 0.0
    if pos_pct >= 85:
        pos_pts = 20
    elif pos_pct >= 50:
        pos_pts = 10
    else:
        pos_pts = 0

    # ── 5. DEVOLUÇÕES ─────────────────────────────────────────────────────────
    returns_row = conn.execute(
        """SELECT COALESCE(SUM(return_value),0) AS returns, COALESCE(SUM(net_value),0) AS net
           FROM fact_vendor_summary WHERE company_id = ? AND competence = ? AND seller_name = ?""",
        (company_id, competence, seller_name),
    ).fetchone()
    returns_val = float(returns_row["returns"]) if returns_row else 0.0
    net_val = float(returns_row["net"]) if returns_row else 0.0
    return_pct = (returns_val / (net_val + returns_val) * 100) if (net_val + returns_val) > 0 else 0.0
    dev_pts = 10 if return_pct <= 4.5 else 0

    # ── 6. EXTRA POSITIVAÇÃO (inativos reativados) ────────────────────────────
    # Clientes PJ da carteira do vendedor que estavam inativos (sem compra ≥2 meses) e compraram este mês ≥R$999
    prev_comp = (first_day_of_competence(competence).replace(day=1) - timedelta(days=1)).strftime("%Y-%m")
    prev2_comp = (first_day_of_competence(prev_comp).replace(day=1) - timedelta(days=1)).strftime("%Y-%m")

    inactive_reactivated = conn.execute(
        """SELECT COUNT(*) AS cnt
           FROM crm_client_summary curr
           WHERE curr.company_id = ? AND curr.competence = ? AND curr.seller_name = ? AND curr.net_value >= 999
             AND NOT EXISTS (
                 SELECT 1 FROM crm_client_summary prev
                 WHERE prev.company_id = curr.company_id AND prev.client_code = curr.client_code
                   AND prev.competence IN (?, ?) AND prev.net_value > 0
             )""",
        (company_id, competence, seller_name, prev_comp, prev2_comp),
    ).fetchone()
    extra_pos = int(inactive_reactivated["cnt"]) if inactive_reactivated else 0
    extra_pts = min(extra_pos, 10)

    # ── 7. TREINAMENTOS EAD (manual — sem integração disponível) ─────────────
    training_pts = 0  # placeholder — integrar quando EAD Passini tiver API
    training_done = False

    # ── 8. LIGAÇÕES ATIVAS (interações registradas no mês) ───────────────────
    calls_row = conn.execute(
        """SELECT COUNT(*) AS cnt FROM crm_interactions
           WHERE company_id = ? AND seller_name = ?
             AND substr(occurred_at, 1, 7) = ?
             AND result_code NOT IN ('NAO_ATENDEU', 'PEDIU_RETORNO')
             AND initiative = 'ATIVO'""",
        (company_id, seller_name, competence),
    ).fetchone()
    calls_actual = int(calls_row["cnt"]) if calls_row else 0
    calls_pts = 10 if calls_actual >= 60 else 0

    # ── 9. REDES SOCIAIS (manual — sem integração disponível) ────────────────
    social_pts = 0  # placeholder
    social_count = 0

    # ── TOTAL ─────────────────────────────────────────────────────────────────
    total_pts = goal_pts + margin_pts + item_pts + pos_pts + dev_pts + extra_pts + training_pts + calls_pts + social_pts
    max_pts = 150

    # ── PREMIAÇÃO ESTIMADA ────────────────────────────────────────────────────
    base_prize = 185 if revenue_goal < 160000 else (380 if revenue_goal < 310000 else 530)
    prize_pct = min(total_pts / 100, 1.5) if total_pts >= 60 else (total_pts / 100)
    estimated_prize = round(base_prize * prize_pct, 2) if total_pts >= 60 else 0.0
    gate_ok = unit_gate_met or seller_overrides_gate
    eligible = goal_pts > 0 and gate_ok  # precisa bater meta própria E gatilho da unidade (ou 105%+)

    return {
        "competence": competence,
        "sellerName": seller_name,
        "totalPoints": total_pts,
        "maxPoints": max_pts,
        "estimatedPrize": estimated_prize,
        "basePrize": base_prize,
        "eligible": eligible,
        "unitGate": {
            "unitName": base_unit,
            "unitGoalPct": round(unit_goal_pct, 1),
            "unitGoal": round(unit_revenue_goal, 2),
            "unitActual": round(unit_revenue_actual, 2),
            "gateMet": unit_gate_met,
            "sellerOverrides": seller_overrides_gate,
            "gateOk": gate_ok,
        },
        "indicators": {
            "goalSales":       {"pts": goal_pts,    "max": 50, "actual": round(goal_pct, 1),   "goal": 100.0,    "unit": "%",  "label": "Meta de Vendas"},
            "margin":          {"pts": margin_pts,  "max": 20, "actual": round(margin_actual,2),"goal": 1.59,     "unit": "x",  "label": "Margem de Venda"},
            "mix":             {"pts": item_pts,    "max": 10, "actual": sku_actual,           "goal": item_goal,"unit": "itens","label": "Mix de Itens"},
            "positivacao":     {"pts": pos_pts,     "max": 20, "actual": round(pos_pct, 1),    "goal": 85.0,     "unit": "%",  "label": "Positivação da Carteira"},
            "returns":         {"pts": dev_pts,     "max": 10, "actual": round(return_pct, 2), "goal": 4.5,      "unit": "%",  "label": "Devoluções"},
            "extraPos":        {"pts": extra_pts,   "max": 10, "actual": extra_pos,            "goal": 10,       "unit": "clientes","label": "Extra Positivação"},
            "training":        {"pts": training_pts,"max": 10, "actual": 0,                    "goal": 100,      "unit": "%",  "label": "Treinamentos EAD"},
            "calls":           {"pts": calls_pts,   "max": 10, "actual": calls_actual,         "goal": 60,       "unit": "lig.","label": "Ligações Ativas"},
            "social":          {"pts": social_pts,  "max": 10, "actual": social_count,         "goal": 10,       "unit": "posts","label": "Redes Sociais"},
        },
    }


def compute_team_score(conn: sqlite3.Connection, company_id: int, user: sqlite3.Row) -> dict[str, Any]:
    """Calcula o score de todos os vendedores visíveis ao gerente/admin no mês corrente."""
    today = date.today()
    competence = today.strftime("%Y-%m")
    comp_start = first_day_of_competence(competence).isoformat()

    # Determinar quais vendedores estão no escopo do usuário
    allowed_units = linked_units_for_user(user) if user["role"] in {"Gerente", "Analista"} else []

    # Buscar todos os vendedores com metas no mês ou com vendas no mês
    sellers_in_scope: list[str] = []
    goal_rows = conn.execute(
        "SELECT DISTINCT seller_name FROM goals_seller WHERE company_id = ? AND competence = ?",
        (company_id, competence),
    ).fetchall()
    vendor_rows = conn.execute(
        "SELECT DISTINCT seller_name FROM fact_vendor_summary WHERE company_id = ? AND competence = ?",
        (company_id, competence),
    ).fetchall()
    all_sellers = {normalize_whitespace(r["seller_name"]) for r in goal_rows + vendor_rows if r["seller_name"]}

    # Filtrar por unidades do gerente se aplicável
    for seller_name in sorted(all_sellers):
        if allowed_units:
            _, base_unit = current_role_and_unit(conn, company_id, seller_name, competence)
            if normalize_unit(base_unit) not in allowed_units:
                continue
        sellers_in_scope.append(seller_name)

    # Pré-carregar metas e vendas em batch
    goal_map: dict[str, float] = {
        normalize_whitespace(r["seller_name"]): float(r["revenue_goal"])
        for r in conn.execute(
            "SELECT seller_name, revenue_goal FROM goals_seller WHERE company_id = ? AND competence = ?",
            (company_id, competence),
        ).fetchall()
    }
    vendor_map: dict[str, dict] = {
        normalize_whitespace(r["seller_name"]): dict(r)
        for r in conn.execute(
            "SELECT seller_name, SUM(net_value) AS net_value, SUM(return_value) AS return_value, AVG(margin_value) AS margin_value FROM fact_vendor_summary WHERE company_id = ? AND competence = ? GROUP BY seller_name",
            (company_id, competence),
        ).fetchall()
    }

    # Pré-carregar positivação
    pos_map: dict[str, int] = {}
    total_clients_map: dict[str, int] = {}
    for r in conn.execute(
        "SELECT seller_name, COUNT(DISTINCT client_code) AS total FROM crm_client_summary WHERE company_id = ? GROUP BY seller_name",
        (company_id,),
    ).fetchall():
        total_clients_map[normalize_whitespace(r["seller_name"])] = int(r["total"])
    for r in conn.execute(
        "SELECT seller_name, COUNT(DISTINCT client_code) AS active FROM crm_client_summary WHERE company_id = ? AND competence = ? AND net_value > 0 GROUP BY seller_name",
        (company_id, competence),
    ).fetchall():
        pos_map[normalize_whitespace(r["seller_name"])] = int(r["active"])

    # Pré-carregar ligações ativas
    calls_map: dict[str, int] = {
        normalize_whitespace(r["seller_name"]): int(r["cnt"])
        for r in conn.execute(
            "SELECT seller_name, COUNT(*) AS cnt FROM crm_interactions WHERE company_id = ? "
            "AND substr(occurred_at,1,7) = ? AND result_code NOT IN ('NAO_ATENDEU', 'PEDIU_RETORNO') "
            "AND initiative = 'ATIVO' GROUP BY seller_name",
            (company_id, competence),
        ).fetchall()
    }

    # Pré-carregar SKUs por vendedor
    sku_map: dict[str, int] = {
        normalize_whitespace(r["seller_name"]): int(r["sku_count"])
        for r in conn.execute(
            "SELECT seller_name, COUNT(DISTINCT sku_key) AS sku_count FROM fact_sales_detail WHERE company_id = ? AND competence = ? GROUP BY seller_name",
            (company_id, competence),
        ).fetchall()
    }

    # Pré-carregar unidades dos vendedores
    unit_item_goals = {"MATRIZ": 650, "LAJEADO": 1000, "PELOTAS": 850, "PORTO ALEGRE": 900, "XANGRI-LA": 750}
    _comp_target = first_day_of_competence(competence).isoformat()
    seller_unit_map: dict[str, str] = {
        normalize_whitespace(r["person_name"]): normalize_unit(r["base_unit"]) or ""
        for r in conn.execute(
            "SELECT person_name, base_unit FROM people_records WHERE company_id = ? AND date(valid_from) <= date(?) AND (valid_to IS NULL OR date(valid_to) >= date(?)) ORDER BY valid_from DESC",
            (company_id, _comp_target, _comp_target),
        ).fetchall()
    }

    results = []
    for seller_name in sellers_in_scope:
        revenue_goal = goal_map.get(seller_name, 0.0)
        vendor = vendor_map.get(seller_name, {})
        revenue_actual = float(vendor.get("net_value") or 0)
        returns_val = float(vendor.get("return_value") or 0)
        margin = float(vendor.get("margin_value") or 0)
        goal_pct = (revenue_actual / revenue_goal * 100) if revenue_goal > 0 else 0.0

        if goal_pct >= 110: goal_pts = 50
        elif goal_pct >= 100: goal_pts = 30
        elif goal_pct >= 95: goal_pts = 15
        elif goal_pct >= 90: goal_pts = 5
        else: goal_pts = 0

        margin_pts = 20 if margin >= 1.59 else (10 if margin >= 1.52 else (5 if margin >= 1.50 else 0))

        base_unit = seller_unit_map.get(seller_name, "")
        item_goal = unit_item_goals.get(base_unit, 800)
        sku_actual = sku_map.get(seller_name, 0)
        item_pts = 10 if sku_actual >= item_goal else 0

        total_clients = total_clients_map.get(seller_name, 0)
        active_clients = pos_map.get(seller_name, 0)
        pos_pct = (active_clients / total_clients * 100) if total_clients > 0 else 0.0
        pos_pts = 20 if pos_pct >= 85 else (10 if pos_pct >= 50 else 0)

        net_val = revenue_actual
        return_pct = (returns_val / (net_val + returns_val) * 100) if (net_val + returns_val) > 0 else 0.0
        dev_pts = 10 if return_pct <= 4.5 else 0

        calls_actual = calls_map.get(seller_name, 0)
        calls_pts = 10 if calls_actual >= 60 else 0

        total_pts = goal_pts + margin_pts + item_pts + pos_pts + dev_pts + calls_pts
        base_prize = 185 if revenue_goal < 160000 else (380 if revenue_goal < 310000 else 530)
        eligible = goal_pts > 0

        results.append({
            "sellerName": seller_name,
            "baseUnit": base_unit,
            "totalPoints": total_pts,
            "eligible": eligible,
            "estimatedPrize": round(base_prize * min(total_pts / 100, 1.5), 2) if total_pts >= 60 else 0.0,
            "goalPct": round(goal_pct, 1),
            "goalPts": goal_pts,
            "positivacaoPct": round(pos_pct, 1),
            "positivacaoPts": pos_pts,
            "callsActual": calls_actual,
            "callsPts": calls_pts,
            "returnPct": round(return_pct, 2),
            "devPts": dev_pts,
            "marginPts": margin_pts,
            "itemPts": item_pts,
            "revenueActual": round(revenue_actual, 2),
            "revenueGoal": round(revenue_goal, 2),
        })

    results.sort(key=lambda r: r["totalPoints"], reverse=True)

    return {
        "competence": competence,
        "sellers": results,
        "summary": {
            "total": len(results),
            "eligible": sum(1 for r in results if r["eligible"]),
            "inPrizeZone": sum(1 for r in results if r["totalPoints"] >= 60),
            "fullPrize": sum(1 for r in results if r["totalPoints"] >= 100),
        },
    }


# Meta diária de contatos por vendedor na Missão do Dia
DAILY_CONTACT_GOAL = 5
# Dias sem interação para considerar que a cobertura falhou
COVERAGE_GAP_DAYS = 7


def sellers_available_for_assignment(
    conn: sqlite3.Connection, company_id: int, user: sqlite3.Row,
    client_city: str | None = None,
) -> list[dict[str, Any]]:
    """Vendedores que o gestor pode escolher ao cobrar um contato.

    Restrito às unidades do usuário. Quando a cidade do cliente é conhecida, os
    vendedores da unidade que atende aquela cidade vêm primeiro — é a escolha
    mais provável.
    """
    competence = crm_latest_competence(conn, company_id) or date.today().strftime("%Y-%m")
    comp_day = first_day_of_competence(competence).isoformat()
    allowed = crm_allowed_units_for_user(conn, user)

    rows = conn.execute(
        """
        SELECT person_name, base_unit
        FROM people_records
        WHERE company_id = ? AND role_classification = 'Vendedor'
          AND date(valid_from) <= date(?)
          AND (valid_to IS NULL OR valid_to = '' OR date(valid_to) >= date(?))
        ORDER BY person_name
        """,
        (company_id, comp_day, comp_day),
    ).fetchall()

    # Unidade sugerida pela cidade do cliente
    preferred_unit = None
    if client_city:
        preferred_unit = resolve_city_unit(conn, company_id, client_city, competence)

    seen: set[str] = set()
    result: list[dict[str, Any]] = []
    for r in rows:
        name = normalize_whitespace(r["person_name"])
        key = normalize_upper(name)
        if not name or key in seen:
            continue
        unit = normalize_unit(r["base_unit"])
        if allowed and unit not in allowed:
            continue
        seen.add(key)
        result.append({
            "sellerName": name,
            "baseUnit": unit,
            "preferred": bool(preferred_unit and unit == preferred_unit),
        })

    # Vendedores da unidade da cidade primeiro
    result.sort(key=lambda s: (not s["preferred"], s["sellerName"]))
    return result


# Sufixos societários que o cadastro carrega e o faturamento nem sempre traz.
# "POWERTECH CAR SERVICE LTDA" no faturamento contra "POWERTECH CAR SERVICE
# LTDA - ME" no cadastro: para o computador são dois clientes diferentes; para
# quem opera, obviamente o mesmo.
COMPANY_SUFFIX_TOKENS = {
    "LTDA", "ME", "EPP", "EIRELI", "MEI", "SA", "S", "A", "CIA", "EI",
    "LIMITADA", "COMERCIO", "E", "DE", "DA", "DO",
}


def company_match_key(nome: str | None) -> str:
    """Chave tolerante para casar cliente entre faturamento e cadastro.

    Tira acento, pontuação, o documento colado no começo do nome e os sufixos
    societários do fim. NÃO substitui a chave exata: é usada só como segunda
    tentativa, e apenas quando aponta para um único cliente — nome parecido
    demais é justamente o caso em que errar dói.
    """
    texto = normalize_upper(strip_accents(nome))
    texto = re.sub(r"[^A-Z0-9 ]", " ", texto)
    texto = re.sub(r"^[0-9 ]+", "", texto)
    tokens = [t for t in texto.split() if t]
    while tokens and tokens[-1] in COMPANY_SUFFIX_TOKENS:
        tokens.pop()
    return " ".join(tokens)


def build_document_client_map(
    conn: sqlite3.Connection, company_id: int
) -> tuple[dict[str, str], dict[str, set[str]]]:
    """CNPJ/CPF → código do cliente, e raiz de CNPJ → códigos.

    O relatório de faturamento às vezes escreve o cliente com o documento
    grudado no nome — "54.719.029 ISMAEL GREGORY", "ROGERIO GAUGER 01860545050".
    O casamento por nome falha nesses casos, e o cliente aparecia como "sem
    cadastro" mesmo estando cadastrado. O documento resolve sozinho boa parte
    deles, sem ninguém precisar conciliar à mão.
    """
    exatos: dict[str, str] = {}
    por_raiz: dict[str, set[str]] = defaultdict(set)
    for row in conn.execute(
        "SELECT client_code, document_number FROM crm_client_profiles "
        "WHERE company_id = ? AND document_number IS NOT NULL AND TRIM(document_number) <> ''",
        (company_id,),
    ).fetchall():
        digitos = only_digits(row["document_number"])
        codigo = normalize_whitespace(row["client_code"])
        if not digitos or not codigo:
            continue
        exatos.setdefault(digitos, codigo)
        if len(digitos) == 14:
            por_raiz[digitos[:8]].add(codigo)
    return exatos, por_raiz


def client_code_from_name_digits(
    nome: str, exatos: dict[str, str], por_raiz: dict[str, set[str]]
) -> str | None:
    """Código do cliente a partir do documento embutido no nome do faturamento.

    A raiz de 8 dígitos só vale quando aponta para UM cliente: filiais dividem a
    mesma raiz, e chutar entre elas seria pior que não resolver.
    """
    digitos = only_digits(nome)
    if len(digitos) in (11, 14):
        return exatos.get(digitos)
    if len(digitos) == 8:
        candidatos = por_raiz.get(digitos) or set()
        return next(iter(candidatos)) if len(candidatos) == 1 else None
    return None


def client_alias_map(conn: sqlite3.Connection, company_id: int) -> dict[str, dict[str, str]]:
    """Nome do faturamento → cliente do cadastro, conciliado à mão.

    Usado SÓ para dar código a quem ficou sem. O casamento automático por nome
    continua sendo a regra: este mapa entra depois dele e apenas preenche as
    lacunas, nunca substitui nem remove nada da lista.
    """
    return {
        r["sales_name_key"]: {"clientCode": r["client_code"], "clientName": r["client_name"]}
        for r in conn.execute(
            "SELECT sales_name_key, client_code, client_name FROM client_name_aliases "
            "WHERE company_id = ?", (company_id,),
        ).fetchall()
    }


CLIENT_LOOKUP_COLUMNS = (
    "client_code, client_name, trade_name, city_name, document_number, "
    "NULLIF(TRIM(internal_seller_name), '') AS seller_name"
)


def find_client_by_code(
    conn: sqlite3.Connection, company_id: int, codigo: str
) -> dict[str, Any] | None:
    """Cliente pelo código do Alfa, tolerante à forma como o código foi gravado.

    Tenta na ordem: igual, sem zeros à esquerda e comparando só os dígitos. O
    cadastro às vezes chega com o código como texto ("076614") e a pessoa
    digita como número — recusar por isso seria implicância do sistema.
    """
    code = normalize_whitespace(codigo)
    if not code:
        return None
    tentativas = [code, code.lstrip("0"), only_digits(code)]
    for alvo in [t for t in dict.fromkeys(tentativas) if t]:
        row = conn.execute(
            f"SELECT {CLIENT_LOOKUP_COLUMNS} FROM crm_client_profiles "
            "WHERE company_id = ? AND (TRIM(client_code) = ? "
            "   OR CAST(TRIM(client_code) AS INTEGER) = CAST(? AS INTEGER))",
            (company_id, alvo, alvo),
        ).fetchone()
        if row:
            return dict(row)
    return None


def search_clients_by_name(
    conn: sqlite3.Connection, company_id: int, termo: str, limite: int = 20
) -> list[dict[str, Any]]:
    """Candidatos por razão social — o caminho quando o código não é conhecido."""
    texto = normalize_whitespace(termo)
    if len(texto) < 3:
        return []
    chave = company_match_key(texto)
    alvo = f"%{normalize_upper(strip_accents(texto))}%"
    linhas = [dict(r) for r in conn.execute(
        f"SELECT {CLIENT_LOOKUP_COLUMNS} FROM crm_client_profiles "
        "WHERE company_id = ? AND (UPPER(client_name) LIKE ? OR UPPER(COALESCE(trade_name,'')) LIKE ?) "
        "ORDER BY client_name LIMIT ?",
        (company_id, alvo, alvo, limite),
    ).fetchall()]
    if linhas or not chave:
        return linhas
    # Nada com o texto inteiro: tenta pela chave sem sufixo societário, que é o
    # caso de "POWERTECH CAR SERVICE LTDA" contra "... LTDA - ME".
    return [r for r in conn.execute(
        f"SELECT {CLIENT_LOOKUP_COLUMNS} FROM crm_client_profiles WHERE company_id = ?",
        (company_id,),
    ).fetchall() if company_match_key(r["client_name"]) == chave][:limite]


def save_client_alias(
    conn: sqlite3.Connection, company_id: int, user_id: int, payload: dict[str, Any]
) -> dict[str, Any]:
    nome_faturamento = normalize_whitespace(payload.get("salesName"))
    codigo = normalize_whitespace(payload.get("clientCode"))
    if not nome_faturamento or not codigo:
        raise ValueError("Informe o nome do faturamento e o código do cliente.")
    cliente = find_client_by_code(conn, company_id, codigo)
    if not cliente:
        raise ValueError(f"Cliente {codigo} não existe no cadastro.")

    conn.execute(
        """
        INSERT INTO client_name_aliases
            (company_id, sales_name_key, sales_name, client_code, client_name,
             created_by_user_id, created_at)
        VALUES (?,?,?,?,?,?,?)
        ON CONFLICT(company_id, sales_name_key)
        DO UPDATE SET client_code = excluded.client_code,
                      client_name = excluded.client_name,
                      created_by_user_id = excluded.created_by_user_id,
                      created_at = excluded.created_at
        """,
        (company_id, normalize_client_key(nome_faturamento), nome_faturamento,
         cliente["client_code"], cliente["client_name"], user_id, now_iso()),
    )
    audit_log(conn, company_id, user_id, "conciliar", "client_name_aliases",
              nome_faturamento, {"clientCode": cliente["client_code"]})
    conn.commit()
    invalidate_crm_cache(company_id)
    return {"message": f"{nome_faturamento} conciliado com {cliente['client_name']} "
                       f"({cliente['client_code']}).",
            "clientCode": cliente["client_code"], "clientName": cliente["client_name"]}


def compute_unassigned_clients(
    conn: sqlite3.Connection, company_id: int, user: sqlite3.Row,
    min_months: int = 2, months_window: int = 6, limit: int = 200,
) -> dict[str, Any]:
    """Clientes que compram com recorrência mas não têm vendedor no cadastro CRM.

    São clientes "no limpo": ninguém responde por eles, então ninguém previne a
    perda. O gestor usa esta lista para atribuir dono. Mostra quem já atendeu o
    cliente, o que ajuda a decidir a atribuição.
    """
    latest = crm_latest_competence(conn, company_id) or date.today().strftime("%Y-%m")
    window_start = shift_competence(latest, -(months_window - 1))

    # Clientes COM dono no cadastro — excluídos do resultado
    owned: set[str] = set()
    for row in conn.execute(
        "SELECT client_name FROM crm_client_profiles WHERE company_id = ? "
        "AND TRIM(COALESCE(internal_seller_name,'')) <> ''",
        (company_id,),
    ).fetchall():
        key = normalize_client_key(row["client_name"])
        if key:
            owned.add(key)

    # Código do cliente por nome — a ficha é aberta pelo código, mas esta lista
    # nasce do faturamento, que só tem o nome. Cliente sem cadastro fica sem
    # código e, nesse caso, a tela não oferece o botão de ficha.
    code_by_name: dict[str, str] = {}
    for row in conn.execute(
        "SELECT client_code, client_name FROM crm_client_profiles WHERE company_id = ?",
        (company_id,),
    ).fetchall():
        key = normalize_client_key(row["client_name"])
        code = normalize_whitespace(row["client_code"])
        if key and code and key not in code_by_name:
            code_by_name[key] = code

    # Segunda tentativa automática: documento embutido no nome do faturamento.
    docs_exatos, docs_raiz = build_document_client_map(conn, company_id)

    # Terceira: nome sem sufixo societário. Guarda o conjunto de códigos por
    # chave para poder recusar quando houver mais de um candidato.
    codes_by_company: dict[str, set[str]] = defaultdict(set)
    for row in conn.execute(
        "SELECT client_code, client_name FROM crm_client_profiles WHERE company_id = ?",
        (company_id,),
    ).fetchall():
        chave_empresa = company_match_key(row["client_name"])
        codigo = normalize_whitespace(row["client_code"])
        if chave_empresa and codigo:
            codes_by_company[chave_empresa].add(codigo)

    # Conciliação manual só PREENCHE lacuna: entra depois do casamento por nome
    # e nunca sobrescreve um código que já foi encontrado sozinho.
    aliases = client_alias_map(conn, company_id)
    for chave, destino in aliases.items():
        code_by_name.setdefault(chave, destino["clientCode"])

    # Dono do cliente por CÓDIGO — usado para tirar da lista quem só não casava
    # pelo nome mas tem vendedor no cadastro.
    dono_por_codigo = {
        normalize_whitespace(r["client_code"])
        for r in conn.execute(
            "SELECT client_code FROM crm_client_profiles WHERE company_id = ? "
            "AND TRIM(COALESCE(internal_seller_name,'')) <> ''",
            (company_id,),
        ).fetchall()
    }

    allowed_units = crm_allowed_units_for_user(conn, user)
    city_unit = build_city_unit_map(conn, company_id, latest)

    rows = conn.execute(
        """
        SELECT client_name,
               MAX(city_name) AS city_name,
               COUNT(DISTINCT competence) AS meses,
               ROUND(SUM(net_value), 2) AS receita,
               MAX(issue_date) AS ultima_compra,
               COUNT(DISTINCT seller_name) AS qtd_vendedores
        FROM fact_sales_detail
        WHERE company_id = ? AND competence >= ? AND net_value > 0
        GROUP BY client_name
        HAVING meses >= ?
        ORDER BY receita DESC
        """,
        (company_id, window_start, min_months),
    ).fetchall()

    # Vendedores que atenderam cada cliente, com receita de cada um
    sellers_by_client: dict[str, list[dict[str, Any]]] = defaultdict(list)
    for r in conn.execute(
        """
        SELECT client_name, seller_name,
               COUNT(DISTINCT competence) AS meses,
               ROUND(SUM(net_value), 2) AS receita
        FROM fact_sales_detail
        WHERE company_id = ? AND competence >= ? AND net_value > 0
        GROUP BY client_name, seller_name
        ORDER BY receita DESC
        """,
        (company_id, window_start),
    ).fetchall():
        sellers_by_client[normalize_client_key(r["client_name"])].append({
            "sellerName": normalize_whitespace(r["seller_name"]),
            "months": int(r["meses"] or 0),
            "revenue": float(r["receita"] or 0.0),
        })

    items: list[dict[str, Any]] = []
    total_revenue = 0.0
    for r in rows:
        key = normalize_client_key(r["client_name"])
        if not key or key in owned:
            continue
        if key not in code_by_name:
            achado = client_code_from_name_digits(r["client_name"], docs_exatos, docs_raiz)
            if not achado:
                candidatos = codes_by_company.get(company_match_key(r["client_name"])) or set()
                achado = next(iter(candidatos)) if len(candidatos) == 1 else None
            if achado:
                code_by_name[key] = achado
        # Casou por documento (ou por conciliação) com um cliente que tem
        # vendedor: a pendência era só o nome, então sai da lista.
        if code_by_name.get(key) in dono_por_codigo:
            continue
        unit = city_unit.get(normalize_upper(r["city_name"]))
        if allowed_units and unit not in allowed_units:
            continue
        sellers = sellers_by_client.get(key, [])[:5]
        receita = float(r["receita"] or 0.0)
        total_revenue += receita
        items.append({
            "clientKey": code_by_name.get(key),
            "clientName": r["client_name"],
            "aliasOf": (aliases.get(key) or {}).get("clientName"),
            "cityName": r["city_name"],
            "unitName": unit,
            "months": int(r["meses"] or 0),
            "revenue": round(receita, 2),
            "avgMonthly": round(safe_div(receita, int(r["meses"] or 1)), 2),
            "lastPurchaseAt": r["ultima_compra"],
            "sellerCount": int(r["qtd_vendedores"] or 0),
            "sellers": sellers,
            "mainSeller": sellers[0]["sellerName"] if sellers else None,
        })

    return {
        "items": items[:limit],
        "total": len(items),
        "totalRevenue": round(total_revenue, 2),
        "criteria": {
            "minMonths": min_months,
            "monthsWindow": months_window,
            "windowStart": window_start,
            "windowEnd": latest,
        },
    }


def compute_manager_mission(
    conn: sqlite3.Connection, company_id: int, user: sqlite3.Row,
    filters: dict[str, str | None], limit: int = 12,
) -> dict[str, Any]:
    """Missão do Dia na ótica de gestão: onde a execução está falhando.

    Diferente da visão do vendedor (fila de quem ligar), o gestor recebe dois
    recortes de risco, ambos com o vendedor responsável ao lado para cobrança:
      - coverageGap: cliente parou de comprar E ninguém o contatou há dias
      - highValueDrop: cliente Diamante/Ouro comprando menos
    """
    rows = crm_base_client_rows_cached(conn, company_id, filters)
    today = date.today()

    def days_since(value: Any) -> int | None:
        dt = parse_datetime_flexible(value) if value else None
        return (today - dt.date()).days if dt else None

    coverage_gap: list[dict[str, Any]] = []
    high_value_drop: list[dict[str, Any]] = []

    for row in rows:
        seller = normalize_whitespace(row.get("assignedSeller")) or "Sem vendedor"
        days_no_contact = days_since(row.get("lastInteractionAt"))
        status = row.get("statusCode")
        item = {
            "clientKey": row.get("clientKey"),
            "clientName": row.get("clientName"),
            "cityName": row.get("cityName"),
            "unitName": row.get("unitName"),
            "assignedSeller": seller,
            "statusCode": status,
            "classCode": row.get("classCode"),
            "daysWithoutPurchase": row.get("daysWithoutPurchase"),
            "daysWithoutContact": days_no_contact,
            "averageRevenue": row.get("averageRevenue"),
            "currentRevenue": row.get("currentRevenue"),
            "dropPct": row.get("dropPct"),
            "phone": row.get("phone"),
        }
        # Bloco 1 — cobertura falha: parou de comprar e ninguém falou com ele
        if status in {"INATIVO", "PRE_INATIVO"} and (days_no_contact is None or days_no_contact >= COVERAGE_GAP_DAYS):
            coverage_gap.append(item)
        # Bloco 2 — cliente grande perdendo volume
        if row.get("classCode") in {"DIAMANTE", "OURO"} and float(row.get("dropPct") or 0) <= -0.1:
            high_value_drop.append(item)

    # Prioriza o que dói mais no bolso: maior média histórica primeiro
    coverage_gap.sort(key=lambda r: float(r.get("averageRevenue") or 0), reverse=True)
    high_value_drop.sort(key=lambda r: float(r.get("averageRevenue") or 0), reverse=True)

    def by_seller(items: list[dict[str, Any]]) -> list[dict[str, Any]]:
        agg: dict[str, int] = defaultdict(int)
        for i in items:
            agg[i["assignedSeller"]] += 1
        return [{"sellerName": k, "count": v} for k, v in sorted(agg.items(), key=lambda x: -x[1])]

    return {
        "coverageGap": coverage_gap[:limit],
        "coverageGapTotal": len(coverage_gap),
        "coverageGapBySeller": by_seller(coverage_gap),
        "highValueDrop": high_value_drop[:limit],
        "highValueDropTotal": len(high_value_drop),
        "highValueDropBySeller": by_seller(high_value_drop),
        "coverageGapDays": COVERAGE_GAP_DAYS,
    }


def compute_team_activity_today(
    conn: sqlite3.Connection, company_id: int, user: sqlite3.Row
) -> dict[str, Any]:
    """Retorna atividade de hoje por vendedor para visão gerencial da Missão do Dia."""
    today_str = date.today().isoformat()
    competence = date.today().strftime("%Y-%m")
    comp_target = first_day_of_competence(competence).isoformat()

    contacts_rows = conn.execute(
        """
        SELECT seller_name,
               SUM(CASE WHEN initiative = 'ATIVO' THEN 1 ELSE 0 END) AS total,
               SUM(CASE WHEN result_code NOT IN ('NAO_ATENDEU','PEDIU_RETORNO')
                         AND initiative = 'ATIVO' THEN 1 ELSE 0 END) AS active
        FROM crm_interactions
        WHERE company_id = ? AND substr(occurred_at,1,10) = ?
        GROUP BY seller_name
        """,
        (company_id, today_str),
    ).fetchall()

    tasks_rows = conn.execute(
        "SELECT seller_name, COUNT(*) AS overdue FROM crm_tasks WHERE company_id = ? AND status = 'ATRASADA' GROUP BY seller_name",
        (company_id,),
    ).fetchall()

    open_rows = conn.execute(
        "SELECT seller_name, COUNT(*) AS abertas FROM crm_tasks WHERE company_id = ? AND status = 'ABERTA' GROUP BY seller_name",
        (company_id,),
    ).fetchall()
    open_map = {normalize_whitespace(r["seller_name"]): int(r["abertas"]) for r in open_rows}

    # Última vez que cada vendedor registrou qualquer interação
    last_interaction_map = {
        normalize_whitespace(r["seller_name"]): r["last_at"]
        for r in conn.execute(
            "SELECT seller_name, MAX(occurred_at) AS last_at FROM crm_interactions WHERE company_id = ? GROUP BY seller_name",
            (company_id,),
        ).fetchall()
    }

    # Vendedores com meta cadastrada no mês atual. A própria meta guarda a unidade
    # base — usar isso evita depender de people_records estar atualizado.
    seller_rows = conn.execute(
        "SELECT seller_name, MAX(base_unit) AS base_unit FROM goals_seller "
        "WHERE company_id = ? AND competence = ? GROUP BY seller_name",
        (company_id, competence),
    ).fetchall()

    seller_unit_map: dict[str, str] = {}
    for r in conn.execute(
        "SELECT person_name, base_unit FROM people_records WHERE company_id = ? AND date(valid_from) <= date(?) AND (valid_to IS NULL OR date(valid_to) >= date(?)) ORDER BY valid_from DESC",
        (company_id, comp_target, comp_target),
    ).fetchall():
        key = normalize_whitespace(r["person_name"])
        if key and key not in seller_unit_map:
            seller_unit_map[key] = normalize_unit(r["base_unit"]) or ""

    # Escopo pelo perfil (não pelo nome do papel): a Missão do Dia é uma visão de CRM,
    # então mesmo "unidade + consolidado" fica restrito às unidades vinculadas.
    allowed_units = crm_allowed_units_for_user(conn, user) or []
    contacts_map = {normalize_whitespace(r["seller_name"]): {"total": int(r["total"]), "active": int(r["active"])} for r in contacts_rows}
    overdue_map = {normalize_whitespace(r["seller_name"]): int(r["overdue"]) for r in tasks_rows}

    results = []
    for row in seller_rows:
        seller_name = normalize_whitespace(row["seller_name"])
        if not seller_name:
            continue
        # Unidade da meta tem prioridade; people_records é o complemento
        unit = normalize_unit(row["base_unit"]) or seller_unit_map.get(seller_name, "")
        if allowed_units and unit not in allowed_units:
            continue
        c = contacts_map.get(seller_name, {"total": 0, "active": 0})
        results.append({
            "sellerName": seller_name,
            "unit": unit,
            "contactsToday": c["active"],
            "totalInteractionsToday": c["total"],
            "overdueTasks": overdue_map.get(seller_name, 0),
            "openTasks": open_map.get(seller_name, 0),
            "lastInteractionAt": last_interaction_map.get(seller_name),
            "dailyGoal": DAILY_CONTACT_GOAL,
        })

    # Quem não trabalhou aparece primeiro — é onde o gerente precisa agir
    results.sort(key=lambda r: (r["contactsToday"], -r["overdueTasks"]))
    team_goal = len(results) * DAILY_CONTACT_GOAL
    total_contacts = sum(r["contactsToday"] for r in results)
    return {
        "date": today_str,
        "sellers": results,
        "totalContactsToday": total_contacts,
        "teamGoal": team_goal,
        "dailyGoalPerSeller": DAILY_CONTACT_GOAL,
        "sellersWithContact": sum(1 for r in results if r["contactsToday"] > 0),
        "sellersWithoutContact": sum(1 for r in results if r["contactsToday"] == 0),
        "totalOverdueTasks": sum(r["overdueTasks"] for r in results),
        "goalPct": round(total_contacts / team_goal * 100, 1) if team_goal > 0 else 0.0,
    }


def compute_portfolio_summary_by_seller(
    conn: sqlite3.Connection, company_id: int, user: sqlite3.Row,
    competence: str | None = None, unit_filter: str | list[str] | None = None,
    person_type_filter: str | None = None,
) -> dict[str, Any]:
    """Retorna resumo da carteira por vendedor para dashboard gerencial.

    Baseia a contagem de clientes em crm_client_profiles (relatório de cadastro),
    independente de importações de faturamento. Calcula status a partir da compra
    mais recente em crm_client_summary (qualquer competência), com fallback para
    last_sale_at do perfil.
    """
    today = date.today()
    # unit_filter aceita string única ou lista (gerente pode ter várias unidades).
    # None = sem restrição.
    if unit_filter is None:
        unit_filter_set = None
    else:
        values = [unit_filter] if isinstance(unit_filter, str) else list(unit_filter)
        unit_filter_set = {normalize_unit(v) for v in values if v}
    # Usa a competência mais recente com dados de CRM como padrão
    if not competence:
        competence = crm_summary_latest_competence(conn, company_id) or today.strftime("%Y-%m")
    # Calcula mês anterior relativo à competência solicitada
    comp_year, comp_month = int(competence[:4]), int(competence[5:7])
    if comp_month == 1:
        prev_competence = f"{comp_year - 1}-12"
    else:
        prev_competence = f"{comp_year}-{comp_month - 1:02d}"

    # Busca todos os clientes do cadastro, enriquecidos com:
    # - data de compra mais recente (pré-agregada via CTE para evitar subconsulta correlacionada)
    # - receita no mês atual e anterior (para comVendaMes e queda)
    rows = conn.execute(
        """
        WITH latest_purchase AS (
            SELECT client_code, MAX(last_purchase_at) AS latest_purchase_at
            FROM crm_client_summary
            WHERE company_id = ?
            GROUP BY client_code
        ),
        seller_units AS (
            SELECT person_name, base_unit,
                ROW_NUMBER() OVER (PARTITION BY company_id, person_name ORDER BY valid_from DESC) AS rn
            FROM people_records
            WHERE company_id = ?
        )
        SELECT
            COALESCE(NULLIF(TRIM(p.internal_seller_name), ''), 'Sem Vendedor') AS seller,
            COALESCE(su.base_unit, '') AS seller_unit,
            COALESCE(lp.latest_purchase_at, p.last_sale_at) AS effective_last_sale,
            COALESCE(cs.net_value, 0) AS current_revenue,
            COALESCE(cs_prev.net_value, 0) AS prev_revenue,
            p.document_number AS document_number,
            CASE WHEN COALESCE(cs.net_value, 0) > 0 THEN 1 ELSE 0 END AS bought_current,
            CASE WHEN COALESCE(cs_prev.net_value, 0) > 0 THEN 1 ELSE 0 END AS bought_prev
        FROM crm_client_profiles p
        LEFT JOIN latest_purchase lp ON lp.client_code = p.client_code
        LEFT JOIN crm_client_summary cs
            ON cs.company_id = p.company_id
            AND cs.client_code = p.client_code
            AND cs.competence = ?
        LEFT JOIN crm_client_summary cs_prev
            ON cs_prev.company_id = p.company_id
            AND cs_prev.client_code = p.client_code
            AND cs_prev.competence = ?
        LEFT JOIN seller_units su
            ON su.person_name = COALESCE(NULLIF(TRIM(p.internal_seller_name), ''), 'Sem Vendedor')
            AND su.rn = 1
        WHERE p.company_id = ?
        """,
        (company_id, company_id, competence, prev_competence, company_id),
    ).fetchall()

    by_seller: dict[str, dict[str, Any]] = {}

    for row in rows:
        seller = row["seller"]
        unit = row["seller_unit"] or ""

        # Aplica filtro por unidade no backend (case-insensitive)
        if unit_filter_set is not None and normalize_unit(unit) not in unit_filter_set:
            continue

        # Aplica filtro por tipo de pessoa (PJ/PF)
        if person_type_filter:
            doc_pt, _ = person_type_from_document(row["document_number"])
            if not doc_pt:
                doc_pt, _, _ = infer_person_type_from_name(row["seller"])
            if normalize_upper(doc_pt) != normalize_upper(person_type_filter):
                continue

        if seller not in by_seller:
            by_seller[seller] = {
                "sellerName": seller, "unit": unit,
                "total": 0, "ativos": 0, "preInativos": 0, "inativos": 0,
                "comVendaMes": 0, "semVendaMes": 0, "comVendaMesAnterior": 0,
                "queda30": 0, "queda20": 0, "queda10": 0,
            }
        d = by_seller[seller]
        d["total"] += 1

        # Status a partir da data de compra mais recente (crm_client_summary ou perfil)
        last_sale = row["effective_last_sale"]
        if last_sale:
            try:
                last_date = date.fromisoformat(str(last_sale)[:10])
                days_since = (today - last_date).days
                if days_since <= 29:
                    d["ativos"] += 1
                elif days_since <= 60:
                    d["preInativos"] += 1
                else:
                    d["inativos"] += 1
            except (ValueError, TypeError):
                d["inativos"] += 1
        else:
            d["inativos"] += 1

        if row["bought_current"]:
            d["comVendaMes"] += 1
        else:
            d["semVendaMes"] += 1

        if row["bought_prev"]:
            d["comVendaMesAnterior"] += 1

        # Queda de faturamento: somente clientes que compraram nos dois meses
        cur_rev = float(row["current_revenue"] or 0)
        prev_rev = float(row["prev_revenue"] or 0)
        if cur_rev > 0 and prev_rev > 0:
            growth_pct = (cur_rev - prev_rev) / prev_rev * 100
            if growth_pct < -30:
                d["queda30"] += 1
            elif growth_pct < -20:
                d["queda20"] += 1
            elif growth_pct < -10:
                d["queda10"] += 1

    result = sorted(by_seller.values(), key=lambda r: r["total"], reverse=True)
    for d in result:
        t = max(d["total"], 1)
        d["pctAtivos"] = round(d["ativos"] / t * 100, 1)
        d["pctPreInativos"] = round(d["preInativos"] / t * 100, 1)
        d["pctInativos"] = round(d["inativos"] / t * 100, 1)
        d["pctComVendaMes"] = round(d["comVendaMes"] / t * 100, 1)

    totals: dict[str, Any] = {k: sum(d[k] for d in result) for k in ("total", "ativos", "preInativos", "inativos", "comVendaMes", "semVendaMes", "comVendaMesAnterior", "queda30", "queda20", "queda10")}
    t2 = max(totals["total"], 1)
    totals["pctAtivos"] = round(totals["ativos"] / t2 * 100, 1)
    totals["pctPreInativos"] = round(totals["preInativos"] / t2 * 100, 1)
    totals["pctInativos"] = round(totals["inativos"] / t2 * 100, 1)
    totals["pctComVendaMes"] = round(totals["comVendaMes"] / t2 * 100, 1)

    return {
        "competence": competence,
        "prevCompetence": prev_competence,
        "sellers": result,
        "totals": totals,
    }


class AppHandler(BaseHTTPRequestHandler):
    server_version = "PassiniDashboard/1.0"

    def log_message(self, format: str, *args) -> None:
        return

    def handle(self) -> None:
        try:
            super().handle()
        except Exception:
            traceback.print_exc()
            raise

    def handle_one_request(self) -> None:
        try:
            super().handle_one_request()
        except Exception:
            traceback.print_exc()
            raise

    def _set_headers(self, status: int = 200, content_type: str = "application/json; charset=utf-8", extra_headers: dict[str, str] | None = None) -> None:
        self.send_response(status)
        self.send_header("Content-Type", content_type)
        self.send_header("Cache-Control", "no-store")
        if extra_headers:
            for key, value in extra_headers.items():
                self.send_header(key, value)
        self.end_headers()

    def _read_json(self) -> dict[str, Any]:
        length = int(self.headers.get("Content-Length", "0"))
        raw = self.rfile.read(length) if length else b"{}"
        return json.loads(raw.decode("utf-8") or "{}")

    def _parse_multipart(self) -> tuple[list[dict[str, Any]], dict[str, str]]:
        environ = {"REQUEST_METHOD": "POST", "CONTENT_TYPE": self.headers.get("Content-Type", "")}
        form = cgi.FieldStorage(fp=self.rfile, headers=self.headers, environ=environ, keep_blank_values=True)
        files: list[dict[str, Any]] = []
        fields: dict[str, str] = {}
        if form.list:
            for item in form.list:
                if item.filename:
                    files.append({"fieldName": item.name, "fileName": item.filename, "content": item.file.read()})
                else:
                    fields[item.name] = item.value
        return files, fields

    def _serve_file(self, file_path: Path, content_type: str) -> None:
        """Serve arquivo estático com validação de cache.

        Sem estes cabeçalhos o navegador guardava app.js indefinidamente e a
        tela continuava na versão antiga depois do deploy — era preciso pedir
        Ctrl+Shift+R para cada usuário. O ETag usa data e tamanho do arquivo:
        enquanto nada muda o navegador reaproveita (resposta 304, leve), e no
        instante em que o arquivo é atualizado a nova versão desce sozinha.
        """
        if not file_path.exists():
            self._set_headers(404)
            self.wfile.write(json_dumps({"error": "Arquivo não encontrado"}))
            return
        stat = file_path.stat()
        etag = f'"{int(stat.st_mtime)}-{stat.st_size}"'
        if self.headers.get("If-None-Match") == etag:
            self.send_response(304)
            self.send_header("ETag", etag)
            self.send_header("Cache-Control", "no-cache, must-revalidate")
            self.end_headers()
            return
        # Cabeçalhos escritos direto: _set_headers força "no-store", que faria o
        # navegador rebaixar o arquivo inteiro toda vez. Com no-cache + ETag ele
        # revalida e recebe 304 quando nada mudou.
        self.send_response(200)
        self.send_header("Content-Type", content_type)
        self.send_header("ETag", etag)
        self.send_header("Cache-Control", "no-cache, must-revalidate")
        self.end_headers()
        self.wfile.write(file_path.read_bytes())

    def _current_user(self) -> dict[str, Any] | None:
        cookie = SimpleCookie(self.headers.get("Cookie"))
        session_cookie = cookie.get(SESSION_COOKIE)
        if not session_cookie:
            return None
        session_id = session_cookie.value
        with closing(get_connection()) as conn:
            row = conn.execute(
                """
                SELECT users.*
                FROM sessions
                JOIN users ON users.id = sessions.user_id
                WHERE sessions.id = ? AND datetime(sessions.expires_at) >= datetime(?)
                """,
                (session_id, now_iso()),
            ).fetchone()
            return dict(row) if row else None

    def _require_auth(self) -> dict[str, Any] | None:
        user = self._current_user()
        if not user:
            self._set_headers(401)
            self.wfile.write(json_dumps({"error": "Sessão expirada"}))
            return None
        return user

    # Módulos que caracterizam a área administrativa — basta ter um deles no perfil
    _ADMIN_AREA_MODULES = {"configuracoes", "administracao", "importacoes", "acessos"}

    def _require_admin_area(self, user: dict[str, Any] | None) -> bool:
        if not user:
            return False
        with closing(get_connection()) as conn:
            profile = get_access_profile_for_user(conn, user)
        if profile is not None:
            if not (set(profile["modules"]) & self._ADMIN_AREA_MODULES):
                self._set_headers(403)
                self.wfile.write(json_dumps(
                    {"error": f"O perfil '{profile['name']}' não tem acesso à área administrativa."}
                ))
                return False
            return True
        # Sem perfil cadastrado: mantém a regra antiga
        if user["role"] == "Vendedor":
            self._set_headers(403)
            self.wfile.write(json_dumps({"error": "Perfil sem acesso a area administrativa"}))
            return False
        return True

    def _require_unit_allowed(self, user: dict[str, Any] | None, unit_name: str | None) -> bool:
        """Impede gravar dados de unidade fora do vínculo do usuário."""
        if not user:
            return False
        with closing(get_connection()) as conn:
            scope = data_scope_for_user(conn, user)
            if scope not in {"unidade", "unidade_consolidado"}:
                return True
            allowed = linked_units_for_user(user)
        target = normalize_unit(unit_name)
        if target and target in allowed:
            return True
        self._set_headers(403)
        self.wfile.write(json_dumps(
            {"error": f"Seu perfil só permite alterar dados de: {', '.join(allowed) or 'nenhuma unidade'}."}
        ))
        return False

    def _require_user_management(self, user: dict[str, Any] | None) -> bool:
        """Gestão de usuários e perfis é restrita a quem tem a permissão no perfil."""
        if not user:
            return False
        with closing(get_connection()) as conn:
            allowed = user_can_manage_users(conn, user)
        if not allowed:
            self._set_headers(403)
            self.wfile.write(json_dumps(
                {"error": "Seu perfil não permite gerenciar usuários. Fale com o Diretor ou Administrador."}
            ))
            return False
        return True

    def do_OPTIONS(self) -> None:
        self.send_response(204)
        self.send_header("Access-Control-Allow-Origin", "*")
        self.send_header("Access-Control-Allow-Methods", "GET, POST, OPTIONS")
        self.send_header("Access-Control-Allow-Headers", "Content-Type")
        self.end_headers()

    def do_GET(self) -> None:
        _t0 = time.time()
        try:
            self._do_GET_inner()
        finally:
            _elapsed = time.time() - _t0
            if _elapsed > 1.0 and "/api/" in self.path:
                print(f"[req] {_elapsed:6.1f}s  GET {self.path}", flush=True)

    def _do_GET_inner(self) -> None:
        try:
            parsed = urlparse(self.path)
            path = parsed.path
            if path == "/":
                self._serve_file(STATIC_DIR / "index.html", "text/html; charset=utf-8")
                return
            if path == "/app.js":
                self._serve_file(STATIC_DIR / "app.js", "application/javascript; charset=utf-8")
                return
            if path == "/styles.css":
                self._serve_file(STATIC_DIR / "styles.css", "text/css; charset=utf-8")
                return
            # Imagens da marca servidas direto de /static. Nome fixo por segurança:
            # aceitar caminho do cliente aqui abriria leitura de arquivo arbitrário.
            if path in ("/logo.png", "/logo.svg", "/logo.jpg", "/logo.webp"):
                nome = path.lstrip("/")
                tipo = {
                    "logo.png": "image/png", "logo.svg": "image/svg+xml",
                    "logo.jpg": "image/jpeg", "logo.webp": "image/webp",
                }[nome]
                self._serve_file(STATIC_DIR / nome, tipo)
                return
            if path == "/api/health":
                self._set_headers(200)
                self.wfile.write(json_dumps({"ok": True, "timestamp": now_iso(), "build": "v20260602-agenda-fix"}))
                return
            if path == "/api/debug/data-summary":
                user = self._require_auth()
                if not user:
                    return
                conn = get_db()
                cid = user["company_id"]
                def q(sql, *p):
                    return [dict(r) for r in conn.execute(sql, p).fetchall()]
                result = {
                    "db_path": str(DB_PATH),
                    "db_size_mb": round(DB_PATH.stat().st_size / 1024 / 1024, 2) if DB_PATH.exists() else 0,
                    "fact_sales_detail": q("SELECT competence, COUNT(DISTINCT client_name) AS clientes, COUNT(DISTINCT seller_name) AS vendedores, ROUND(SUM(net_value),2) AS total FROM fact_sales_detail WHERE company_id=? GROUP BY competence ORDER BY competence", cid),
                    "fact_vendor_summary": q("SELECT competence, COUNT(*) AS linhas, ROUND(SUM(sale_value),2) AS total FROM fact_vendor_summary WHERE company_id=? GROUP BY competence ORDER BY competence", cid),
                    "crm_client_summary": q("SELECT competence, COUNT(DISTINCT client_code) AS clientes, SUM(CASE WHEN net_value>0 THEN 1 ELSE 0 END) AS com_valor, ROUND(SUM(net_value),2) AS total FROM crm_client_summary WHERE company_id=? GROUP BY competence ORDER BY competence", cid),
                    "crm_client_profiles": q("SELECT COUNT(*) AS total FROM crm_client_profiles WHERE company_id=?", cid),
                    "fact_unit_summary": q("SELECT competence, COUNT(*) AS linhas FROM fact_unit_summary WHERE company_id=? GROUP BY competence ORDER BY competence", cid),
                    "imports_recentes": q("SELECT file_type, competence, status, created_at FROM imports WHERE company_id=? ORDER BY created_at DESC LIMIT 20", cid),
                }
                conn.close()
                self._set_headers(200)
                self.wfile.write(json_dumps(result))
                return
            if path == "/api/session":
                user = self._current_user()
                if not user:
                    self._set_headers(200)
                    self.wfile.write(json_dumps({"authenticated": False}))
                    return
                with closing(get_connection()) as conn:
                    profile = get_access_profile_for_user(conn, user)
                self._set_headers(200)
                self.wfile.write(
                    json_dumps(
                        {
                            "authenticated": True,
                            "user": {
                                "username": user["username"],
                                "fullName": user["full_name"],
                                "linkedPersonName": user["linked_person_name"],
                                "linkedUnits": linked_units_for_user(user),
                                "role": user["role"],
                                # Permissões vindas do perfil — a tela usa isso para montar o menu
                                "profileName": profile["name"] if profile else user["role"],
                                "modules": profile["modules"] if profile else [],
                                "dataScope": profile["dataScope"] if profile else "todos",
                                "canManageUsers": profile["canManageUsers"] if profile else False,
                            },
                        }
                    )
                )
                return
            if path == "/api/options":
                user = self._require_auth()
                if not user:
                    return
                with closing(get_connection()) as conn:
                    company_id = user["company_id"]
                    competences = query_competences(conn, company_id)
                    if user["role"] == "Vendedor":
                        seller_name = seller_identity_for_user(user)
                        _, base_unit = current_role_and_unit(conn, company_id, seller_name, competences[0] if competences else date.today().strftime("%Y-%m"))
                        units = [normalize_unit(base_unit)] if base_unit else []
                        sellers = [seller_name]
                        cities = [
                            row["city_name"]
                            for row in conn.execute(
                                "SELECT DISTINCT city_name FROM fact_sales_detail WHERE company_id = ? AND seller_name = ? AND city_name IS NOT NULL AND city_name <> '' ORDER BY city_name",
                                (company_id, seller_name),
                            ).fetchall()
                        ]
                    elif user["role"] in {"Gerente", "Analista"}:
                        linked_units = linked_units_for_user(user)
                        if linked_units:
                            units = linked_units
                            seller_competence = competences[0] if competences else date.today().strftime("%Y-%m")
                            sellers = []
                            for row in conn.execute(
                                "SELECT DISTINCT seller_name FROM fact_vendor_summary WHERE company_id = ? ORDER BY seller_name",
                                (company_id,),
                            ).fetchall():
                                seller_name = normalize_whitespace(row["seller_name"])
                                _, seller_base_unit = current_role_and_unit(conn, company_id, seller_name, seller_competence)
                                if normalize_unit(seller_base_unit) in linked_units:
                                    sellers.append(seller_name)
                            cities = active_mapped_cities_for_units(conn, company_id, linked_units)
                        else:
                            units = []
                            sellers = []
                            cities = []
                    else:
                        _db_units = {normalize_unit(row["unit_name"]) for row in conn.execute("SELECT DISTINCT unit_name FROM fact_unit_summary WHERE company_id = ?", (company_id,)).fetchall() if row["unit_name"]}
                        _all_units = list(dict.fromkeys([u for u in CANONICAL_UNITS if u in _db_units or True] + sorted(_db_units - set(CANONICAL_UNITS))))
                        units = _all_units
                        sellers = [row["seller_name"] for row in conn.execute("SELECT DISTINCT seller_name FROM fact_vendor_summary WHERE company_id = ? ORDER BY seller_name", (company_id,)).fetchall()]
                        cities = [row["city_name"] for row in conn.execute("SELECT DISTINCT city_name FROM fact_sales_detail WHERE company_id = ? AND city_name IS NOT NULL AND city_name <> '' ORDER BY city_name", (company_id,)).fetchall()]
                    # Mapeia vendedor → unidade base via people_records
                    seller_unit_map: dict[str, str | None] = {}
                    for r in conn.execute(
                        "SELECT person_name, base_unit FROM people_records WHERE company_id = ? AND valid_to IS NULL ORDER BY valid_from DESC",
                        (company_id,),
                    ).fetchall():
                        pname = normalize_whitespace(r["person_name"])
                        if pname and pname not in seller_unit_map:
                            seller_unit_map[pname] = normalize_whitespace(r["base_unit"])
                    sellers_with_units = [{"name": s, "unit": seller_unit_map.get(s)} for s in sellers]
                    self._set_headers(200)
                    self.wfile.write(json_dumps({"competences": competences, "units": units, "sellers": sellers, "sellersWithUnits": sellers_with_units, "cities": cities}))
                return
            if path == "/api/dashboard":
                user = self._require_auth()
                if not user:
                    return
                query = parse_qs(parsed.query)
                with closing(get_connection()) as conn:
                    filters = scoped_filters_for_user(conn, user["company_id"], user, build_filters_from_query(query))
                    data = get_dashboard_data_cached(conn, user["company_id"], filters)
                self._set_headers(200)
                self.wfile.write(json_dumps(data))
                return
            if path == "/api/audit/integrity":
                user = self._require_auth()
                if not user:
                    return
                query = parse_qs(parsed.query)
                with closing(get_connection()) as conn:
                    competence = normalize_whitespace(query.get("competence", [None])[0]) or (
                        query_competences(conn, user["company_id"])[0] if query_competences(conn, user["company_id"]) else date.today().strftime("%Y-%m")
                    )
                    audit_result = build_integrity_audit(conn, user["company_id"], competence)
                self._set_headers(200)
                self.wfile.write(json_dumps(audit_result))
                return
            if path == "/api/crm/options":
                user = self._require_auth()
                if not user:
                    return
                self._set_headers(200)
                self.wfile.write(
                    json_dumps(
                        {
                            "contactTypes": [{"code": code, "label": label} for code, label in CRM_CONTACT_TYPES],
                            "contactResults": [
                                {
                                    "code": code,
                                    "label": label,
                                    "generatesFollowup": bool(generates_followup),
                                    "requiresFollowupDate": bool(requires_followup_date),
                                }
                                for code, label, generates_followup, requires_followup_date in CRM_CONTACT_RESULTS
                            ],
                        }
                    )
                )
                return
            if path == "/api/crm/seller-score":
                user = self._require_auth()
                if not user:
                    return
                query = parse_qs(parsed.query)
                req_comp = normalize_whitespace(query.get("competence", [None])[0])
                with closing(get_connection()) as conn:
                    score_data = compute_seller_score(conn, user["company_id"], user, req_comp or None)
                self._set_headers(200)
                self.wfile.write(json_dumps(score_data))
                return
            if path == "/api/crm/team-score":
                user = self._require_auth()
                if not user:
                    return
                try:
                    with closing(get_connection()) as conn:
                        team_data = compute_team_score(conn, user["company_id"], user)
                    self._set_headers(200)
                    self.wfile.write(json_dumps(team_data))
                except Exception as _e:
                    traceback.print_exc()
                    self._set_headers(500)
                    self.wfile.write(json_dumps({"error": str(_e), "sellers": [], "summary": {"total": 0, "eligible": 0, "inPrizeZone": 0, "fullPrize": 0}, "competence": ""}))
                return
            if path == "/api/crm/team-activity-today":
                # Painel de gestão da Missão do Dia. Gerente precisa dele, então não é
                # tratado como área administrativa — o recorte por unidade já limita.
                user = self._require_auth()
                if not user:
                    return
                try:
                    with closing(get_connection()) as conn:
                        if data_scope_for_user(conn, user) == "proprio":
                            self._set_headers(403)
                            self.wfile.write(json_dumps({"error": "Visão exclusiva de gestão."}))
                            return
                        data = compute_team_activity_today(conn, user["company_id"], user)
                        query = parse_qs(parsed.query)
                        filters = crm_scoped_filters_for_user(
                            conn, user["company_id"], user, build_filters_from_query(query)
                        )
                        data["risk"] = compute_manager_mission(conn, user["company_id"], user, filters)
                        data["scopeUnits"] = crm_allowed_units_for_user(conn, user)
                    self._set_headers(200)
                    self.wfile.write(json_dumps(data))
                except Exception as _e:
                    traceback.print_exc()
                    self._set_headers(500)
                    self.wfile.write(json_dumps({"error": str(_e), "sellers": [], "totalContactsToday": 0, "teamGoal": 0}))
                return
            if path == "/api/auto-import/status":
                user = self._require_auth()
                if not user:
                    return
                if not self._require_admin_area(user):
                    return
                try:
                    with closing(get_connection()) as conn:
                        rows = conn.execute(
                            "SELECT ran_at, folder, scope, competence, status, message, files_json "
                            "FROM auto_import_log ORDER BY ran_at DESC LIMIT 50"
                        ).fetchall()
                    logs = [
                        {
                            "ranAt": r["ran_at"], "folder": r["folder"], "scope": r["scope"],
                            "competence": r["competence"], "status": r["status"],
                            "message": r["message"], "files": json.loads(r["files_json"] or "[]"),
                        }
                        for r in rows
                    ]
                    folders_info = []
                    for cfg in AUTO_IMPORT_FOLDERS:
                        p = AUTO_IMPORT_BASE / cfg["folder"]
                        pending = [f.name for f in p.glob("*.csv")] if p.exists() else []
                        folders_info.append({
                            "folder": cfg["folder"], "label": cfg["label"],
                            "scope": cfg["scope"], "path": str(p),
                            "hint": cfg.get("hint", ""),
                            "pendingFiles": pending,
                        })
                    self._set_headers(200)
                    self.wfile.write(json_dumps({
                        "logs": logs,
                        "folders": folders_info,
                        "intervalMinutes": AUTO_IMPORT_INTERVAL // 60,
                        "running": _AUTO_IMPORT_RUNNING.is_set(),
                    }))
                except Exception as exc:
                    traceback.print_exc()
                    self._set_headers(500)
                    self.wfile.write(json_dumps({"error": str(exc)}))
                return
            if path == "/api/crm/unassigned-clients":
                # Clientes recorrentes sem vendedor no cadastro — visão de gestão
                user = self._require_auth()
                if not user:
                    return
                try:
                    query = parse_qs(parsed.query)
                    min_months = max(1, min(int(query.get("minMonths", ["2"])[0]), 12))
                    window = max(2, min(int(query.get("window", ["6"])[0]), 24))
                    with closing(get_connection()) as conn:
                        if data_scope_for_user(conn, user) == "proprio":
                            self._set_headers(403)
                            self.wfile.write(json_dumps({"error": "Visão exclusiva de gestão."}))
                            return
                        data = compute_unassigned_clients(
                            conn, user["company_id"], user,
                            min_months=min_months, months_window=window,
                        )
                    self._set_headers(200)
                    self.wfile.write(json_dumps(data))
                except Exception as exc:
                    traceback.print_exc()
                    self._set_headers(500)
                    self.wfile.write(json_dumps({"error": str(exc), "items": [], "total": 0}))
                return
            if path == "/api/kpi-thresholds":
                user = self._require_auth()
                if not user:
                    return
                with closing(get_connection()) as conn:
                    thresholds = load_kpi_thresholds(conn, user["company_id"])
                    can_edit = user_can_manage_users(conn, user)
                self._set_headers(200)
                self.wfile.write(json_dumps({
                    "metrics": [thresholds[m["id"]] for m in KPI_METRICS if m["id"] in thresholds],
                    "levels": FAROL_LEVELS,
                    "canEdit": can_edit,
                }))
                return
            if path == "/api/prospects":
                user = self._require_auth()
                if not user:
                    return
                query = parse_qs(parsed.query)
                with closing(get_connection()) as conn:
                    competencia = (normalize_whitespace(query.get("competence", [""])[0])
                                   or crm_latest_competence(conn, user["company_id"])
                                   or date.today().strftime("%Y-%m"))
                    escopo = data_scope_for_user(conn, user)
                    permitidas = crm_allowed_units_for_user(conn, user)
                    if escopo == "proprio":
                        minha_unidade = seller_unit_name(
                            conn, user["company_id"], seller_identity_for_user(user), competencia)
                    else:
                        minha_unidade = (permitidas or [""])[0] if permitidas else ""
                    payload = {
                        "prospects": list_prospects(
                            conn, user["company_id"], user,
                            status=normalize_upper(query.get("status", [""])[0]),
                            search=normalize_whitespace(query.get("q", [""])[0]),
                            seller=normalize_whitespace(query.get("seller", [""])[0])),
                        "funnel": prospect_funnel(conn, user["company_id"], user),
                        "statuses": PROSPECT_STATUSES,
                        "triggers": PROSPECT_TRIGGERS,
                        "metrics": ACTIVITY_METRICS,
                        "competence": competencia,
                        "unitName": minha_unidade,
                        "unitPhase": get_unit_phase(conn, user["company_id"], minha_unidade),
                        "activity": activity_progress(
                            conn, user["company_id"], competencia, minha_unidade,
                            seller_identity_for_user(user) if escopo == "proprio" else ""),
                        "canManage": escopo != "proprio",
                        "canSetPhase": user_can_manage_users(conn, user),
                        "units": permitidas if permitidas is not None else [],
                        "phases": list_unit_phases(conn, user["company_id"]) if user_can_manage_users(conn, user) else [],
                        "sellers": ([s["sellerName"] for s in
                                     sellers_available_for_assignment(conn, user["company_id"], user)]
                                    if escopo != "proprio" else []),
                        "myName": seller_identity_for_user(user) if escopo == "proprio" else meeting_person_identity(user),
                    }
                self._set_headers(200)
                self.wfile.write(json_dumps(payload))
                return
            if path == "/api/help":
                user = self._require_auth()
                if not user:
                    return
                with closing(get_connection()) as conn:
                    dados = assistant_payload(conn, user["company_id"], user)
                self._set_headers(200)
                self.wfile.write(json_dumps(dados))
                return
            if path == "/api/visits":
                user = self._require_auth()
                if not user:
                    return
                query = parse_qs(parsed.query)
                with closing(get_connection()) as conn:
                    pode_gerir = user_can_manage_visits(conn, user)
                    payload = {
                        "visits": list_visits(
                            conn, user["company_id"], user,
                            client_key=normalize_whitespace(query.get("client", [""])[0]),
                            status=normalize_upper(query.get("status", [""])[0])),
                        # O gestor vê a fila do que precisa responder; o vendedor vê
                        # TODOS os pedidos dele, inclusive os recusados — senão ele
                        # nunca descobre por que a visita não aconteceu.
                        "requests": list_visit_requests(
                            conn, user["company_id"], user,
                            status=normalize_upper(query.get("requestStatus", [""])[0])
                                   or ("" if not pode_gerir else "PENDENTE")),
                        "types": VISIT_TYPES,
                        "statuses": VISIT_STATUSES,
                        "canManage": pode_gerir,
                        "myName": meeting_person_identity(user),
                        "sellers": ([s["sellerName"] for s in
                                     sellers_available_for_assignment(conn, user["company_id"], user)]
                                    if pode_gerir else []),
                    }
                self._set_headers(200)
                self.wfile.write(json_dumps(payload))
                return
            if path == "/api/visits/client-search":
                user = self._require_auth()
                if not user:
                    return
                query = parse_qs(parsed.query)
                with closing(get_connection()) as conn:
                    if not user_can_manage_visits(conn, user):
                        self._set_headers(403)
                        self.wfile.write(json_dumps({"error": "Busca disponível para a gestão."}))
                        return
                    itens = search_clients_for_visit(
                        conn, user["company_id"], user,
                        normalize_whitespace(query.get("q", [""])[0]))
                self._set_headers(200)
                self.wfile.write(json_dumps({"clients": itens}))
                return
            if path == "/api/visits/suggestions":
                user = self._require_auth()
                if not user:
                    return
                query = parse_qs(parsed.query)
                with closing(get_connection()) as conn:
                    if not user_can_manage_visits(conn, user):
                        self._set_headers(403)
                        self.wfile.write(json_dumps({"error": "Roteiro é da gestão."}))
                        return
                    dados = suggest_visits(
                        conn, user["company_id"], user,
                        city=normalize_whitespace(query.get("city", [""])[0]),
                        include_relationship=query.get("relationship", ["1"])[0] == "1",
                    )
                self._set_headers(200)
                self.wfile.write(json_dumps(dados))
                return
            if path == "/api/feedback":
                user = self._require_auth()
                if not user:
                    return
                query = parse_qs(parsed.query)
                with closing(get_connection()) as conn:
                    pode_dar = user_can_give_feedback(conn, user)
                    competencias = query_competences(conn, user["company_id"])
                    payload = {
                        "feedbacks": list_feedbacks(
                            conn, user["company_id"], user,
                            kind=normalize_upper(query.get("kind", [""])[0]),
                            competence=normalize_whitespace(query.get("competence", [""])[0]),
                            person=normalize_whitespace(query.get("person", [""])[0]),
                        ),
                        "notes": list_feedback_notes(
                            conn, user["company_id"], user,
                            competence=normalize_whitespace(query.get("competence", [""])[0])),
                        "kinds": FEEDBACK_KINDS,
                        "noteKinds": FEEDBACK_NOTE_KINDS,
                        "levels": _mec_content().FEEDBACK_LEVELS,
                        "pdiStatuses": PDI_STATUSES,
                        "pdiMaxActive": PDI_MAX_ACTIVE,
                        "canGive": pode_dar,
                        "canGiveManagerFeedback": user_can_manage_users(conn, user),
                        "canReadConfidential": user_can_read_confidential(conn, user),
                        "pendingCount": (count_pending_feedback_ack(conn, user["company_id"], user)
                                         + count_pending_note_ack(conn, user["company_id"], user)),
                        "pendingFeedbackCount": count_pending_feedback_ack(conn, user["company_id"], user),
                        "myName": meeting_person_identity(user),
                        "people": list_meeting_people(conn, user["company_id"], user) if pode_dar else [],
                        "units": meeting_units_for_user(conn, user)["units"] if pode_dar else [],
                        # query_competences já devolve da mais recente para a mais
                        # antiga — a atual é a PRIMEIRA, não a última.
                        "competences": competencias,
                        "latestCompetence": competencias[0] if competencias else "",
                    }
                self._set_headers(200)
                self.wfile.write(json_dumps(payload))
                return
            if path == "/api/feedback/pdi":
                user = self._require_auth()
                if not user:
                    return
                query = parse_qs(parsed.query)
                nome = normalize_whitespace(query.get("person", [""])[0])
                with closing(get_connection()) as conn:
                    if not nome and data_scope_for_user(conn, user) == "proprio":
                        nome = meeting_person_identity(user)
                    itens = list_pdi_items(conn, user["company_id"], person_key(nome)) if nome else []
                self._set_headers(200)
                self.wfile.write(json_dumps({"items": itens, "personName": nome}))
                return
            if path == "/api/meetings":
                user = self._require_auth()
                if not user:
                    return
                query = parse_qs(parsed.query)
                with closing(get_connection()) as conn:
                    itens = list_meetings(
                        conn, user["company_id"], user,
                        search=normalize_whitespace(query.get("q", [""])[0]),
                        kind=normalize_upper(query.get("kind", [""])[0]),
                        date_from=normalize_whitespace(query.get("from", [""])[0]),
                        date_to=normalize_whitespace(query.get("to", [""])[0]),
                        only_mine=query.get("mine", ["0"])[0] == "1",
                    )
                    pode_gerir = user_can_manage_meetings(conn, user)
                    escopo = meeting_units_for_user(conn, user) if pode_gerir else {
                        "units": [], "canBeCorporate": False, "defaultUnit": ""
                    }
                    payload = {
                        "meetings": itens,
                        "kinds": MEETING_KINDS,
                        "canManage": pode_gerir,
                        "pendingCount": count_pending_acknowledgements(conn, user["company_id"], user),
                        "myName": meeting_person_identity(user),
                        "people": list_meeting_people(conn, user["company_id"], user) if pode_gerir else [],
                        "units": escopo["units"],
                        "canBeCorporate": escopo["canBeCorporate"],
                        "defaultUnit": escopo["defaultUnit"],
                        "maxAttachmentMb": MEETING_ATTACHMENT_MAX_BYTES // (1024 * 1024),
                    }
                self._set_headers(200)
                self.wfile.write(json_dumps(payload))
                return
            if path.startswith("/api/meetings/attachment/"):
                user = self._require_auth()
                if not user:
                    return
                try:
                    attachment_id = int(path.rsplit("/", 1)[-1])
                except ValueError:
                    self._set_headers(400)
                    self.wfile.write(json_dumps({"error": "Anexo inválido."}))
                    return
                with closing(get_connection()) as conn:
                    row = conn.execute(
                        """
                        SELECT a.file_name, a.stored_name, a.content_type, m.id AS meeting_id, m.status
                        FROM meeting_attachments a
                        JOIN meetings m ON m.id = a.meeting_id
                        WHERE a.id = ? AND m.company_id = ?
                        """,
                        (attachment_id, user["company_id"]),
                    ).fetchone()
                    # Vendedor só baixa anexo de ata publicada em que ele consta.
                    liberado = bool(row)
                    if row and data_scope_for_user(conn, user) == "proprio":
                        chaves = user_person_keys(user)
                        marcadores = ",".join("?" for _ in chaves) or "''"
                        liberado = row["status"] == "PUBLICADA" and bool(conn.execute(
                            "SELECT 1 FROM meeting_participants "
                            f"WHERE meeting_id = ? AND (user_id = ? OR person_key IN ({marcadores}))",
                            (row["meeting_id"], user["id"], *chaves),
                        ).fetchone())
                if not row or not liberado:
                    self._set_headers(404)
                    self.wfile.write(json_dumps({"error": "Anexo não encontrado."}))
                    return
                caminho = meeting_files_dir() / row["stored_name"]
                if not caminho.exists():
                    self._set_headers(404)
                    self.wfile.write(json_dumps({"error": "Arquivo removido do servidor."}))
                    return
                conteudo = caminho.read_bytes()
                # inline=1 abre no navegador (PDF e imagem); sem o parâmetro, baixa.
                # Ler a apostila sem precisar salvar o arquivo importa no celular,
                # onde o vendedor abre a ata durante a reunião.
                query_anexo = parse_qs(parsed.query)
                disposicao = "inline" if query_anexo.get("inline", ["0"])[0] == "1" else "attachment"
                self.send_response(200)
                self.send_header("Content-Type", row["content_type"] or "application/octet-stream")
                self.send_header("Content-Length", str(len(conteudo)))
                self.send_header(
                    "Content-Disposition",
                    f'{disposicao}; filename="{Path(row["file_name"]).name}"',
                )
                self.end_headers()
                self.wfile.write(conteudo)
                return
            if path == "/api/content":
                # Biblioteca de conteúdo. Leitura liberada a todos os autenticados —
                # é material de trabalho do vendedor.
                user = self._require_auth()
                if not user:
                    return
                query = parse_qs(parsed.query)
                category = normalize_whitespace(query.get("category", [None])[0]) or None
                situation = normalize_upper(query.get("situation", [None])[0]) or None
                include_inactive = query.get("all", ["0"])[0] == "1"
                with closing(get_connection()) as conn:
                    can_edit = user_can_manage_users(conn, user)
                    items = list_content_library(
                        conn, user["company_id"], category, situation,
                        only_active=not (include_inactive and can_edit),
                    )
                self._set_headers(200)
                self.wfile.write(json_dumps({
                    "items": items,
                    "categories": CONTENT_CATEGORIES,
                    "situations": CONTENT_SITUATIONS,
                    "canEdit": can_edit,
                }))
                return
            if path == "/api/crm/portfolio-summary":
                # Visão de carteira por vendedor: gerente precisa dela. Não é área
                # administrativa — o recorte por unidade é aplicado abaixo.
                user = self._require_auth()
                if not user:
                    return
                try:
                    query = parse_qs(parsed.query)
                    req_competence = query.get("competence", [None])[0] or None
                    req_unit = query.get("unit", [None])[0] or None
                    req_person_type = normalize_upper(query.get("personType", [None])[0]) or None
                    with closing(get_connection()) as conn:
                        req_unit = crm_unit_filter_for_user(conn, user, req_unit)
                        data = compute_portfolio_summary_by_seller(
                            conn, user["company_id"], user,
                            competence=req_competence, unit_filter=req_unit,
                            person_type_filter=req_person_type,
                        )
                    self._set_headers(200)
                    self.wfile.write(json_dumps(data))
                except Exception as _e:
                    traceback.print_exc()
                    self._set_headers(500)
                    self.wfile.write(json_dumps({"error": str(_e), "sellers": [], "totals": {}}))
                return
            if path == "/api/crm/summary":
                user = self._require_auth()
                if not user:
                    return
                query = parse_qs(parsed.query)
                with closing(get_connection()) as conn:
                    filters = crm_scoped_filters_for_user(conn, user["company_id"], user, build_filters_from_query(query))
                    data = crm_summary_for_user(conn, user["company_id"], user, filters)
                self._set_headers(200)
                self.wfile.write(json_dumps(data))
                return
            if path == "/api/crm/agenda":
                user = self._require_auth()
                if not user:
                    return
                query = parse_qs(parsed.query)
                limit = max(1, min(int(query.get("limit", ["20"])[0]), 50))
                estatisticas: dict[str, Any] = {}
                with closing(get_connection()) as conn:
                    filters = crm_scoped_filters_for_user(conn, user["company_id"], user, build_filters_from_query(query))
                    clients = list_crm_clients(conn, user["company_id"], filters, limit,
                                               exclude_contacted_today=True, stats=estatisticas)
                self._set_headers(200)
                self.wfile.write(
                    json_dumps(
                        {
                            "top5": clients[:5],
                            "extended": clients[5:limit],
                            "total": len(clients),
                            "rotation": estatisticas,
                        }
                    )
                )
                return
            if path == "/api/crm/clients":
                user = self._require_auth()
                if not user:
                    return
                query = parse_qs(parsed.query)
                with closing(get_connection()) as conn:
                    filters = crm_scoped_filters_for_user(conn, user["company_id"], user, build_filters_from_query(query))
                    clients = query_crm_clients_page(
                        conn,
                        user["company_id"],
                        filters,
                        parse_int(query.get("page", [1])[0]) or 1,
                        parse_int(query.get("pageSize", [50])[0]) or 50,
                    )
                print(
                    "[CRM CLIENTS DEBUG]",
                    {
                        "total": clients["total"],
                        "page": clients["page"],
                        "pageSize": clients["pageSize"],
                        "totalPages": clients["totalPages"],
                        "rowsReturned": len(clients["rows"]),
                    },
                )
                self._set_headers(200)
                self.wfile.write(json_dumps(clients))
                return
            if path == "/api/crm/client":
                user = self._require_auth()
                if not user:
                    return
                query = parse_qs(parsed.query)
                client_key = normalize_client_key(query.get("clientKey", [None])[0])
                if not client_key:
                    self._set_headers(400)
                    self.wfile.write(json_dumps({"error": "Informe clientKey"}))
                    return
                fora = query.get("outside", ["0"])[0] in {"1", "true", "sim"}
                with closing(get_connection()) as conn:
                    filters = scoped_filters_for_user(conn, user["company_id"], user, build_filters_from_query(query))
                    data = get_crm_client_360(conn, user["company_id"], filters, client_key,
                                              allow_outside=fora)
                    if data:
                        dono = conn.execute(
                            "SELECT NULLIF(TRIM(internal_seller_name), '') v "
                            "FROM crm_client_profiles WHERE company_id = ? AND client_code = ?",
                            (user["company_id"], client_key),
                        ).fetchone()
                        data["ownerName"] = normalize_whitespace(dono["v"]) if dono else ""
                        data["isOwnClient"] = not client_is_outside_own_portfolio(
                            conn, user["company_id"], user, client_key)
                if not data:
                    self._set_headers(404)
                    self.wfile.write(json_dumps({"error": "Cliente nao encontrado"}))
                    return
                payload = json_dumps(data)
                self._set_headers(200)
                self.wfile.write(payload)
                return
            if path == "/api/crm/client/summary":
                user = self._require_auth()
                if not user:
                    return
                query = parse_qs(parsed.query)
                client_key = normalize_client_key(query.get("clientKey", [None])[0])
                if not client_key:
                    self._set_headers(400)
                    self.wfile.write(json_dumps({"error": "Informe clientKey"}))
                    return
                with closing(get_connection()) as conn:
                    filters = crm_scoped_filters_for_user(conn, user["company_id"], user, build_filters_from_query(query))
                    data = get_crm_client_summary(
                        conn, user["company_id"], filters, client_key,
                        seller_name=user["full_name"] or user["username"],
                        allow_outside=query.get("outside", ["0"])[0] in {"1", "true", "sim"},
                    )
                    if data:
                        data["isOwnClient"] = not client_is_outside_own_portfolio(
                            conn, user["company_id"], user, client_key)
                        # Gestão pode cobrar contato de qualquer cliente da ficha.
                        # A lista de vendedores acompanha para o caso de o cliente
                        # não ter dono definido.
                        is_manager = data_scope_for_user(conn, user) != "proprio"
                        data["canAssignTask"] = is_manager
                        if is_manager:
                            data["assignableSellers"] = sellers_available_for_assignment(
                                conn, user["company_id"], user,
                                (data.get("summary") or {}).get("cityName"),
                            )
                if not data:
                    self._set_headers(404)
                    self.wfile.write(json_dumps({"error": "Cliente nao encontrado"}))
                    return
                payload = json_dumps(data)
                self._set_headers(200)
                self.wfile.write(payload)
                return
            if path == "/api/crm/client/interactions":
                user = self._require_auth()
                if not user:
                    return
                query = parse_qs(parsed.query)
                client_key = normalize_client_key(query.get("clientKey", [None])[0])
                if not client_key:
                    self._set_headers(400)
                    self.wfile.write(json_dumps({"error": "Informe clientKey"}))
                    return
                with closing(get_connection()) as conn:
                    filters = scoped_filters_for_user(conn, user["company_id"], user, build_filters_from_query(query))
                    data = get_crm_client_interactions(
                        conn,
                        user["company_id"],
                        filters,
                        client_key,
                        parse_int(query.get("page", [1])[0]) or 1,
                        parse_int(query.get("pageSize", [20])[0]) or 20,
                    )
                if not data:
                    self._set_headers(404)
                    self.wfile.write(json_dumps({"error": "Cliente nao encontrado"}))
                    return
                self._set_headers(200)
                self.wfile.write(json_dumps(data))
                return
            if path == "/api/crm/client/purchases":
                user = self._require_auth()
                if not user:
                    return
                query = parse_qs(parsed.query)
                client_key = normalize_client_key(query.get("clientKey", [None])[0])
                if not client_key:
                    self._set_headers(400)
                    self.wfile.write(json_dumps({"error": "Informe clientKey"}))
                    return
                with closing(get_connection()) as conn:
                    filters = scoped_filters_for_user(conn, user["company_id"], user, build_filters_from_query(query))
                    data = get_crm_client_purchases(conn, user["company_id"], filters, client_key)
                if data is None:
                    self._set_headers(404)
                    self.wfile.write(json_dumps({"error": "Cliente nao encontrado"}))
                    return
                self._set_headers(200)
                self.wfile.write(json_dumps({"rows": data}))
                return
            if path == "/api/crm/client/items":
                user = self._require_auth()
                if not user:
                    return
                query = parse_qs(parsed.query)
                client_key = normalize_client_key(query.get("clientKey", [None])[0])
                if not client_key:
                    self._set_headers(400)
                    self.wfile.write(json_dumps({"error": "Informe clientKey"}))
                    return
                with closing(get_connection()) as conn:
                    filters = scoped_filters_for_user(conn, user["company_id"], user, build_filters_from_query(query))
                    data = get_crm_client_items(
                        conn,
                        user["company_id"],
                        filters,
                        client_key,
                        parse_int(query.get("page", [1])[0]) or 1,
                        parse_int(query.get("pageSize", [20])[0]) or 20,
                    )
                if not data:
                    self._set_headers(404)
                    self.wfile.write(json_dumps({"error": "Cliente nao encontrado"}))
                    return
                self._set_headers(200)
                self.wfile.write(json_dumps(data))
                return
            if path == "/api/crm/client/tasks":
                user = self._require_auth()
                if not user:
                    return
                query = parse_qs(parsed.query)
                client_key = normalize_client_key(query.get("clientKey", [None])[0])
                if not client_key:
                    self._set_headers(400)
                    self.wfile.write(json_dumps({"error": "Informe clientKey"}))
                    return
                with closing(get_connection()) as conn:
                    filters = scoped_filters_for_user(conn, user["company_id"], user, build_filters_from_query(query))
                    data = get_crm_client_tasks(conn, user["company_id"], filters, client_key)
                if data is None:
                    self._set_headers(404)
                    self.wfile.write(json_dumps({"error": "Cliente nao encontrado"}))
                    return
                self._set_headers(200)
                self.wfile.write(json_dumps({"rows": data}))
                return
            if path == "/api/crm/tasks":
                user = self._require_auth()
                if not user:
                    return
                query = parse_qs(parsed.query)
                with closing(get_connection()) as conn:
                    pode_criar = data_scope_for_user(conn, user) != "proprio"
                    rows = list_crm_tasks(
                        conn, user["company_id"], user,
                        status=normalize_upper(query.get("status", ["ABERTAS"])[0]),
                        seller=normalize_whitespace(query.get("seller", [""])[0]),
                        date_from=normalize_whitespace(query.get("from", [""])[0]),
                        date_to=normalize_whitespace(query.get("to", [""])[0]),
                        origin=normalize_upper(query.get("origin", [""])[0]),
                        search=normalize_whitespace(query.get("q", [""])[0]),
                    )
                    payload = {
                        "rows": rows,
                        "total": len(rows),
                        "counters": crm_task_counters(conn, user["company_id"], user),
                        "origins": TASK_ORIGINS,
                        "priorities": TASK_PRIORITIES,
                        "statusFilters": TASK_STATUS_FILTERS,
                        "canCreate": pode_criar,
                        "people": task_assignable_people(conn, user["company_id"], user) if pode_criar else [],
                        "sellers": task_visible_sellers(conn, user["company_id"], user) or [],
                        "myName": meeting_person_identity(user),
                    }
                self._set_headers(200)
                self.wfile.write(json_dumps(payload))
                return
            if path == "/api/crm/agenda/actions":
                user = self._require_auth()
                if not user:
                    return
                if self.command != "POST":
                    self._set_headers(405)
                    self.wfile.write(json_dumps({"error": "Metodo nao permitido"}))
                    return
            if path == "/api/admin/all":
                # Payload usado por várias telas (pessoas, metas, mapeamentos). Perfis sem
                # gestão de usuários recebem a versão sem dados sensíveis.
                user = self._require_auth()
                if not user:
                    return
                with closing(get_connection()) as conn:
                    if data_scope_for_user(conn, user) == "proprio":
                        self._set_headers(403)
                        self.wfile.write(json_dumps({"error": "Perfil sem acesso a esses dados."}))
                        return
                    data = list_admin_data(conn, user["company_id"])
                    data = filter_admin_data_for_user(conn, user["company_id"], user, data)
                    if not user_can_manage_users(conn, user):
                        for sensitive in ("users", "profiles", "audit"):
                            data[sensitive] = []
                self._set_headers(200)
                self.wfile.write(json_dumps(data))
                return
            if path == "/api/crm/contacts":
                user = self._require_auth()
                if not user:
                    return
                query = parse_qs(parsed.query)
                filtros = {k: query.get(k, [""])[0] for k in
                           ("start", "end", "seller", "type", "result", "initiative", "search", "limit")}
                with closing(get_connection()) as conn:
                    dados = contact_history(conn, user["company_id"], user, filtros)
                self._set_headers(200)
                self.wfile.write(json_dumps(dados))
                return
            if path == "/api/crm/coverages":
                user = self._require_auth()
                if not user:
                    return
                with closing(get_connection()) as conn:
                    escopo = data_scope_for_user(conn, user)
                    if escopo == "proprio":
                        itens = active_coverages_for_seller(
                            conn, user["company_id"], seller_identity_for_user(user))
                        payload = {"mine": itens, "canManage": False}
                    else:
                        payload = {"all": list_coverages(conn, user["company_id"], user),
                                   "canManage": True,
                                   "sellers": sellers_available_for_assignment(
                                       conn, user["company_id"], user)}
                self._set_headers(200)
                self.wfile.write(json_dumps(payload))
                return
            if path == "/api/crm/client/support":
                user = self._require_auth()
                if not user:
                    return
                with closing(get_connection()) as conn:
                    achado = support_client_view(
                        conn, user["company_id"], user,
                        parse_qs(parsed.query).get("code", [""])[0])
                if not achado:
                    self._set_headers(404)
                    self.wfile.write(json_dumps(
                        {"error": "Nenhum cliente com este código no cadastro."}))
                    return
                self._set_headers(200)
                self.wfile.write(json_dumps({"client": achado}))
                return
            if path == "/api/crm/clients/by-code":
                user = self._require_auth()
                if not user:
                    return
                with closing(get_connection()) as conn:
                    if data_scope_for_user(conn, user) == "proprio":
                        self._set_headers(403)
                        self.wfile.write(json_dumps({"error": "Consulta da gestão."}))
                        return
                consulta = parse_qs(parsed.query)
                achado = find_client_by_code(
                    conn, user["company_id"], consulta.get("code", [""])[0])
                candidatos = ([] if achado else
                              search_clients_by_name(conn, user["company_id"],
                                                     consulta.get("name", [""])[0]))
                self._set_headers(200)
                self.wfile.write(json_dumps({"client": achado,
                                             "candidates": [dict(c) for c in candidatos]}))
                return
            if path == "/api/admin/territories":
                user = self._require_auth()
                if not user:
                    return
                with closing(get_connection()) as conn:
                    if data_scope_for_user(conn, user) == "proprio":
                        self._set_headers(403)
                        self.wfile.write(json_dumps({"error": "Perfil sem acesso."}))
                        return
                    cidade = normalize_whitespace(parse_qs(parsed.query).get("city", [""])[0])
                    itens = list_territory_mappings(conn, user["company_id"], cidade)
                    cobertura = territory_coverage_report(conn, user["company_id"])
                self._set_headers(200)
                self.wfile.write(json_dumps({
                    "territories": itens,
                    "coverage": cobertura,
                    "units": CANONICAL_UNITS,
                    "sharedLabel": TERRITORIO_COMPARTILHADO,
                    "cityWide": TERRITORIO_CIDADE_INTEIRA,
                    "defaultValidFrom": territory_default_valid_from(),
                }))
                return
            if path.startswith("/api/templates/"):
                user = self._require_auth()
                if not user or not self._require_admin_area(user):
                    return
                kind = path.split("/")[-1]
                content = csv_template(kind)
                self._set_headers(
                    200,
                    "text/csv; charset=utf-8",
                    {"Content-Disposition": f'attachment; filename="{kind}_template.csv"'},
                )
                self.wfile.write(content)
                return
            if path == "/api/export.xlsx":
                user = self._require_auth()
                if not user:
                    return
                query = parse_qs(parsed.query)
                with closing(get_connection()) as conn:
                    filters = scoped_filters_for_user(conn, user["company_id"], user, build_filters_from_query(query))
                    data = get_dashboard_data(conn, user["company_id"], filters)
                content = export_dashboard_xlsx(data)
                self._set_headers(
                    200,
                    "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    {"Content-Disposition": 'attachment; filename="dashboard_passini.xlsx"'},
                )
                self.wfile.write(content)
                return
            if path == "/api/export.pdf":
                user = self._require_auth()
                if not user:
                    return
                query = parse_qs(parsed.query)
                with closing(get_connection()) as conn:
                    filters = scoped_filters_for_user(conn, user["company_id"], user, build_filters_from_query(query))
                    data = get_dashboard_data(conn, user["company_id"], filters)
                content = export_dashboard_pdf(data)
                self._set_headers(200, "application/pdf", {"Content-Disposition": 'attachment; filename="dashboard_passini.pdf"'})
                self.wfile.write(content)
                return
            if path == "/api/backup/database":
                user = self._require_auth()
                if not user or not self._require_admin_area(user):
                    return
                timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
                self._set_headers(
                    200,
                    "application/octet-stream",
                    {"Content-Disposition": f'attachment; filename="passini_dashboard_backup_{timestamp}.db"'},
                )
                self.wfile.write(DB_PATH.read_bytes())
                return
            self._set_headers(404)
            self.wfile.write(json_dumps({"error": "Rota não encontrada"}))
        except Exception as exc:
            traceback.print_exc()
            self._set_headers(500)
            self.wfile.write(json_dumps({"error": f"Erro interno: {exc}"}))

    def do_POST(self) -> None:
        try:
            parsed = urlparse(self.path)
            path = parsed.path
            if path == "/api/login":
                payload = self._read_json()
                username = payload.get("username", "")
                password = payload.get("password", "")
                with closing(get_connection()) as conn:
                    user = conn.execute("SELECT * FROM users WHERE username = ? AND is_active = 1", (username,)).fetchone()
                    if not user or not verify_password(password, user["password_hash"], user["password_salt"]):
                        self._set_headers(401)
                        self.wfile.write(json_dumps({"error": "Usuário ou senha inválidos"}))
                        return
                    user = dict(user)
                    session_id = secrets.token_hex(24)
                    expires_at = (datetime.now() + timedelta(hours=SESSION_TTL_HOURS)).isoformat(timespec="seconds")
                    conn.execute("INSERT INTO sessions (id, user_id, created_at, expires_at) VALUES (?, ?, ?, ?)", (session_id, user["id"], now_iso(), expires_at))
                    conn.commit()
                    profile = get_access_profile_for_user(conn, user)
                headers = {"Set-Cookie": f"{SESSION_COOKIE}={session_id}; HttpOnly; Path=/; SameSite=Lax"}
                self._set_headers(200, extra_headers=headers)
                self.wfile.write(
                    json_dumps(
                        {
                            "ok": True,
                            "user": {
                                "username": user["username"],
                                "fullName": user["full_name"],
                                "linkedPersonName": user["linked_person_name"],
                                "linkedUnits": linked_units_for_user(user),
                                "role": user["role"],
                                "profileName": profile["name"] if profile else user["role"],
                                "modules": profile["modules"] if profile else [],
                                "dataScope": profile["dataScope"] if profile else "todos",
                                "canManageUsers": profile["canManageUsers"] if profile else False,
                            },
                        }
                    )
                )
                return
            if path == "/api/logout":
                user = self._current_user()
                cookie = SimpleCookie(self.headers.get("Cookie"))
                session_cookie = cookie.get(SESSION_COOKIE)
                if session_cookie:
                    with closing(get_connection()) as conn:
                        conn.execute("DELETE FROM sessions WHERE id = ?", (session_cookie.value,))
                        conn.commit()
                headers = {"Set-Cookie": f"{SESSION_COOKIE}=deleted; HttpOnly; Path=/; Max-Age=0; SameSite=Lax"}
                self._set_headers(200, extra_headers=headers)
                self.wfile.write(json_dumps({"ok": True}))
                return
            if path == "/api/crm/interactions":
                user = self._require_auth()
                if not user:
                    return
                payload = self._read_json()
                try:
                    with closing(get_connection()) as conn:
                        result = create_crm_interaction(conn, user["company_id"], user, payload)
                    # Derruba o cache para o cliente sair da fila imediatamente,
                    # tanto para o vendedor quanto para a visão do gerente.
                    invalidate_crm_cache(user["company_id"])
                except ValueError as exc:
                    self._set_headers(400)
                    self.wfile.write(json_dumps({"error": str(exc)}))
                    return
                self._set_headers(200)
                self.wfile.write(json_dumps({"ok": True, "result": result}))
                return
            if path == "/api/crm/client/contact":
                user = self._require_auth()
                if not user:
                    return
                payload = self._read_json()
                try:
                    with closing(get_connection()) as conn:
                        result = save_crm_client_contact(conn, user["company_id"], user, payload)
                except ValueError as exc:
                    self._set_headers(400)
                    self.wfile.write(json_dumps({"error": str(exc)}))
                    return
                self._set_headers(200)
                self.wfile.write(json_dumps({"ok": True, "result": result}))
                return
            if path == "/api/crm/agenda/actions":
                user = self._require_auth()
                if not user:
                    return
                payload = self._read_json()
                try:
                    with closing(get_connection()) as conn:
                        result = save_crm_agenda_action(conn, user["company_id"], user, payload)
                except ValueError as exc:
                    self._set_headers(400)
                    self.wfile.write(json_dumps({"error": str(exc)}))
                    return
                self._set_headers(200)
                self.wfile.write(json_dumps({"ok": True, "result": result}))
                return
            if path == "/api/crm/tasks/assign":
                # Gestor cria uma tarefa atribuída ao vendedor responsável pelo cliente,
                # transformando a cobrança verbal em pendência rastreável.
                user = self._require_auth()
                if not user:
                    return
                payload = self._read_json()
                try:
                    with closing(get_connection()) as conn:
                        # Vendedor pode agendar contato para SI MESMO; a gestão pode
                        # atribuir a qualquer vendedor da unidade. O nome é forçado
                        # no caso do vendedor para ele não criar tarefa para outro.
                        is_seller = data_scope_for_user(conn, user) == "proprio"
                        if is_seller:
                            payload["sellerName"] = seller_identity_for_user(user)
                        client_key = normalize_whitespace(payload.get("clientKey"))
                        client_name = normalize_whitespace(payload.get("clientName"))
                        seller_name = normalize_whitespace(payload.get("sellerName"))
                        title = normalize_whitespace(payload.get("title")) or "Contatar cliente"
                        description = normalize_whitespace(payload.get("description"))
                        due_at = normalize_whitespace(payload.get("dueAt")) or date.today().isoformat()
                        # O vendedor escolhido precisa estar nas unidades do gestor —
                        # sem isso, um gerente poderia criar tarefa para outra equipe.
                        if seller_name and not is_seller:
                            permitidos = {
                                normalize_upper(s["sellerName"])
                                for s in sellers_available_for_assignment(conn, user["company_id"], user)
                            }
                            if permitidos and normalize_upper(seller_name) not in permitidos:
                                self._set_headers(403)
                                self.wfile.write(json_dumps(
                                    {"error": f"{seller_name} não está entre os vendedores que você gerencia."}
                                ))
                                return
                        if not client_key or not seller_name:
                            self._set_headers(400)
                            self.wfile.write(json_dumps({"error": "Cliente e vendedor são obrigatórios."}))
                            return
                        # Evita duplicar cobrança para o mesmo cliente/vendedor
                        existing = conn.execute(
                            "SELECT id FROM crm_tasks WHERE company_id = ? AND client_key = ? "
                            "AND seller_name = ? AND status IN ('ABERTA','ATRASADA')",
                            (user["company_id"], client_key, seller_name),
                        ).fetchone()
                        if existing:
                            self._set_headers(200)
                            self.wfile.write(json_dumps({
                                "ok": True, "duplicated": True,
                                "message": ("Você já tem uma tarefa aberta para este cliente."
                                            if is_seller
                                            else f"{seller_name} já tem uma tarefa aberta para este cliente."),
                            }))
                            return
                        conn.execute(
                            """
                            INSERT INTO crm_tasks
                                (company_id, client_key, client_name, seller_name, title, description,
                                 due_at, status, origin, created_by_name, created_by_user_id, created_at)
                            VALUES (?, ?, ?, ?, ?, ?, ?, 'ABERTA', ?, ?, ?, ?)
                            """,
                            (user["company_id"], client_key, client_name, seller_name,
                             title, description, due_at,
                             "FOLLOWUP" if is_seller else "COBRANCA",
                             meeting_person_identity(user), user["id"], now_iso()),
                        )
                        audit_log(conn, user["company_id"], user["id"], "criar", "crm_tasks", client_key,
                                  {"seller": seller_name,
                                   "origem": "agendamento_vendedor" if is_seller else "cobranca_gestor"})
                        conn.commit()
                    self._set_headers(200)
                    self.wfile.write(json_dumps({
                        "ok": True,
                        "message": ("Contato agendado. A tarefa está em Tarefas."
                                    if is_seller else f"Tarefa criada para {seller_name}."),
                    }))
                except Exception as exc:
                    traceback.print_exc()
                    self._set_headers(400)
                    self.wfile.write(json_dumps({"error": str(exc)}))
                return
            if path == "/api/crm/tasks/complete":
                user = self._require_auth()
                if not user:
                    return
                payload = self._read_json()
                try:
                    with closing(get_connection()) as conn:
                        complete_crm_task(conn, user["company_id"], user, int(payload.get("taskId") or 0))
                except (ValueError, TypeError) as exc:
                    self._set_headers(400)
                    self.wfile.write(json_dumps({"error": str(exc)}))
                    return
                self._set_headers(200)
                self.wfile.write(json_dumps({"ok": True}))
                return
            if path == "/api/crm/tasks/reschedule":
                user = self._require_auth()
                if not user:
                    return
                payload = self._read_json()
                try:
                    with closing(get_connection()) as conn:
                        reschedule_crm_task(
                            conn,
                            user["company_id"],
                            user,
                            int(payload.get("taskId") or 0),
                            normalize_whitespace(payload.get("dueAt") or "").replace("T", " "),
                        )
                except (ValueError, TypeError) as exc:
                    self._set_headers(400)
                    self.wfile.write(json_dumps({"error": str(exc)}))
                    return
                self._set_headers(200)
                self.wfile.write(json_dumps({"ok": True}))
                return
            if path in ("/api/admin/goals/seller", "/api/admin/goals/unit"):
                user = self._require_auth()
                if not user or not self._require_admin_area(user):
                    return
                body = json.loads(self.rfile.read(int(self.headers.get("Content-Length", 0))) or b"{}")
                competence = normalize_whitespace(body.get("competence", ""))
                if not competence:
                    self._set_headers(400)
                    self.wfile.write(json_dumps({"error": "Competência obrigatória"}))
                    return
                try:
                    datetime.strptime(competence[:7], "%Y-%m")
                    competence = competence[:7]
                except ValueError:
                    self._set_headers(400)
                    self.wfile.write(json_dumps({"error": f"Competência inválida: {competence}. Use YYYY-MM."}))
                    return
                with closing(get_connection()) as conn:
                    if path == "/api/admin/goals/seller":
                        seller_name = normalize_whitespace(body.get("seller_name", ""))
                        base_unit = normalize_unit(body.get("base_unit", ""))
                        revenue_goal = float(body.get("revenue_goal") or 0)
                        if not seller_name:
                            self._set_headers(400)
                            self.wfile.write(json_dumps({"error": "Vendedor obrigatório"}))
                            return
                        if not self._require_unit_allowed(user, base_unit):
                            return
                        conn.execute(
                            """
                            INSERT INTO goals_seller
                                (company_id, competence, seller_name, base_unit, revenue_goal, returns_goal, created_at)
                            VALUES (?, ?, ?, ?, ?, 0, ?)
                            ON CONFLICT(company_id, competence, seller_name)
                            DO UPDATE SET
                                base_unit = excluded.base_unit,
                                revenue_goal = excluded.revenue_goal,
                                created_at = excluded.created_at
                            """,
                            (user["company_id"], competence, seller_name, base_unit, revenue_goal, now_iso()),
                        )
                        conn.commit()
                    else:
                        unit_name = normalize_unit(body.get("unit_name", ""))
                        revenue_goal = float(body.get("revenue_goal") or 0)
                        if not unit_name:
                            self._set_headers(400)
                            self.wfile.write(json_dumps({"error": "Unidade obrigatória"}))
                            return
                        if not self._require_unit_allowed(user, unit_name):
                            return
                        conn.execute(
                            """
                            INSERT INTO goals_unit
                                (company_id, competence, unit_name, revenue_goal, returns_goal, created_at)
                            VALUES (?, ?, ?, ?, 0, ?)
                            ON CONFLICT(company_id, competence, unit_name)
                            DO UPDATE SET
                                revenue_goal = excluded.revenue_goal,
                                created_at = excluded.created_at
                            """,
                            (user["company_id"], competence, unit_name, revenue_goal, now_iso()),
                        )
                        conn.commit()
                invalidate_dashboard_cache(user["company_id"])
                self._set_headers(200)
                self.wfile.write(json_dumps({"ok": True}))
                return
            if path == "/api/admin/goals/seller/delete":
                user = self._require_auth()
                if not user or not self._require_admin_area(user):
                    return
                body = json.loads(self.rfile.read(int(self.headers.get("Content-Length", 0))) or b"{}")
                competence = normalize_whitespace(body.get("competence", ""))
                seller_name = normalize_whitespace(body.get("seller_name", ""))
                with closing(get_connection()) as conn:
                    conn.execute(
                        "DELETE FROM goals_seller WHERE company_id = ? AND competence = ? AND seller_name = ?",
                        (user["company_id"], competence, seller_name),
                    )
                    conn.commit()
                invalidate_dashboard_cache(user["company_id"])
                self._set_headers(200)
                self.wfile.write(json_dumps({"ok": True}))
                return
            if path == "/api/admin/goals/unit/delete":
                user = self._require_auth()
                if not user or not self._require_admin_area(user):
                    return
                body = json.loads(self.rfile.read(int(self.headers.get("Content-Length", 0))) or b"{}")
                competence = normalize_whitespace(body.get("competence", ""))
                unit_name = normalize_unit(body.get("unit_name", ""))
                with closing(get_connection()) as conn:
                    conn.execute(
                        "DELETE FROM goals_unit WHERE company_id = ? AND competence = ? AND unit_name = ?",
                        (user["company_id"], competence, unit_name),
                    )
                    conn.commit()
                invalidate_dashboard_cache(user["company_id"])
                self._set_headers(200)
                self.wfile.write(json_dumps({"ok": True}))
                return
            if path == "/api/auto-import/run":
                user = self._require_auth()
                if not user or not self._require_admin_area(user):
                    return
                if _AUTO_IMPORT_RUNNING.is_set():
                    self._set_headers(409)
                    self.wfile.write(json_dumps({"error": "Uma importação já está em andamento. Aguarde."}))
                    return
                try:
                    _AUTO_IMPORT_WARNINGS.clear()
                    auto_import_tick()
                    self._set_headers(200)
                    self.wfile.write(json_dumps({
                        "ok": True,
                        "message": "Verificação de importação executada.",
                        "warnings": list(_AUTO_IMPORT_WARNINGS),
                    }))
                except Exception as exc:
                    traceback.print_exc()
                    self._set_headers(500)
                    self.wfile.write(json_dumps({"error": str(exc)}))
                return
            if path == "/api/admin/issues/resolve":
                user = self._require_auth()
                if not user or not self._require_admin_area(user):
                    return
                payload = self._read_json()
                try:
                    with closing(get_connection()) as conn:
                        result = resolve_import_issue(conn, user["company_id"], user["id"], payload)
                except ValueError as exc:
                    self._set_headers(400)
                    self.wfile.write(json_dumps({"error": str(exc)}))
                    return
                self._set_headers(200)
                self.wfile.write(json_dumps(result))
                return
            if path == "/api/admin/vacation":
                user = self._require_auth()
                if not user or not self._require_admin_area(user):
                    return
                payload = self._read_json()
                person_name = normalize_whitespace(payload.get("person_name") or "")
                start_date = normalize_whitespace(payload.get("start_date") or "")
                end_date = normalize_whitespace(payload.get("end_date") or "")
                notes = normalize_whitespace(payload.get("notes") or "") or None
                if not person_name or not start_date or not end_date:
                    self._set_headers(400)
                    self.wfile.write(json_dumps({"error": "Nome, data inicial e data final são obrigatórios"}))
                    return
                with closing(get_connection()) as conn:
                    # Verificar sobreposição de datas com outras entradas do mesmo colaborador
                    overlap = conn.execute(
                        """SELECT id, start_date, end_date FROM vacations
                           WHERE company_id = ? AND person_name = ?
                             AND date(end_date) >= date(?) AND date(start_date) <= date(?)""",
                        (user["company_id"], person_name, start_date, end_date),
                    ).fetchall()
                    if overlap:
                        overlap_info = [f"{r['start_date']} a {r['end_date']}" for r in overlap]
                        self._set_headers(409)
                        self.wfile.write(json_dumps({"error": f"Já existe férias para {person_name} no período sobreposto: {'; '.join(overlap_info)}. Edite ou exclua a entrada existente."}))
                        return
                    conn.execute(
                        "INSERT INTO vacations (company_id, person_name, start_date, end_date, notes, created_at) VALUES (?, ?, ?, ?, ?, ?)",
                        (user["company_id"], person_name, start_date, end_date, notes, now_iso()),
                    )
                    audit_log(conn, user["company_id"], user["id"], "criar", "vacations", "", {"person_name": person_name, "start_date": start_date, "end_date": end_date})
                    conn.commit()
                invalidate_calendar_cache(user["company_id"])
                invalidate_dashboard_cache(user["company_id"])
                self._set_headers(200)
                self.wfile.write(json_dumps({"ok": True}))
                return
            if path == "/api/admin/vacation/update":
                user = self._require_auth()
                if not user or not self._require_admin_area(user):
                    return
                payload = self._read_json()
                vac_id = int(payload.get("id") or 0)
                person_name = normalize_whitespace(payload.get("person_name") or "")
                start_date = normalize_whitespace(payload.get("start_date") or "")
                end_date = normalize_whitespace(payload.get("end_date") or "")
                notes = normalize_whitespace(payload.get("notes") or "") or None
                if not vac_id or not person_name or not start_date or not end_date:
                    self._set_headers(400)
                    self.wfile.write(json_dumps({"error": "ID, nome, data inicial e data final são obrigatórios"}))
                    return
                with closing(get_connection()) as conn:
                    row = conn.execute("SELECT id FROM vacations WHERE id = ? AND company_id = ?", (vac_id, user["company_id"])).fetchone()
                    if not row:
                        self._set_headers(404)
                        self.wfile.write(json_dumps({"error": "Férias não encontradas"}))
                        return
                    conn.execute(
                        "UPDATE vacations SET person_name = ?, start_date = ?, end_date = ?, notes = ? WHERE id = ? AND company_id = ?",
                        (person_name, start_date, end_date, notes, vac_id, user["company_id"]),
                    )
                    audit_log(conn, user["company_id"], user["id"], "editar", "vacations", str(vac_id), {"person_name": person_name, "start_date": start_date, "end_date": end_date})
                    conn.commit()
                invalidate_calendar_cache(user["company_id"])
                invalidate_dashboard_cache(user["company_id"])
                self._set_headers(200)
                self.wfile.write(json_dumps({"ok": True}))
                return
            if path == "/api/admin/vacation/delete":
                user = self._require_auth()
                if not user or not self._require_admin_area(user):
                    return
                payload = self._read_json()
                vac_id = int(payload.get("id") or 0)
                if not vac_id:
                    self._set_headers(400)
                    self.wfile.write(json_dumps({"error": "ID obrigatório"}))
                    return
                with closing(get_connection()) as conn:
                    row = conn.execute("SELECT id FROM vacations WHERE id = ? AND company_id = ?", (vac_id, user["company_id"])).fetchone()
                    if not row:
                        self._set_headers(404)
                        self.wfile.write(json_dumps({"error": "Férias não encontradas"}))
                        return
                    conn.execute("DELETE FROM vacations WHERE id = ? AND company_id = ?", (vac_id, user["company_id"]))
                    audit_log(conn, user["company_id"], user["id"], "excluir", "vacations", str(vac_id), {})
                    conn.commit()
                invalidate_calendar_cache(user["company_id"])
                invalidate_dashboard_cache(user["company_id"])
                self._set_headers(200)
                self.wfile.write(json_dumps({"ok": True}))
                return
            if path == "/api/import/preview":
                user = self._require_auth()
                if not user:
                    return
                if not self._require_admin_area(user):
                    return
                files_payload, fields = self._parse_multipart()
                import_scope = normalize_import_scope(fields.get("importScope"))
                preview = preview_import_package(files_payload, import_scope)
                self._set_headers(200)
                self.wfile.write(json_dumps(preview))
                return
            if path == "/api/import/package":
                user = self._require_auth()
                if not user:
                    return
                if not self._require_admin_area(user):
                    return
                files_payload, fields = self._parse_multipart()
                import_scope = normalize_import_scope(fields.get("importScope"))
                preview = preview_import_package(files_payload, import_scope)
                if not preview.get("isValid"):
                    unsupported = preview.get("unsupportedFiles", [])
                    if unsupported:
                        error_message = "Formato invalido para importacao operacional. Use CSV. Arquivos: " + ", ".join(
                            f["fileName"] for f in unsupported
                        )
                    else:
                        error_message = "Importacao invalida: " + str(preview.get("errors", ["Erro desconhecido"])[0] if preview.get("errors") else "Erro desconhecido")
                    self._set_headers(400)
                    self.wfile.write(json_dumps({"error": error_message}))
                    return
                competence_raw = normalize_whitespace(fields.get("competence") or preview.get("detectedCompetence") or "")
                if not competence_raw:
                    self._set_headers(400)
                    self.wfile.write(json_dumps({"error": "Competencia nao informada"}))
                    return
                # Valida formato YYYY-MM
                try:
                    datetime.strptime(competence_raw[:7], "%Y-%m")
                    competence = competence_raw[:7]
                except ValueError:
                    self._set_headers(400)
                    self.wfile.write(json_dumps({"error": f"Competencia invalida: {competence_raw}. Use formato YYYY-MM."}))
                    return
                import_action = fields.get("importAction") or ""
                try:
                    with closing(get_connection()) as conn:
                        result = import_package(conn, user["company_id"], user["id"], competence, import_action, import_scope, preview, files_payload)
                except Exception as exc:
                    self._set_headers(500)
                    self.wfile.write(json_dumps({"error": str(exc)}))
                    return
                invalidate_crm_cache(user["company_id"])
                self._set_headers(200)
                self.wfile.write(json_dumps(result))
                return
            if path == "/api/admin/import-file/clientes":
                user = self._require_auth()
                if not user:
                    return
                if not self._require_admin_area(user):
                    return
                files_payload, _ = self._parse_multipart()
                if not files_payload:
                    self._set_headers(400)
                    self.wfile.write(json_dumps({"error": "Nenhum arquivo enviado"}))
                    return
            if path == "/api/kpi-thresholds/save":
                user = self._require_auth()
                if not user or not self._require_user_management(user):
                    return
                payload = self._read_json()
                try:
                    metric_id = normalize_whitespace(payload.get("metricId"))
                    if metric_id not in KPI_METRIC_BY_ID:
                        raise ValueError("Indicador desconhecido.")
                    good_at = float(payload.get("goodAt"))
                    warn_at = float(payload.get("warnAt"))
                    direction = KPI_METRIC_BY_ID[metric_id]["direction"]
                    # Coerência: em "maior é melhor" o limite verde fica ACIMA do amarelo;
                    # em "menor é melhor", abaixo. Sem isso o farol nunca acende verde.
                    if direction == "higher" and good_at <= warn_at:
                        raise ValueError("Para este indicador, o limite de 'No ritmo' deve ser maior que o de 'Atenção'.")
                    if direction == "lower" and good_at >= warn_at:
                        raise ValueError("Para este indicador, o limite de 'No ritmo' deve ser menor que o de 'Atenção'.")
                    with closing(get_connection()) as conn:
                        conn.execute(
                            """
                            INSERT INTO kpi_thresholds (company_id, metric_id, good_at, warn_at, is_active, created_at)
                            VALUES (?, ?, ?, ?, ?, ?)
                            ON CONFLICT(company_id, metric_id) DO UPDATE SET
                                good_at = excluded.good_at,
                                warn_at = excluded.warn_at,
                                is_active = excluded.is_active,
                                updated_at = ?
                            """,
                            (user["company_id"], metric_id, good_at, warn_at,
                             1 if payload.get("isActive", True) else 0, now_iso(), now_iso()),
                        )
                        audit_log(conn, user["company_id"], user["id"], "editar", "kpi_thresholds",
                                  metric_id, {"goodAt": good_at, "warnAt": warn_at})
                        conn.commit()
                        invalidate_dashboard_cache(user["company_id"])
                    self._set_headers(200)
                    self.wfile.write(json_dumps({"message": "Limites atualizados."}))
                except ValueError as exc:
                    self._set_headers(400)
                    self.wfile.write(json_dumps({"error": str(exc)}))
                except Exception as exc:
                    traceback.print_exc()
                    self._set_headers(400)
                    self.wfile.write(json_dumps({"error": str(exc)}))
                return
            if path in ("/api/help/ask", "/api/help/tour", "/api/help/question/answer",
                        "/api/help/article/save", "/api/help/article/delete",
                        "/api/help/tip/save", "/api/help/tip/delete"):
                user = self._require_auth()
                if not user:
                    return
                payload = self._read_json()
                try:
                    with closing(get_connection()) as conn:
                        if path == "/api/help/ask":
                            papel = help_role_for_user(conn, user)
                            pergunta = normalize_whitespace(payload.get("question"))
                            achados = search_help(conn, user["company_id"], papel, pergunta)
                            registrada = None
                            if not achados:
                                registrada = register_help_question(conn, user["company_id"], user, pergunta)
                            resultado = {
                                "results": achados,
                                "registered": bool(registrada),
                                "question": pergunta,
                            }
                        elif path == "/api/help/tour":
                            mark_tour_seen(conn, user,
                                           normalize_whitespace(payload.get("tourKey")),
                                           bool(payload.get("skipped")))
                            resultado = {}
                        elif path == "/api/help/question/answer":
                            resultado = answer_help_question(conn, user["company_id"], user, payload)
                        elif path == "/api/help/article/save":
                            resultado = save_help_article(conn, user["company_id"], user, payload)
                        elif path == "/api/help/article/delete":
                            delete_help_article(conn, user["company_id"], user,
                                                int(payload.get("articleId") or 0))
                            resultado = {}
                        elif path == "/api/help/tip/save":
                            resultado = save_assistant_tip(conn, user["company_id"], user, payload)
                        else:
                            delete_assistant_tip(conn, user["company_id"], user,
                                                 int(payload.get("tipId") or 0))
                            resultado = {}
                except PermissionError as exc:
                    self._set_headers(403)
                    self.wfile.write(json_dumps({"error": str(exc)}))
                    return
                except ValueError as exc:
                    self._set_headers(400)
                    self.wfile.write(json_dumps({"error": str(exc)}))
                    return
                self._set_headers(200)
                self.wfile.write(json_dumps({"ok": True, **resultado}))
                return
            if path in ("/api/prospects/save", "/api/prospects/delete", "/api/prospects/lost",
                        "/api/prospects/link", "/api/prospects/reconcile",
                        "/api/prospects/phase", "/api/prospects/activity-goal"):
                user = self._require_auth()
                if not user:
                    return
                payload = self._read_json()
                try:
                    with closing(get_connection()) as conn:
                        if path == "/api/prospects/save":
                            resultado = save_prospect(conn, user["company_id"], user, payload)
                        elif path == "/api/prospects/delete":
                            delete_prospect(conn, user["company_id"], user, int(payload.get("prospectId") or 0))
                            resultado = {}
                        elif path == "/api/prospects/lost":
                            mark_prospect_lost(conn, user["company_id"], user,
                                               int(payload.get("prospectId") or 0),
                                               payload.get("reason") or "")
                            resultado = {}
                        elif path == "/api/prospects/link":
                            resultado = link_prospect_to_client(
                                conn, user["company_id"], int(payload.get("prospectId") or 0),
                                normalize_whitespace(payload.get("clientCode")))
                        elif path == "/api/prospects/reconcile":
                            if data_scope_for_user(conn, user) == "proprio":
                                raise PermissionError("Apenas gestão roda a reconciliação.")
                            vinculados = reconcile_prospects(conn, user["company_id"])
                            primeiras = refresh_prospect_first_purchases(conn, user["company_id"])
                            resultado = {"linked": vinculados, "firstPurchases": primeiras}
                        elif path == "/api/prospects/phase":
                            resultado = save_unit_phase(conn, user["company_id"], user, payload)
                        else:
                            resultado = save_activity_goal(conn, user["company_id"], user, payload)
                except PermissionError as exc:
                    self._set_headers(403)
                    self.wfile.write(json_dumps({"error": str(exc)}))
                    return
                except ValueError as exc:
                    self._set_headers(400)
                    self.wfile.write(json_dumps({"error": str(exc)}))
                    return
                self._set_headers(200)
                self.wfile.write(json_dumps({"ok": True, **resultado}))
                return
            if path in ("/api/crm/tasks/create", "/api/crm/tasks/delete"):
                user = self._require_auth()
                if not user:
                    return
                payload = self._read_json()
                try:
                    with closing(get_connection()) as conn:
                        if path == "/api/crm/tasks/create":
                            resultado = create_crm_tasks(conn, user["company_id"], user, payload)
                        else:
                            delete_crm_task(conn, user["company_id"], user, int(payload.get("taskId") or 0))
                            resultado = {}
                except PermissionError as exc:
                    self._set_headers(403)
                    self.wfile.write(json_dumps({"error": str(exc)}))
                    return
                except ValueError as exc:
                    self._set_headers(400)
                    self.wfile.write(json_dumps({"error": str(exc)}))
                    return
                self._set_headers(200)
                self.wfile.write(json_dumps({"ok": True, **resultado}))
                return
            if path in ("/api/visits/save", "/api/visits/delete", "/api/visits/request",
                        "/api/visits/request/resolve", "/api/visits/client"):
                user = self._require_auth()
                if not user:
                    return
                payload = self._read_json()
                try:
                    with closing(get_connection()) as conn:
                        if path == "/api/visits/save":
                            resultado = save_visit(conn, user["company_id"], user, payload)
                        elif path == "/api/visits/delete":
                            delete_visit(conn, user["company_id"], user, int(payload.get("visitId") or 0))
                            resultado = {}
                        elif path == "/api/visits/request":
                            resultado = create_visit_request(conn, user["company_id"], user, payload)
                        elif path == "/api/visits/request/resolve":
                            resultado = resolve_visit_request(
                                conn, user["company_id"], user,
                                int(payload.get("requestId") or 0),
                                bool(payload.get("accept")), payload.get("note") or "")
                        else:
                            resultado = client_contact_effect(
                                conn, user["company_id"], normalize_whitespace(payload.get("clientKey")))
                except PermissionError as exc:
                    self._set_headers(403)
                    self.wfile.write(json_dumps({"error": str(exc)}))
                    return
                except ValueError as exc:
                    self._set_headers(400)
                    self.wfile.write(json_dumps({"error": str(exc)}))
                    return
                self._set_headers(200)
                self.wfile.write(json_dumps({"ok": True, **resultado}))
                return
            if path in ("/api/feedback/save", "/api/feedback/publish", "/api/feedback/detail",
                        "/api/feedback/acknowledge", "/api/feedback/delete",
                        "/api/feedback/pdi/save", "/api/feedback/pdi/delete",
                        "/api/feedback/preview", "/api/feedback/note/save",
                        "/api/feedback/note/acknowledge", "/api/feedback/note/delete"):
                user = self._require_auth()
                if not user:
                    return
                payload = self._read_json()
                try:
                    with closing(get_connection()) as conn:
                        if path == "/api/feedback/save":
                            resultado = save_feedback(conn, user["company_id"], user, payload)
                        elif path == "/api/feedback/publish":
                            resultado = publish_feedback(
                                conn, user["company_id"], user, int(payload.get("feedbackId") or 0))
                        elif path == "/api/feedback/detail":
                            fb = load_feedback(
                                conn, user["company_id"], int(payload.get("feedbackId") or 0), user)
                            if fb and data_scope_for_user(conn, user) == "proprio":
                                # Vendedor abre só o próprio, e só publicado.
                                if not fb["isMe"] or fb["status"] != "PUBLICADO":
                                    fb = None
                            if not fb:
                                self._set_headers(404)
                                self.wfile.write(json_dumps({"error": "Feedback não encontrado."}))
                                return
                            resultado = {"feedback": fb}
                        elif path == "/api/feedback/acknowledge":
                            resultado = acknowledge_feedback(
                                conn, user["company_id"], user,
                                int(payload.get("feedbackId") or 0),
                                payload.get("note") or "", payload.get("confidential") or "")
                        elif path == "/api/feedback/delete":
                            delete_feedback(conn, user["company_id"], user, int(payload.get("feedbackId") or 0))
                            resultado = {}
                        elif path == "/api/feedback/pdi/save":
                            resultado = save_pdi_item(conn, user["company_id"], user, payload)
                        elif path == "/api/feedback/pdi/delete":
                            delete_pdi_item(conn, user["company_id"], user, int(payload.get("pdiId") or 0))
                            resultado = {}
                        elif path == "/api/feedback/note/save":
                            resultado = save_feedback_note(conn, user["company_id"], user, payload)
                        elif path == "/api/feedback/note/acknowledge":
                            resultado = acknowledge_feedback_note(
                                conn, user["company_id"], user,
                                int(payload.get("noteId") or 0), payload.get("note") or "")
                        elif path == "/api/feedback/note/delete":
                            delete_feedback_note(conn, user["company_id"], user, int(payload.get("noteId") or 0))
                            resultado = {}
                        else:
                            # Prévia dos indicadores antes de abrir o feedback: o gerente
                            # olha os números da pessoa sem precisar criar registro.
                            if not user_can_give_feedback(conn, user):
                                raise PermissionError("Apenas gestão consulta a prévia.")
                            kind = normalize_upper(payload.get("kind")) or "VENDEDOR"
                            competence = normalize_whitespace(payload.get("competence"))
                            indicadores = safe_feedback_indicators(
                                conn, user["company_id"], kind,
                                normalize_whitespace(payload.get("personName")),
                                normalize_whitespace(payload.get("unitName")), competence)
                            resultado = {
                                "indicators": indicadores,
                                "guidance": feedback_guidance(indicadores) if kind == "VENDEDOR" else [],
                                "items": feedback_items_for_kind(kind),
                                "groups": feedback_groups_for_kind(kind),
                                "script": (_mec_content().FEEDBACK_SCRIPT_MANAGER if kind == "GERENTE"
                                           else _mec_content().FEEDBACK_SCRIPT_SELLER),
                                "pdi": list_pdi_items(
                                    conn, user["company_id"],
                                    person_key(payload.get("personName") or "")),
                                "notes": list_feedback_notes(
                                    conn, user["company_id"], user,
                                    person_key(payload.get("personName") or ""), competence),
                            }
                except PermissionError as exc:
                    self._set_headers(403)
                    self.wfile.write(json_dumps({"error": str(exc)}))
                    return
                except ValueError as exc:
                    self._set_headers(400)
                    self.wfile.write(json_dumps({"error": str(exc)}))
                    return
                self._set_headers(200)
                self.wfile.write(json_dumps({"ok": True, **resultado}))
                return
            if path == "/api/meetings/save":
                user = self._require_auth()
                if not user:
                    return
                payload = self._read_json()
                try:
                    with closing(get_connection()) as conn:
                        resultado = save_meeting(conn, user["company_id"], user, payload)
                except PermissionError as exc:
                    self._set_headers(403)
                    self.wfile.write(json_dumps({"error": str(exc)}))
                    return
                except ValueError as exc:
                    self._set_headers(400)
                    self.wfile.write(json_dumps({"error": str(exc)}))
                    return
                self._set_headers(200)
                self.wfile.write(json_dumps({"ok": True, **resultado}))
                return
            if path == "/api/meetings/publish":
                user = self._require_auth()
                if not user:
                    return
                payload = self._read_json()
                try:
                    with closing(get_connection()) as conn:
                        resultado = publish_meeting(
                            conn, user["company_id"], user, int(payload.get("meetingId") or 0)
                        )
                except PermissionError as exc:
                    self._set_headers(403)
                    self.wfile.write(json_dumps({"error": str(exc)}))
                    return
                except ValueError as exc:
                    self._set_headers(400)
                    self.wfile.write(json_dumps({"error": str(exc)}))
                    return
                self._set_headers(200)
                self.wfile.write(json_dumps({"ok": True, **resultado}))
                return
            if path == "/api/meetings/acknowledge":
                user = self._require_auth()
                if not user:
                    return
                payload = self._read_json()
                try:
                    with closing(get_connection()) as conn:
                        resultado = acknowledge_meeting(
                            conn, user["company_id"], user,
                            int(payload.get("meetingId") or 0),
                            payload.get("feedback") or "",
                        )
                except ValueError as exc:
                    self._set_headers(400)
                    self.wfile.write(json_dumps({"error": str(exc)}))
                    return
                self._set_headers(200)
                self.wfile.write(json_dumps({"ok": True, **resultado}))
                return
            if path == "/api/meetings/detail":
                user = self._require_auth()
                if not user:
                    return
                payload = self._read_json()
                meeting_id = int(payload.get("meetingId") or 0)
                with closing(get_connection()) as conn:
                    ata = load_meeting(conn, user["company_id"], meeting_id)
                    if ata:
                        # Quem é "eu" nesta lista é decidido AQUI, comparando chaves e
                        # a conta vinculada. A tela não tem como fazer isso: ela só
                        # conhece o nome do login, que quase nunca é igual ao nome do
                        # cadastro — era por isso que o botão de ciência não aparecia.
                        minhas = set(user_person_keys(user))
                        def sou_eu(p: dict[str, Any]) -> bool:
                            return p["personKey"] in minhas or p.get("userId") == user["id"]

                        ata["participants"] = [{**p, "isMe": sou_eu(p)} for p in ata["participants"]]
                        ata["iAmParticipant"] = any(p["isMe"] for p in ata["participants"])
                        ata["myAcknowledgedAt"] = next(
                            (p["acknowledgedAt"] for p in ata["participants"] if p["isMe"]), ""
                        )
                        if data_scope_for_user(conn, user) == "proprio":
                            if ata["status"] != "PUBLICADA" or not ata["iAmParticipant"]:
                                ata = None
                            else:
                                # O vendedor não enxerga o feedback dos colegas — o
                                # canal é dele com o gestor, não um mural público.
                                ata["participants"] = [
                                    {**p, "feedback": p["feedback"] if p["isMe"] else ""}
                                    for p in ata["participants"]
                                ]
                if not ata:
                    self._set_headers(404)
                    self.wfile.write(json_dumps({"error": "Ata não encontrada."}))
                    return
                self._set_headers(200)
                self.wfile.write(json_dumps({"meeting": ata}))
                return
            if path == "/api/meetings/delete":
                user = self._require_auth()
                if not user:
                    return
                payload = self._read_json()
                try:
                    with closing(get_connection()) as conn:
                        delete_meeting(conn, user["company_id"], user, int(payload.get("meetingId") or 0))
                except PermissionError as exc:
                    self._set_headers(403)
                    self.wfile.write(json_dumps({"error": str(exc)}))
                    return
                self._set_headers(200)
                self.wfile.write(json_dumps({"ok": True}))
                return
            if path == "/api/meetings/attachment/upload":
                user = self._require_auth()
                if not user:
                    return
                arquivos, campos = self._parse_multipart()
                try:
                    meeting_id = int(campos.get("meetingId") or 0)
                    salvos = []
                    with closing(get_connection()) as conn:
                        for arquivo in arquivos:
                            salvos.append(save_meeting_attachment(
                                conn, user["company_id"], user, meeting_id,
                                arquivo["fileName"], arquivo["content"],
                            ))
                except PermissionError as exc:
                    self._set_headers(403)
                    self.wfile.write(json_dumps({"error": str(exc)}))
                    return
                except ValueError as exc:
                    self._set_headers(400)
                    self.wfile.write(json_dumps({"error": str(exc)}))
                    return
                self._set_headers(200)
                self.wfile.write(json_dumps({"ok": True, "attachments": salvos}))
                return
            if path == "/api/meetings/attachment/delete":
                user = self._require_auth()
                if not user:
                    return
                payload = self._read_json()
                try:
                    with closing(get_connection()) as conn:
                        delete_meeting_attachment(
                            conn, user["company_id"], user, int(payload.get("attachmentId") or 0)
                        )
                except PermissionError as exc:
                    self._set_headers(403)
                    self.wfile.write(json_dumps({"error": str(exc)}))
                    return
                except ValueError as exc:
                    self._set_headers(400)
                    self.wfile.write(json_dumps({"error": str(exc)}))
                    return
                self._set_headers(200)
                self.wfile.write(json_dumps({"ok": True}))
                return
            if path == "/api/content/save":
                user = self._require_auth()
                if not user or not self._require_user_management(user):
                    return
                payload = self._read_json()
                try:
                    with closing(get_connection()) as conn:
                        res = upsert_content_item(conn, user["company_id"], user["id"], payload)
                        conn.commit()
                    self._set_headers(200)
                    self.wfile.write(json_dumps({
                        "message": "Conteúdo criado." if res.get("created") else "Conteúdo atualizado.",
                        **res,
                    }))
                except ValueError as exc:
                    self._set_headers(400)
                    self.wfile.write(json_dumps({"error": str(exc)}))
                except Exception as exc:
                    traceback.print_exc()
                    self._set_headers(400)
                    self.wfile.write(json_dumps({"error": str(exc)}))
                return
            if path == "/api/content/delete":
                user = self._require_auth()
                if not user or not self._require_user_management(user):
                    return
                payload = self._read_json()
                try:
                    with closing(get_connection()) as conn:
                        delete_content_item(conn, user["company_id"], user["id"], payload.get("id"))
                        conn.commit()
                    self._set_headers(200)
                    self.wfile.write(json_dumps({"message": "Conteúdo excluído."}))
                except ValueError as exc:
                    self._set_headers(400)
                    self.wfile.write(json_dumps({"error": str(exc)}))
                return
            if path == "/api/admin/profiles":
                user = self._require_auth()
                if not user or not self._require_user_management(user):
                    return
                payload = self._read_json()
                try:
                    with closing(get_connection()) as conn:
                        res = upsert_access_profile(conn, user["company_id"], user["id"], payload)
                        conn.commit()
                    self._set_headers(200)
                    self.wfile.write(json_dumps({
                        "message": "Perfil criado." if res.get("created") else "Perfil atualizado.",
                        **res,
                    }))
                except ValueError as exc:
                    self._set_headers(400)
                    self.wfile.write(json_dumps({"error": str(exc)}))
                except Exception as exc:
                    traceback.print_exc()
                    self._set_headers(400)
                    self.wfile.write(json_dumps({"error": str(exc)}))
                return
            if path == "/api/admin/profiles/delete":
                user = self._require_auth()
                if not user or not self._require_user_management(user):
                    return
                payload = self._read_json()
                try:
                    with closing(get_connection()) as conn:
                        delete_access_profile(conn, user["company_id"], user["id"], payload.get("id"))
                        conn.commit()
                    self._set_headers(200)
                    self.wfile.write(json_dumps({"message": "Perfil excluído."}))
                except ValueError as exc:
                    self._set_headers(400)
                    self.wfile.write(json_dumps({"error": str(exc)}))
                except Exception as exc:
                    traceback.print_exc()
                    self._set_headers(400)
                    self.wfile.write(json_dumps({"error": str(exc)}))
                return
            if path == "/api/admin/users":
                user = self._require_auth()
                if not user or not self._require_user_management(user):
                    return
                payload = self._read_json()
                try:
                    with closing(get_connection()) as conn:
                        res = upsert_user(conn, user["company_id"], user["id"], payload)
                        conn.commit()
                    self._set_headers(200)
                    self.wfile.write(json_dumps({"message": "Usuário criado com sucesso." if res.get("created") else "Usuário atualizado.", **res}))
                except Exception as exc:
                    traceback.print_exc()
                    self._set_headers(400)
                    self.wfile.write(json_dumps({"error": str(exc)}))
                return
            if path == "/api/admin/users/password":
                user = self._require_auth()
                if not user or not self._require_user_management(user):
                    return
                payload = self._read_json()
                try:
                    with closing(get_connection()) as conn:
                        set_user_password(conn, user["company_id"], user["id"], payload.get("id"), payload.get("password"))
                        conn.commit()
                    self._set_headers(200)
                    self.wfile.write(json_dumps({"message": "Senha atualizada com sucesso."}))
                except Exception as exc:
                    traceback.print_exc()
                    self._set_headers(400)
                    self.wfile.write(json_dumps({"error": str(exc)}))
                return
            if path == "/api/admin/users/delete":
                user = self._require_auth()
                if not user or not self._require_user_management(user):
                    return
                payload = self._read_json()
                try:
                    with closing(get_connection()) as conn:
                        delete_user_record(conn, user["company_id"], user["id"], payload.get("id"))
                        conn.commit()
                    self._set_headers(200)
                    self.wfile.write(json_dumps({"message": "Usuário excluído."}))
                except Exception as exc:
                    traceback.print_exc()
                    self._set_headers(400)
                    self.wfile.write(json_dumps({"error": str(exc)}))
                return
            if path in {"/api/crm/coverages/save", "/api/crm/coverages/delete"}:
                user = self._require_auth()
                if not user:
                    return
                payload = self._read_json()
                try:
                    with closing(get_connection()) as conn:
                        if data_scope_for_user(conn, user) == "proprio":
                            self._set_headers(403)
                            self.wfile.write(json_dumps(
                                {"error": "Só a gestão define cobertura de carteira."}))
                            return
                        if path.endswith("/delete"):
                            res = delete_coverage(conn, user["company_id"], user["id"],
                                                  payload.get("id"))
                        else:
                            res = save_coverage(conn, user["company_id"], user["id"], payload)
                    self._set_headers(200)
                    self.wfile.write(json_dumps(res))
                except Exception as exc:
                    traceback.print_exc()
                    self._set_headers(400)
                    self.wfile.write(json_dumps({"error": str(exc)}))
                return
            if path == "/api/crm/clients/alias":
                user = self._require_auth()
                if not user:
                    return
                payload = self._read_json()
                try:
                    with closing(get_connection()) as conn:
                        if data_scope_for_user(conn, user) == "proprio":
                            self._set_headers(403)
                            self.wfile.write(json_dumps({"error": "Conciliação é da gestão."}))
                            return
                        res = save_client_alias(conn, user["company_id"], user["id"], payload)
                    self._set_headers(200)
                    self.wfile.write(json_dumps(res))
                except Exception as exc:
                    traceback.print_exc()
                    self._set_headers(400)
                    self.wfile.write(json_dumps({"error": str(exc)}))
                return
            if path == "/api/admin/issues/cities/bulk":
                user = self._require_auth()
                if not user or not self._require_admin_area(user):
                    return
                payload = self._read_json()
                try:
                    with closing(get_connection()) as conn:
                        res = resolve_city_issues_bulk(conn, user["company_id"], user["id"], payload)
                    self._set_headers(200)
                    self.wfile.write(json_dumps(res))
                except Exception as exc:
                    traceback.print_exc()
                    self._set_headers(400)
                    self.wfile.write(json_dumps({"error": str(exc)}))
                return
            if path in {"/api/admin/territories/save", "/api/admin/territories/delete"}:
                user = self._require_auth()
                if not user or not self._require_admin_area(user):
                    return
                payload = self._read_json()
                try:
                    with closing(get_connection()) as conn:
                        if path.endswith("/save"):
                            res = save_territory_mapping(conn, user["company_id"], user["id"], payload)
                        else:
                            res = delete_territory_mapping(conn, user["company_id"], user["id"],
                                                           payload.get("id"))
                    self._set_headers(200)
                    self.wfile.write(json_dumps(res))
                except Exception as exc:
                    traceback.print_exc()
                    self._set_headers(400)
                    self.wfile.write(json_dumps({"error": str(exc)}))
                return
            if path == "/api/admin/people/search":
                user = self._require_auth()
                if not user:
                    return
                payload = self._read_json()
                with closing(get_connection()) as conn:
                    if data_scope_for_user(conn, user) == "proprio":
                        self._set_headers(403)
                        self.wfile.write(json_dumps({"error": "Perfil sem acesso."}))
                        return
                    itens = search_person_candidates(
                        conn, user["company_id"], normalize_whitespace(payload.get("q")))
                self._set_headers(200)
                self.wfile.write(json_dumps({"candidates": itens}))
                return
            if path == "/api/admin/people/terminate":
                user = self._require_auth()
                if not user:
                    return
                payload = self._read_json()
                try:
                    with closing(get_connection()) as conn:
                        if not user_can_manage_users(conn, user):
                            self._set_headers(403)
                            self.wfile.write(json_dumps({"error": "Apenas a diretoria desliga pessoas."}))
                            return
                        resultado = terminate_person(conn, user["company_id"], user["id"], payload)
                except ValueError as exc:
                    self._set_headers(400)
                    self.wfile.write(json_dumps({"error": str(exc)}))
                    return
                self._set_headers(200)
                self.wfile.write(json_dumps({"ok": True, **resultado}))
                return
            if path == "/api/admin/people/save":
                user = self._require_auth()
                if not user:
                    return
                payload = self._read_json()
                try:
                    with closing(get_connection()) as conn:
                        if data_scope_for_user(conn, user) == "proprio":
                            self._set_headers(403)
                            self.wfile.write(json_dumps({"error": "Perfil sem acesso."}))
                            return
                        resultado = save_person_record(conn, user["company_id"], user["id"], payload)
                except ValueError as exc:
                    self._set_headers(400)
                    self.wfile.write(json_dumps({"error": str(exc)}))
                    return
                self._set_headers(200)
                self.wfile.write(json_dumps({"ok": True, **resultado}))
                return
            if path == "/api/admin/people/update-unit":
                user = self._require_auth()
                if not user or not self._require_admin_area(user):
                    return
                payload = self._read_json()
                try:
                    with closing(get_connection()) as conn:
                        unit = update_person_unit(conn, user["company_id"], user["id"], payload.get("person_name"), payload.get("base_unit"))
                        conn.commit()
                    self._set_headers(200)
                    self.wfile.write(json_dumps({"message": f"Vendedor '{normalize_whitespace(payload.get('person_name'))}' ajustado para {unit}."}))
                except Exception as exc:
                    traceback.print_exc()
                    self._set_headers(400)
                    self.wfile.write(json_dumps({"error": str(exc)}))
                return
            if path == "/api/admin/city/update-unit":
                user = self._require_auth()
                if not user or not self._require_admin_area(user):
                    return
                payload = self._read_json()
                try:
                    with closing(get_connection()) as conn:
                        unit = update_city_unit(conn, user["company_id"], user["id"], payload.get("city_name"), payload.get("principal_unit"))
                        conn.commit()
                    self._set_headers(200)
                    self.wfile.write(json_dumps({"message": f"Cidade '{normalize_upper(payload.get('city_name'))}' ajustada para {unit}."}))
                except Exception as exc:
                    traceback.print_exc()
                    self._set_headers(400)
                    self.wfile.write(json_dumps({"error": str(exc)}))
                return
            if path.startswith("/api/admin/import/"):
                user = self._require_auth()
                if not user or not self._require_admin_area(user):
                    return
                import_type = path.split("/api/admin/import/")[-1]
                if import_type == "cidade-unidade":
                    files_payload, _ = self._parse_multipart()
                    file_entry = next((f for f in files_payload if f.get("content")), None)
                    if not file_entry:
                        self._set_headers(400)
                        self.wfile.write(json_dumps({"error": "Nenhum arquivo encontrado na requisição."}))
                        return
                    try:
                        with closing(get_connection()) as conn:
                            res = import_city_mappings_csv(conn, user["company_id"], user["id"], file_entry["content"])
                            audit_log(conn, user["company_id"], user["id"], "importar", "city_mappings", import_type, res)
                            conn.commit()
                        self._set_headers(200)
                        self.wfile.write(json_dumps({"message": f"{res['updated']} cidade(s) atualizada(s); {res['resolved']} pendência(s) de cidade resolvida(s)."}))
                    except Exception as exc:
                        traceback.print_exc()
                        self._set_headers(500)
                        self.wfile.write(json_dumps({"error": str(exc)}))
                    return
                TABLE_MAP = {
                    "people":       "people_records",
                    "vacations":    "vacations",
                    "holidays":     "holidays",
                    "goals-seller": "goals_seller",
                    "goals-unit":   "goals_unit",
                    "clients":      "client_registry",
                }
                table_name = TABLE_MAP.get(import_type)
                if not table_name:
                    self._set_headers(400)
                    self.wfile.write(json_dumps({"error": f"Tipo de importação desconhecido: {import_type}"}))
                    return
                files_payload, _ = self._parse_multipart()
                file_entry = next((f for f in files_payload if f.get("content")), None)
                if not file_entry:
                    self._set_headers(400)
                    self.wfile.write(json_dumps({"error": "Nenhum arquivo encontrado na requisição."}))
                    return
                try:
                    with closing(get_connection()) as conn:
                        total = import_admin_csv(conn, user["company_id"], user["id"], table_name, file_entry["content"])
                        audit_log(conn, user["company_id"], user["id"], "importar", table_name, import_type, {"rows": total})
                        conn.commit()
                    self._set_headers(200)
                    self.wfile.write(json_dumps({"message": f"{total} registro(s) importado(s) com sucesso."}))
                except Exception as exc:
                    traceback.print_exc()
                    self._set_headers(500)
                    self.wfile.write(json_dumps({"error": str(exc)}))
                return
            self._set_headers(404)
            self.wfile.write(json_dumps({"error": "rota nao encontrada"}))
        except Exception as exc:
            traceback.print_exc()
            self._set_headers(500)
            self.wfile.write(json_dumps({"error": f"Erro interno: {exc}"}))


_AUTO_IMPORT_LOCK = threading.Lock()
_AUTO_IMPORT_RUNNING = threading.Event()
# Avisos da última execução (ex.: base de clientes encolheu) — lidos pelo botão manual
_AUTO_IMPORT_WARNINGS: list[str] = []


def _auto_import_get_admin_user(conn: sqlite3.Connection) -> sqlite3.Row | None:
    """Retorna o primeiro usuário Admin/Diretor da empresa para usar nos imports automáticos."""
    return conn.execute(
        "SELECT u.id, u.company_id FROM users u WHERE u.role IN ('Administrador','Admin','Diretor') ORDER BY u.id LIMIT 1"
    ).fetchone()


def _auto_import_extract_competence(filename: str) -> str | None:
    """Extrai YYYY-MM do nome do arquivo. Ex: faturamento_2025-06.csv → 2025-06."""
    m = re.search(r"(\d{4}[-/]\d{2})|\b(\d{2}[-/]\d{4})\b", filename)
    if not m:
        return None
    raw = m.group(1) or m.group(2)
    parts = re.split(r"[-/]", raw)
    if len(parts[0]) == 4:
        return f"{parts[0]}-{parts[1]}"   # YYYY-MM
    return f"{parts[1]}-{parts[0]}"       # MM/YYYY → YYYY-MM


def _auto_import_log(conn: sqlite3.Connection, folder: str, scope: str,
                     competence: str | None, status: str, message: str,
                     files: list[str]) -> None:
    conn.execute(
        "INSERT INTO auto_import_log (ran_at, folder, scope, competence, status, message, files_json) VALUES (?,?,?,?,?,?,?)",
        (now_iso(), folder, scope, competence, status, message, json.dumps(files)),
    )
    conn.commit()


def _auto_import_header_cols(content: bytes) -> list[str]:
    """Retorna os nomes das colunas (primeira linha) de um CSV."""
    try:
        first_line = content.decode("latin-1", errors="replace").splitlines()[0]
        sep = ";" if ";" in first_line else ","
        return [c.strip().strip('"') for c in first_line.split(sep)]
    except Exception:
        return []


def _auto_import_detect_cost_field(content: bytes) -> str:
    """Detecta se um CSV de custo é de unidade ou vendedor pela primeira coluna."""
    cols = _auto_import_header_cols(content)
    first_col = cols[0].upper() if cols else ""
    if first_col in {"VENDEDOR", "VENDEDOR CONSOLIDADO"}:
        return "import-cost-vendor-file"
    return "import-cost-unit-file"


def _auto_import_detect_crm_field(content: bytes, filename: str = "") -> str:
    """Detecta se um CSV de CRM é cadastro de clientes ou faturamento consolidado."""
    cols_raw = _auto_import_header_cols(content)
    cols_upper = {c.strip().upper() for c in cols_raw}
    # Cadastro de clientes: colunas típicas
    cadastro_markers = {"RAZAO SOCIAL/NOME", "RAZAO SOCIAL", "NOME FANTASIA", "CPF/CNPJ", "CODIGO CLIENTE"}
    if cadastro_markers & cols_upper:
        return "import-crm-clients-file"
    # Faturamento consolidado: colunas típicas
    consolidado_markers = {"ULT.COMPRA", "ULTIMA COMPRA", "FATURAMENTO", "VALOR LIQUIDO"}
    if consolidado_markers & cols_upper:
        return "import-crm-summary-file"
    # Fallback por nome de arquivo
    stem = Path(filename).stem.lower()
    if any(k in stem for k in ("pessoa", "cadastro", "cliente", "crm", "perfil")):
        return "import-crm-clients-file"
    return "import-crm-summary-file"


def _auto_import_build_payload(files: list[Path], scope: str) -> list[dict[str, Any]]:
    """
    Monta o files_payload como lista de dicts com fieldName correto.
    Escopos com pasta dedicada (crm_clients / crm_summary) dispensam detecção por
    conteúdo — a pasta já define o tipo, evitando classificação errada.
    """
    payload = []
    for f in files:
        content = f.read_bytes()
        if scope == "crm_clients":
            field_name = "import-crm-clients-file"
        elif scope == "crm_summary":
            field_name = "import-crm-summary-file"
        elif scope == "warranty":
            field_name = "import-warranty-file"
        elif scope == "cost":
            field_name = _auto_import_detect_cost_field(content)
        elif scope == "crm":
            field_name = _auto_import_detect_crm_field(content, f.name)
        else:
            field_name = "import-sales-file"  # faturamento: override garante tipo faturamento_detalhado
        payload.append({"fieldName": field_name, "fileName": f.name, "content": content})
    return payload


def auto_import_tick() -> None:
    """Verifica as pastas de auto-import e processa CSVs pendentes."""
    if not _AUTO_IMPORT_LOCK.acquire(blocking=False):
        print("[auto-import] já em execução, ignorando chamada concorrente")
        return
    _AUTO_IMPORT_RUNNING.set()
    try:
        _auto_import_tick_inner()
    finally:
        _AUTO_IMPORT_RUNNING.clear()
        _AUTO_IMPORT_LOCK.release()


def _auto_import_tick_inner() -> None:
    imported_any = False
    for cfg in AUTO_IMPORT_FOLDERS:
        folder_path = AUTO_IMPORT_BASE / cfg["folder"]
        if not folder_path.exists():
            folder_path.mkdir(parents=True, exist_ok=True)
            continue

        csv_files = sorted(folder_path.glob("*.csv"))
        if not csv_files:
            continue

        scope = cfg["scope"]

        # Cadastro de clientes: base mestre do Alfa, SEM competência.
        # O Alfa não exporta a base inteira em um arquivo — são DUAS (ou mais)
        # exportações COMPLEMENTARES. Portanto TODOS os CSVs da pasta entram em
        # UM ÚNICO import: o delete da base acontece uma vez e as linhas de todos
        # os arquivos são inseridas juntas (upsert por client_code).
        # Se cada arquivo virasse um import próprio, o segundo apagaria o primeiro.
        if scope == "crm_clients":
            # Só processa quando os arquivos estão "estáveis" (nenhuma escrita recente),
            # evitando importar meia base enquanto a segunda exportação ainda é copiada.
            _now_ts = time.time()
            _unstable = [f for f in csv_files if (_now_ts - f.stat().st_mtime) < CRM_CLIENTS_SETTLE_SECONDS]
            if _unstable:
                print(f"[auto-import] crm/clientes: {len(_unstable)} arquivo(s) ainda sendo gravado(s), "
                      f"aguardando {CRM_CLIENTS_SETTLE_SECONDS}s de estabilidade")
                continue
            by_competence: dict[str, list[Path]] = {"_clients__all": list(csv_files)}
        else:
            # Agrupa arquivos por competência extraída do nome
            by_competence = {}
            no_comp: list[Path] = []
            for f in csv_files:
                comp = _auto_import_extract_competence(f.name)
                if comp:
                    by_competence.setdefault(comp, []).append(f)
                else:
                    no_comp.append(f)

            # Faturamento: competência lida do conteúdo — aceita sem data no nome.
            # IMPORTANTE: cada arquivo vai em seu próprio grupo (_from_content__N).
            # Agrupar vários arquivos sob uma única chave fazia a competência de UM
            # arquivo ser aplicada a TODOS, gravando dados no mês errado e apagando
            # (import "substituir") o mês correto.
            if no_comp:
                # sales e warranty derivam a competência da data de cada linha,
                # então dispensam o mês no nome do arquivo.
                if scope in {"sales", "warranty"}:
                    for _idx, _f in enumerate(no_comp):
                        by_competence[f"_from_content__{_idx}"] = [_f]
                else:
                    _exemplo = "2026-07" if scope == "crm_summary" else "2026-06"
                    with closing(get_connection()) as conn:
                        for f in no_comp:
                            _auto_import_log(
                                conn, cfg["folder"], scope, None, "erro",
                                f"Competência não encontrada no nome do arquivo '{f.name}'. "
                                f"Inclua o mês no nome, ex: {_exemplo}_{f.name}",
                                [f.name],
                            )
                            print(f"[auto-import] REJEITADO {f.name}: sem competência no nome")

        required_kinds = IMPORT_SCOPE_REQUIREMENTS[scope]

        for competence_key, files in by_competence.items():
            # Monta payload com fieldName correto por escopo/conteúdo
            files_payload = _auto_import_build_payload(files, scope)
            try:
                preview = preview_import_package(files_payload, scope)
            except Exception as exc:
                with closing(get_connection()) as conn:
                    _auto_import_log(conn, cfg["folder"], scope, competence_key,
                                     "erro", f"Erro ao analisar arquivos: {exc}",
                                     [f.name for f in files])
                continue

            # Para faturamento sem data no nome, usa a competência sugerida pelo conteúdo
            if competence_key.startswith("_from_content"):
                suggested = preview.get("suggestedCompetence")
                if not suggested:
                    with closing(get_connection()) as conn:
                        _auto_import_log(conn, cfg["folder"], scope, None, "erro",
                                         "Não foi possível determinar a competência pelo conteúdo do arquivo.",
                                         [f.name for f in files])
                    continue
                competence = suggested
            elif competence_key.startswith("_clients__"):
                # Cadastro de clientes não tem competência; usa o mês atual como rótulo
                competence = date.today().strftime("%Y-%m")
            else:
                competence = competence_key

            detected_kinds = {fi["fileType"] for fi in preview.get("files", []) if fi.get("fileType")}
            if not detected_kinds:
                print(f"[auto-import] {cfg['folder']}/{competence}: nenhum arquivo reconhecido, aguardando")
                continue
            # Escopo "crm" (legado, pasta única) aceita import parcial.
            # Escopos individuais e demais: exige os tipos necessários.
            if scope != "crm":
                missing = required_kinds - detected_kinds
                if missing:
                    print(f"[auto-import] {cfg['folder']}/{competence}: aguardando {missing}")
                    continue

            # Executa o import — pasta de destino espelha a origem para rastreabilidade
            _dest_sub = cfg["folder"].replace("/", "-")
            dest_ok  = AUTO_IMPORT_BASE / "processados" / _dest_sub / competence
            dest_err = AUTO_IMPORT_BASE / "erros" / _dest_sub / competence
            try:
                with closing(get_connection()) as conn:
                    user = _auto_import_get_admin_user(conn)
                    if not user:
                        _auto_import_log(conn, cfg["folder"], scope, competence,
                                         "erro", "Nenhum usuário Admin/Diretor encontrado.", [f.name for f in files])
                        continue
                    result = import_package(
                        conn, user["company_id"], user["id"],
                        competence, "substituir", scope, preview, files_payload,
                    )
                    invalidate_crm_cache(user["company_id"])
                    imported_any = True
                    msg = result.get("message", "OK")
                    _warning = result.get("warning")
                    _auto_import_log(conn, cfg["folder"], scope, competence,
                                     "alerta" if _warning else "sucesso",
                                     f"{msg} — {_warning}" if _warning else msg,
                                     [f.name for f in files])
                    print(f"[auto-import] {cfg['folder']}/{competence}: {msg}")
                    if _warning:
                        print(f"[auto-import] ⚠ {_warning}")
                        _AUTO_IMPORT_WARNINGS.append(_warning)

                dest_ok.mkdir(parents=True, exist_ok=True)
                for f in files:
                    _target = dest_ok / f.name
                    if _target.exists():
                        _stamp = datetime.now().strftime("%Y%m%d-%H%M%S")
                        _target = dest_ok / f"{f.stem}__{_stamp}{f.suffix}"
                    shutil.move(str(f), str(_target))

            except Exception as exc:
                traceback.print_exc()
                with closing(get_connection()) as conn:
                    _auto_import_log(conn, cfg["folder"], scope, competence,
                                     "erro", str(exc), [f.name for f in files])
                dest_err.mkdir(parents=True, exist_ok=True)
                for f in files:
                    try:
                        _target = dest_err / f.name
                        if _target.exists():
                            _stamp = datetime.now().strftime("%Y%m%d-%H%M%S")
                            _target = dest_err / f"{f.stem}__{_stamp}{f.suffix}"
                        shutil.move(str(f), str(_target))
                    except Exception:
                        pass

    # Importou algo: recalcula o dashboard de todas as competências para que a
    # primeira visualização já venha pronta.
    if imported_any:
        warm_dashboard_cache()


def _auto_import_loop() -> None:
    """Thread de background que roda auto_import_tick periodicamente."""
    time.sleep(10)  # Aguarda o servidor inicializar
    while True:
        try:
            auto_import_tick()
        except Exception:
            traceback.print_exc()
        time.sleep(AUTO_IMPORT_INTERVAL)


def _migrate_legacy_crm_folder() -> None:
    """Move CSVs soltos na antiga pasta crm/ para as novas subpastas dedicadas.

    A pasta crm/ passou a ter subpastas (clientes / faturamento-consolidado). Arquivos
    que ficaram na raiz são classificados por conteúdo e movidos uma única vez.
    """
    legacy = AUTO_IMPORT_BASE / "crm"
    if not legacy.exists():
        return
    for f in sorted(legacy.glob("*.csv")):
        try:
            field = _auto_import_detect_crm_field(f.read_bytes(), f.name)
            sub = "clientes" if field == "import-crm-clients-file" else "faturamento-consolidado"
            dest_dir = legacy / sub
            dest_dir.mkdir(parents=True, exist_ok=True)
            shutil.move(str(f), str(dest_dir / f.name))
            print(f"[auto-import] migrado {f.name} -> crm/{sub}/")
        except Exception:
            traceback.print_exc()


def run_server(host: str = DEFAULT_HOST, port: int = DEFAULT_PORT) -> None:
    # Cria pastas de auto-import se não existirem
    for cfg in AUTO_IMPORT_FOLDERS:
        (AUTO_IMPORT_BASE / cfg["folder"]).mkdir(parents=True, exist_ok=True)
    (AUTO_IMPORT_BASE / "processados").mkdir(parents=True, exist_ok=True)
    (AUTO_IMPORT_BASE / "erros").mkdir(parents=True, exist_ok=True)
    _migrate_legacy_crm_folder()

    t = threading.Thread(target=_auto_import_loop, daemon=True, name="auto-import")
    t.start()

    # Pré-aquece o cache do dashboard em segundo plano — o servidor já sobe atendendo
    threading.Thread(target=warm_dashboard_cache, daemon=True, name="dashboard-warmup").start()

    server = ThreadingHTTPServer((host, port), AppHandler)
    print(f"Servidor rodando em http://{host}:{port}")
    server.serve_forever()

if __name__ == "__main__":
    init_db()
    run_server()
